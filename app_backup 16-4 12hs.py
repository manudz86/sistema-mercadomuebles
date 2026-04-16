"""
SISTEMA INVENTARIO CANNON - VERSIÓN SIMPLIFICADA
Todo el código en un solo archivo para evitar problemas de imports en Windows
"""

import os
import json
import requests  # pip install requests
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from dotenv import load_dotenv
import pymysql
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from utils import log_evento, crear_tabla_sistema_logs

import threading

_ml_rate_lock = threading.Lock()
_ml_last_request = 0.0
_ML_MIN_INTERVAL = 0.7  # ~1.4 requests/segundo máximo

ML_SELLER_ID = 29563319

# Cargar configuración
load_dotenv('config/.env')

# Configurar Flask
app = Flask(__name__, template_folder='templates')
app.secret_key = os.getenv('SECRET_KEY', '')

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="memory://",
)

# Filtro personalizado para dashboard visual
@app.template_filter('zero_dash')
def zero_dash(value):
    """Convierte 0 en '-' para el dashboard visual"""
    return '-' if value == 0 or value is None else value

# ============================================================================
# FLASK-LOGIN SETUP
# ============================================================================
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '⚠️ Debés iniciar sesión para acceder'
login_manager.login_message_category = 'warning'

@app.context_processor
def inject_alertas_pendientes():
    """Inyecta el contador de alertas en todos los templates"""
    try:
        from flask_login import current_user
        if current_user.is_authenticated:
            # Contar alertas que tienen al menos una fila visible en el template
            # (tipo_procesado != 'ambos' significa que alguna acción queda pendiente)
            result = query_one("""
                SELECT COUNT(*) as total FROM alertas_stock a
                WHERE a.estado = 'pendiente'
                AND (a.tipo_procesado IS NULL OR a.tipo_procesado NOT IN ('ambos'))
                AND EXISTS (
                    SELECT 1 FROM sku_mla_mapeo m 
                    WHERE (m.sku = a.sku OR m.sku = CONCAT(a.sku, 'Z'))
                    AND m.activo = TRUE
                )
            """)
            return {'alertas_pendientes_count': result['total'] if result else 0, 'now': datetime.now()}
    except:
        pass
    return {'alertas_pendientes_count': 0, 'now': datetime.now()}

class User(UserMixin):
    def __init__(self, id, username, rol, activo):
        self.id = id
        self.username = username
        self.rol = rol
        self.activo = activo

@login_manager.user_loader
def load_user(user_id):
    row = query_one('SELECT * FROM usuarios WHERE id = %s AND activo = TRUE', (user_id,))
    if row:
        return User(row['id'], row['username'], row['rol'], row['activo'])
    return None

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.rol != 'admin':
            flash('❌ No tenés permisos para acceder a esta sección', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def vendedor_required(f):
    """Admin y vendedor pueden acceder. Solo viewer no puede."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.rol not in ('admin', 'vendedor'):
            flash('❌ No tenés permisos para acceder a esta sección', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def agencia_only(f):
    """Solo agencia puede acceder — redirige a /agencia si intenta acceder a otra ruta."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if current_user.rol == 'agencia':
            return redirect(url_for('agencia_dashboard'))
        return f(*args, **kwargs)
    return decorated

# Configuración de base de datos
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'port': int(os.getenv('DB_PORT', 3306)),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', ''),
    'database': os.getenv('DB_NAME', 'inventario_cannon'),
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

# ============================================================================
# FUNCIONES DE BASE DE DATOS
# ============================================================================

def get_db_connection():
    """Crear conexión a la base de datos"""
    return pymysql.connect(**DB_CONFIG)

def query_db(query, params=None):
    """Ejecutar query SELECT y retornar resultados"""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(query, params or ())
            return cursor.fetchall()
    finally:
        conn.close()

def query_one(query, params=None):
    """Ejecutar query SELECT y retornar un solo resultado"""
    results = query_db(query, params)
    return results[0] if results else None

def execute_db(query, params=None):
    """Ejecutar query INSERT/UPDATE/DELETE"""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(query, params or ())
            conn.commit()
            return cursor.lastrowid
    finally:
        conn.close()

# ============================================================================
# RUTAS - PÁGINAS
# ============================================================================

# ============================================================================
# RUTAS DE LOGIN / LOGOUT
# ============================================================================
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute; 30 per hour")
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user_row = query_one('SELECT * FROM usuarios WHERE username = %s AND activo = TRUE', (username,))
        if user_row and check_password_hash(user_row['password_hash'], password):
            user = User(user_row['id'], user_row['username'], user_row['rol'], user_row['activo'])
            login_user(user, remember=True)
            if user_row['rol'] == 'agencia':
                return redirect(url_for('agencia_dashboard'))
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        flash('❌ Usuario o contraseña incorrectos', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('✅ Sesión cerrada', 'success')
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    """Dashboard principal"""
    stats = {'ventas_activas': 0, 'ventas_en_proceso': 0, 'alertas_pendientes': 0}
    stock_critico = []
    
    try:
        # Contar ventas activas
        result = query_one("SELECT COUNT(*) as total FROM ventas WHERE estado_entrega = 'pendiente'")
        stats['ventas_activas'] = result['total'] if result else 0
        
        # Contar ventas en proceso
        result = query_one("SELECT COUNT(*) as total FROM ventas WHERE estado_entrega = 'en_proceso'")
        stats['ventas_en_proceso'] = result['total'] if result else 0
        
        # Contar alertas
        result = query_one("SELECT COUNT(*) as total FROM alertas_stock WHERE estado = 'pendiente'")
        stats['alertas_pendientes'] = result['total'] if result else 0
        
        # Stock crítico
        stock_critico = query_db("SELECT * FROM stock_disponible_ml WHERE estado_stock = 'SIN_STOCK' LIMIT 10")
        
    except Exception as e:
        flash(f'Error al cargar dashboard: {str(e)}', 'error')
    
    return render_template('dashboard.html', stats=stats, stock_critico=stock_critico, now=datetime.now)




# ============================================================================
# FUNCIONES AUXILIARES: VERIFICACIÓN DE STOCK
# Agregar ANTES de las rutas de /ventas/activas en app.py
# ============================================================================

def verificar_stock_disponible(cursor, items, ubicacion_despacho):
    """
    Verifica si hay stock suficiente para todos los items de una venta.
    Retorna (True, []) si hay stock suficiente
    Retorna (False, [lista de errores]) si falta stock
    """
    errores = []
    
    for item in items:
        sku = item['sku']
        cantidad = item['cantidad']
        
        # Verificar si es un combo
        cursor.execute('SELECT id FROM productos_compuestos WHERE sku = %s', (sku,))
        combo = cursor.fetchone()
        
        if combo:
            # Es combo: verificar componentes
            cursor.execute('''
                SELECT pb.sku, pb.nombre, pb.tipo, c.cantidad_necesaria
                FROM componentes c
                JOIN productos_base pb ON c.producto_base_id = pb.id
                WHERE c.producto_compuesto_id = %s
            ''', (combo['id'],))
            componentes = cursor.fetchall()
            
            for comp in componentes:
                sku_comp = comp['sku']
                nombre_comp = comp['nombre']
                cant_necesaria = comp['cantidad_necesaria'] * cantidad
                tipo_comp = comp['tipo']
                
                # Verificar stock del componente
                stock_disponible = obtener_stock_disponible(cursor, sku_comp, tipo_comp, ubicacion_despacho)
                
                if stock_disponible < cant_necesaria:
                    errores.append(f"{nombre_comp} (SKU: {sku_comp}): Necesitas {cant_necesaria}, disponible {stock_disponible}")
        
        else:
            # Es producto simple
            cursor.execute('SELECT nombre, tipo FROM productos_base WHERE sku = %s', (sku,))
            prod = cursor.fetchone()
            
            if not prod:
                errores.append(f"Producto {sku} no encontrado en base de datos")
                continue
            
            nombre = prod['nombre']
            tipo = prod['tipo']
            
            # Verificar stock
            stock_disponible = obtener_stock_disponible(cursor, sku, tipo, ubicacion_despacho)
            
            if stock_disponible < cantidad:
                errores.append(f"{nombre} (SKU: {sku}): Necesitas {cantidad}, disponible {stock_disponible}")
    
    if errores:
        return False, errores
    else:
        return True, []


def obtener_stock_disponible(cursor, sku, tipo, ubicacion_despacho):
    """
    Obtiene el stock disponible de un producto según su tipo y ubicación.
    """
    # COMPAC: tiene _DEP y _FULL
    if '_DEP' in sku or '_FULL' in sku:
        if ubicacion_despacho == 'FULL':
            sku_real = sku.replace('_DEP', '_FULL')
        else:
            sku_real = sku.replace('_FULL', '_DEP')
        
        cursor.execute('SELECT stock_actual FROM productos_base WHERE sku = %s', (sku_real,))
        prod = cursor.fetchone()
        return prod['stock_actual'] if prod else 0
    
    # ALMOHADAS: tienen stock_actual (DEP) y stock_full (FULL)
    elif tipo == 'almohada':
        cursor.execute('SELECT stock_actual, stock_full FROM productos_base WHERE sku = %s', (sku,))
        prod = cursor.fetchone()
        
        if not prod:
            return 0
        
        if ubicacion_despacho == 'FULL':
            return prod['stock_full']
        else:
            return prod['stock_actual']
    
    # BASES CHICAS (80200, 90200, 100200): usar stock directo
    elif tipo == 'base' and any(x in sku for x in ['80200', '90200', '100200']):
        cursor.execute('SELECT stock_actual FROM productos_base WHERE sku = %s', (sku,))
        prod = cursor.fetchone()
        return prod['stock_actual'] if prod else 0
    
    # BASES GRANDES (160, 180, 200): dividir stock de bases chicas entre 2
    elif tipo == 'base' and any(x in sku for x in ['160', '180', '200']):
        # Determinar SKU de bases chicas
        if '160' in sku:
            sku_chica = sku.replace('160', '80200')
        elif '180' in sku:
            sku_chica = sku.replace('180', '90200')
        elif '200' in sku:
            sku_chica = sku.replace('200', '100200')
        else:
            return 0
        
        cursor.execute('SELECT stock_actual FROM productos_base WHERE sku = %s', (sku_chica,))
        prod = cursor.fetchone()
        
        if not prod:
            return 0
        
        # Una base grande = 2 bases chicas
        # Stock disponible de bases grandes = stock_chicas / 2
        return prod['stock_actual'] // 2
    
    # OTROS: stock_actual normal
    else:
        cursor.execute('SELECT stock_actual FROM productos_base WHERE sku = %s', (sku,))
        prod = cursor.fetchone()
        return prod['stock_actual'] if prod else 0



# Porcentajes ML por defecto
PORCENTAJES_ML_DEFAULT = {
    'cuota_simple':  5.00,
    'cuotas_3':      9.40,
    'cuotas_6':     15.10,
    'cuotas_9':     20.70,
    'cuotas_12':    25.90,
}

@app.route('/configuracion/porcentajes-ml', methods=['GET'])
@login_required
def get_porcentajes_ml():
    """Retorna los porcentajes de ML guardados en DB"""
    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'porcentajes_ml'")
        if row:
            return jsonify({'ok': True, 'porcentajes': json.loads(row['valor'])})
        return jsonify({'ok': True, 'porcentajes': PORCENTAJES_ML_DEFAULT})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/configuracion/porcentajes-ml', methods=['POST'])
@login_required
def guardar_porcentajes_ml():
    """Guarda los porcentajes de ML en DB"""
    try:
        data = request.get_json()
        porcentajes = data.get('porcentajes', {})
        valor = json.dumps(porcentajes)
        execute_db(
            "INSERT INTO configuracion (clave, valor) VALUES ('porcentajes_ml', %s) "
            "ON DUPLICATE KEY UPDATE valor = %s, actualizado_at = NOW()",
            (valor, valor)
        )
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# ============================================================================
# FUNCIÓN CORREGIDA: EXCLUIR VENTA ACTUAL
# Reemplazar detectar_alertas_stock_bajo() en app.py
# ============================================================================

def detectar_alertas_stock_bajo(cursor, items_vendidos, venta_id_actual=None):
    """
    Detecta productos con stock disponible <= 0.
    
    Args:
        cursor: Cursor de BD
        items_vendidos: Lista de items de la venta actual
        venta_id_actual: ID de la venta actual para excluirla del cálculo
    
    Returns:
        Lista de productos sin stock
    """
    productos_sin_stock = []
    
    try:
        print("\n" + "="*60)
        print("🔍 DEBUG - DETECTAR ALERTAS")
        print("="*60)
        print(f"Items vendidos: {items_vendidos}")
        print(f"Venta ID actual (excluir): {venta_id_actual}")
        
        # ============================================
        # 1. SKUs A VERIFICAR
        # ============================================
        skus_a_verificar = set()
        cantidades_venta_actual = {}
        
        for item in items_vendidos:
            sku = item['sku']
            cantidad = item['cantidad']
            
            print(f"\n📦 Procesando item: {sku} x{cantidad}")
            
            cursor.execute('SELECT id FROM productos_compuestos WHERE sku = %s', (sku,))
            es_combo = cursor.fetchone()
            
            if es_combo:
                print(f"  → Es COMBO")
                cursor.execute('''
                    SELECT pb.sku, c.cantidad_necesaria
                    FROM componentes c
                    JOIN productos_base pb ON c.producto_base_id = pb.id
                    WHERE c.producto_compuesto_id = %s
                ''', (es_combo['id'],))
                
                componentes = cursor.fetchall()
                for comp in componentes:
                    comp_sku = comp['sku']
                    comp_cant = comp['cantidad_necesaria'] * cantidad
                    skus_a_verificar.add(comp_sku)
                    cantidades_venta_actual[comp_sku] = cantidades_venta_actual.get(comp_sku, 0) + comp_cant
            else:
                print(f"  → Es PRODUCTO BASE")
                skus_a_verificar.add(sku)
                cantidades_venta_actual[sku] = cantidades_venta_actual.get(sku, 0) + cantidad
        
        print(f"\n📋 SKUs a verificar: {skus_a_verificar}")
        print(f"📋 Cantidades venta actual: {cantidades_venta_actual}")
        
        # ============================================
        # 2. VENTAS ACTIVAS (EXCLUYENDO LA ACTUAL)
        # ============================================
        if venta_id_actual:
            cursor.execute('''
                SELECT 
                    COALESCE(pb_comp.sku, iv.sku) as sku,
                    SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
                FROM items_venta iv
                JOIN ventas v ON iv.venta_id = v.id
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
                LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
                WHERE v.estado_entrega = 'pendiente'
                AND v.id != %s
                GROUP BY sku
            ''', (venta_id_actual,))
        else:
            cursor.execute('''
                SELECT 
                    COALESCE(pb_comp.sku, iv.sku) as sku,
                    SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
                FROM items_venta iv
                JOIN ventas v ON iv.venta_id = v.id
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
                LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
                WHERE v.estado_entrega = 'pendiente'
                GROUP BY sku
            ''')
        
        ventas_activas = cursor.fetchall()
        ventas_dict = {v['sku']: int(v['vendido']) for v in ventas_activas}
        
        print(f"\n📊 Ventas activas en BD (sin venta actual): {ventas_dict}")
        
        # ============================================
        # 3. VERIFICAR CADA SKU
        # ============================================
        skus_sin_stock = []
        
        for sku in skus_a_verificar:
            print(f"\n🔎 Verificando: {sku}")
            
            cursor.execute('''
                SELECT sku, nombre, stock_actual, COALESCE(stock_full, 0) as stock_full, tipo
                FROM productos_base
                WHERE sku = %s
            ''', (sku,))
            
            prod = cursor.fetchone()
            
            if not prod:
                print(f"  ❌ NO EXISTE")
                continue
            
            print(f"  ✅ Encontrado: {prod['nombre']}")
            
            stock_fisico = prod['stock_actual'] + prod['stock_full']
            vendido_anterior = ventas_dict.get(sku, 0)
            vendido_actual = cantidades_venta_actual.get(sku, 0)
            vendido_total = vendido_anterior + vendido_actual
            stock_disponible = stock_fisico - vendido_total
            
            print(f"  📊 Stock físico: {stock_fisico}")
            print(f"  📊 Vendido anterior: {vendido_anterior}")
            print(f"  📊 Vendido actual: {vendido_actual}")
            print(f"  📊 Vendido TOTAL: {vendido_total}")
            print(f"  📊 Disponible: {stock_disponible}")
            
            if stock_disponible <= 0:
                print(f"  ⚠️ SIN STOCK - DEBE ALERTAR")
                
                skus_sin_stock.append(sku)
                
                productos_sin_stock.append({
                    'sku': sku,
                    'nombre': prod['nombre'],
                    'stock_fisico': stock_fisico,
                    'vendido': vendido_total,
                    'stock_disponible': stock_disponible,
                    'tipo_producto': 'base'
                })
                
                # Guardar alerta
                try:
                    cursor.execute('SELECT id FROM alertas_stock WHERE sku = %s AND estado = "pendiente"', (sku,))
                    
                    if not cursor.fetchone():
                        print(f"  💾 Guardando alerta...")
                        cursor.execute('''
                            INSERT INTO alertas_stock 
                            (sku, nombre_producto, stock_fisico, stock_vendido, stock_disponible, tipo_alerta, estado)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku, prod['nombre'], stock_fisico, vendido_total, stock_disponible, 'SIN_STOCK', 'pendiente'))
                        print(f"  ✅ Alerta guardada")
                    else:
                        print(f"  ℹ️ Alerta ya existe")
                except Exception as e:
                    print(f"  ❌ Error al guardar alerta: {str(e)}")
            else:
                print(f"  ✅ Stock OK")
        
        # ============================================
        # 4. BUSCAR COMBOS AFECTADOS
        # ============================================
        combos_afectados = []
        
        if skus_sin_stock:
            for sku_sin_stock in skus_sin_stock:
                try:
                    cursor.execute('''
                        SELECT DISTINCT pc.sku as combo_sku, pc.nombre as combo_nombre
                        FROM productos_compuestos pc
                        JOIN componentes c ON pc.id = c.producto_compuesto_id
                        JOIN productos_base pb ON c.producto_base_id = pb.id
                        WHERE pb.sku = %s AND pc.activo = 1
                    ''', (sku_sin_stock,))
                    
                    combos_que_usan = cursor.fetchall()
                    
                    for combo in combos_que_usan:
                        combo_sku = combo['combo_sku']
                        combo_nombre = combo['combo_nombre']
                        
                        if not any(p['sku'] == combo_sku for p in productos_sin_stock):
                            combos_afectados.append({
                                'sku': combo_sku,
                                'nombre': combo_nombre,
                                'componente_faltante': sku_sin_stock,
                                'tipo_producto': 'combo'
                            })
                            
                            try:
                                cursor.execute('SELECT id FROM alertas_stock WHERE sku = %s AND estado = "pendiente"', (combo_sku,))
                                
                                if not cursor.fetchone():
                                    cursor.execute('''
                                        INSERT INTO alertas_stock 
                                        (sku, nombre_producto, stock_fisico, stock_vendido, stock_disponible, tipo_alerta, estado, mlas_afectados)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                                    ''', (combo_sku, combo_nombre, 0, 0, 0, 'COMBO_SIN_COMPONENTE', 'pendiente', sku_sin_stock))
                            except Exception as e:
                                print(f"  ❌ Error al guardar alerta combo: {str(e)}")
                                
                except Exception as e:
                    print(f"⚠️ Error al buscar combos: {str(e)}")
        
        print(f"\n📋 Total productos sin stock: {len(productos_sin_stock) + len(combos_afectados)}")
        print("="*60 + "\n")
        
        return productos_sin_stock + combos_afectados
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return []



@app.route('/ventas/activas/exportar-excel')
@login_required
def exportar_ventas_activas_excel():
    from flask import make_response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    filtro_buscar       = request.args.get('buscar', '').strip()
    filtro_tipo_entrega = request.args.get('tipo_entrega', '')
    filtro_metodo_envio = request.args.get('metodo_envio', '')
    filtro_zona         = request.args.get('zona', '')
    filtro_canal        = request.args.get('canal', '')
    filtro_estado_pago  = request.args.get('estado_pago', '')
    query = '''SELECT v.id, v.numero_venta, v.fecha_venta, v.canal, v.mla_code,
        v.nombre_cliente, v.telefono_cliente, v.tipo_entrega, v.metodo_envio, v.zona_envio,
        v.direccion_entrega, v.costo_flete, v.metodo_pago, v.importe_total, v.importe_abonado,
        v.estado_pago, v.notas, v.cancelada_en_ml FROM ventas v WHERE v.estado_entrega = 'pendiente' '''
    params = []
    if filtro_buscar:
        query += ' AND (v.mla_code LIKE %s OR v.nombre_cliente LIKE %s OR v.id IN (SELECT venta_id FROM items_venta WHERE sku LIKE %s))'
        b = f'%{filtro_buscar}%'; params.extend([b, b, b])
    if filtro_tipo_entrega: query += ' AND v.tipo_entrega = %s'; params.append(filtro_tipo_entrega)
    if filtro_metodo_envio: query += ' AND v.metodo_envio = %s'; params.append(filtro_metodo_envio)
    if filtro_zona: query += ' AND v.zona_envio = %s'; params.append(filtro_zona)
    if filtro_canal: query += ' AND v.canal = %s'; params.append(filtro_canal)
    if filtro_estado_pago:
        if filtro_estado_pago == 'pagado': query += ' AND v.importe_abonado >= v.importe_total'
        elif filtro_estado_pago == 'pendiente': query += ' AND v.importe_abonado = 0'
        elif filtro_estado_pago == 'parcial': query += ' AND v.importe_abonado > 0 AND v.importe_abonado < v.importe_total'
    query += " ORDER BY CASE WHEN v.metodo_envio = 'Turbo' THEN 0 ELSE 1 END, v.fecha_venta DESC, v.id DESC"
    ventas = query_db(query, tuple(params) if params else None)
    for venta in ventas:
        items = query_db('''SELECT iv.sku, iv.cantidad, iv.precio_unitario,
            COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
            FROM items_venta iv LEFT JOIN productos_base pb ON iv.sku = pb.sku
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            WHERE iv.venta_id = %s ORDER BY iv.id''', (venta['id'],))
        venta['items'] = items
    wb = Workbook(); ws = wb.active; ws.title = 'Ventas Activas'
    hf = Font(bold=True, color='FFFFFF'); hfill = PatternFill('solid', fgColor='2563EB')
    center = Alignment(horizontal='center', vertical='center')
    headers = ['ID','N° Venta','Fecha','Hora','Canal','MLA/Código','Cliente','Teléfono',
               'Tipo Entrega','Método Envío','Zona','Dirección','Costo Flete',
               'Método Pago','Total','Abonado','Estado Pago','Productos','Notas']
    ws.append(headers)
    for i, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=i); c.font = hf; c.fill = hfill; c.alignment = center
    for v in ventas:
        f = v['fecha_venta']
        prods = ' | '.join(f"{i['nombre_producto']} x{i['cantidad']}" for i in (v.get('items') or []))
        ws.append([v['id'], v['numero_venta'] or '', f.strftime('%d/%m/%Y') if f else '',
            f.strftime('%H:%M') if f else '', v['canal'] or '', v['mla_code'] or '',
            v['nombre_cliente'] or '', v['telefono_cliente'] or '', v['tipo_entrega'] or '',
            v['metodo_envio'] or '', v['zona_envio'] or '', v['direccion_entrega'] or '',
            float(v['costo_flete']) if v['costo_flete'] else 0, v['metodo_pago'] or '',
            float(v['importe_total']) if v['importe_total'] else 0,
            float(v['importe_abonado']) if v['importe_abonado'] else 0,
            v['estado_pago'] or '', prods, v['notas'] or ''])
    for i, w in enumerate([6,12,12,8,12,16,24,14,12,14,12,30,12,14,12,12,12,50,40], 1):
        ws.column_dimensions[ws.cell(row=1,column=i).column_letter].width = w
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    from datetime import datetime
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename=ventas_activas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return resp


@app.route('/ventas/activas')
@login_required
def ventas_activas():
    """Lista de ventas activas con filtros de búsqueda"""
    try:
        # ========================================
        # OBTENER FILTROS
        # ========================================
        filtro_buscar = request.args.get('buscar', '').strip()
        filtro_tipo_entrega = request.args.get('tipo_entrega', '')
        filtro_metodo_envio = request.args.get('metodo_envio', '')
        filtro_zona = request.args.get('zona', '')
        filtro_canal = request.args.get('canal', '')
        filtro_estado_pago = request.args.get('estado_pago', '')
        
        # ========================================
        # CONSTRUIR QUERY CON FILTROS
        # ========================================
        query = '''
            SELECT 
                id, numero_venta, fecha_venta, canal, mla_code,
                nombre_cliente, telefono_cliente,
                tipo_entrega, metodo_envio, ubicacion_despacho,
                zona_envio, direccion_entrega, costo_flete,
metodo_pago, importe_total, importe_abonado,
                pago_mercadopago, pago_efectivo,
                estado_entrega, estado_pago, notas, cancelada_en_ml,
                fecha_entrega_estimada
            FROM ventas
            WHERE estado_entrega = 'pendiente'
        '''
        params = []
        
        # Filtro: Búsqueda de texto (apodo, nombre, productos)
        if filtro_buscar:
            query += '''
                AND (
                    mla_code LIKE %s 
                    OR nombre_cliente LIKE %s
                    OR id IN (
                        SELECT venta_id FROM items_venta WHERE sku LIKE %s
                    )
                )
            '''
            busqueda = f'%{filtro_buscar}%'
            params.extend([busqueda, busqueda, busqueda])
        
        # Filtro: Tipo de entrega
        if filtro_tipo_entrega:
            query += ' AND tipo_entrega = %s'
            params.append(filtro_tipo_entrega)
        
        # Filtro: Método de envío
        if filtro_metodo_envio:
            query += ' AND metodo_envio = %s'
            params.append(filtro_metodo_envio)
        
        # Filtro: Zona
        if filtro_zona:
            query += ' AND zona_envio = %s'
            params.append(filtro_zona)
        
        # Filtro: Canal
        if filtro_canal:
            query += ' AND canal = %s'
            params.append(filtro_canal)
        
        # Filtro: Estado de pago
        if filtro_estado_pago:
            if filtro_estado_pago == 'pagado':
                query += ' AND importe_abonado >= importe_total'
            elif filtro_estado_pago == 'pendiente':
                query += ' AND importe_abonado = 0'
            elif filtro_estado_pago == 'parcial':
                query += ' AND importe_abonado > 0 AND importe_abonado < importe_total'
        
        # Ordenar: más recientes arriba por fecha real de compra
        query += " ORDER BY CASE WHEN metodo_envio = 'Turbo' THEN 0 ELSE 1 END, fecha_venta DESC, id DESC"
        
        # Ejecutar query
        ventas = query_db(query, tuple(params) if params else None)
        
        # ========================================
        # OBTENER ITEMS DE CADA VENTA
        # ========================================
        for venta in ventas:
            items = query_db('''
                SELECT 
                    iv.sku, 
                    iv.cantidad, 
                    iv.precio_unitario,
                    COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
                ORDER BY iv.id
            ''', (venta['id'],))
            venta['items'] = items
            # Hora de compra para mostrar en listado
            if venta.get('fecha_venta') and hasattr(venta['fecha_venta'], 'strftime'):
                venta['hora_venta_str'] = venta['fecha_venta'].strftime('%H:%M')
            else:
                venta['hora_venta_str'] = ''
        
        # Leer estado auto-import
        try:
            row_ai = query_db("SELECT valor FROM configuracion WHERE clave = 'auto_import_activo' LIMIT 1")
            auto_import_activo = (row_ai[0]['valor'] != '0') if row_ai else True
        except Exception:
            auto_import_activo = True

        return render_template('ventas_activas.html', 
                             ventas=ventas,
                             filtro_buscar=filtro_buscar,
                             filtro_tipo_entrega=filtro_tipo_entrega,
                             filtro_metodo_envio=filtro_metodo_envio,
                             filtro_zona=filtro_zona,
                             filtro_canal=filtro_canal,
                             filtro_estado_pago=filtro_estado_pago,
                             hora_corte_colecta=session.get('hora_corte_colecta', '14:00'),
                             auto_import_activo=auto_import_activo)
        
    except Exception as e:
        flash(f'Error al cargar ventas activas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('index'))

@app.route('/ventas/activas/<int:venta_id>/etiqueta-ml')
@login_required
def etiqueta_ml(venta_id):
    """Descarga etiqueta de envío ML (PDF o ZPL) para una venta."""
    from flask import Response
    import requests as _req
    formato = request.args.get('formato', 'pdf')  # 'pdf' o 'zpl'
    venta = query_one("SELECT numero_venta FROM ventas WHERE id = %s", (venta_id,))
    if not venta:
        flash('Venta no encontrada', 'error')
        return redirect(url_for('ventas_activas'))
    numero = venta['numero_venta']
    orden_id = numero.replace('ML-', '').strip()
    access_token = cargar_ml_token()
    if not access_token:
        flash('Sin token ML', 'error')
        return redirect(url_for('ventas_activas'))
    headers = {'Authorization': f'Bearer {access_token}'}
    # Obtener shipping_id
    r = _req.get(f'https://api.mercadolibre.com/orders/{orden_id}', headers=headers, timeout=10)
    if r.status_code != 200:
        flash(f'Error obteniendo orden ML: {r.status_code}', 'error')
        return redirect(url_for('ventas_activas'))
    shipping_id = r.json().get('shipping', {}).get('id')
    if not shipping_id:
        flash('Esta venta no tiene envío ML asociado', 'error')
        return redirect(url_for('ventas_activas'))
    # Calcular copias ZPL según SKUs de la venta (solo para sommiers)
    def _copias_zpl(venta_id):
        import re as _re
        items = query_db("SELECT sku, cantidad FROM items_venta WHERE venta_id = %s", (venta_id,))
        copias = 1
        for item in items:
            sku = item['sku'].upper()
            if not sku.startswith('S'):  # solo sommiers
                continue
            nums = _re.findall(r'\d+', sku)
            if not nums:
                continue
            primer_num = nums[0]
            if len(primer_num) >= 4:
                ancho = int(primer_num[:3]) if int(primer_num[:3]) in (140, 150, 160, 180, 200) else int(primer_num[:2])
            else:
                ancho = int(primer_num)
            if ancho in (160, 180, 200):
                copias = max(copias, 4)
            elif ancho in (80, 90, 100, 140, 150):
                copias = max(copias, 3)
        return copias
    # Obtener etiqueta
    fmt = 'pdf' if formato == 'pdf' else 'zpl2'
    label_r = _req.get(
        f'https://api.mercadolibre.com/shipment_labels?shipment_ids={shipping_id}&response_type={fmt}',
        headers=headers, timeout=20
    )
    if label_r.status_code != 200:
        flash(f'Error obteniendo etiqueta: {label_r.status_code} — {label_r.text[:200]}', 'error')
        return redirect(url_for('ventas_activas'))
    if formato == 'pdf':
        return Response(label_r.content, headers={
            'Content-Type': 'application/pdf',
            'Content-Disposition': f'attachment; filename="etiqueta_{numero}.pdf"'
        })
    else:
        import zipfile, io as _io
        try:
            z = zipfile.ZipFile(_io.BytesIO(label_r.content))
            zpl_content = z.read(z.namelist()[0])
        except Exception:
            zpl_content = label_r.content
        # Multiplicar etiqueta según tipo de sommier
        copias = _copias_zpl(venta_id)
        if copias > 1:
            zpl_str = zpl_content.decode('utf-8', errors='replace')
            zpl_content = (zpl_str * copias).encode('utf-8')
        return Response(zpl_content, headers={
            'Content-Type': 'application/octet-stream',
            'Content-Disposition': f'attachment; filename="etiqueta_{numero}.zpl"'
        })


@app.route('/ventas/etiquetas-ml-masivo', methods=['POST'])
@login_required
def etiquetas_ml_masivo():
    """Descarga etiquetas ML para múltiples ventas en un solo ZPL/PDF."""
    from flask import Response
    import requests as _req, re as _re, zipfile, io as _io
    formato = request.form.get('formato', 'pdf')
    venta_ids = request.form.getlist('venta_ids')
    if not venta_ids:
        flash('No hay ventas seleccionadas', 'error')
        return redirect(url_for('ventas_activas'))
    access_token = cargar_ml_token()
    if not access_token:
        flash('Sin token ML', 'error')
        return redirect(url_for('ventas_activas'))
    headers = {'Authorization': f'Bearer {access_token}'}

    def copias_para_venta(vid):
        items = query_db("SELECT sku FROM items_venta WHERE venta_id = %s", (int(vid),))
        copias = 1
        for item in items:
            sku = item['sku'].upper()
            if not sku.startswith('S'):
                continue
            nums = _re.findall(r'\d+', sku)
            if not nums: continue
            primer_num = nums[0]
            if len(primer_num) >= 4:
                ancho = int(primer_num[:3]) if int(primer_num[:3]) in (140,150,160,180,200) else int(primer_num[:2])
            else:
                ancho = int(primer_num)
            if ancho in (160, 180, 200): copias = max(copias, 4)
            elif ancho in (80, 90, 100, 140, 150): copias = max(copias, 3)
        return copias

    # Recopilar shipping_ids y copias por venta
    ventas_info = []  # [(shipping_id, copias)]
    for vid in venta_ids[:50]:
        venta = query_one("SELECT numero_venta FROM ventas WHERE id = %s", (int(vid),))
        if not venta: continue
        orden_id = venta['numero_venta'].replace('ML-', '').strip()
        try:
            r = _req.get(f'https://api.mercadolibre.com/orders/{orden_id}', headers=headers, timeout=10)
            sid = r.json().get('shipping', {}).get('id') if r.status_code == 200 else None
            if sid:
                ventas_info.append((str(sid), copias_para_venta(vid)))
        except Exception:
            continue

    if not ventas_info:
        flash('No se encontraron shipments válidos', 'error')
        return redirect(url_for('ventas_activas'))

    fmt = 'pdf' if formato == 'pdf' else 'zpl2'
    ids_str = ','.join(sid for sid, _ in ventas_info)
    label_r = _req.get(
        f'https://api.mercadolibre.com/shipment_labels?shipment_ids={ids_str}&response_type={fmt}',
        headers=headers, timeout=30
    )
    if label_r.status_code != 200:
        flash(f'Error obteniendo etiquetas: {label_r.status_code}', 'error')
        return redirect(url_for('ventas_activas'))

    if formato == 'pdf':
        return Response(label_r.content, headers={
            'Content-Type': 'application/pdf',
            'Content-Disposition': 'attachment; filename="etiquetas_ml.pdf"'
        })
    else:
        # ML devuelve un ZIP — extraer y multiplicar por sommier
        try:
            z = zipfile.ZipFile(_io.BytesIO(label_r.content))
            # ML devuelve un archivo por shipment en orden
            archivos = z.namelist()
            zpl_parts = []
            for i, (sid, copias) in enumerate(ventas_info):
                if i < len(archivos):
                    contenido = z.read(archivos[i]).decode('utf-8', errors='replace')
                else:
                    break
                zpl_parts.append(contenido * copias)
            zpl_content = ''.join(zpl_parts).encode('utf-8')
        except Exception:
            zpl_content = label_r.content

        return Response(zpl_content, headers={
            'Content-Type': 'application/octet-stream',
            'Content-Disposition': 'attachment; filename="etiquetas_ml.zpl"'
        })


@app.route('/ventas/activas/<int:venta_id>/orden-retiro')
@login_required
def orden_retiro_pdf(venta_id):
    """Generar PDF de orden de retiro (2 copias: cliente + archivo)"""
    from flask import make_response
    from io import BytesIO
    from datetime import date
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    venta = query_one('SELECT * FROM ventas WHERE id = %s', (venta_id,))
    if not venta:
        flash('Venta no encontrada', 'danger')
        return redirect(url_for('ventas_activas'))

    items = query_db('''
        SELECT iv.sku, iv.cantidad, iv.precio_unitario,
               COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
        FROM items_venta iv
        LEFT JOIN productos_base pb ON iv.sku = pb.sku
        LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
        WHERE iv.venta_id = %s ORDER BY iv.id
    ''', (venta_id,))

    # Datos del documento
    nombre    = venta.get('nombre_cliente') or ''
    telefono  = venta.get('telefono_cliente') or ''
    metodo    = venta.get('metodo_pago') or ''
    nro_op    = venta.get('numero_venta') or venta.get('mla_code') or str(venta['id'])
    total     = venta.get('importe_total') or 0
    abonado   = venta.get('importe_abonado') or 0

    # DNI / CUIT
    doc_label = 'DNI/CUIT'
    doc_num   = ''
    if venta.get('factura_doc_number'):
        doc_num = str(venta['factura_doc_number'])
        doc_type = venta.get('factura_doc_type', '')
        doc_label = doc_type if doc_type else 'DNI/CUIT'
    elif venta.get('dni_cliente'):
        doc_num = str(venta['dni_cliente'])

    fecha_hoy = date.today().strftime('%d/%m/%Y')

    def dibujar_copia(c, y_offset, tipo_copia):
        W, H = A4
        m  = 15 * mm   # margen izquierdo
        mR = 15 * mm   # margen derecho
        ancho = W - m - mR

        y = y_offset

        # ── ENCABEZADO ────────────────────────────────────────────────
        # Caja título izquierda
        c.setFillColor(colors.HexColor('#1a1a2e'))
        c.roundRect(m, y - 18*mm, ancho * 0.65, 22*mm, 3, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 13)
        c.drawString(m + 4*mm, y - 7*mm, 'Recibo de entrega de producto')
        c.setFont('Helvetica', 8)
        c.drawString(m + 4*mm, y - 13*mm, f'COPIA {tipo_copia}')

        # Caja logo/empresa derecha
        x_logo = m + ancho * 0.67
        c.setFillColor(colors.HexColor('#f0f4ff'))
        c.roundRect(x_logo, y - 18*mm, ancho * 0.33, 22*mm, 3, fill=1, stroke=0)
        c.setFillColor(colors.HexColor('#1a1a2e'))
        c.setFont('Helvetica-Bold', 9)
        c.drawCentredString(x_logo + (ancho * 0.33) / 2, y - 6*mm, 'MERCADOMUEBLES')
        c.setFont('Helvetica', 7)
        c.drawCentredString(x_logo + (ancho * 0.33) / 2, y - 10*mm, 'Cimater SRL')
        c.setFont('Helvetica', 7)
        c.drawCentredString(x_logo + (ancho * 0.33) / 2, y - 14*mm, 'Emilio Lamarca 1870, CABA')

        y -= 24 * mm

        # ── AVISO ─────────────────────────────────────────────────────
        c.setFillColor(colors.HexColor('#fff8e1'))
        c.rect(m, y - 7*mm, ancho, 7*mm, fill=1, stroke=0)
        c.setFillColor(colors.HexColor('#5d4037'))
        c.setFont('Helvetica-Oblique', 7.5)
        c.drawCentredString(W / 2, y - 4.5*mm, 'Por favor completar en letra mayúscula de imprenta')
        y -= 10 * mm

        # ── SECCIÓN 1: DATOS CLIENTE ──────────────────────────────────
        c.setFillColor(colors.HexColor('#e8eaf6'))
        c.rect(m, y - 6*mm, ancho, 6*mm, fill=1, stroke=0)
        c.setFillColor(colors.HexColor('#1a237e'))
        c.setFont('Helvetica-Bold', 8)
        c.drawString(m + 2*mm, y - 4*mm, '1. Datos de quien retira el artículo')
        y -= 9 * mm

        def campo(label, valor, x, cy, w, h=7*mm):
            c.setStrokeColor(colors.HexColor('#cccccc'))
            c.setFillColor(colors.white)
            c.rect(x, cy - h, w, h, fill=1, stroke=1)
            c.setFillColor(colors.HexColor('#666666'))
            c.setFont('Helvetica', 6.5)
            c.drawString(x + 1.5*mm, cy - 2.5*mm, label)
            c.setFillColor(colors.black)
            c.setFont('Helvetica-Bold', 8.5)
            c.drawString(x + 1.5*mm, cy - 5.5*mm, str(valor)[:60])

        # Nombre (fila completa)
        campo('Nombre completo', nombre, m, y, ancho)
        y -= 9 * mm

        # DNI | Tel
        campo(doc_label, doc_num, m, y, ancho * 0.48)
        campo('Teléfono', telefono, m + ancho * 0.52, y, ancho * 0.48)
        y -= 9 * mm

        # Dirección
        campo('Dirección de retiro', 'Emilio Lamarca 1870, CABA', m, y, ancho)
        y -= 12 * mm

        # ── SECCIÓN 2: PRODUCTOS ──────────────────────────────────────
        c.setFillColor(colors.HexColor('#e8eaf6'))
        c.rect(m, y - 6*mm, ancho, 6*mm, fill=1, stroke=0)
        c.setFillColor(colors.HexColor('#1a237e'))
        c.setFont('Helvetica-Bold', 8)
        c.drawString(m + 2*mm, y - 4*mm, '2. Productos')
        y -= 9 * mm

        # Header tabla
        c.setFillColor(colors.HexColor('#37474f'))
        c.rect(m, y - 6*mm, ancho, 6*mm, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 7.5)
        c.drawString(m + 2*mm, y - 4*mm, 'Descripción')
        c.drawRightString(m + ancho * 0.72, y - 4*mm, 'Cant.')
        c.drawRightString(m + ancho * 0.87, y - 4*mm, 'P. Unit.')
        c.drawRightString(m + ancho, y - 4*mm, 'Subtotal')
        y -= 7 * mm

        for i, item in enumerate(items):
            bg = colors.HexColor('#f5f5f5') if i % 2 == 0 else colors.white
            c.setFillColor(bg)
            c.rect(m, y - 11*mm, ancho, 11*mm, fill=1, stroke=0)
            c.setStrokeColor(colors.HexColor('#eeeeee'))
            c.line(m, y - 11*mm, m + ancho, y - 11*mm)
            nombre_prod = str(item['nombre_producto'])[:55]
            c.setFillColor(colors.black)
            c.setFont('Helvetica', 8)
            c.drawString(m + 2*mm, y - 6*mm, nombre_prod)
            c.setFont('Helvetica-Bold', 8)
            c.drawRightString(m + ancho * 0.72, y - 6*mm, str(item['cantidad']))
            pu = float(item['precio_unitario'] or 0)
            st = pu * int(item['cantidad'])
            c.drawRightString(m + ancho * 0.87, y - 6*mm, f'${pu:,.0f}')
            c.drawRightString(m + ancho, y - 6*mm, f'${st:,.0f}')
            y -= 12 * mm

        # Totales
        c.setStrokeColor(colors.HexColor('#cccccc'))
        c.line(m, y, m + ancho, y)
        y -= 5 * mm
        c.setFillColor(colors.HexColor('#37474f'))
        c.setFont('Helvetica-Bold', 8)
        c.drawRightString(m + ancho * 0.87, y, 'TOTAL:')
        c.drawRightString(m + ancho, y, f'${float(total):,.0f}')
        y -= 5 * mm
        if float(abonado) > 0:
            c.setFont('Helvetica', 7.5)
            c.setFillColor(colors.HexColor('#2e7d32'))
            c.drawRightString(m + ancho * 0.87, y, 'Abonado:')
            c.drawRightString(m + ancho, y, f'${float(abonado):,.0f}')
            y -= 5 * mm
        y -= 5 * mm

        # ── SECCIÓN 3: PAGO Y OPERACIÓN ──────────────────────────────
        c.setFillColor(colors.HexColor('#e8eaf6'))
        c.rect(m, y - 6*mm, ancho, 6*mm, fill=1, stroke=0)
        c.setFillColor(colors.HexColor('#1a237e'))
        c.setFont('Helvetica-Bold', 8)
        c.drawString(m + 2*mm, y - 4*mm, '3. Información de pago')
        y -= 9 * mm

        campo('Abonó con', metodo, m, y, ancho * 0.48)
        campo('N° de operación', nro_op, m + ancho * 0.52, y, ancho * 0.48)
        y -= 12 * mm

        # ── SECCIÓN 4: FIRMA ──────────────────────────────────────────
        c.setFillColor(colors.HexColor('#e8eaf6'))
        c.rect(m, y - 6*mm, ancho, 6*mm, fill=1, stroke=0)
        c.setFillColor(colors.HexColor('#1a237e'))
        c.setFont('Helvetica-Bold', 8)
        c.drawString(m + 2*mm, y - 4*mm, '4. Conformidad de retiro')
        y -= 9 * mm

        col_w = ancho * 0.48
        alto_firma = 35*mm

        if tipo_copia == 'ARCHIVO':
            # Columna izquierda: firma cliente
            c.setStrokeColor(colors.HexColor('#cccccc'))
            c.setFillColor(colors.white)
            c.rect(m, y - alto_firma, col_w, alto_firma, fill=1, stroke=1)
            c.setFillColor(colors.HexColor('#666666'))
            c.setFont('Helvetica', 6.5)
            c.drawString(m + 1.5*mm, y - 5*mm, 'Firma:')
            c.line(m + 12*mm, y - 5.5*mm, m + col_w - 2*mm, y - 5.5*mm)
            c.drawString(m + 1.5*mm, y - 16*mm, 'Aclaración:')
            c.line(m + 16*mm, y - 16.5*mm, m + col_w - 2*mm, y - 16.5*mm)
            c.drawString(m + 1.5*mm, y - 27*mm, 'DNI:')
            c.line(m + 9*mm, y - 27.5*mm, m + col_w - 2*mm, y - 27.5*mm)
            x_emp = m + ancho * 0.52
        else:
            # CLIENTE: recuadro empresa ocupa todo el ancho
            x_emp = m
            col_w = ancho

        # Recuadro empresa
        c.setFillColor(colors.HexColor('#f8f9fa'))
        c.rect(x_emp, y - alto_firma, col_w, alto_firma, fill=1, stroke=1)
        c.setFillColor(colors.HexColor('#1a1a2e'))
        c.setFont('Helvetica-Bold', 8.5)
        c.drawCentredString(x_emp + col_w / 2, y - 9*mm, 'Mercadomuebles - Cimater SRL')

        # ENTREGADO en rojo
        c.setFillColor(colors.HexColor('#c62828'))
        c.setFont('Helvetica-Bold', 16)
        c.drawCentredString(x_emp + col_w / 2, y - 20*mm, 'ENTREGADO')

        c.setFillColor(colors.HexColor('#444444'))
        c.setFont('Helvetica', 7)
        c.drawCentredString(x_emp + col_w / 2, y - 29*mm, f'Lugar y fecha: {fecha_hoy}')

        # ── FOOTER ───────────────────────────────────────────────────
        footer_y = y - alto_firma - 5*mm
        c.setStrokeColor(colors.HexColor('#cccccc'))
        c.line(m, footer_y, m + ancho, footer_y)
        c.setFillColor(colors.HexColor('#888888'))
        c.setFont('Helvetica', 6.5)
        footer_txt = 'Emilio Lamarca 1870, Floresta, CABA  —  www.mercadomuebles.com.ar  —  Mercadomuebles de Cimater SRL  —  Tel: (011) 4639-7370  —  WhatsApp: +5491126275185'
        c.drawCentredString(W / 2, footer_y - 4*mm, footer_txt)

    # ── GENERAR PDF ───────────────────────────────────────────────────
    buffer = BytesIO()
    c = rl_canvas.Canvas(buffer, pagesize=A4)
    W, H = A4

    # Página 1: CLIENTE
    dibujar_copia(c, H - 5*mm, 'CLIENTE')
    c.showPage()
    # Página 2: ARCHIVO
    dibujar_copia(c, H - 5*mm, 'ARCHIVO')
    c.save()

    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    nombre_archivo = f'orden_retiro_{venta_id}_{fecha_hoy.replace("/", "-")}.pdf'
    response.headers['Content-Disposition'] = f'inline; filename={nombre_archivo}'
    return response


@app.route('/ventas/activas/<int:venta_id>/proceso', methods=['POST'])
@login_required
def pasar_a_proceso(venta_id):
    """Pasar venta a proceso de envío (descuenta stock con verificación)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_activas'))
        
        # Obtener items
        cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
        items = cursor.fetchall()
        
        # ========================================
        # VERIFICAR STOCK ANTES DE DESCONTAR
        # ========================================
        hay_stock, errores = verificar_stock_disponible(cursor, items, venta['ubicacion_despacho'])
        
        if not hay_stock:
            # No hay stock suficiente - mostrar modal de error
            mensaje_html = f'''
                <p><strong>No se puede procesar la venta {venta["numero_venta"]}</strong></p>
                <p>Los siguientes productos no tienen stock suficiente:</p>
                <ul class="list-unstyled">
            '''
            for error in errores:
                mensaje_html += f'<li class="text-danger mb-2"><i class="bi bi-x-circle-fill"></i> {error}</li>'
            
            mensaje_html += '''
                </ul>
                <div class="alert alert-info mt-3">
                    <i class="bi bi-info-circle"></i> 
                    Por favor, <strong>carga más stock</strong> antes de procesar esta venta.
                </div>
            '''
            flash(mensaje_html, 'error_stock')
            return redirect(url_for('ventas_activas'))
        
        # ========================================
        # HAY STOCK - PROCEDER CON DESCUENTO
        # ========================================
        for item in items:
            descontar_stock_item(cursor, item, venta['ubicacion_despacho'])
        
        # Actualizar estado
        cursor.execute('''
            UPDATE ventas 
            SET estado_entrega = 'en_proceso',
                fecha_modificacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        conn.commit()
        flash(f'✅ Venta {venta["numero_venta"]} pasada a Proceso de Envío. Stock descontado.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al pasar a proceso: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_activas'))


@app.route('/ventas/activas/<int:venta_id>/entregada', methods=['POST'])
@login_required
def marcar_entregada(venta_id):
    """Marcar venta como entregada (descuenta stock si no se descontó, con verificación)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_activas'))
        
        # Si está pendiente, necesita descontar stock
        if venta['estado_entrega'] == 'pendiente':
            cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
            items = cursor.fetchall()
            
            # ========================================
            # VERIFICAR STOCK ANTES DE DESCONTAR
            # ========================================
            hay_stock, errores = verificar_stock_disponible(cursor, items, venta['ubicacion_despacho'])
            
            if not hay_stock:
                # No hay stock suficiente - mostrar modal de error
                mensaje_html = f'''
                    <p><strong>No se puede marcar como entregada la venta {venta["numero_venta"]}</strong></p>
                    <p>Los siguientes productos no tienen stock suficiente:</p>
                    <ul class="list-unstyled">
                '''
                for error in errores:
                    mensaje_html += f'<li class="text-danger mb-2"><i class="bi bi-x-circle-fill"></i> {error}</li>'
                
                mensaje_html += '''
                    </ul>
                    <div class="alert alert-info mt-3">
                        <i class="bi bi-info-circle"></i> 
                        Por favor, <strong>carga más stock</strong> antes de marcar como entregada.
                    </div>
                '''
                flash(mensaje_html, 'error_stock')
                return redirect(url_for('ventas_activas'))
            
            # HAY STOCK - DESCONTAR
            for item in items:
                descontar_stock_item(cursor, item, venta['ubicacion_despacho'])
        
        # Actualizar estado Y FECHA DE ENTREGA
        cursor.execute('''
            UPDATE ventas 
            SET estado_entrega = 'entregada',
                fecha_entrega = NOW(),
                fecha_modificacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        conn.commit()
        log_evento('INFO', 'entrega', 'venta_entregada',
            f"Venta {venta['numero_venta']} marcada como entregada. Canal: {venta.get('canal','')}",
            venta_id=venta_id, usuario=current_user.username if current_user.is_authenticated else 'Sistema')
        flash(f'✅ Venta {venta["numero_venta"]} marcada como Entregada.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al marcar como entregada: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_activas'))


@app.route('/ventas/activas/<int:venta_id>/cancelar', methods=['POST'])
@login_required
def cancelar_venta(venta_id):
    """Cancelar venta (NO descuenta stock)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT numero_venta FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_activas'))
        
        # Actualizar estado
        cursor.execute('''
            UPDATE ventas 
            SET estado_entrega = 'cancelada',
                fecha_modificacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        conn.commit()

        # Actualizar publis ML — la cancelación libera disponible
        try:
            items_venta = query_db(
                "SELECT sku, cantidad FROM items_venta WHERE venta_id = %s", (venta_id,)
            )
            skus_afectados = _extraer_skus_base_de_items(
                [{'sku': i['sku'], 'cantidad': i['cantidad']} for i in items_venta]
            )
            if skus_afectados:
                import threading
                def _cancel_ml_bg():
                    try:
                        actualizar_publicaciones_ml_con_progreso(skus_afectados)
                    except Exception as e_ml:
                        print(f"[AUTO-ML] Error actualizando ML tras cancelación: {e_ml}")
                threading.Thread(target=_cancel_ml_bg, daemon=True).start()
        except Exception as e_ml:
            print(f"[AUTO-ML] Error actualizando ML tras cancelación: {e_ml}")

        flash(f'✅ Venta {venta["numero_venta"]} cancelada. No se descontó stock.', 'info')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al cancelar venta: {str(e)}', 'error')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_activas'))

@app.route('/ventas/activas/<int:venta_id>/eliminar', methods=['POST'])
@login_required
def eliminar_venta(venta_id):
    """
    Eliminar venta completamente de la base de datos
    NO descuenta stock (igual que cancelar)
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener info de la venta antes de borrar
        cursor.execute('SELECT numero_venta FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('❌ Venta no encontrada', 'error')
            return redirect(url_for('ventas_activas'))
        
        numero_venta = venta['numero_venta']
        
        # Obtener items antes de borrar para actualizar ML después
        cursor.execute('SELECT sku, cantidad FROM items_venta WHERE venta_id = %s', (venta_id,))
        items_venta = cursor.fetchall()

        # 1. Eliminar items de venta
        cursor.execute('DELETE FROM items_venta WHERE venta_id = %s', (venta_id,))
        items_eliminados = cursor.rowcount

        # 2. Eliminar venta
        cursor.execute('DELETE FROM ventas WHERE id = %s', (venta_id,))

        conn.commit()

        # Actualizar ML — la eliminación libera disponible
        try:
            import threading
            skus_afectados = _extraer_skus_base_de_items(
                [{'sku': i['sku'], 'cantidad': i['cantidad']} for i in items_venta]
            )
            if skus_afectados:
                def _elim_ml_bg():
                    try:
                        actualizar_publicaciones_ml_con_progreso(skus_afectados)
                    except Exception as e_ml:
                        print(f"[AUTO-ML] Error actualizando ML tras eliminación: {e_ml}")
                threading.Thread(target=_elim_ml_bg, daemon=True).start()
        except Exception as e_ml2:
            print(f"[AUTO-ML] Error iniciando thread eliminación: {e_ml2}")
        
        flash(f'✅ Venta {numero_venta} eliminada correctamente ({items_eliminados} items borrados). No se descontó stock.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al eliminar venta: {str(e)}', 'error')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_activas'))




# ============================================================================
# RUTAS: ACCIONES MÚLTIPLES EN VENTAS ACTIVAS
# Agregar en app.py después de las rutas individuales de activas
# ============================================================================

@app.route('/ventas/activas/pasar-proceso-multiple', methods=['POST'])
@login_required
def pasar_a_proceso_multiple():
    """
    Pasar múltiples ventas a proceso de envío
    Descuenta stock con verificación
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_activas'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        ventas_sin_stock = []
        
        for venta_id in venta_ids:
            try:
                # Obtener venta
                cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta or venta['estado_entrega'] != 'pendiente':
                    continue
                
                # Obtener items
                cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
                items = cursor.fetchall()
                
                # VERIFICAR STOCK ANTES DE DESCONTAR
                hay_stock, errores = verificar_stock_disponible(cursor, items, venta['ubicacion_despacho'])
                
                if not hay_stock:
                    # No hay stock - agregar a lista de errores
                    ventas_sin_stock.append({
                        'numero': venta['numero_venta'],
                        'errores': errores
                    })
                    continue
                
                # HAY STOCK - PROCEDER CON DESCUENTO
                for item in items:
                    descontar_stock_item(cursor, item, venta['ubicacion_despacho'])
                
                # Actualizar estado
                cursor.execute('''
                    UPDATE ventas 
                    SET estado_entrega = 'en_proceso',
                        fecha_modificacion = NOW()
                    WHERE id = %s
                ''', (venta_id,))
                
                ventas_procesadas += 1
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # MENSAJES
        if ventas_procesadas > 0:
            flash(f'✅ {ventas_procesadas} venta(s) pasadas a Proceso de Envío. Stock descontado.', 'success')
        
        if ventas_sin_stock:
            mensaje_html = '<div class="alert alert-warning"><strong>⚠️ Algunas ventas no se pudieron procesar por falta de stock:</strong><ul class="mt-2">'
            for v in ventas_sin_stock:
                mensaje_html += f'<li><strong>{v["numero"]}</strong><ul class="list-unstyled ms-3">'
                for error in v['errores']:
                    mensaje_html += f'<li class="text-danger"><small>{error}</small></li>'
                mensaje_html += '</ul></li>'
            mensaje_html += '</ul></div>'
            flash(mensaje_html, 'error_stock')
        
        if ventas_procesadas == 0 and not ventas_sin_stock:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        
        # Mantener filtros
        filtros = {}
        for key in ['buscar', 'zona', 'metodo_envio', 'tipo_entrega', 'estado_pago']:
            if request.form.get(key):
                filtros[key] = request.form.get(key)
        
        return redirect(url_for('ventas_activas', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_activas'))


@app.route('/ventas/activas/marcar-entregadas-multiple', methods=['POST'])
@login_required
def marcar_entregadas_multiple():
    """
    Marcar múltiples ventas como entregadas
    Descuenta stock si no se descontó, con verificación
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_activas'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        ventas_sin_stock = []
        
        for venta_id in venta_ids:
            try:
                # Obtener venta
                cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta or venta['estado_entrega'] != 'pendiente':
                    continue
                
                # Si está pendiente, necesita descontar stock
                cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
                items = cursor.fetchall()
                
                # VERIFICAR STOCK ANTES DE DESCONTAR
                hay_stock, errores = verificar_stock_disponible(cursor, items, venta['ubicacion_despacho'])
                
                if not hay_stock:
                    # No hay stock - agregar a lista de errores
                    ventas_sin_stock.append({
                        'numero': venta['numero_venta'],
                        'errores': errores
                    })
                    continue
                
                # HAY STOCK - DESCONTAR
                for item in items:
                    descontar_stock_item(cursor, item, venta['ubicacion_despacho'])
                
                # Actualizar estado Y FECHA DE ENTREGA
                cursor.execute('''
                    UPDATE ventas 
                    SET estado_entrega = 'entregada',
                        fecha_entrega = NOW(),
                        fecha_modificacion = NOW()
                    WHERE id = %s
                ''', (venta_id,))
                
                ventas_procesadas += 1
                log_evento('INFO', 'entrega', 'venta_entregada',
                    f"Venta {venta['numero_venta']} marcada como entregada (masivo). Canal: {venta.get('canal','')}",
                    venta_id=int(venta_id), usuario=current_user.username if current_user.is_authenticated else 'Sistema')
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # MENSAJES
        if ventas_procesadas > 0:
            flash(f'✅ {ventas_procesadas} venta(s) marcadas como Entregadas.', 'success')
        
        if ventas_sin_stock:
            mensaje_html = '<div class="alert alert-warning"><strong>⚠️ Algunas ventas no se pudieron marcar como entregadas por falta de stock:</strong><ul class="mt-2">'
            for v in ventas_sin_stock:
                mensaje_html += f'<li><strong>{v["numero"]}</strong><ul class="list-unstyled ms-3">'
                for error in v['errores']:
                    mensaje_html += f'<li class="text-danger"><small>{error}</small></li>'
                mensaje_html += '</ul></li>'
            mensaje_html += '</ul></div>'
            flash(mensaje_html, 'error_stock')
        
        if ventas_procesadas == 0 and not ventas_sin_stock:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        
        # Mantener filtros
        filtros = {}
        for key in ['buscar', 'zona', 'metodo_envio', 'tipo_entrega', 'estado_pago']:
            if request.form.get(key):
                filtros[key] = request.form.get(key)
        
        return redirect(url_for('ventas_activas', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_activas'))


@app.route('/ventas/activas/cancelar-multiple', methods=['POST'])
@login_required
def cancelar_ventas_multiple():
    """
    Cancelar múltiples ventas
    NO descuenta stock (porque son ventas pendientes)
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_activas'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        
        for venta_id in venta_ids:
            try:
                # Verificar que exista y esté pendiente
                cursor.execute('SELECT id FROM ventas WHERE id = %s AND estado_entrega = %s', (venta_id, 'pendiente'))
                venta = cursor.fetchone()
                
                if not venta:
                    continue
                
                # Actualizar estado
                cursor.execute('''
                    UPDATE ventas 
                    SET estado_entrega = 'cancelada',
                        fecha_modificacion = NOW()
                    WHERE id = %s
                ''', (venta_id,))
                
                ventas_procesadas += 1
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        if ventas_procesadas == 0:
            flash('❌ No se pudieron cancelar las ventas seleccionadas', 'error')
        else:
            flash(f'✅ {ventas_procesadas} venta(s) canceladas. No se descontó stock.', 'info')
        
        # Mantener filtros
        filtros = {}
        for key in ['buscar', 'zona', 'metodo_envio', 'tipo_entrega', 'estado_pago']:
            if request.form.get(key):
                filtros[key] = request.form.get(key)
        
        return redirect(url_for('ventas_activas', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_activas'))



# ============================================================================
# PROCESO DE ENVÍO - RUTAS ACTUALIZADAS
# Incluye: Volver a Activas + Motivo de Cancelación
# Reemplazar las rutas de /ventas/proceso en app.py
# ============================================================================

@app.route('/ventas/proceso')
@login_required
def ventas_proceso():
    """Lista de ventas en proceso de envío con filtros"""
    try:
        # ========================================
        # OBTENER FILTROS
        # ========================================
        filtro_buscar = request.args.get('buscar', '').strip()
        filtro_tipo_entrega = request.args.get('tipo_entrega', '')
        filtro_metodo_envio = request.args.get('metodo_envio', '')
        filtro_zona = request.args.get('zona', '')
        filtro_canal = request.args.get('canal', '')
        filtro_estado_pago = request.args.get('estado_pago', '')
        
        # ========================================
        # CONSTRUIR QUERY CON FILTROS
        # ========================================
        query = '''
            SELECT 
                id, numero_venta, fecha_venta, canal, mla_code,
                nombre_cliente, telefono_cliente,
                tipo_entrega, metodo_envio, ubicacion_despacho,
                zona_envio, direccion_entrega, costo_flete,
metodo_pago, importe_total, importe_abonado,
                pago_mercadopago, pago_efectivo,
                estado_entrega, estado_pago, notas
            FROM ventas
            WHERE estado_entrega = 'en_proceso'
        '''
        params = []
        
        # Filtro: Búsqueda de texto
        if filtro_buscar:
            query += '''
                AND (
                    mla_code LIKE %s 
                    OR nombre_cliente LIKE %s
                    OR id IN (
                        SELECT venta_id FROM items_venta WHERE sku LIKE %s
                    )
                )
            '''
            busqueda = f'%{filtro_buscar}%'
            params.extend([busqueda, busqueda, busqueda])
        
        # Filtro: Tipo de entrega
        if filtro_tipo_entrega:
            query += ' AND tipo_entrega = %s'
            params.append(filtro_tipo_entrega)
        
        # Filtro: Método de envío
        if filtro_metodo_envio:
            query += ' AND metodo_envio = %s'
            params.append(filtro_metodo_envio)
        
        # Filtro: Zona
        if filtro_zona:
            query += ' AND zona_envio = %s'
            params.append(filtro_zona)
        
        # Filtro: Canal
        if filtro_canal:
            query += ' AND canal = %s'
            params.append(filtro_canal)
        
        # Filtro: Estado de pago
        if filtro_estado_pago:
            if filtro_estado_pago == 'pagado':
                query += ' AND importe_abonado >= importe_total'
            elif filtro_estado_pago == 'pendiente':
                query += ' AND importe_abonado = 0'
            elif filtro_estado_pago == 'parcial':
                query += ' AND importe_abonado > 0 AND importe_abonado < importe_total'
        
        # Ordenar: más recientes arriba por fecha real de compra
        query += " ORDER BY CASE WHEN metodo_envio = 'Turbo' THEN 0 ELSE 1 END, fecha_venta DESC, id DESC"
        
        # Ejecutar query
        ventas = query_db(query, tuple(params) if params else None)
        
        # ========================================
        # OBTENER ITEMS DE CADA VENTA
        # ========================================
        for venta in ventas:
            items = query_db('''
                SELECT 
                    iv.sku, 
                    iv.cantidad, 
                    iv.precio_unitario,
                    COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
                ORDER BY iv.id
            ''', (venta['id'],))
            venta['items'] = items
            if venta.get('fecha_venta') and hasattr(venta['fecha_venta'], 'strftime'):
                venta['hora_venta_str'] = venta['fecha_venta'].strftime('%H:%M')
            else:
                venta['hora_venta_str'] = ''
        
        return render_template('proceso_envio.html', 
                             ventas=ventas,
                             filtro_buscar=filtro_buscar,
                             filtro_tipo_entrega=filtro_tipo_entrega,
                             filtro_metodo_envio=filtro_metodo_envio,
                             filtro_zona=filtro_zona,
                             filtro_canal=filtro_canal,
                             filtro_estado_pago=filtro_estado_pago)
        
    except Exception as e:
        flash(f'Error al cargar proceso de envío: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('index'))


@app.route('/ventas/proceso/<int:venta_id>/volver_activas', methods=['POST'])
@login_required
def proceso_volver_activas(venta_id):
    """Volver venta de proceso a activas (devuelve stock)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_proceso'))
        
        if venta['estado_entrega'] != 'en_proceso':
            flash('La venta no está en proceso', 'warning')
            return redirect(url_for('ventas_proceso'))
        
        # Obtener items de la venta
        cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
        items = cursor.fetchall()
        
        # DEVOLVER STOCK (porque ya se había descontado)
        for item in items:
            devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
        
        # Actualizar estado a 'pendiente' (volver a activas)
        cursor.execute('''
            UPDATE ventas 
            SET estado_entrega = 'pendiente',
                fecha_modificacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        conn.commit()
        flash(f'✅ Venta {venta["numero_venta"]} devuelta a Ventas Activas. Stock restaurado.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al volver a activas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_proceso'))


@app.route('/ventas/proceso/<int:venta_id>/entregada', methods=['POST'])
@login_required
def proceso_marcar_entregada(venta_id):
    """Marcar venta en proceso como entregada (stock ya descontado)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT numero_venta, estado_entrega FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_proceso'))
        
        if venta['estado_entrega'] != 'en_proceso':
            flash('La venta no está en proceso', 'warning')
            return redirect(url_for('ventas_proceso'))
        
        # Actualizar estado Y FECHA DE ENTREGA (NO descuenta stock, ya se descontó)
        cursor.execute('''
            UPDATE ventas 
            SET estado_entrega = 'entregada',
                fecha_entrega = NOW(),
                fecha_modificacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        conn.commit()
        flash(f'✅ Venta {venta["numero_venta"]} marcada como Entregada.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_proceso'))


@app.route('/ventas/proceso/<int:venta_id>/cancelar', methods=['POST'])
@login_required
def proceso_cancelar_devolver(venta_id):
    """Cancelar venta en proceso y DEVOLVER stock descontado (con motivo opcional)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_proceso'))
        
        if venta['estado_entrega'] != 'en_proceso':
            flash('La venta no está en proceso', 'warning')
            return redirect(url_for('ventas_proceso'))
        
        # Obtener motivo de cancelación (opcional)
        motivo_cancelacion = request.form.get('motivo_cancelacion', '').strip()
        
        # Obtener items de la venta
        cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
        items = cursor.fetchall()
        
        # DEVOLVER STOCK (lo opuesto a descontar)
        for item in items:
            devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
        
        # Actualizar estado y agregar motivo si existe
        if motivo_cancelacion:
            # Agregar motivo a las notas existentes
            notas_actuales = venta.get('notas', '') or ''
            notas_nuevas = f"{notas_actuales}\n[CANCELACIÓN] {motivo_cancelacion}".strip()
            
            cursor.execute('''
                UPDATE ventas 
                SET estado_entrega = 'cancelada',
                    notas = %s,
                    fecha_modificacion = NOW()
                WHERE id = %s
            ''', (notas_nuevas, venta_id))
        else:
            cursor.execute('''
                UPDATE ventas 
                SET estado_entrega = 'cancelada',
                    fecha_modificacion = NOW()
                WHERE id = %s
            ''', (venta_id,))
        
        conn.commit()
        
        mensaje = f'✅ Venta {venta["numero_venta"]} cancelada. Stock devuelto correctamente.'
        if motivo_cancelacion:
            mensaje += f' Motivo: {motivo_cancelacion}'
        
        flash(mensaje, 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al cancelar y devolver stock: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_proceso'))

# ============================================================================
# RUTAS: ACCIONES MÚLTIPLES EN VENTAS EN PROCESO
# Agregar en app.py después de las rutas individuales de proceso
# ============================================================================

@app.route('/ventas/proceso/volver-activas-multiple', methods=['POST'])
@login_required
def proceso_volver_activas_multiple():
    """
    Volver múltiples ventas en proceso a activas
    Devuelve stock de todas
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_proceso'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        
        for venta_id in venta_ids:
            try:
                # Obtener venta
                cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta or venta['estado_entrega'] != 'en_proceso':
                    continue
                
                # Obtener items
                cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
                items = cursor.fetchall()
                
                # DEVOLVER STOCK (porque ya se había descontado)
                for item in items:
                    devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
                
                # Actualizar estado a 'pendiente' (volver a activas)
                cursor.execute('''
                    UPDATE ventas 
                    SET estado_entrega = 'pendiente',
                        fecha_modificacion = NOW()
                    WHERE id = %s
                ''', (venta_id,))
                
                ventas_procesadas += 1
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        if ventas_procesadas == 0:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        else:
            flash(f'✅ {ventas_procesadas} venta(s) devueltas a Ventas Activas. Stock restaurado.', 'success')
        
        # Mantener filtros
        filtros = {}
        if request.form.get('buscar'):
            filtros['buscar'] = request.form.get('buscar')
        if request.form.get('zona'):
            filtros['zona'] = request.form.get('zona')
        if request.form.get('metodo_envio'):
            filtros['metodo_envio'] = request.form.get('metodo_envio')
        
        return redirect(url_for('ventas_proceso', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_proceso'))


@app.route('/ventas/proceso/marcar-entregadas-multiple', methods=['POST'])
@login_required
def proceso_marcar_entregadas_multiple():
    """
    Marcar múltiples ventas en proceso como entregadas
    NO toca stock (ya estaba descontado)
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_proceso'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        
        for venta_id in venta_ids:
            try:
                # Verificar que esté en proceso
                cursor.execute('SELECT estado_entrega FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta or venta['estado_entrega'] != 'en_proceso':
                    continue
                
                # Actualizar estado Y FECHA DE ENTREGA (NO descuenta stock)
                cursor.execute('''
                    UPDATE ventas 
                    SET estado_entrega = 'entregada',
                        fecha_entrega = NOW(),
                        fecha_modificacion = NOW()
                    WHERE id = %s
                ''', (venta_id,))
                
                ventas_procesadas += 1
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        if ventas_procesadas == 0:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        else:
            flash(f'✅ {ventas_procesadas} venta(s) marcadas como Entregadas.', 'success')
        
        # Mantener filtros
        filtros = {}
        if request.form.get('buscar'):
            filtros['buscar'] = request.form.get('buscar')
        if request.form.get('zona'):
            filtros['zona'] = request.form.get('zona')
        if request.form.get('metodo_envio'):
            filtros['metodo_envio'] = request.form.get('metodo_envio')
        
        return redirect(url_for('ventas_proceso', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_proceso'))


@app.route('/ventas/proceso/cancelar-multiple', methods=['POST'])
@login_required
def proceso_cancelar_multiple():
    """
    Cancelar múltiples ventas en proceso y DEVOLVER stock
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        motivo_cancelacion = request.form.get('motivo_cancelacion', '').strip()
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_proceso'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        
        for venta_id in venta_ids:
            try:
                # Obtener venta
                cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta or venta['estado_entrega'] != 'en_proceso':
                    continue
                
                # Obtener items
                cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
                items = cursor.fetchall()
                
                # DEVOLVER STOCK
                for item in items:
                    devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
                
                # Actualizar estado y agregar motivo si existe
                if motivo_cancelacion:
                    notas_actuales = venta.get('notas', '') or ''
                    notas_nuevas = f"{notas_actuales}\n[CANCELACIÓN MÚLTIPLE] {motivo_cancelacion}".strip()
                    
                    cursor.execute('''
                        UPDATE ventas 
                        SET estado_entrega = 'cancelada',
                            notas = %s,
                            fecha_modificacion = NOW()
                        WHERE id = %s
                    ''', (notas_nuevas, venta_id))
                else:
                    cursor.execute('''
                        UPDATE ventas 
                        SET estado_entrega = 'cancelada',
                            fecha_modificacion = NOW()
                        WHERE id = %s
                    ''', (venta_id,))
                
                ventas_procesadas += 1
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        if ventas_procesadas == 0:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        else:
            mensaje = f'✅ {ventas_procesadas} venta(s) canceladas. Stock devuelto correctamente.'
            if motivo_cancelacion:
                mensaje += f' Motivo: {motivo_cancelacion}'
            flash(mensaje, 'success')
        
        # Mantener filtros
        filtros = {}
        if request.form.get('buscar'):
            filtros['buscar'] = request.form.get('buscar')
        if request.form.get('zona'):
            filtros['zona'] = request.form.get('zona')
        if request.form.get('metodo_envio'):
            filtros['metodo_envio'] = request.form.get('metodo_envio')
        
        return redirect(url_for('ventas_proceso', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_proceso'))


@app.route('/ventas/editar/<int:venta_id>', methods=['GET', 'POST'])
@login_required
def editar_venta(venta_id):
    """Editar una venta activa"""
    
    if request.method == 'GET':
        # Obtener datos de la venta
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Venta principal
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_activas'))
        
        # CORREGIDO: Convertir fecha a formato string para input type="date"
        if venta['fecha_venta']:
            from datetime import datetime
            if isinstance(venta['fecha_venta'], datetime):
                venta['fecha_venta_str'] = venta['fecha_venta'].strftime('%Y-%m-%d')
            elif hasattr(venta['fecha_venta'], 'strftime'):  # date object
                venta['fecha_venta_str'] = venta['fecha_venta'].strftime('%Y-%m-%d')
            else:
                venta['fecha_venta_str'] = str(venta['fecha_venta'])
        else:
            venta['fecha_venta_str'] = ''
        
        # Items de la venta
        cursor.execute('''
            SELECT iv.*, 
                   COALESCE(pb.nombre, pc.nombre) as nombre_producto
            FROM items_venta iv
            LEFT JOIN productos_base pb ON iv.sku = pb.sku
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            WHERE iv.venta_id = %s
        ''', (venta_id,))
        items = cursor.fetchall()
        
        # Productos disponibles para agregar
        cursor.execute('''
            SELECT sku, nombre, tipo FROM productos_base
            UNION
            SELECT sku, nombre, 'sommier' as tipo FROM productos_compuestos WHERE activo = 1
            ORDER BY nombre
        ''')
        productos = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return render_template('editar_venta.html', 
                             venta=venta, 
                             items=items,
                             productos=productos)
    
    # POST - Guardar cambios
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener items ANTES del cambio (para comparar)
        cursor.execute('SELECT sku, cantidad FROM items_venta WHERE venta_id = %s', (venta_id,))
        items_anteriores = {item['sku']: item['cantidad'] for item in cursor.fetchall()}
        
        # ========================================
        # 1. ACTUALIZAR DATOS GENERALES
        # ========================================
        numero_venta = request.form.get('numero_venta')
        fecha_venta_form = request.form.get('fecha_venta')
        canal = request.form.get('canal')
        mla_code = request.form.get('mla_code', '').strip()
        nombre_cliente = request.form.get('nombre_cliente', '').strip()
        telefono_cliente = request.form.get('telefono_cliente', '')

        # Preservar hora original — solo actualizar la fecha, no la hora
        if fecha_venta_form:
            cursor.execute('SELECT fecha_venta FROM ventas WHERE id = %s', (venta_id,))
            row_fecha = cursor.fetchone()
            hora_original = row_fecha['fecha_venta'].strftime('%H:%M:%S') if row_fecha and row_fecha['fecha_venta'] else '00:00:00'
            fecha_venta = f"{fecha_venta_form} {hora_original}"
        else:
            fecha_venta = None
        
        # Entrega
        tipo_entrega = request.form.get('tipo_entrega')
        direccion_entrega = request.form.get('direccion_entrega', '')
        metodo_envio = request.form.get('metodo_envio', '')
        zona_envio = request.form.get('zona_envio', '')
        
        # Calcular ubicación de despacho
        if metodo_envio == 'Full':
            ubicacion_despacho = 'FULL'
        else:
            ubicacion_despacho = 'DEP'
        
        responsable_entrega = request.form.get('responsable_entrega', '')
        costo_flete = float(request.form.get('costo_flete') or 0)
        
        # ========================================
        # 2. PRODUCTOS Y CÁLCULO DE TOTAL
        # ========================================
        # Borrar items existentes
        cursor.execute('DELETE FROM items_venta WHERE venta_id = %s', (venta_id,))
        
        # Insertar items nuevos y calcular total
        productos = request.form.to_dict(flat=False)
        items_nuevos = {}
        importe_total = 0  # CORREGIDO: Calcular desde productos
        
        for key in productos.keys():
            if key.startswith('productos[') and key.endswith('[sku]'):
                index = key.split('[')[1].split(']')[0]
                sku = productos.get(f'productos[{index}][sku]', [None])[0]
                cantidad = int(productos.get(f'productos[{index}][cantidad]', [0])[0])
                precio = float(productos.get(f'productos[{index}][precio]', [0])[0])
                
                if sku and cantidad > 0:
                    # Insertar item
                    cursor.execute('''
                        INSERT INTO items_venta (venta_id, sku, cantidad, precio_unitario)
                        VALUES (%s, %s, %s, %s)
                    ''', (venta_id, sku, cantidad, precio))
                    
                    items_nuevos[sku] = cantidad
                    
                    # CORREGIDO: Sumar al total
                    importe_total += (cantidad * precio)
        
        # ========================================
        # 3. PAGO (INDEPENDIENTE DEL TOTAL)
        # ========================================
        metodo_pago = request.form.get('metodo_pago')
        # CORREGIDO: Los pagos son independientes del total calculado
        pago_mercadopago = float(request.form.get('pago_mercadopago', 0))
        pago_efectivo = float(request.form.get('pago_efectivo', 0))
        importe_abonado = pago_mercadopago + pago_efectivo
        
        # Estado pago
        if importe_abonado >= importe_total:
            estado_pago = 'pagado'
        elif importe_abonado > 0:
            estado_pago = 'pago_parcial'
        else:
            estado_pago = 'pago_pendiente'
        
        # Notas
        notas = request.form.get('notas', '')
        
        # ========================================
        # 4. ACTUALIZAR VENTA
        # ========================================
        cursor.execute('''
            UPDATE ventas SET
                numero_venta = %s,
                fecha_venta = %s,
                canal = %s,
                mla_code = %s,
                nombre_cliente = %s,
                telefono_cliente = %s,
                tipo_entrega = %s,
                metodo_envio = %s,
                ubicacion_despacho = %s,
                zona_envio = %s,
                direccion_entrega = %s,
                responsable_entrega = %s,
                costo_flete = %s,
                metodo_pago = %s,
                importe_total = %s,
                importe_abonado = %s,
                pago_mercadopago = %s,
                pago_efectivo = %s,
                estado_pago = %s,
                notas = %s
            WHERE id = %s
        ''', (
            numero_venta, fecha_venta, canal, mla_code,
            nombre_cliente, telefono_cliente,
            tipo_entrega, metodo_envio, ubicacion_despacho,
            zona_envio, direccion_entrega, responsable_entrega,
            costo_flete, metodo_pago, importe_total, importe_abonado,
            pago_mercadopago, pago_efectivo,
            estado_pago, notas,
            venta_id
        ))
        
        # ========================================
        # 5. DETECTAR ALERTAS (Solo si cambiaron items)
        # ========================================
        items_cambiaron = items_anteriores != items_nuevos
        
        if items_cambiaron:
            try:
                items_vendidos_lista = [{'sku': sku, 'cantidad': cant} for sku, cant in items_nuevos.items()]
                if items_vendidos_lista:
                    productos_sin_stock = detectar_alertas_stock_bajo(cursor, items_vendidos_lista, venta_id)
                    
                    if productos_sin_stock:
                        # Mostrar alerta simplificada
                        productos_base = [p for p in productos_sin_stock if p.get('tipo_producto') == 'base']
                        if productos_base:
                            nombres = ', '.join([p['nombre'] for p in productos_base[:3]])
                            flash(f'⚠️ Productos sin stock: {nombres}', 'warning')
            except Exception as e:
                print(f"Error al detectar alertas: {str(e)}")
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash(f'✅ Venta {numero_venta} actualizada correctamente. Total: ${importe_total:,.0f}', 'success')
        return redirect(url_for('ventas_activas'))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        import traceback
        traceback.print_exc()
        flash(f'❌ Error al actualizar venta: {str(e)}', 'error')
        return redirect(url_for('ventas_activas'))



# ============================================================================
# FUNCIÓN AUXILIAR: DEVOLVER STOCK (OPUESTO A DESCONTAR)
# ============================================================================

def devolver_stock_item(cursor, item, ubicacion_despacho):
    """
    Devuelve stock de un item (lo opuesto a descontar_stock_item)
    Considera:
    - Ubicación de despacho (DEP o FULL)
    - Descomposición de combos
    - Bases grandes (160, 180, 200) devuelven 2 bases chicas
    """
    sku = item['sku']
    cantidad = item['cantidad']
    
    # Verificar si es un combo
    cursor.execute('SELECT id FROM productos_compuestos WHERE sku = %s', (sku,))
    combo = cursor.fetchone()
    
    if combo:
        # Es combo: descomponer y devolver componentes
        cursor.execute('''
            SELECT pb.sku, pb.tipo, c.cantidad_necesaria
            FROM componentes c
            JOIN productos_base pb ON c.producto_base_id = pb.id
            WHERE c.producto_compuesto_id = %s
        ''', (combo['id'],))
        componentes = cursor.fetchall()
        
        for comp in componentes:
            sku_comp = comp['sku']
            cant_comp = comp['cantidad_necesaria'] * cantidad
            tipo_comp = comp['tipo']
            
            # Devolver componente según ubicación
            devolver_stock_simple(cursor, sku_comp, cant_comp, tipo_comp, ubicacion_despacho)
    
    else:
        # Es producto simple
        cursor.execute('SELECT tipo FROM productos_base WHERE sku = %s', (sku,))
        prod = cursor.fetchone()
        tipo = prod['tipo'] if prod else 'colchon'
        
        devolver_stock_simple(cursor, sku, cantidad, tipo, ubicacion_despacho)


def devolver_stock_simple(cursor, sku, cantidad, tipo, ubicacion_despacho):
    """
    Devuelve stock de un producto simple según ubicación (SUMA en lugar de RESTAR)
    """
    # COMPAC: tiene _DEP y _FULL
    if '_DEP' in sku or '_FULL' in sku:
        if ubicacion_despacho == 'FULL':
            # Devolver a _FULL
            sku_real = sku.replace('_DEP', '_FULL')
        else:
            # Devolver a _DEP
            sku_real = sku.replace('_FULL', '_DEP')
        
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual + %s 
            WHERE sku = %s
        ''', (cantidad, sku_real))
    
    # ALMOHADAS: tienen stock_actual (DEP) y stock_full (FULL)
    elif tipo == 'almohada':
        if ubicacion_despacho == 'FULL':
            cursor.execute('''
                UPDATE productos_base 
                SET stock_full = stock_full + %s 
                WHERE sku = %s
            ''', (cantidad, sku))
        else:
            cursor.execute('''
                UPDATE productos_base 
                SET stock_actual = stock_actual + %s 
                WHERE sku = %s
            ''', (cantidad, sku))
    
    # BASES CHICAS (80200, 90200, 100200): devolver directamente
    elif tipo == 'base' and any(x in sku for x in ['80200', '90200', '100200']):
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual + %s 
            WHERE sku = %s
        ''', (cantidad, sku))
    
    # BASES GRANDES (160, 180, 200): devuelven 2 bases chicas
    elif tipo == 'base' and any(x in sku for x in ['160', '180', '200']):
        # Determinar SKU de bases chicas
        if '160' in sku:
            sku_chica = sku.replace('160', '80200')
            cant_bases = cantidad * 2
        elif '180' in sku:
            sku_chica = sku.replace('180', '90200')
            cant_bases = cantidad * 2
        elif '200' in sku:
            sku_chica = sku.replace('200', '100200')
            cant_bases = cantidad * 2
        else:
            return
        
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual + %s 
            WHERE sku = %s
        ''', (cant_bases, sku_chica))
    
    # OTROS: devolver a stock_actual
    else:
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual + %s 
            WHERE sku = %s
        ''', (cantidad, sku))



# ============================================================================
# FUNCIÓN AUXILIAR: DESCONTAR STOCK
# ============================================================================

def descontar_stock_item(cursor, item, ubicacion_despacho):
    """
    Descuenta stock de un item considerando:
    - Ubicación de despacho (DEP o FULL)
    - Descomposición de combos
    - Bases grandes (160, 180, 200) descuentan 2 bases chicas
    """
    sku = item['sku']
    cantidad = item['cantidad']
    
    # Verificar si es un combo
    cursor.execute('SELECT id FROM productos_compuestos WHERE sku = %s', (sku,))
    combo = cursor.fetchone()
    
    if combo:
        # Es combo: descomponer y descontar componentes
        cursor.execute('''
            SELECT pb.sku, pb.tipo, c.cantidad_necesaria
            FROM componentes c
            JOIN productos_base pb ON c.producto_base_id = pb.id
            WHERE c.producto_compuesto_id = %s
        ''', (combo['id'],))
        componentes = cursor.fetchall()
        
        for comp in componentes:
            sku_comp = comp['sku']
            cant_comp = comp['cantidad_necesaria'] * cantidad
            tipo_comp = comp['tipo']
            
            # Descontar componente según ubicación
            descontar_stock_simple(cursor, sku_comp, cant_comp, tipo_comp, ubicacion_despacho)
    
    else:
        # Es producto simple
        cursor.execute('SELECT tipo FROM productos_base WHERE sku = %s', (sku,))
        prod = cursor.fetchone()
        tipo = prod['tipo'] if prod else 'colchon'
        
        descontar_stock_simple(cursor, sku, cantidad, tipo, ubicacion_despacho)


def descontar_stock_simple(cursor, sku, cantidad, tipo, ubicacion_despacho):
    """Descuenta stock de un producto simple según ubicación y registra en movimientos_stock"""

    def _descontar_y_registrar(sku_real, cant):
        cursor.execute('SELECT stock_actual, nombre FROM productos_base WHERE sku = %s', (sku_real,))
        prod = cursor.fetchone()
        stock_anterior = int(prod['stock_actual'] or 0) if prod else 0
        nombre_prod = prod['nombre'] if prod else sku_real
        stock_nuevo = stock_anterior - cant
        cursor.execute(
            'UPDATE productos_base SET stock_actual = stock_actual - %s WHERE sku = %s',
            (cant, sku_real))
        cursor.execute("""
            INSERT INTO movimientos_stock
                (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo, usuario)
            VALUES (%s, %s, 'venta', %s, %s, %s, 'Descuento por entrega', 'Sistema')
        """, (sku_real, nombre_prod, cant, stock_anterior, stock_nuevo))

    # COMPAC: tiene _DEP y _FULL
    if '_DEP' in sku or '_FULL' in sku:
        sku_real = sku.replace('_DEP', '_FULL') if ubicacion_despacho == 'FULL' else sku.replace('_FULL', '_DEP')
        _descontar_y_registrar(sku_real, cantidad)

    # ALMOHADAS: tienen stock_actual (DEP) y stock_full (FULL)
    elif tipo == 'almohada':
        if ubicacion_despacho == 'FULL':
            cursor.execute('SELECT stock_full, nombre FROM productos_base WHERE sku = %s', (sku,))
            prod = cursor.fetchone()
            stock_anterior = int(prod['stock_full'] or 0) if prod else 0
            nombre_prod = prod['nombre'] if prod else sku
            cursor.execute('UPDATE productos_base SET stock_full = stock_full - %s WHERE sku = %s', (cantidad, sku))
        else:
            cursor.execute('SELECT stock_actual, nombre FROM productos_base WHERE sku = %s', (sku,))
            prod = cursor.fetchone()
            stock_anterior = int(prod['stock_actual'] or 0) if prod else 0
            nombre_prod = prod['nombre'] if prod else sku
            cursor.execute('UPDATE productos_base SET stock_actual = stock_actual - %s WHERE sku = %s', (cantidad, sku))
        cursor.execute("""
            INSERT INTO movimientos_stock
                (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo, usuario)
            VALUES (%s, %s, 'venta', %s, %s, %s, 'Descuento por entrega', 'Sistema')
        """, (sku, nombre_prod, cantidad, stock_anterior, stock_anterior - cantidad))

    # BASES CHICAS (80200, 90200, 100200): descontar directamente
    elif tipo == 'base' and any(x in sku for x in ['80200', '90200', '100200']):
        _descontar_y_registrar(sku, cantidad)

    # BASES GRANDES (160, 180, 200): descuentan 2 bases chicas
    elif tipo == 'base' and any(x in sku for x in ['160', '180', '200']):
        if '160' in sku:
            sku_chica = sku.replace('160', '80200')
        elif '180' in sku:
            sku_chica = sku.replace('180', '90200')
        else:
            sku_chica = sku.replace('200', '100200')
        _descontar_y_registrar(sku_chica, cantidad * 2)

    # OTROS: descontar de stock_actual
    else:
        _descontar_y_registrar(sku, cantidad)




# ─── FUNCIÓN HELPER: Actualizar stock en ML ───
def actualizar_stock_ml(mla_id, cantidad, access_token):
    """
    Actualizar stock de una publicación en Mercado Libre
    
    Args:
        mla_id: ID de la publicación (ej: MLA603027006)
        cantidad: Nueva cantidad de stock (0 para pausar ventas)
        access_token: Token de ML
    
    Returns:
        (success: bool, message: str)
    """
    try:
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        data = {
            "available_quantity": cantidad
        }
        
        response = requests.put(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers,
            json=data
        )
        
        if response.status_code == 200:
            return True, f"Stock actualizado a {cantidad} en ML"
        else:
            error_data = response.json()
            error_msg = error_data.get('message', 'Error desconocido')
            return False, f"Error ML: {error_msg}"
    
    except Exception as e:
        return False, f"Error: {str(e)}"


# ─── FUNCIÓN HELPER: Pausar publicación en ML ───
def pausar_publicacion_ml(mla_id, access_token):
    """
    Pausar una publicación en Mercado Libre
    
    Args:
        mla_id: ID de la publicación
        access_token: Token de ML
    
    Returns:
        (success: bool, message: str)
    """
    try:
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        data = {
            "status": "paused"
        }
        
        response = requests.put(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers,
            json=data
        )
        
        if response.status_code == 200:
            return True, "Publicación pausada en ML"
        else:
            error_data = response.json()
            error_msg = error_data.get('message', 'Error desconocido')
            return False, f"Error ML: {error_msg}"
    
    except Exception as e:
        return False, f"Error: {str(e)}"


# ─── RUTA NUEVA: Sincronizar stock con ML desde alertas ───
@app.route('/alertas/<int:alerta_id>/sincronizar-ml', methods=['POST'])
@login_required
def sincronizar_ml_desde_alerta(alerta_id):
    """
    Poner stock en 0 en las publicaciones NORMALES (sin Z)
    Procesa solo la parte NORMAL de la alerta
    """
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay token de ML configurado', 'error')
        return redirect(url_for('alertas_ml'))
    
    try:
        # Obtener la alerta
        alerta = query_db('SELECT * FROM alertas_stock WHERE id = %s', (alerta_id,))
        if alerta:
            alerta = alerta[0]
        
        if not alerta:
            flash('❌ Alerta no encontrada', 'error')
            return redirect(url_for('alertas_ml'))
        
        sku = alerta['sku']
        
        # Obtener publicaciones NORMALES (sin Z)
        publicaciones = query_db(
            'SELECT * FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE',
            (sku,)
        )
        
        if not publicaciones:
            flash(f'⚠️ No hay publicaciones de ML mapeadas para el SKU {sku}', 'warning')
            return redirect(url_for('alertas_ml'))
        
        # Actualizar stock en cada publicación
        resultados = []
        errores = []
        
        for pub in publicaciones:
            mla_id = pub['mla_id']
            success, message = actualizar_stock_ml(mla_id, 0, access_token)
            
            if success:
                resultados.append(f"{mla_id}: {message}")
            else:
                errores.append(f"{mla_id}: {message}")
        
        # ✅ LÓGICA DE PROCESAMIENTO INDEPENDIENTE
        sku_con_z = f"{sku}Z"
        tiene_variante_z = len(query_db(
            'SELECT 1 FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE LIMIT 1',
            (sku_con_z,)
        )) > 0
        
        tipo_procesado_actual = alerta.get('tipo_procesado')
        
        if tiene_variante_z:
            # Tiene variante Z - solo marcar que procesamos la parte normal
            if tipo_procesado_actual == 'z':
                # Ya se había procesado Z → ahora marcar como ambos y cerrar alerta
                execute_db(
                    "UPDATE alertas_stock SET tipo_procesado = 'ambos', estado = 'procesada', fecha_procesada = NOW() WHERE id = %s",
                    (alerta_id,)
                )
            else:
                # Solo marcar que se procesó la parte normal
                execute_db(
                    "UPDATE alertas_stock SET tipo_procesado = 'normal' WHERE id = %s",
                    (alerta_id,)
                )
        else:
            # No tiene variante Z - cerrar la alerta directamente
            execute_db(
                "UPDATE alertas_stock SET tipo_procesado = 'ambos', estado = 'procesada', fecha_procesada = NOW() WHERE id = %s",
                (alerta_id,)
            )
        
        # Mostrar resultados
        if resultados:
            flash(f'✅ Stock actualizado en ML: {", ".join(resultados)}', 'success')
        
        if errores:
            flash(f'❌ Errores: {", ".join(errores)}', 'error')
    
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('alertas_ml'))


# ─── RUTA ACTUALIZADA: alertas_ml con info de publicaciones ───
@app.route('/alertas')
@login_required
def alertas_ml():
    """Ver alertas de stock pendientes con info de publicaciones ML (normales y con Z)"""
    alertas = []
    try:
        alertas_raw = query_db('''
            SELECT * FROM alertas_stock 
            WHERE estado = 'pendiente'
            ORDER BY stock_disponible ASC, fecha_creacion DESC
        ''')
        
        # Enriquecer cada alerta con info de publicaciones ML
        for alerta in alertas_raw:
            sku = alerta['sku']
            
            # Publicaciones normales (sin Z)
            publicaciones = query_db(
                'SELECT mla_id, titulo_ml FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE',
                (sku,)
            )
            
            # Publicaciones con variante Z
            sku_con_z = f"{sku}Z"
            publicaciones_z = query_db(
                'SELECT mla_id, titulo_ml FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE',
                (sku_con_z,)
            )
            
            alerta['publicaciones_ml'] = publicaciones
            alerta['tiene_ml'] = len(publicaciones) > 0
            alerta['publicaciones_ml_z'] = publicaciones_z
            alerta['tiene_ml_z'] = len(publicaciones_z) > 0
            alertas.append(alerta)
    
    except Exception as e:
        flash(f'Error al cargar alertas: {str(e)}', 'error')
    
    return render_template('alertas.html', alertas=alertas)



@app.route('/alertas/<int:alerta_id>/procesar', methods=['POST'])
@login_required
def marcar_alerta_procesada(alerta_id):
    """Marcar una alerta como procesada"""
    try:
        execute_db(
            "UPDATE alertas_stock SET estado = 'procesada', fecha_procesada = NOW() WHERE id = %s",
            (alerta_id,)
        )
        flash('✅ Alerta marcada como procesada', 'success')
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
    
    return redirect(url_for('alertas_ml'))


@app.route('/alertas/marcar-todas-procesadas', methods=['POST'])
@login_required
def marcar_todas_procesadas():
    """Marcar TODAS las alertas pendientes como procesadas"""
    try:
        result = execute_db(
            "UPDATE alertas_stock SET estado = 'procesada', fecha_procesada = NOW() WHERE estado = 'pendiente'"
        )
        flash('✅ Todas las alertas fueron marcadas como procesadas', 'success')
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
    
    return redirect(url_for('alertas_ml'))

@app.route('/stock')
@login_required
def ver_stock():
    """Ver stock disponible con filtros - PRODUCTOS BASE + COMBOS"""
    productos = []
    filtro_estado = request.args.get('estado', 'TODOS')
    filtro_tipo = request.args.get('tipo', 'TODOS')
    filtro_modelo = request.args.get('modelo', 'TODOS')
    filtro_medida = request.args.get('medida', 'TODAS')
    buscar = request.args.get('buscar', '').strip()
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ============================================
        # 1. VENTAS ACTIVAS
        # ============================================
        cursor.execute('''
            SELECT 
                COALESCE(pb_comp.sku, iv.sku) as sku,
                SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
            FROM items_venta iv
            JOIN ventas v ON iv.venta_id = v.id
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
            LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
            WHERE v.estado_entrega = 'pendiente'
            GROUP BY sku
        ''')
        ventas_activas = cursor.fetchall()
        ventas_dict = {v['sku']: int(v['vendido']) for v in ventas_activas}
        
        # ============================================
        # 2. PRODUCTOS BASE
        # ============================================
        query_productos = """
            SELECT 
                sku, nombre, tipo, medida, modelo,
                stock_actual, COALESCE(stock_full, 0) as stock_full
            FROM productos_base
            WHERE 1=1
        """
        params = []
        
        if filtro_tipo != 'TODOS' and filtro_tipo in ('colchon', 'base', 'almohada'):
            query_productos += " AND tipo = %s"
            params.append(filtro_tipo)
        
        if filtro_medida != 'TODAS':
            query_productos += " AND medida = %s"
            params.append(filtro_medida)
        
        if filtro_modelo != 'TODOS':
            query_productos += " AND (nombre LIKE %s OR modelo LIKE %s)"
            params.append(f'%{filtro_modelo}%')
            params.append(f'%{filtro_modelo}%')
        
        if buscar:
            query_productos += " AND (sku LIKE %s OR nombre LIKE %s)"
            params.append(f'%{buscar}%')
            params.append(f'%{buscar}%')
        
        cursor.execute(query_productos, tuple(params) if params else None)
        productos_base = cursor.fetchall()
        
        for prod in productos_base:
            sku = prod['sku']
            stock_fisico = prod['stock_actual'] + prod['stock_full']
            vendido = ventas_dict.get(sku, 0)
            stock_disponible = stock_fisico - vendido
            
            if stock_disponible <= 0:
                estado_stock = 'SIN_STOCK'
            elif stock_disponible <= 2:
                estado_stock = 'POCO_STOCK'
            else:
                estado_stock = 'DISPONIBLE'
            
            if filtro_estado != 'TODOS' and estado_stock != filtro_estado:
                continue
            
            productos.append({
                'sku': sku,
                'nombre': prod['nombre'],
                'tipo': prod['tipo'],
                'medida': prod['medida'],
                'modelo': prod['modelo'],
                'stock_fisico': stock_fisico,
                'stock_disponible': stock_disponible,
                'estado_stock': estado_stock,
                'es_combo': False
            })
        
        # ============================================
        # 3. COMBOS
        # ============================================
        if filtro_tipo == 'TODOS' or filtro_tipo == 'sommier':
            # Consultar solo columnas que existen
            query_combos = """
                SELECT id, sku, nombre
                FROM productos_compuestos
                WHERE activo = 1
            """
            params_combos = []
            
            # Filtros básicos
            if filtro_modelo != 'TODOS':
                query_combos += " AND nombre LIKE %s"
                params_combos.append(f'%{filtro_modelo}%')
            
            if buscar:
                query_combos += " AND (sku LIKE %s OR nombre LIKE %s)"
                params_combos.append(f'%{buscar}%')
                params_combos.append(f'%{buscar}%')
            
            cursor.execute(query_combos, tuple(params_combos) if params_combos else None)
            combos = cursor.fetchall()
            
            for combo in combos:
                combo_id = combo['id']
                combo_sku = combo['sku']
                combo_nombre = combo['nombre']
                
                # Extraer medida del SKU (ej: SEX140 -> 140)
                medida_combo = None
                if combo_sku[-3:].isdigit():
                    medida_num = combo_sku[-3:]
                    if medida_num in ['080', '090', '100', '140', '150', '160', '180', '200']:
                        medida_combo = medida_num.lstrip('0') + 'x190'  # Ej: 140x190
                
                # Filtro por medida (si aplica)
                if filtro_medida != 'TODAS':
                    if not medida_combo or not medida_combo.startswith(filtro_medida.split('x')[0]):
                        continue
                
                # Obtener componentes
                cursor.execute('''
                    SELECT pb.sku, pb.stock_actual, COALESCE(pb.stock_full, 0) as stock_full,
                           c.cantidad_necesaria
                    FROM componentes c
                    JOIN productos_base pb ON c.producto_base_id = pb.id
                    WHERE c.producto_compuesto_id = %s
                ''', (combo_id,))
                
                componentes = cursor.fetchall()
                
                if not componentes:
                    continue
                
                # Calcular stock disponible
                stock_disponible_combo = float('inf')
                
                for comp in componentes:
                    comp_sku = comp['sku']
                    comp_stock_fisico = comp['stock_actual'] + comp['stock_full']
                    comp_vendido = ventas_dict.get(comp_sku, 0)
                    comp_stock_disponible = comp_stock_fisico - comp_vendido
                    combos_posibles = comp_stock_disponible // comp['cantidad_necesaria']
                    stock_disponible_combo = min(stock_disponible_combo, combos_posibles)
                
                stock_disponible_combo = int(stock_disponible_combo) if stock_disponible_combo != float('inf') else 0
                
                # Estado
                if stock_disponible_combo <= 0:
                    estado_stock = 'SIN_STOCK'
                elif stock_disponible_combo <= 2:
                    estado_stock = 'POCO_STOCK'
                else:
                    estado_stock = 'DISPONIBLE'
                
                # Filtro por estado
                if filtro_estado != 'TODOS' and estado_stock != filtro_estado:
                    continue
                
                productos.append({
                    'sku': combo_sku,
                    'nombre': combo_nombre,
                    'tipo': 'sommier',
                    'medida': medida_combo,
                    'modelo': None,
                    'stock_fisico': '-',
                    'stock_disponible': stock_disponible_combo,
                    'estado_stock': estado_stock,
                    'es_combo': True
                })
        
        # Ordenar
        productos.sort(key=lambda x: (x['tipo'], x['nombre'], x['medida'] or ''))
        
        cursor.close()
        conn.close()
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    
    return render_template('stock.html', 
                         productos=productos, 
                         filtro_estado=filtro_estado,
                         filtro_tipo=filtro_tipo,
                         filtro_modelo=filtro_modelo,
                         filtro_medida=filtro_medida,
                         buscar=buscar)


@app.route('/ventas/historicas')
@login_required
def ventas_historicas():
    """Lista de ventas históricas (entregadas y canceladas) con filtros"""
    try:
        # ========================================
        # OBTENER FILTROS
        # ========================================
        filtro_buscar = request.args.get('buscar', '').strip()
        filtro_estado = request.args.get('estado', '')  # '' = Todos, 'entregada', 'cancelada'
        filtro_periodo = request.args.get('periodo', 'todo')  # 'hoy', 'semana', 'mes', 'trimestre', 'todo'
        filtro_metodo_envio = request.args.get('metodo_envio', '')
        filtro_zona = request.args.get('zona', '')
        filtro_canal = request.args.get('canal', '')
        
        # ========================================
        # CONSTRUIR QUERY CON FILTROS
        # ========================================
        query = '''
            SELECT 
                id, numero_venta, fecha_venta, fecha_entrega, canal, mla_code,
                nombre_cliente, telefono_cliente,
                tipo_entrega, metodo_envio, ubicacion_despacho,
                zona_envio, direccion_entrega, costo_flete,
metodo_pago, importe_total, importe_abonado,
                pago_mercadopago, pago_efectivo,
                estado_entrega, estado_pago, notas,
    factura_generada, factura_fecha_generacion
            FROM ventas
            WHERE estado_entrega IN ('entregada', 'cancelada')
        '''
        params = []
        
        # Filtro: Estado (entregada, cancelada, o ambas)
        if filtro_estado:
            query += ' AND estado_entrega = %s'
            params.append(filtro_estado)
        
        # Filtro: Período (por fecha de entrega)
        if filtro_periodo == 'hoy':
            query += ' AND DATE(COALESCE(fecha_entrega, fecha_modificacion)) = CURDATE()'
        elif filtro_periodo == 'semana':
            query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 7 DAY)'
        elif filtro_periodo == 'mes':
            query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 30 DAY)'
        elif filtro_periodo == 'trimestre':
            query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 90 DAY)'
        
        # Filtro: Búsqueda de texto
        if filtro_buscar:
            query += '''
                AND (
                    mla_code LIKE %s 
                    OR nombre_cliente LIKE %s
                    OR id IN (
                        SELECT venta_id FROM items_venta WHERE sku LIKE %s
                    )
                )
            '''
            busqueda = f'%{filtro_buscar}%'
            params.extend([busqueda, busqueda, busqueda])
        
        # Filtro: Método de envío
        if filtro_metodo_envio:
            query += ' AND metodo_envio = %s'
            params.append(filtro_metodo_envio)
        
        # Filtro: Zona
        if filtro_zona:
            query += ' AND zona_envio = %s'
            params.append(filtro_zona)
        
        # Filtro: Canal
        if filtro_canal:
            query += ' AND canal = %s'
            params.append(filtro_canal)
        
        # Ordenar: Más recientes arriba (por fecha de entrega, o fecha_modificacion si no hay fecha_entrega)
        query += " ORDER BY CASE WHEN metodo_envio = 'Turbo' THEN 0 ELSE 1 END, fecha_venta DESC, id DESC"
        
        # Ejecutar query
        ventas = query_db(query, tuple(params) if params else None)
        
        # ========================================
        # OBTENER ITEMS DE CADA VENTA
        # ========================================
        for venta in ventas:
            items = query_db('''
                SELECT 
                    iv.sku, 
                    iv.cantidad, 
                    iv.precio_unitario,
                    COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
                ORDER BY iv.id
            ''', (venta['id'],))
            venta['items'] = items
            if venta.get('fecha_venta') and hasattr(venta['fecha_venta'], 'strftime'):
                venta['hora_venta_str'] = venta['fecha_venta'].strftime('%H:%M')
            else:
                venta['hora_venta_str'] = ''
        
        # ========================================
        # CONTAR ENTREGADAS Y CANCELADAS
        # ========================================
        stats_query = '''
            SELECT 
                estado_entrega,
                COUNT(*) as total
            FROM ventas
            WHERE estado_entrega IN ('entregada', 'cancelada')
        '''
        if filtro_periodo != 'todo':
            if filtro_periodo == 'hoy':
                stats_query += ' AND DATE(COALESCE(fecha_entrega, fecha_modificacion)) = CURDATE()'
            elif filtro_periodo == 'semana':
                stats_query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 7 DAY)'
            elif filtro_periodo == 'mes':
                stats_query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 30 DAY)'
            elif filtro_periodo == 'trimestre':
                stats_query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 90 DAY)'
        
        stats_query += ' GROUP BY estado_entrega'
        stats = query_db(stats_query)
        
        entregadas = 0
        canceladas = 0
        for stat in stats:
            if stat['estado_entrega'] == 'entregada':
                entregadas = stat['total']
            elif stat['estado_entrega'] == 'cancelada':
                canceladas = stat['total']
        
        return render_template('ventas_historicas.html', 
                             ventas=ventas,
                             entregadas=entregadas,
                             canceladas=canceladas,
                             filtro_buscar=filtro_buscar,
                             filtro_estado=filtro_estado,
                             filtro_periodo=filtro_periodo,
                             filtro_metodo_envio=filtro_metodo_envio,
                             filtro_zona=filtro_zona,
                             filtro_canal=filtro_canal)
        
    except Exception as e:
        flash(f'Error al cargar ventas históricas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('index'))

@app.route('/ventas/historicas/<int:venta_id>/volver_activas', methods=['POST'])
@login_required
def historicas_volver_activas(venta_id):
    """Volver venta histórica (entregada o cancelada) a ventas activas"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_historicas'))
        
        estado_anterior = venta['estado_entrega']
        numero_venta = venta['numero_venta']
        
        # Verificar que sea histórica (entregada o cancelada)
        if estado_anterior not in ['entregada', 'cancelada']:
            flash(f'La venta {numero_venta} no es histórica', 'warning')
            return redirect(url_for('ventas_historicas'))
        
        # ========================================
        # LÓGICA SEGÚN ESTADO ANTERIOR
        # ========================================
        
        if estado_anterior == 'cancelada':
            # CANCELADA → ACTIVA
            # NO devolver stock (porque nunca se descontó)
            cursor.execute('''
                UPDATE ventas 
                SET estado_entrega = 'pendiente',
                    fecha_modificacion = NOW()
                WHERE id = %s
            ''', (venta_id,))
            
            mensaje = f'✅ Venta {numero_venta} devuelta a Ventas Activas (sin cambios en stock)'
        
        elif estado_anterior == 'entregada':
            # ENTREGADA → ACTIVA
            # SÍ devolver stock (porque se descontó cuando se entregó)
            cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
            items = cursor.fetchall()
            
            # Devolver stock de cada item
            for item in items:
                devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
            
            # Cambiar estado a pendiente
            cursor.execute('''
                UPDATE ventas 
                SET estado_entrega = 'pendiente',
                    fecha_modificacion = NOW()
                WHERE id = %s
            ''', (venta_id,))
            
            mensaje = f'✅ Venta {numero_venta} devuelta a Ventas Activas. Stock restaurado.'
        
        conn.commit()
        flash(mensaje, 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al volver a activas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_historicas'))

@app.route('/ventas/historicas/volver-activas-multiple', methods=['POST'])
@login_required
def historicas_volver_activas_multiple():
    """
    Volver múltiples ventas históricas a ventas activas
    Usa la misma lógica que historicas_volver_activas() individual
    Mantiene filtros después de la acción
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        ventas_con_stock_devuelto = 0
        ventas_sin_stock = 0
        
        for venta_id in venta_ids:
            try:
                # Obtener venta
                cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta:
                    continue
                
                estado_anterior = venta['estado_entrega']
                
                # Verificar que sea histórica
                if estado_anterior not in ['entregada', 'cancelada']:
                    continue
                
                # ========================================
                # LÓGICA SEGÚN ESTADO ANTERIOR
                # (igual que historicas_volver_activas)
                # ========================================
                
                if estado_anterior == 'cancelada':
                    # CANCELADA → ACTIVA
                    # NO devolver stock (porque nunca se descontó)
                    cursor.execute('''
                        UPDATE ventas 
                        SET estado_entrega = 'pendiente',
                            fecha_modificacion = NOW()
                        WHERE id = %s
                    ''', (venta_id,))
                    
                    ventas_procesadas += 1
                    ventas_sin_stock += 1
                
                elif estado_anterior == 'entregada':
                    # ENTREGADA → ACTIVA
                    # SÍ devolver stock (porque se descontó cuando se entregó)
                    cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
                    items = cursor.fetchall()
                    
                    # Devolver stock de cada item usando la función existente
                    for item in items:
                        devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
                    
                    # Cambiar estado a pendiente
                    cursor.execute('''
                        UPDATE ventas 
                        SET estado_entrega = 'pendiente',
                            fecha_modificacion = NOW()
                        WHERE id = %s
                    ''', (venta_id,))
                    
                    ventas_procesadas += 1
                    ventas_con_stock_devuelto += 1
            
            except Exception as e:
                # Si falla una venta, continuar con las demás
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # ========================================
        # MENSAJE DE ÉXITO
        # ========================================
        if ventas_procesadas == 0:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        else:
            mensaje = f'✅ {ventas_procesadas} venta(s) devueltas a Ventas Activas.'
            
            if ventas_con_stock_devuelto > 0:
                mensaje += f' Stock restaurado en {ventas_con_stock_devuelto} venta(s).'
            
            if ventas_sin_stock > 0:
                mensaje += f' {ventas_sin_stock} cancelada(s) sin cambios en stock.'
            
            flash(mensaje, 'success')
        
        # ========================================
        # ✅ MANTENER FILTROS
        # ========================================
        filtros = {}
        if request.form.get('buscar'):
            filtros['buscar'] = request.form.get('buscar')
        if request.form.get('estado'):
            filtros['estado'] = request.form.get('estado')
        if request.form.get('periodo'):
            filtros['periodo'] = request.form.get('periodo')
        if request.form.get('metodo_envio'):
            filtros['metodo_envio'] = request.form.get('metodo_envio')
        if request.form.get('zona'):
            filtros['zona'] = request.form.get('zona')
        if request.form.get('canal'):
            filtros['canal'] = request.form.get('canal')
        
        return redirect(url_for('ventas_historicas', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))

# ============================================================================
# MAPEO DE PROVINCIA A CÓDIGO AFIP
# ============================================================================

def provincia_a_codigo(provincia_str):
    """Convierte nombre de provincia a código AFIP, sin importar tildes/mayúsculas"""
    import unicodedata

    def normalizar(s):
        if not s:
            return ''
        s = str(s).upper().strip()
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return s

    MAPA = {
        'CDAD. AUTONOMA DE BS AS': 1,
        'CIUDAD AUTONOMA DE BUENOS AIRES': 1,
        'CABA': 1,
        'CAPITAL FEDERAL': 1,
        'C.A.B.A.': 1,
        'CIUDAD DE BUENOS AIRES': 1,
        'BUENOS AIRES': 2,
        'PROVINCIA DE BUENOS AIRES': 2,
        'PBA': 2,
        'CATAMARCA': 3,
        'CORDOBA': 4,
        'CORRIENTES': 5,
        'ENTRE RIOS': 6,
        'JUJUY': 7,
        'MENDOZA': 8,
        'LA RIOJA': 9,
        'SALTA': 10,
        'SAN JUAN': 11,
        'SAN LUIS': 12,
        'SANTA FE': 13,
        'SANTIAGO DEL ESTERO': 14,
        'TUCUMAN': 15,
        'CHACO': 16,
        'CHUBUT': 17,
        'FORMOSA': 18,
        'MISIONES': 19,
        'NEUQUEN': 20,
        'LA PAMPA': 21,
        'RIO NEGRO': 22,
        'SANTA CRUZ': 23,
        'TIERRA DEL FUEGO': 24,
        'EXTERIOR': 25,
    }

    norm = normalizar(provincia_str)
    # Búsqueda exacta primero
    if norm in MAPA:
        return MAPA[norm]
    # Búsqueda parcial (ej: "Buenos Aires" puede venir como zona_envio)
    for key, codigo in MAPA.items():
        if key in norm or norm in key:
            return codigo
    # Default: Capital Federal
    return 1


@app.route('/ventas/historicas/<int:venta_id>/generar-factura-excel')
@login_required
def generar_factura_excel(venta_id):
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    try:
        venta = query_one('SELECT * FROM ventas WHERE id = %s', (venta_id,))

        if not venta:
            flash('❌ Venta no encontrada', 'error')
            return redirect(url_for('ventas_historicas'))

        items = list(query_db('''
            SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
            FROM items_venta iv
            LEFT JOIN productos_base pb ON iv.sku = pb.sku
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            WHERE iv.venta_id = %s
        ''', (venta_id,)))

        costo_flete = float(venta.get('costo_flete') or 0)

        incluir_flete_param = request.args.get('incluir_flete', 'false').lower()
        incluir_flete = (incluir_flete_param == 'true' and costo_flete > 0)

        cant_slots = len(items) + (1 if incluir_flete else 0)

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación"

        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia', 'rubro']
        for idx in range(1, cant_slots + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        headers.append('importe total')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        def mapear_iva(valor):
            m = {
                'IVA Exento': 'Exento',
                'IVA Responsable Inscripto': 'Responsable Inscripto',
                'Monotributo': 'Responsable Monotributo',
            }
            return m.get(valor, valor or 'Consumidor Final')

        row_data = []

        if venta.get('mla_code'):
            id_venta = venta['mla_code']
        else:
            id_venta = venta.get('factura_business_name') or venta['nombre_cliente']
        row_data.append(id_venta)

        row_data.append(mapear_iva(venta.get('factura_taxpayer_type')))

        row_data.append(venta.get('factura_business_name') or venta['nombre_cliente'])

        if venta.get('factura_doc_number'):
            dni = str(venta['factura_doc_number'])
        elif venta.get('dni_cliente'):
            dni = str(venta['dni_cliente'])
        else:
            dni = '99999999'
        row_data.append(dni)

        if venta.get('factura_street'):
            direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
        else:
            direccion = venta.get('direccion_entrega') or ''
        row_data.append(direccion)

        provincia_str = (venta.get('factura_state') or venta.get('provincia_cliente') or
                         venta.get('zona_envio') or 'Capital Federal')
        row_data.append(provincia_a_codigo(provincia_str))

        rubro = 'F' if len(dni.replace('.', '').replace('-', '').strip()) == 11 else 'R'
        row_data.append(rubro)

        for item in items:
            row_data.append(item['sku'])
            row_data.append(int(item['cantidad']))
            row_data.append(float(item['precio_unitario']))

        if incluir_flete:
            row_data.append('FLETE')
            row_data.append(1)
            row_data.append(costo_flete)

        total_excel = float(venta['importe_total']) + (costo_flete if incluir_flete else 0)
        row_data.append(total_excel)

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=2, column=col_idx, value=value)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        nombre_archivo = f"factura_{venta['numero_venta'].replace('/', '-')}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        execute_db('''
            UPDATE ventas SET factura_generada = TRUE, factura_fecha_generacion = NOW()
            WHERE id = %s
        ''', (venta_id,))

        return response

    except Exception as e:
        flash(f'❌ Error al generar factura: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))


# ============================================================================
# FUNCIÓN 2: FACTURA MÚLTIPLE
# ============================================================================

@app.route('/ventas/historicas/facturar-multiple-excel')
@login_required
def facturar_multiple_excel():
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    try:
        ids_str = request.args.get('ids', '')
        if not ids_str:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_historicas'))

        venta_ids = [int(id) for id in ids_str.split(',') if id.strip()]

        if not venta_ids:
            flash('❌ No se seleccionaron ventas válidas', 'error')
            return redirect(url_for('ventas_historicas'))

        ids_con_flete_str = request.args.get('ids_con_flete', '')
        ids_con_flete = set()
        if ids_con_flete_str:
            ids_con_flete = {int(i) for i in ids_con_flete_str.split(',') if i.strip()}

        placeholders = ', '.join(['%s'] * len(venta_ids))
        ventas = query_db(f'SELECT * FROM ventas WHERE id IN ({placeholders}) ORDER BY id DESC', tuple(venta_ids))

        if not ventas:
            flash('❌ No se encontraron ventas', 'error')
            return redirect(url_for('ventas_historicas'))

        ventas_preparadas = []
        max_slots = 0

        for venta in ventas:
            items = list(query_db('''
                SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
            ''', (venta['id'],)))

            costo_flete = float(venta.get('costo_flete') or 0)
            incluir_flete = (venta['id'] in ids_con_flete and costo_flete > 0)

            cant_slots = len(items) + (1 if incluir_flete else 0)
            if cant_slots > max_slots:
                max_slots = cant_slots

            ventas_preparadas.append({
                'venta': venta,
                'items': items,
                'costo_flete': costo_flete,
                'incluir_flete': incluir_flete,
                'cant_slots': cant_slots
            })

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación Múltiple"

        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia', 'rubro']
        for idx in range(1, max_slots + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        headers.append('importe total')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row = 2

        def mapear_iva(valor):
            m = {
                'IVA Exento': 'Exento',
                'IVA Responsable Inscripto': 'Responsable Inscripto',
                'Monotributo': 'Responsable Monotributo',
            }
            return m.get(valor, valor or 'Consumidor Final')

        for venta_prep in ventas_preparadas:
            venta = venta_prep['venta']
            items = venta_prep['items']
            costo_flete = venta_prep['costo_flete']
            incluir_flete = venta_prep['incluir_flete']
            cant_slots = venta_prep['cant_slots']

            row_data = []

            if venta.get('mla_code'):
                id_venta = venta['mla_code']
            else:
                id_venta = venta.get('factura_business_name') or venta['nombre_cliente']
            row_data.append(id_venta)

            row_data.append(mapear_iva(venta.get('factura_taxpayer_type')))
            row_data.append(venta.get('factura_business_name') or venta['nombre_cliente'])

            if venta.get('factura_doc_number'):
                dni = str(venta['factura_doc_number'])
            elif venta.get('dni_cliente'):
                dni = str(venta['dni_cliente'])
            else:
                dni = '99999999'
            row_data.append(dni)

            if venta.get('factura_street'):
                direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
            else:
                direccion = venta.get('direccion_entrega') or ''
            row_data.append(direccion)

            provincia_str = (venta.get('factura_state') or venta.get('provincia_cliente') or
                             venta.get('zona_envio') or 'Capital Federal')
            row_data.append(provincia_a_codigo(provincia_str))

            rubro = 'F' if len(dni.replace('.', '').replace('-', '').strip()) == 11 else 'R'
            row_data.append(rubro)

            for item in items:
                row_data.append(item['sku'])
                row_data.append(int(item['cantidad']))
                row_data.append(float(item['precio_unitario']))

            if incluir_flete:
                row_data.append('FLETE')
                row_data.append(1)
                row_data.append(costo_flete)

            for _ in range(max_slots - cant_slots):
                row_data.extend(['', '', ''])

            total_excel = float(venta['importe_total']) + (costo_flete if incluir_flete else 0)
            row_data.append(total_excel)

            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_idx, value=value)

            current_row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"facturas_multiple_{fecha_str}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        placeholders = ', '.join(['%s'] * len(venta_ids))
        execute_db(f'''
            UPDATE ventas SET factura_generada = TRUE, factura_fecha_generacion = NOW()
            WHERE id IN ({placeholders})
        ''', tuple(venta_ids))

        return response

    except Exception as e:
        flash(f'❌ Error al generar facturas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))




# ============================================================================
# VENTAS HISTÓRICAS - VOLVER A ACTIVAS
# Agregar en app.py después de la ruta de ventas_historicas()
# ============================================================================



@app.route('/ventas/historicas/facturar-multiple')
@login_required
def facturar_multiple():
    """
    Generar UN SOLO archivo .txt con TODAS las ventas seleccionadas
    """
    from flask import make_response
    from datetime import datetime
    
    try:
        # Obtener IDs
        ids_str = request.args.get('ids', '')
        if not ids_str:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        venta_ids = [int(id) for id in ids_str.split(',') if id.strip()]
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas válidas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        # ============================================
        # OBTENER TODAS LAS VENTAS
        # ============================================
        placeholders = ', '.join(['%s'] * len(venta_ids))
        query = f'''
            SELECT * FROM ventas 
            WHERE id IN ({placeholders})
            ORDER BY fecha_venta DESC
        '''
        ventas = query_db(query, tuple(venta_ids))
        
        if not ventas:
            flash('❌ No se encontraron ventas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        # ============================================
        # GENERAR CONTENIDO DEL ARCHIVO TXT
        # ============================================
        
        lineas = []
        lineas.append("="*80)
        lineas.append("FACTURACIÓN MÚLTIPLE - DATOS PARA FACTURAR")
        lineas.append("="*80)
        lineas.append(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        lineas.append(f"Total de ventas: {len(ventas)}")
        lineas.append("="*80)
        lineas.append("")
        
        total_general = 0
        
        # Procesar cada venta
        for idx, venta in enumerate(ventas, 1):
            lineas.append("")
            lineas.append("="*80)
            lineas.append(f"VENTA #{idx} - {venta['numero_venta']}")
            lineas.append("="*80)
            lineas.append("")
            
            # DATOS DE LA VENTA
            lineas.append(f"  Fecha: {venta['fecha_venta'].strftime('%d/%m/%Y')}")
            if venta.get('mla_code'):
                lineas.append(f"  ML Code: {venta['mla_code']}")
            lineas.append("")
            
            # DATOS DEL COMPRADOR
            lineas.append("-"*80)
            lineas.append("DATOS DEL COMPRADOR:")
            lineas.append("-"*80)
            
            if venta.get('factura_business_name'):
                # Tiene datos de facturación
                lineas.append(f"  Razón Social: {venta['factura_business_name']}")
                lineas.append(f"  {venta.get('factura_doc_type', 'Documento')}: {venta.get('factura_doc_number', 'N/A')}")
                lineas.append(f"  Condición IVA: {venta.get('factura_taxpayer_type', 'N/A')}")
                lineas.append("")
                lineas.append("  DOMICILIO FISCAL:")
                lineas.append(f"    Calle: {venta.get('factura_street', 'N/A')}")
                lineas.append(f"    Ciudad: {venta.get('factura_city', 'N/A')}")
                lineas.append(f"    Provincia: {venta.get('factura_state', 'N/A')}")
                lineas.append(f"    CP: {venta.get('factura_zip_code', 'N/A')}")
            else:
                # Consumidor Final
                lineas.append(f"  Nombre: {venta['nombre_cliente']}")
                if venta.get('telefono_cliente'):
                    lineas.append(f"  Teléfono: {venta['telefono_cliente']}")
                lineas.append(f"  Condición IVA: Consumidor Final")
                
                if venta.get('direccion_entrega'):
                    lineas.append("")
                    lineas.append("  DIRECCIÓN DE ENTREGA:")
                    lineas.append(f"    {venta['direccion_entrega']}")
                    if venta.get('zona_envio'):
                        lineas.append(f"    Zona: {venta['zona_envio']}")
            
            lineas.append("")
            
            # PRODUCTOS
            items = query_db('''
                SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
            ''', (venta['id'],))
            
            lineas.append("-"*80)
            lineas.append("PRODUCTOS:")
            lineas.append("-"*80)
            lineas.append("")
            lineas.append(f"{'Cant':<6} {'SKU':<15} {'Descripción':<35} {'P.Unit':<12} {'Subtotal':<12}")
            lineas.append("-"*80)
            
            total_items = 0
            for item in items:
                cant = item['cantidad']
                sku = item['sku']
                nombre = item['nombre_producto']
                precio = item['precio_unitario']
                subtotal = cant * precio
                
                lineas.append(f"{cant:<6} {sku:<15} {nombre:<35} ${precio:<11,.2f} ${subtotal:<11,.2f}")
                total_items += subtotal
            
            lineas.append("-"*80)
            lineas.append("")
            
            # TOTALES
            lineas.append("TOTALES:")
            lineas.append(f"  Subtotal Productos: ${total_items:,.2f}")
            
            if venta.get('costo_flete') and venta['costo_flete'] > 0:
                lineas.append(f"  Costo de Envío: ${venta['costo_flete']:,.2f}")
                lineas.append(f"  TOTAL: ${venta['importe_total']:,.2f}")
            else:
                lineas.append(f"  TOTAL: ${venta['importe_total']:,.2f}")
            
            lineas.append("")
            lineas.append("-"*80)
            
            # MÉTODO DE PAGO
            lineas.append("MÉTODO DE PAGO:")
            lineas.append(f"  {venta.get('metodo_pago', 'N/A')}")
            if venta.get('pago_mercadopago') and venta['pago_mercadopago'] > 0:
                lineas.append(f"    Mercadopago: ${venta['pago_mercadopago']:,.2f}")
            if venta.get('pago_efectivo') and venta['pago_efectivo'] > 0:
                lineas.append(f"    Efectivo: ${venta['pago_efectivo']:,.2f}")
            
            lineas.append("")
            
            total_general += venta['importe_total']
        
        # RESUMEN FINAL
        lineas.append("")
        lineas.append("="*80)
        lineas.append("RESUMEN GENERAL")
        lineas.append("="*80)
        lineas.append(f"Total de ventas facturadas: {len(ventas)}")
        lineas.append(f"TOTAL GENERAL: ${total_general:,.2f}")
        lineas.append("="*80)
        
        # ============================================
        # CREAR RESPUESTA CON ARCHIVO
        # ============================================
        
        contenido = "\n".join(lineas)
        
        response = make_response(contenido)
        response.headers['Content-Type'] = 'text/plain; charset=utf-8'
        
        # Nombre del archivo
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"facturas_multiple_{fecha_str}.txt"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        
        # ============================================
        # MARCAR TODAS COMO GENERADAS
        # ============================================
        
        placeholders = ', '.join(['%s'] * len(venta_ids))
        execute_db(f'''
            UPDATE ventas 
            SET factura_generada = TRUE,
                factura_fecha_generacion = NOW()
            WHERE id IN ({placeholders})
        ''', tuple(venta_ids))
        
        return response
        
    except Exception as e:
        flash(f'❌ Error al generar facturas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))




# ============================================================================
# FORMULARIOS VISUALES - SIN SQL
# ============================================================================

# ============================================================================
# API PARA CARRITO
# ============================================================================

@app.route('/api/productos')
@login_required
def api_productos():
    """API para el buscador de productos en templates"""
    from flask import jsonify
    
    # Productos base
    productos_base = query_db('SELECT sku, nombre, tipo, stock_actual FROM productos_base ORDER BY nombre')
    
    # Combos - tabla: productos_compuestos
    combos = query_db('SELECT sku, nombre FROM productos_compuestos ORDER BY nombre')
    
    todos = []
    
    for p in productos_base:
        todos.append({
            'sku': p['sku'],
            'nombre': p['nombre'],
            'tipo': p['tipo'],
            'stock_actual': p['stock_actual']
        })
    
    for c in combos:
        todos.append({
            'sku': c['sku'],
            'nombre': c['nombre'],
            'tipo': 'combo',
            'stock_actual': 0
        })
    
    return jsonify(todos)

@app.route('/cargar-stock', methods=['GET', 'POST'])
@login_required
def cargar_stock():
    """Formulario para cargar/agregar stock de productos"""
    from flask import jsonify
    
    if request.method == 'POST' and request.is_json:
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            data = request.get_json()
            items = data.get('items', [])
            
            for item in items:
                sku = item['sku']
                cantidad = item['cantidad']
                ubicacion = item['ubicacion']
                motivo = item.get('motivo', 'Carga de stock')
                
                # Obtener datos actuales del producto
                cursor.execute('SELECT nombre, stock_actual, COALESCE(stock_full, 0) as stock_full FROM productos_base WHERE sku = %s', (sku,))
                prod = cursor.fetchone()
                
                if not prod:
                    continue
                
                nombre_producto = prod['nombre']
                
                # Actualizar stock y registrar movimiento
                if ubicacion == 'stock_actual':
                    stock_anterior = prod['stock_actual']
                    stock_nuevo = stock_anterior + cantidad
                    
                    cursor.execute('UPDATE productos_base SET stock_actual = stock_actual + %s WHERE sku = %s', (cantidad, sku))
                    
                    # Registrar en historial
                    cursor.execute('''
                        INSERT INTO movimientos_stock 
                        (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ''', (sku, nombre_producto, 'carga', cantidad, stock_anterior, stock_nuevo, motivo))
                    
                else:  # stock_full
                    stock_anterior = prod['stock_full']
                    stock_nuevo = stock_anterior + cantidad
                    
                    cursor.execute('UPDATE productos_base SET stock_full = stock_full + %s WHERE sku = %s', (cantidad, sku))
                    
                    # Registrar en historial
                    cursor.execute('''
                        INSERT INTO movimientos_stock 
                        (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ''', (sku, nombre_producto, 'carga', cantidad, stock_anterior, stock_nuevo, motivo))
            
            conn.commit()
            cursor.close()
            conn.close()

            # Disparar actualización ML en background con progreso
            import threading
            skus_cargados = {item['sku'] for item in items}
            def _update_ml_bg():
                try:
                    actualizar_publicaciones_ml_con_progreso(skus_cargados)
                except Exception as e_ml:
                    print(f"[AUTO-ML] Error actualizando ML tras carga stock: {e_ml}")
            threading.Thread(target=_update_ml_bg, daemon=True).start()

            return jsonify({'success': True, 'message': f'Stock cargado: {len(items)} productos. Actualizando publicaciones ML en background...'})
        except Exception as e:
            conn.rollback()
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # Si es GET, mostrar formulario
    productos = []
    try:
        productos = query_db('''
            SELECT * FROM productos_base 
            ORDER BY 
                CASE tipo
                    WHEN 'colchon' THEN 1
                    WHEN 'base' THEN 2
                    WHEN 'almohada' THEN 3
                END,
                nombre
        ''')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return render_template('cargar_stock.html', productos=productos)

@app.route('/cargar-stock/guardar', methods=['POST'])
@login_required
def guardar_stock():
    """Agregar stock (SUMA al stock actual) y registrar movimientos"""
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # Procesar cada producto
        productos_cargados = 0
        skus_cargados = set()
        for key, value in request.form.items():
            if key.startswith('agregar_'):
                sku = key.replace('agregar_', '')
                cantidad_agregar = int(value) if value else 0
                
                # Solo procesar si hay cantidad a agregar
                if cantidad_agregar > 0:
                    # Obtener stock ACTUAL de la base de datos
                    cursor.execute('SELECT stock_actual, nombre FROM productos_base WHERE sku = %s', (sku,))
                    resultado = cursor.fetchone()
                    
                    if resultado:
                        stock_anterior = resultado['stock_actual']
                        nombre_producto = resultado['nombre']
                        stock_nuevo = stock_anterior + cantidad_agregar
                        
                        # Actualizar stock (SUMAR)
                        cursor.execute('''
                            UPDATE productos_base 
                            SET stock_actual = stock_actual + %s 
                            WHERE sku = %s
                        ''', (cantidad_agregar, sku))
                        
                        # Registrar movimiento
                        cursor.execute('''
                            INSERT INTO movimientos_stock 
                            (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku, nombre_producto, 'carga', cantidad_agregar, stock_anterior, stock_nuevo, 'Carga de stock nuevo'))

                        log_evento('INFO', 'stock', 'carga_stock',
                            f"Carga manual: {nombre_producto} ({sku}) +{cantidad_agregar} unidades. Stock: {stock_anterior} → {stock_nuevo}",
                            sku=sku, usuario=current_user.username if current_user.is_authenticated else 'Sistema')

                        productos_cargados += 1
                        skus_cargados.add(sku)
        
        conn.commit()
        
        if productos_cargados > 0:
            flash(f'✅ Stock cargado correctamente ({productos_cargados} productos)', 'success')
            # Actualizar publicaciones ML con los SKUs base cargados
            try:
                actualizar_publicaciones_ml(skus_cargados)
            except Exception as e_ml:
                print(f"[AUTO-ML] Error actualizando ML tras carga stock: {e_ml}")
        else:
            flash('ℹ️ No se agregó stock. Ingresá cantidades mayores a 0.', 'info')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('cargar_stock'))


@app.route('/bajar-stock')
@login_required
def bajar_stock():
    """Formulario para dar de baja stock"""
    productos = []
    try:
        productos = query_db('''
            SELECT sku, nombre, tipo, stock_actual
            FROM productos_base 
            WHERE stock_actual > 0 
            AND tipo IN ('colchon', 'base', 'almohada')
            ORDER BY 
                CASE tipo
                    WHEN 'colchon' THEN 1
                    WHEN 'base' THEN 2
                    WHEN 'almohada' THEN 3
                END,
                nombre
        ''')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return render_template('bajar_stock.html', productos=productos)


@app.route('/bajar-stock/guardar', methods=['POST'])
@login_required
def bajar_stock_guardar():
    """Guardar bajas de stock - acepta form o JSON"""
    from flask import jsonify as _jsonify
    es_json = request.is_json
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        if es_json:
            data = request.get_json()
            motivo = data.get('motivo', 'Baja de stock manual')
            bajas = [(item['cantidad'], item['sku']) for item in data.get('items', []) if item.get('cantidad', 0) > 0]
        else:
            motivo = request.form.get('motivo', 'Baja de stock manual')
            bajas = []
            for key, value in request.form.items():
                if key.startswith('baja_'):
                    sku = key.replace('baja_', '')
                    cantidad_baja = int(value) if value else 0
                    if cantidad_baja > 0:
                        bajas.append((cantidad_baja, sku))

        for cantidad_baja, sku in bajas:
            cursor.execute('SELECT stock_actual, nombre FROM productos_base WHERE sku = %s', (sku,))
            resultado = cursor.fetchone()
            stock_anterior = resultado['stock_actual'] if resultado else 0
            nombre_producto = resultado['nombre'] if resultado else ''
            stock_nuevo = stock_anterior - cantidad_baja
            cursor.execute('UPDATE productos_base SET stock_actual = stock_actual - %s WHERE sku = %s', (cantidad_baja, sku))
            cursor.execute('''INSERT INTO movimientos_stock
                (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (sku, nombre_producto, 'baja', cantidad_baja, stock_anterior, stock_nuevo, motivo))
            log_evento('INFO', 'stock', 'baja_stock',
                f"Baja manual: {nombre_producto} ({sku}) -{cantidad_baja} unidades. Stock: {stock_anterior} → {stock_nuevo}. Motivo: {motivo}",
                sku=sku, usuario=current_user.username if current_user.is_authenticated else 'Sistema')

        conn.commit()

        if bajas:
            import threading
            skus_bajados = {sku for _, sku in bajas}
            def _update_ml_bg():
                try:
                    actualizar_publicaciones_ml_con_progreso(skus_bajados)
                except Exception as e_ml:
                    print(f"[AUTO-ML] Error ML tras baja stock: {e_ml}")
            threading.Thread(target=_update_ml_bg, daemon=True).start()

        if es_json:
            return _jsonify({'success': True, 'message': f'Baja registrada: {len(bajas)} productos. Actualizando ML...'})

        motivo_msg = f' - {motivo}' if motivo else ''
        flash(f'✅ Stock dado de baja correctamente ({len(bajas)} productos){motivo_msg}', 'success')

    except Exception as e:
        conn.rollback()
        if es_json:
            return _jsonify({'success': False, 'error': str(e)}), 500
        flash(f'❌ Error al dar de baja stock: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('bajar_stock'))


@app.route('/historial-stock')
@login_required
def historial_stock():
    """Ver historial de movimientos de stock"""
    from datetime import datetime, timedelta
    
    # Obtener filtros
    filtro_tipo = request.args.get('tipo', '')
    filtro_sku = request.args.get('sku', '').strip()
    filtro_fecha = request.args.get('fecha', 'todo')
    
    try:
        # Query base
        query = """
            SELECT 
                id,
                sku,
                nombre_producto,
                tipo_movimiento,
                cantidad,
                stock_anterior,
                stock_nuevo,
                motivo,
                usuario,
                fecha_movimiento
            FROM movimientos_stock
            WHERE 1=1
        """
        params = []
        
        # Aplicar filtros
        if filtro_tipo:
            query += " AND tipo_movimiento = %s"
            params.append(filtro_tipo)
        
        if filtro_sku:
            query += " AND (sku LIKE %s OR nombre_producto LIKE %s)"
            params.append(f'%{filtro_sku}%')
            params.append(f'%{filtro_sku}%')
        
        # Filtro de fecha
        if filtro_fecha == 'hoy':
            query += " AND DATE(fecha_movimiento) = CURDATE()"
        elif filtro_fecha == 'semana':
            query += " AND fecha_movimiento >= DATE_SUB(NOW(), INTERVAL 7 DAY)"
        elif filtro_fecha == 'mes':
            query += " AND fecha_movimiento >= DATE_SUB(NOW(), INTERVAL 30 DAY)"
        
        query += " ORDER BY fecha_movimiento DESC LIMIT 500"
        
        movimientos = query_db(query, tuple(params) if params else None)
        
        # Estadísticas
        stats_query = """
            SELECT 
                tipo_movimiento,
                COUNT(*) as total_movimientos,
                SUM(cantidad) as total_unidades
            FROM movimientos_stock
            GROUP BY tipo_movimiento
        """
        estadisticas = query_db(stats_query)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        movimientos = []
        estadisticas = []
    
    return render_template('historial_stock.html', 
                         movimientos=movimientos,
                         estadisticas=estadisticas,
                         filtro_tipo=filtro_tipo,
                         filtro_sku=filtro_sku,
                         filtro_fecha=filtro_fecha)


@app.route('/transferir-stock')
@login_required
def transferir_stock():
    """Formulario para transferir stock de Depósito a Full (Compac y Almohadas)"""
    productos_compac = []
    productos_almohadas = []
    
    try:
        # Productos Compac con ubicación _DEP que tengan stock
        productos_compac = query_db('''
            SELECT sku, nombre, stock_actual
            FROM productos_base 
            WHERE sku IN ('CCO80_DEP', 'CCO100_DEP', 'CCO140_DEP', 'CCO160_DEP',
                          'CCP80_DEP', 'CCP100_DEP', 'CCP140_DEP', 'CCP160_DEP')
            AND stock_actual > 0
            ORDER BY sku
        ''')
        
        # Almohadas que tengan stock en Depósito (stock_actual > 0)
        productos_almohadas = query_db('''
            SELECT sku, nombre, stock_actual, stock_full
            FROM productos_base 
            WHERE tipo = 'almohada'
            AND stock_actual > 0
            ORDER BY nombre
        ''')
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return render_template('transferir_stock.html', 
                         productos_compac=productos_compac,
                         productos_almohadas=productos_almohadas)


@app.route('/transferir-stock/guardar', methods=['POST'])
@login_required
def transferir_stock_guardar():
    """Procesar transferencia de stock de Depósito a Full (Compac y Almohadas)"""
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        transferencias = 0
        
        # Procesar COMPAC (usan _DEP y _FULL en SKU)
        for key, value in request.form.items():
            if key.startswith('transferir_compac_'):
                sku_dep = key.replace('transferir_compac_', '')
                cantidad_transferir = int(value) if value else 0
                
                if cantidad_transferir > 0:
                    sku_full = sku_dep.replace('_DEP', '_FULL')
                    
                    cursor.execute('SELECT stock_actual, nombre FROM productos_base WHERE sku = %s', (sku_dep,))
                    dep = cursor.fetchone()
                    cursor.execute('SELECT stock_actual, nombre FROM productos_base WHERE sku = %s', (sku_full,))
                    full = cursor.fetchone()
                    
                    if dep and full:
                        stock_dep_anterior = dep['stock_actual']
                        stock_full_anterior = full['stock_actual']
                        nombre_base = dep['nombre'].replace(' (Depósito)', '')
                        
                        if cantidad_transferir > stock_dep_anterior:
                            flash(f'❌ {sku_dep}: No hay suficiente stock (disponible: {stock_dep_anterior})', 'error')
                            continue
                        
                        # Descontar de depósito
                        cursor.execute('''
                            UPDATE productos_base 
                            SET stock_actual = stock_actual - %s 
                            WHERE sku = %s
                        ''', (cantidad_transferir, sku_dep))
                        
                        # Sumar a Full
                        cursor.execute('''
                            UPDATE productos_base 
                            SET stock_actual = stock_actual + %s 
                            WHERE sku = %s
                        ''', (cantidad_transferir, sku_full))
                        
                        motivo_baja = 'Transferencia a Full (' + sku_full + ')'
                        motivo_carga = 'Transferencia desde Deposito (' + sku_dep + ')'
                        
                        cursor.execute('''
                            INSERT INTO movimientos_stock 
                            (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku_dep, nombre_base, 'baja', cantidad_transferir, 
                              stock_dep_anterior, stock_dep_anterior - cantidad_transferir, 
                              motivo_baja))
                        
                        cursor.execute('''
                            INSERT INTO movimientos_stock 
                            (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku_full, nombre_base, 'carga', cantidad_transferir, 
                              stock_full_anterior, stock_full_anterior + cantidad_transferir, 
                              motivo_carga))
                        
                        transferencias += 1
        
        # Procesar ALMOHADAS (usan stock_actual y stock_full en misma fila)
        for key, value in request.form.items():
            if key.startswith('transferir_alm_'):
                sku_alm = key.replace('transferir_alm_', '')
                cantidad_transferir = int(value) if value else 0
                
                if cantidad_transferir > 0:
                    cursor.execute('SELECT stock_actual, stock_full, nombre FROM productos_base WHERE sku = %s', (sku_alm,))
                    alm = cursor.fetchone()
                    
                    if alm:
                        stock_dep_anterior = alm['stock_actual']
                        stock_full_anterior = alm.get('stock_full', 0)
                        nombre = alm['nombre']
                        
                        if cantidad_transferir > stock_dep_anterior:
                            flash(f'❌ {nombre}: No hay suficiente stock en Depósito (disponible: {stock_dep_anterior})', 'error')
                            continue
                        
                        # Transferir: resta de stock_actual, suma a stock_full
                        cursor.execute('''
                            UPDATE productos_base 
                            SET stock_actual = stock_actual - %s,
                                stock_full = stock_full + %s
                            WHERE sku = %s
                        ''', (cantidad_transferir, cantidad_transferir, sku_alm))
                        
                        motivo_baja = 'Transferencia a Full ML'
                        motivo_carga = 'Transferencia desde Deposito'
                        
                        cursor.execute('''
                            INSERT INTO movimientos_stock 
                            (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku_alm, nombre, 'baja', cantidad_transferir, 
                              stock_dep_anterior, stock_dep_anterior - cantidad_transferir, 
                              motivo_baja))
                        
                        cursor.execute('''
                            INSERT INTO movimientos_stock 
                            (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku_alm, nombre, 'carga', cantidad_transferir, 
                              stock_full_anterior, stock_full_anterior + cantidad_transferir, 
                              motivo_carga))
                        
                        transferencias += 1
        
        conn.commit()

        if transferencias > 0:
            flash(f'✅ Transferencia completada ({transferencias} productos)', 'success')
            # Actualizar stock DEP en ML para compacs transferidos
            try:
                import threading
                skus_dep_transferidos = set()
                for key, value in request.form.items():
                    if key.startswith('transferir_compac_') and (int(value) if value else 0) > 0:
                        skus_dep_transferidos.add(key.replace('transferir_compac_', ''))
                if skus_dep_transferidos:
                    def _transfer_ml_bg():
                        try:
                            actualizar_publicaciones_ml_con_progreso(skus_dep_transferidos)
                        except Exception as e_ml:
                            print(f"[AUTO-ML] Error actualizando ML tras transferencia: {e_ml}")
                    threading.Thread(target=_transfer_ml_bg, daemon=True).start()
            except Exception as e_t:
                print(f"[AUTO-ML] Error iniciando thread transferencia: {e_t}")
        else:
            flash('ℹ️ No se realizaron transferencias', 'info')

    except Exception as e:
        conn.rollback()
        flash(f'❌ Error: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('transferir_stock'))


# ============================================================================
# HELPER PARA JSON (agregar después de las funciones de BD)
# ============================================================================

def decimal_to_float(obj):
    """Convertir Decimals a float para JSON serialization"""
    from decimal import Decimal
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, dict):
        return {k: decimal_to_float(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [decimal_to_float(item) for item in obj]
    return obj


# ============================================================================
# NUEVA VENTA CORREGIDO - Reemplazar desde línea 266 hasta antes de dashboard_visual
# ============================================================================

@app.route('/nueva-venta')
@login_required
def nueva_venta():
    """Formulario para registrar venta con stock disponible y ubicaciones"""
    from datetime import date
    import json
    
    try:
        # ========================================
        # 1. GENERAR SIGUIENTE NÚMERO DE VENTA
        # ========================================
        ultima_venta = query_db('SELECT numero_venta FROM ventas ORDER BY id DESC LIMIT 1')
        siguiente_numero = 'VENTA-001'
        
        if ultima_venta and len(ultima_venta) > 0:
            num_venta = ultima_venta[0].get('numero_venta')
            if num_venta and '-' in num_venta:
                try:
                    ultimo_num = int(num_venta.split('-')[1])
                    siguiente_numero = f'VENTA-{ultimo_num + 1:03d}'
                except (ValueError, IndexError):
                    siguiente_numero = 'VENTA-001'
        
        # ========================================
        # 2. OBTENER STOCK FÍSICO (con ubicaciones)
        # ========================================
        productos_base = query_db('''
            SELECT 
                sku, 
                nombre, 
                tipo, 
                stock_actual,
                COALESCE(stock_full, 0) as stock_full
            FROM productos_base 
            ORDER BY tipo, nombre
        ''')
        
        # ========================================
        # 3. OBTENER VENTAS ACTIVAS (DESCOMPONIENDO COMBOS)
        # ========================================
        ventas_activas = query_db('''
            SELECT 
                COALESCE(pb_comp.sku, iv.sku) as sku,
                SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
            FROM items_venta iv
            JOIN ventas v ON iv.venta_id = v.id
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
            LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
            WHERE v.estado_entrega = 'pendiente'
            GROUP BY sku
        ''')
        
        # Convertir a diccionario
        ventas_dict = {v['sku']: int(v['vendido']) for v in ventas_activas}
        
        # ========================================
        # 4. CALCULAR STOCK DISPONIBLE
        # ========================================
        productos_procesados = []
        
        for prod in productos_base:
            sku = prod['sku']
            vendido = ventas_dict.get(sku, 0)
            
            # Para productos con ubicaciones
            if '_DEP' in sku or '_FULL' in sku:
                # Compac: tiene _DEP y _FULL separados
                stock_fisico = int(prod['stock_actual'])
                stock_disponible_dep = stock_fisico - vendido
                
                productos_procesados.append({
                    'sku': sku,
                    'nombre': prod['nombre'],
                    'tipo': prod['tipo'],
                    'stock': stock_fisico,
                    'stock_disponible': stock_disponible_dep,
                    'tiene_ubicaciones': True,
                    'ubicacion': 'DEP' if '_DEP' in sku else 'FULL',
                    'precio': 0
                })
                
            elif prod['tipo'] == 'almohada':
                # Almohadas: tienen stock_actual (DEP) y stock_full (FULL)
                stock_dep = int(prod['stock_actual'])
                stock_full = int(prod['stock_full'])
                stock_total = stock_dep + stock_full
                stock_disponible = stock_total - vendido
                
                productos_procesados.append({
                    'sku': sku,
                    'nombre': prod['nombre'],
                    'tipo': prod['tipo'],
                    'stock': stock_total,
                    'stock_dep': stock_dep,
                    'stock_full': stock_full,
                    'stock_disponible': stock_disponible,
                    'tiene_ubicaciones': True,
                    'precio': 0
                })
                
            else:
                # Otros productos: solo stock_actual
                stock_fisico = int(prod['stock_actual'])
                stock_disponible = stock_fisico - vendido
                
                productos_procesados.append({
                    'sku': sku,
                    'nombre': prod['nombre'],
                    'tipo': prod['tipo'],
                    'stock': stock_fisico,
                    'stock_disponible': stock_disponible,
                    'tiene_ubicaciones': False,
                    'precio': 0
                })
        
        # ========================================
        # 5. OBTENER COMBOS/SOMMIERS
        # ========================================
        try:
            productos_combos = query_db('''
                SELECT 
                    sku, 
                    nombre, 
                    'combo' as tipo,
                    0 as stock,
                    0 as stock_disponible,
                    0 as precio
                FROM productos_compuestos
                WHERE activo = 1
                ORDER BY nombre
            ''')
            
            # Para cada combo, calcular su disponibilidad según componentes
            if productos_combos:
                for combo in productos_combos:
                    # Obtener componentes del combo usando IDs
                    componentes = query_db('''
                        SELECT pb.sku, c.cantidad_necesaria 
                        FROM componentes c
                        JOIN productos_base pb ON c.producto_base_id = pb.id
                        JOIN productos_compuestos pc ON c.producto_compuesto_id = pc.id
                        WHERE pc.sku = %s
                    ''', (combo['sku'],))
                    
                    # Calcular cuántos combos se pueden armar con el stock disponible
                    stock_disponible_combo = 999999  # Empezar con infinito
                    
                    for comp in componentes:
                        sku_comp = comp['sku']
                        cant_necesaria = int(comp['cantidad_necesaria'])
                        
                        # Buscar el stock disponible de este componente
                        prod_comp = next((p for p in productos_procesados if p['sku'] == sku_comp), None)
                        if prod_comp:
                            stock_disp_comp = prod_comp['stock_disponible']
                            # Cuántos combos se pueden hacer con este componente
                            combos_posibles = stock_disp_comp // cant_necesaria if cant_necesaria > 0 else 0
                            # El mínimo define cuántos combos se pueden armar
                            stock_disponible_combo = min(stock_disponible_combo, combos_posibles)
                        else:
                            # Si no existe el componente, no se puede armar el combo
                            stock_disponible_combo = 0
                            break
                    
                    # Si no hay componentes o todos dan infinito, poner 0
                    if stock_disponible_combo == 999999 or stock_disponible_combo < 0:
                        stock_disponible_combo = 0
                    
                    productos_procesados.append({
                        'sku': combo['sku'],
                        'nombre': combo['nombre'],
                        'tipo': combo['tipo'],
                        'stock': 0,  # Los combos no tienen stock físico
                        'stock_disponible': stock_disponible_combo,
                        'tiene_ubicaciones': False,
                        'precio': float(combo.get('precio', 0))
                    })
            
        except Exception as e:
            print(f"Nota: No se pudieron cargar combos - {str(e)}")
            import traceback
            traceback.print_exc()
        
        # ========================================
        # 6. CONVERTIR A JSON (SIN DECIMALS)
        # ========================================
        productos_json = json.dumps(productos_procesados)
        
        return render_template('nueva_venta.html',
                             siguiente_numero=siguiente_numero,
                             fecha_hoy=date.today().strftime('%Y-%m-%d'),
                             productos_json=productos_json)
                             
    except Exception as e:
        import traceback
        error_completo = traceback.format_exc()
        flash(f'Error al cargar Nueva Venta: {str(e)}', 'error')
        print(f"ERROR en nueva_venta:\n{error_completo}")
        return redirect(url_for('index'))



@app.route('/nueva-venta/guardar', methods=['POST'])
@login_required
def guardar_venta():
    """Guardar venta SIN descontar stock (solo registra la venta)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        from flask import session
        from datetime import datetime
        
        # ========================================
        # FECHA DE VENTA
        # ========================================
        fecha_venta_iso = session.get('ml_fecha_venta')
        if fecha_venta_iso:
            fecha_venta = datetime.fromisoformat(fecha_venta_iso)
            print(f"✅ Usando fecha de ML: {fecha_venta}")
        else:
            from datetime import timezone, timedelta
            tz_arg = timezone(timedelta(hours=-3))
            ahora_arg = datetime.now(tz_arg).replace(tzinfo=None)
            fecha_venta_form = request.form.get('fecha_venta')
            fecha_venta = datetime.strptime(fecha_venta_form + ' ' + ahora_arg.strftime('%H:%M:%S'), '%Y-%m-%d %H:%M:%S') if fecha_venta_form else ahora_arg
            print(f"✅ Usando fecha del formulario: {fecha_venta}")
        
        # ========================================
        # 1. DATOS GENERALES
        # ========================================
        numero_venta = request.form.get('numero_venta')
        canal = request.form.get('canal', 'Mercado Libre')
        mla_code = request.form.get('mla_code', '').strip()
        nombre_cliente = request.form.get('nombre_cliente', '').strip()
        
        if not nombre_cliente:
            nombre_cliente = mla_code if mla_code else 'Cliente sin especificar'
        
        telefono_cliente = request.form.get('telefono_cliente', '')
        
        # ========================================
        # 2. ENTREGA
        # ========================================
        tipo_entrega = request.form.get('tipo_entrega')
        direccion_entrega = request.form.get('direccion_entrega', '')
        metodo_envio = request.form.get('metodo_envio', '')
        zona_envio = request.form.get('zona_envio', '')
        
        if metodo_envio == 'Full':
            ubicacion_despacho = 'FULL'
        else:
            ubicacion_despacho = 'DEP'
        
        responsable_entrega = request.form.get('responsable_entrega', '')
        costo_flete = float(request.form.get('costo_flete') or 0)
        
        # ========================================
        # 3. PRODUCTOS - calcular importe_total desde items reales
        # ========================================
        productos_form = request.form.to_dict(flat=False)
        importe_total = 0.0
        
        for key in productos_form.keys():
            if key.startswith('productos[') and key.endswith('[sku]'):
                index = key.split('[')[1].split(']')[0]
                sku = productos_form.get(f'productos[{index}][sku]', [None])[0]
                cantidad = int(productos_form.get(f'productos[{index}][cantidad]', [0])[0])
                precio = float(productos_form.get(f'productos[{index}][precio]', [0])[0])
                if sku and cantidad > 0:
                    importe_total += cantidad * precio
        
        print(f"✅ importe_total calculado desde items: ${importe_total}")
        
        # ========================================
        # 4. PAGO
        # ========================================
        metodo_pago = request.form.get('metodo_pago')
        pago_mercadopago = float(request.form.get('pago_mercadopago', 0))
        pago_efectivo = float(request.form.get('pago_efectivo', 0))
        
        # Para Flete Propio y Zippin de ML, el cliente paga productos + flete por MP
        if canal == 'Mercado Libre' and metodo_envio in ['Flete Propio', 'Zippin'] and costo_flete > 0:
            pago_mercadopago += costo_flete
            print(f"✅ Sumando flete a pago_mercadopago: +${costo_flete} → total ${pago_mercadopago}")
        
        importe_abonado = pago_mercadopago + pago_efectivo
        
        print(f"✅ pago_mercadopago: ${pago_mercadopago}")
        print(f"✅ pago_efectivo: ${pago_efectivo}")
        print(f"✅ importe_abonado: ${importe_abonado}")
        
        # ========================================
        # 5. DATOS DE FACTURACIÓN
        # ========================================
        ml_billing_data = session.get('ml_billing_data')
        billing_info = {
            'business_name': None,
            'doc_type': None,
            'doc_number': None,
            'taxpayer_type': None,
            'city': None,
            'street': None,
            'state': None,
            'zip_code': None
        }
        
        if ml_billing_data:
            billing_info = extraer_billing_info_ml(ml_billing_data)
        
        # ========================================
        # 6. OBSERVACIONES Y ESTADO
        # ========================================
        notas = request.form.get('notas', '')
        estado_entrega = 'pendiente'
        estado_pago = 'pago_pendiente' if importe_abonado < importe_total else 'pagado'
        
        # ========================================
        # 7. INSERTAR VENTA
        # ========================================
        cursor.execute('''
            INSERT INTO ventas (
                numero_venta, fecha_venta, canal, mla_code,
                nombre_cliente, telefono_cliente,
                tipo_entrega, metodo_envio, ubicacion_despacho,
                zona_envio, direccion_entrega, responsable_entrega,
                costo_flete, metodo_pago, importe_total, importe_abonado,
                pago_mercadopago, pago_efectivo,
                estado_entrega, estado_pago, notas,
                factura_business_name, factura_doc_type, factura_doc_number,
                factura_taxpayer_type, factura_city, factura_street,
                factura_state, factura_zip_code
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s
            )
        ''', (
            numero_venta, fecha_venta, canal, mla_code,
            nombre_cliente, telefono_cliente,
            tipo_entrega, metodo_envio, ubicacion_despacho,
            zona_envio, direccion_entrega, responsable_entrega,
            costo_flete, metodo_pago, importe_total, importe_abonado,
            pago_mercadopago, pago_efectivo,
            estado_entrega, estado_pago, notas,
            billing_info['business_name'],
            billing_info['doc_type'],
            billing_info['doc_number'],
            billing_info['taxpayer_type'],
            billing_info['city'],
            billing_info['street'],
            billing_info['state'],
            billing_info['zip_code']
        ))
        
        venta_id = cursor.lastrowid
        
        # ========================================
        # 8. GUARDAR PRODUCTOS
        # ========================================
        items_agregados = 0
        
        for key in productos_form.keys():
            if key.startswith('productos[') and key.endswith('[sku]'):
                index = key.split('[')[1].split(']')[0]
                sku = productos_form.get(f'productos[{index}][sku]', [None])[0]
                cantidad = int(productos_form.get(f'productos[{index}][cantidad]', [0])[0])
                precio = float(productos_form.get(f'productos[{index}][precio]', [0])[0])
                
                if sku and cantidad > 0:
                    cursor.execute('''
                        INSERT INTO items_venta (venta_id, sku, cantidad, precio_unitario)
                        VALUES (%s, %s, %s, %s)
                    ''', (venta_id, sku, cantidad, precio))
                    items_agregados += 1
        
        conn.commit()
        
        # ========================================
        # 9. DETECTAR ALERTAS DE STOCK
        # ========================================
        productos_sin_stock = []
        try:
            items_vendidos_lista = []
            for key in productos_form.keys():
                if key.startswith('productos[') and key.endswith('[sku]'):
                    index = key.split('[')[1].split(']')[0]
                    sku = productos_form.get(f'productos[{index}][sku]', [None])[0]
                    cantidad = int(productos_form.get(f'productos[{index}][cantidad]', [0])[0])
                    if sku and cantidad > 0:
                        items_vendidos_lista.append({'sku': sku, 'cantidad': cantidad})
            
            if items_vendidos_lista:
                productos_sin_stock = detectar_alertas_stock_bajo(cursor, items_vendidos_lista, venta_id)
        except Exception as e_alertas:
            print(f"⚠️ Error al detectar alertas: {str(e_alertas)}")
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # ========================================
        # 10. LIMPIAR SESIÓN DE ML
        # ========================================
        session.pop('ml_orden_id', None)
        session.pop('ml_items', None)
        session.pop('ml_comprador_nombre', None)
        session.pop('ml_comprador_nickname', None)
        session.pop('ml_shipping', None)
        session.pop('ml_fecha_venta', None)
        session.pop('ml_billing_data', None)
        
        # ========================================
        # 11. MENSAJE Y REDIRECCIÓN
        # ========================================
        log_evento('INFO', 'venta', 'nueva_venta',
            f"Nueva venta {numero_venta} registrada. Canal: {canal}. Cliente: {nombre_cliente}. Total: ${importe_total}",
            venta_id=venta_id, usuario=current_user.username if current_user.is_authenticated else 'Sistema',
            ip=request.remote_addr)
        if productos_sin_stock:
            productos_base = [p for p in productos_sin_stock if p.get('tipo_producto') == 'base']
            combos_afectados = [p for p in productos_sin_stock if p.get('tipo_producto') == 'combo']
            
            mensaje_html = f'''
                <div class="alert alert-success mb-3">
                    <strong>✅ Venta {numero_venta} registrada correctamente</strong>
                </div>
                <div class="alert alert-warning">
                    <h5><i class="bi bi-exclamation-triangle-fill"></i> ⚠️ Alerta de Stock ML</h5>
            '''
            
            if productos_base:
                mensaje_html += '<p class="mb-2"><strong>Productos base sin stock disponible:</strong></p><ul class="mb-3">'
                for prod in productos_base:
                    mensaje_html += f'''
                        <li><strong>{prod['nombre']}</strong> (SKU: {prod['sku']})<br>
                            <small>Stock físico: {prod['stock_fisico']} | Vendido: {prod['vendido']} | Disponible: <span class="text-danger">{prod['stock_disponible']}</span></small>
                        </li>
                    '''
                mensaje_html += '</ul>'
            
            if combos_afectados:
                mensaje_html += '<p class="mb-2"><strong>Combos/Sommiers que NO se pueden armar:</strong></p><ul class="mb-3">'
                for combo in combos_afectados:
                    mensaje_html += f'''
                        <li><strong>{combo['nombre']}</strong> (SKU: {combo['sku']})<br>
                            <small class="text-muted">Falta componente: {combo.get('componente_faltante', 'N/A')}</small>
                        </li>
                    '''
                mensaje_html += '</ul>'
            
            mensaje_html += '<p class="mb-0"><strong>Recordá:</strong> Pausá las publicaciones en ML o cargá más stock.</p></div>'
            flash(mensaje_html, 'alerta_stock')
        else:
            mensaje = f'✅ Venta {numero_venta} registrada'
            if ubicacion_despacho == 'FULL':
                mensaje += ' - Se despachará desde FULL ML'
            else:
                mensaje += ' - Se despachará desde Depósito'
            flash(mensaje, 'success')
        
        return redirect(url_for('ventas_activas'))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        import traceback
        error_completo = traceback.format_exc()
        print(f"ERROR al guardar venta:\n{error_completo}")
        flash(f'❌ Error al guardar venta: {str(e)}', 'error')
        return redirect(url_for('nueva_venta'))



# ============================================================================
# DASHBOARD VISUAL - CORREGIDO CON LÓGICA DE BASES GRANDES
# ============================================================================

@app.route('/dashboard-visual')
@login_required
def dashboard_visual():
    """Dashboard visual - Stock Físico y Ventas Activas con lógica de bases grandes"""
    
    try:
        # =================================
        # 1. OBTENER STOCK FÍSICO (con stock_full para almohadas)
        # =================================
        query_stock = """
            SELECT sku, stock_actual, stock_full, tipo
            FROM productos_base
            ORDER BY tipo, sku
        """
        productos_stock = query_db(query_stock)
        
        # Crear diccionario de stock
        stock_dict = {}
        stock_full_dict = {}
        for prod in productos_stock:
            stock_dict[prod['sku']] = prod['stock_actual'] or 0
            stock_full_dict[prod['sku']] = prod.get('stock_full', 0) or 0
        
        # =================================
        # 2. OBTENER VENTAS ACTIVAS (PENDIENTES)
        # =================================
        query_ventas_activas = """
    SELECT 
        COALESCE(pb_comp.sku, iv.sku) as sku,
        SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as cantidad_vendida
    FROM items_venta iv
    JOIN ventas v ON iv.venta_id = v.id
    LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
    LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
    LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
    WHERE v.estado_entrega = 'pendiente'
    GROUP BY sku
        """
        ventas_activas = query_db(query_ventas_activas)
        
        # Convertir ventas activas a diccionario
        ventas_dict = {}
        for venta in ventas_activas:
            ventas_dict[venta['sku']] = venta['cantidad_vendida']
        
        # =================================
        # 3. FUNCIÓN HELPER PARA OBTENER DATOS
        # =================================
        def get_datos(sku):
            """Obtener stock y vendido para un SKU, con lógica de bases grandes"""
            stock = stock_dict.get(sku, 0)
            vendido = ventas_dict.get(sku, 0)
            
            # LÓGICA DE BASES GRANDES (160, 180, 200)
            # Bases x200: 2 bases físicas = 1 en dashboard
            if sku in ['BASE_CHOC160', 'BASE_GRIS160', 'BASE_SUBL160']:
                sku_base = sku.replace('160', '80200')
                stock = stock_dict.get(sku_base, 0) // 2
                vendido = ventas_dict.get(sku_base, 0) // 2
            elif sku in ['BASE_CHOC180', 'BASE_GRIS180', 'BASE_SUBL180']:
                sku_base = sku.replace('180', '90200')
                stock = stock_dict.get(sku_base, 0) // 2
                vendido = ventas_dict.get(sku_base, 0) // 2
            elif sku in ['BASE_CHOC200', 'BASE_GRIS200', 'BASE_SUBL200']:
                sku_base = sku.replace('200', '100200')
                stock = stock_dict.get(sku_base, 0) // 2
                vendido = ventas_dict.get(sku_base, 0) // 2
            
            return {'stock': stock, 'vendido': vendido}
        
        # =================================
        # 4. LÍNEA ESPUMA
        # =================================
        datos_espuma = {}
        for medida in ['80', '90', '100', '140', '150', '160', '180', '200']:
            datos_espuma[medida] = {
                'CPR20': get_datos(f'CPR{medida}20'),
                'BASE_SAB': get_datos(f'BASE_SAB{medida}'),
                'CEX': get_datos(f'CEX{medida}'),
                'CEXP': get_datos(f'CEXP{medida}'),
                'BASE_CHOC': get_datos(f'BASE_CHOC{medida}'),
                'CPR23': get_datos(f'CPR{medida}23'),
                'CRE': get_datos(f'CRE{medida}'),
                'CREP': get_datos(f'CREP{medida}'),
                'BASE_GRIS': get_datos(f'BASE_GRIS{medida}')
            }
        
        # =================================
        # 5. LÍNEA RESORTE
        # =================================
        datos_resorte = {}
        for medida in ['80', '90', '100', '140', '150', '160', '180', '200']:
            datos_resorte[medida] = {
                'CSO': get_datos(f'CSO{medida}'),
                'BASE_SAB': get_datos(f'BASE_SAB{medida}'),
                'CDO': get_datos(f'CDO{medida}'),
                'CDOP': get_datos(f'CDOP{medida}'),
                'BASE_GRIS': get_datos(f'BASE_GRIS{medida}'),
                'CSUP': get_datos(f'CSUP{medida}'),
                'BASE_SUBL': get_datos(f'BASE_SUBL{medida}')
            }
        
        # =================================
        # 6. OTROS PRODUCTOS
        # =================================
        datos_otros = {}
        for medida in ['80', '90', '100', '140', '150', '160', '180', '200']:
            # Tropical: Solo 80, 90, 100
            tropical = {'stock': 0, 'vendido': 0}
            if medida in ['80', '90', '100']:
                tropical = get_datos(f'CTR{medida}')
            
            # Compac: 80, 100, 140, 160 (CON UBICACIONES _DEP y _FULL)
            compac_dep = {'stock': 0, 'vendido': 0}
            compac_full = {'stock': 0, 'vendido': 0}
            if medida in ['80', '100', '140', '160']:
                compac_dep = get_datos(f'CCO{medida}_DEP')
                compac_full = get_datos(f'CCO{medida}_FULL')
            
            # Compac Plus: 80, 100, 140, 160 (CON UBICACIONES _DEP y _FULL)
            compac_plus_dep = {'stock': 0, 'vendido': 0}
            compac_plus_full = {'stock': 0, 'vendido': 0}
            if medida in ['80', '100', '140', '160']:
                compac_plus_dep = get_datos(f'CCP{medida}_DEP')
                compac_plus_full = get_datos(f'CCP{medida}_FULL')
            
            datos_otros[medida] = {
                'TROPICAL': tropical,
                'COMPAC_DEP': compac_dep,
                'COMPAC_FULL': compac_full,
                'COMPAC_PLUS_DEP': compac_plus_dep,
                'COMPAC_PLUS_FULL': compac_plus_full
            }
        
        # =================================
        # 7. ALMOHADAS (CON UBICACIONES DEP Y FULL)
        # =================================
        almohadas = []
        almohadas_skus = ['CERVICAL', 'CLASICA', 'DORAL', 'DUAL', 'EXCLUSIVE', 'PLATINO', 'RENOVATION', 'SUBLIME']
        
        for sku in almohadas_skus:
            stock_dep = stock_dict.get(sku, 0)
            stock_full = stock_full_dict.get(sku, 0)
            vendido = ventas_dict.get(sku, 0)
            
            # Mostrar si hay stock en cualquier ubicación o si hay ventas
            if stock_dep >= 0 or stock_full >= 0 or vendido >= 0:
                nombre_map = {
                    'CERVICAL': 'Almohada Visco Cervical',
                    'CLASICA': 'Almohada Visco Clásica',
                    'DORAL': 'Almohada Doral',
                    'DUAL': 'Almohada Dual Refreshing',
                    'EXCLUSIVE': 'Almohada Exclusive',
                    'PLATINO': 'Almohada Platino',
                    'RENOVATION': 'Almohada Renovation',
                    'SUBLIME': 'Almohada Sublime'
                }
                almohadas.append({
                    'nombre': nombre_map.get(sku, sku),
                    'stock_dep': stock_dep,
                    'stock_full': stock_full,
                    'vendido': vendido  # Total de ventas (no separadas por ubicación)
                })
        
        # =================================
        # 8. FECHA ACTUAL
        # =================================
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        # =================================
        # 9. RENDERIZAR TEMPLATE
        # =================================
        return render_template('dashboard_visual.html',
                             datos_espuma=datos_espuma,
                             datos_resorte=datos_resorte,
                             datos_otros=datos_otros,
                             almohadas=almohadas,
                             fecha_actual=fecha_actual)
    
    except Exception as e:
        flash(f'Error al cargar dashboard visual: {str(e)}', 'error')
        return redirect(url_for('index'))


# ============================================================================
# CONFIGURACIÓN - RESETEO DE SISTEMA
# ============================================================================

@app.route('/configuracion/resetear')
@login_required
def resetear_sistema():
    """Página para resetear el sistema (requiere contraseña)"""
    return render_template('resetear_sistema.html')


@app.route('/configuracion/resetear/ejecutar', methods=['POST'])
@login_required
def ejecutar_reseteo():
    """Ejecutar reseteo completo del sistema"""
    from flask import jsonify
    
    # Verificar contraseña
    password = request.form.get('password', '')
    confirmar = request.form.get('confirmar', '')
    
    # Contraseña configurada (CAMBIAR ESTO por tu contraseña)
    PASSWORD_CORRECTA = '32267845'
    
    if password != PASSWORD_CORRECTA:
        flash('❌ Contraseña incorrecta', 'error')
        return redirect(url_for('resetear_sistema'))
    
    if confirmar != 'RESETEAR':
        flash('❌ Debes escribir "RESETEAR" para confirmar', 'error')
        return redirect(url_for('resetear_sistema'))
    
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # ==================================
        # 1. BORRAR TODAS LAS VENTAS
        # ==================================
        cursor.execute('DELETE FROM items_venta')
        items_borrados = cursor.rowcount
        
        cursor.execute('DELETE FROM ventas')
        ventas_borradas = cursor.rowcount
        
        # ==================================
        # 2. BORRAR MOVIMIENTOS DE STOCK
        # ==================================
        cursor.execute('DELETE FROM movimientos_stock')
        movimientos_borrados = cursor.rowcount
        
        # ==================================
        # 3. RESETEAR STOCK A 0
        # ==================================
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = 0, 
                stock_full = 0
        ''')
        productos_reseteados = cursor.rowcount
        
        # ==================================
        # 4. BORRAR ALERTAS DE STOCK
        # ==================================
        try:
            cursor.execute('DELETE FROM alertas_stock')
            alertas_borradas = cursor.rowcount
        except:
            alertas_borradas = 0
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # Mensaje de confirmación
        mensaje = f'''
        ✅ <strong>Sistema reseteado correctamente</strong><br><br>
        📊 <strong>Operaciones realizadas:</strong><br>
        • {ventas_borradas} ventas eliminadas<br>
        • {items_borrados} items de venta eliminados<br>
        • {movimientos_borrados} movimientos de stock eliminados<br>
        • {productos_reseteados} productos reseteados a stock 0<br>
        • {alertas_borradas} alertas eliminadas<br><br>
        🎯 El sistema está listo para comenzar de cero.
        '''
        
        flash(mensaje, 'success')
        return redirect(url_for('index'))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error al resetear: {str(e)}', 'error')
        return redirect(url_for('resetear_sistema'))

# ============================================================================
# FUNCIONES AUXILIARES: MERCADO LIBRE
# Agregar ANTES del if __name__ == '__main__' al final de app.py
# ============================================================================

import time  # Agregar este import si no lo tenés

def refresh_ml_token():
    """Renovar el access_token usando el refresh_token guardado en DB"""
    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'ml_token'")
        if not row:
            return None
        data = json.loads(row['valor'])

        refresh_token = data.get('refresh_token')
        client_id = data.get('client_id')
        client_secret = data.get('client_secret')

        if not refresh_token or not client_id or not client_secret:
            print("⚠️  No hay refresh_token o credenciales guardadas")
            return None

        print("🔄 Renovando access_token de ML...")
        response = requests.post(
            "https://api.mercadolibre.com/oauth/token",
            data={
                "grant_type": "refresh_token",
                "client_id": client_id,
                "client_secret": client_secret,
                "refresh_token": refresh_token
            }
        )

        if response.status_code == 200:
            new_data = response.json()
            data['access_token'] = new_data.get('access_token')
            data['refresh_token'] = new_data.get('refresh_token', refresh_token)
            data['expires_at'] = time.time() + new_data.get('expires_in', 21600) - 300
            execute_db(
                "INSERT INTO configuracion (clave, valor) VALUES ('ml_token', %s) "
                "ON DUPLICATE KEY UPDATE valor = %s, actualizado_at = NOW()",
                (json.dumps(data), json.dumps(data))
            )
            print("✅ Token renovado automáticamente!")
            return data['access_token']
        else:
            print(f"❌ Error renovando token: {response.status_code} - {response.json()}")
            return None
    except Exception as e:
        print(f"Error en refresh_ml_token: {e}")
        return None

def cargar_ml_token():
    """Cargar token ML desde la base de datos. Si está vencido, lo renueva."""
    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'ml_token'")
        if not row:
            return None
        data = json.loads(row['valor'])
        access_token = data.get('access_token')
        expires_at = data.get('expires_at', 0)
        if expires_at and time.time() > expires_at:
            print("⚠️  Token ML vencido, renovando automáticamente...")
            access_token = refresh_ml_token()
        return access_token
    except Exception as e:
        print(f"Error cargando token ML: {e}")
        return None


def guardar_ml_token(token_data):
    """Guardar token ML en la base de datos (persiste en Railway)"""
    try:
        # Preservar refresh_token y credenciales si ya existen en DB
        existing_json = query_one("SELECT valor FROM configuracion WHERE clave = 'ml_token'")
        if existing_json:
            existing = json.loads(existing_json['valor'])
            if 'refresh_token' not in token_data and 'refresh_token' in existing:
                token_data['refresh_token'] = existing['refresh_token']
            if 'client_id' not in token_data and 'client_id' in existing:
                token_data['client_id'] = existing['client_id']
            if 'client_secret' not in token_data and 'client_secret' in existing:
                token_data['client_secret'] = existing['client_secret']

        if 'refresh_token' not in token_data:
            token_data.pop('expires_at', None)

        execute_db(
            "INSERT INTO configuracion (clave, valor) VALUES ('ml_token', %s) "
            "ON DUPLICATE KEY UPDATE valor = %s, actualizado_at = NOW()",
            (json.dumps(token_data), json.dumps(token_data))
        )
        return True
    except Exception as e:
        print(f"Error guardando token ML: {e}")
        return False


def obtener_ordenes_ml(access_token, limit=20):
    """
    Obtener órdenes de Mercado Libre
    Retorna: (success, data_o_error)
    """
    headers = {'Authorization': f'Bearer {access_token}'}
    
    try:
        # 1. Obtener USER_ID
        user_response = requests.get('https://api.mercadolibre.com/users/me', headers=headers)
        
        if user_response.status_code != 200:
            return False, f"Error obteniendo usuario: {user_response.status_code}"
        
        user_id = user_response.json()['id']
        
        # 2. Buscar órdenes como vendedor
        orders_url = f"https://api.mercadolibre.com/orders/search?seller={user_id}&sort=date_desc&limit={limit}"
        orders_response = requests.get(orders_url, headers=headers)
        
        if orders_response.status_code != 200:
            return False, f"Error obteniendo órdenes: {orders_response.status_code}"
        
        orders_data = orders_response.json()
        return True, orders_data['results']
        
    except Exception as e:
        return False, str(e)


def obtener_shipping_details(access_token, shipping_id):
    """
    Obtener detalles del envío desde ML
    Retorna: (success, data_o_error)
    """
    if not shipping_id:
        return False, "No hay shipping_id"
    
    headers = {'Authorization': f'Bearer {access_token}'}
    
    try:
        response = requests.get(f'https://api.mercadolibre.com/shipments/{shipping_id}', headers=headers)
        
        if response.status_code == 200:
            return True, response.json()
        else:
            return False, f"Error {response.status_code}"
    except Exception as e:
        return False, str(e)

# ─── FUNCIÓN 4: Mapeo automático al importar - quita la Z del SKU ───
def normalizar_sku_ml(sku_ml):
    """
    Normaliza SKUs de ML que difieren del SKU en la BD.
    Retorna (sku_normalizado, cantidad_override).
    cantidad_override != 0 cuando el SKU de ML representa múltiples unidades.
    Ejemplos:
      CEX140Z       → ('CEX140', 0)
      RENOVACIONAL  → ('RENOVATION', 0)
      CLASICAX2     → ('CLASICA', 2)
    """
    if not sku_ml:
        return sku_ml, 0

    sku_up = sku_ml.strip().upper()

    # Mapeos fijos ML → BD
    SKU_MAP = {
        'RENOVATIONAL': ('RENOVATION', 0),
        'CLASICAX2':    ('CLASICA',    2),
    }
    if sku_up in SKU_MAP:
        return SKU_MAP[sku_up]

    # Quitar Z del final
    if sku_up.endswith('Z'):
        return sku_ml[:-1], 0

    return sku_ml, 0


# REEMPLAZAR la función procesar_orden_ml completa
def procesar_orden_ml(orden):
    """
    Procesar orden de ML SIN obtener detalles de shipping
    Usar al LISTAR órdenes (más rápido)
    CAPTURA FECHA REAL DE VENTA y COSTO DE ENVÍO
    """
    # Fecha REAL de la venta en ML
    from datetime import timezone, timedelta as _td
    _dt_utc = datetime.fromisoformat(orden['date_created'].replace('Z', '+00:00'))
    fecha = _dt_utc.astimezone(timezone(_td(hours=-3))).replace(tzinfo=None)
    
    # Items/Productos
    items = []
    for item in orden['order_items']:
        items.append({
            'sku': item['item'].get('seller_sku', ''),
            'titulo': item['item']['title'],
            'cantidad': item['quantity'],
            'precio': item['unit_price']
        })
    
    # Comprador
    buyer = orden.get('buyer', {})
    comprador_nombre = f"{buyer.get('first_name', '')} {buyer.get('last_name', '')}".strip()
    comprador_nickname = buyer.get('nickname', '')
    
    # Total
    total = orden['total_amount']
    paid_amount = orden.get('paid_amount', total)  # incluye flete
    
    # Estado
    estado = orden['status']
    
    # Shipping (solo ID, sin detalles) + COSTO DE ENVÍO
    shipping = orden.get('shipping', {})
    shipping_id = shipping.get('id', '')
    costo_envio = shipping.get('shipping_cost', 0)  # ✅ NUEVO: Capturar costo de envío
    
    shipping_data = {
        'tiene_envio': bool(shipping_id),
        'shipping_id': shipping_id,
        'costo_envio': costo_envio,  # ✅ NUEVO
        'metodo_envio': '',
        'direccion': '',
        'ciudad': '',
        'provincia': '',
        'codigo_postal': '',
        'zona': ''
    }
    
    return {
        'id': orden['id'],
        'fecha': fecha,  # ✅ Fecha real de ML
        'fecha_iso': orden['date_created'],  # ISO string para pasar a shipping
        'comprador_nombre': comprador_nombre,
        'comprador_nickname': comprador_nickname,
        'items': items,
        'total': total,
        'paid_amount': paid_amount,
        'estado': estado,
        'shipping': shipping_data
    }

# Códigos postales con cobertura de Flete Propio
# Fuente: tariff-locations-1288.csv — si el CP no está en esta lista → Zippin
CPS_FLETE_PROPIO = {
    '1001','1002','1003','1004','1005','1006','1007','1008','1009','1010',
    '1011','1012','1013','1014','1015','1016','1017','1018','1019','1020',
    '1021','1022','1023','1024','1025','1026','1027','1028','1029','1030',
    '1031','1032','1033','1034','1035','1036','1037','1038','1039','1040',
    '1041','1042','1043','1044','1045','1046','1047','1048','1049','1050',
    '1051','1052','1053','1054','1055','1056','1057','1058','1059','1060',
    '1061','1062','1063','1064','1065','1066','1067','1068','1069','1070',
    '1071','1072','1073','1074','1075','1076','1077','1078','1079','1080',
    '1081','1082','1083','1084','1085','1086','1087','1088','1089','1090',
    '1091','1092','1093','1094','1095','1096','1097','1098','1099','1100',
    '1101','1102','1103','1104','1105','1106','1107','1108','1109','1110',
    '1111','1112','1113','1114','1115','1116','1117','1118','1119','1120',
    '1121','1122','1123','1124','1125','1126','1127','1128','1129','1130',
    '1131','1132','1133','1134','1135','1136','1137','1138','1139','1140',
    '1141','1142','1143','1144','1145','1146','1147','1148','1149','1150',
    '1151','1152','1153','1154','1155','1156','1157','1158','1159','1160',
    '1161','1162','1163','1164','1165','1166','1167','1168','1169','1170',
    '1171','1172','1173','1174','1175','1176','1177','1178','1179','1180',
    '1181','1182','1183','1184','1185','1186','1187','1188','1189','1190',
    '1191','1192','1193','1194','1195','1196','1197','1198','1199','1200',
    '1201','1202','1203','1204','1205','1206','1207','1208','1209','1210',
    '1211','1212','1213','1214','1215','1216','1217','1218','1219','1220',
    '1221','1222','1223','1224','1225','1226','1227','1228','1229','1230',
    '1231','1232','1233','1234','1235','1236','1237','1238','1239','1240',
    '1241','1242','1243','1244','1245','1246','1247','1248','1249','1250',
    '1251','1252','1253','1254','1255','1256','1257','1258','1259','1260',
    '1261','1262','1263','1264','1265','1266','1267','1268','1269','1270',
    '1271','1272','1273','1274','1275','1276','1277','1278','1279','1280',
    '1281','1282','1283','1284','1285','1286','1287','1288','1289','1290',
    '1291','1292','1293','1294','1295','1296','1298','1300','1301','1302',
    '1303','1304','1305','1306','1307','1308','1309','1310','1311','1312',
    '1313','1314','1315','1316','1317','1318','1319','1320','1321','1322',
    '1323','1324','1325','1326','1327','1328','1329','1330','1331','1332',
    '1333','1334','1335','1336','1337','1338','1339','1340','1341','1342',
    '1343','1344','1345','1346','1347','1348','1349','1350','1351','1352',
    '1353','1354','1355','1356','1357','1358','1359','1360','1361','1362',
    '1363','1364','1365','1366','1367','1368','1369','1370','1371','1372',
    '1373','1374','1375','1376','1377','1378','1379','1380','1381','1382',
    '1383','1384','1385','1386','1387','1388','1389','1390','1391','1392',
    '1393','1394','1395','1396','1397','1398','1399','1400','1401','1402',
    '1405','1406','1407','1408','1409','1410','1411','1412','1413','1414',
    '1416','1417','1419','1420','1421','1422','1423','1424','1425','1426',
    '1427','1428','1429','1430','1431','1432','1433','1434','1435','1436',
    '1437','1439','1440','1453','1599','1602','1603','1605','1606','1607',
    '1609','1611','1612','1613','1614','1615','1617','1618','1621','1636',
    '1638','1640','1642','1643','1644','1646','1648','1649','1650','1651',
    '1653','1655','1657','1659','1661','1663','1664','1665','1667','1670',
    '1672','1674','1676','1678','1682','1684','1686','1688','1702','1704',
    '1706','1708','1712','1713','1714','1716','1718','1722','1723','1742',
    '1744','1752','1754','1755','1757','1759','1761','1763','1765','1766',
    '1768','1770','1771','1772','1773','1774','1776','1778','1802','1804',
    '1806','1812','1822','1824','1825','1826','1828','1829','1831','1832',
    '1834','1835','1836','1838','1842','1846','1847','1849','1852','1854',
    '1856','1870','1871','1872','1874','1875','1876','1878','1879','1881','1882',
}


def obtener_shipping_completo(shipping_id, access_token, fecha_orden_iso=''):
    """
    Obtener detalles completos de shipping desde ML
    MAPEO CORREGIDO según tipos reales de ML
    ✅ NUEVO: Captura COSTO DE ENVÍO del shipment
    """
    shipping_data = {
        'tiene_envio': True,
        'shipping_id': shipping_id,
        'metodo_envio': '',
        'metodo_envio_ml': '',
        'logistic_type_ml': '',
        'costo_envio': 0,
        'fecha_entrega_ml': '',
        'turbo_rango': '',
        'direccion': '',
        'ciudad': '',
        'provincia': '',
        'codigo_postal': '',
        'zona': ''
    }
    
    if not shipping_id or not access_token:
        return shipping_data
    
    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        response = requests.get(f'https://api.mercadolibre.com/shipments/{shipping_id}', headers=headers)
        
        if response.status_code != 200:
            print(f"⚠️ Error al obtener shipment {shipping_id}: {response.status_code}")
            return shipping_data
        
        shipment = response.json()
        
        # ✅ NUEVO: Capturar COSTO DE ENVÍO
        # Puede estar en varios lugares según el tipo de envío
        costo_envio = 0
        
        # Opción 1: base_cost
        if 'base_cost' in shipment:
            costo_envio = shipment.get('base_cost', 0)
            print(f"💰 Costo envío (base_cost): ${costo_envio}")
        
        # Opción 2: shipping_cost
        elif 'shipping_cost' in shipment:
            costo_envio = shipment.get('shipping_cost', 0)
            print(f"💰 Costo envío (shipping_cost): ${costo_envio}")
        
        # Opción 3: cost
        elif 'cost' in shipment:
            costo_envio = shipment.get('cost', 0)
            print(f"💰 Costo envío (cost): ${costo_envio}")
        
        # Opción 4: shipping_option -> cost
        elif 'shipping_option' in shipment:
            shipping_option = shipment.get('shipping_option', {})
            if 'cost' in shipping_option:
                costo_envio = shipping_option.get('cost', 0)
                print(f"💰 Costo envío (shipping_option.cost): ${costo_envio}")
        
        shipping_data['costo_envio'] = costo_envio
        
        # Método de envío
        shipping_option = shipment.get('shipping_option', {})
        shipping_mode = shipping_option.get('shipping_method_id', '')
        logistic_type = shipment.get('logistic_type', '')
        
        # 🔍 DEBUGGING
        print(f"\n🚚 SHIPPING ID: {shipping_id}")
        print(f"📦 shipping_method_id: {shipping_mode}")
        print(f"📦 logistic_type: {logistic_type}")
        print(f"💰 COSTO TOTAL: ${costo_envio}")
        
        # Guardar valores originales
        shipping_data['metodo_envio_ml'] = shipping_mode
        shipping_data['logistic_type_ml'] = logistic_type

        # Fecha estimada de entrega
        try:
            from datetime import datetime, timedelta
            lt = logistic_type  # alias corto

            if lt in ('cross_docking', 'xd_drop_off', 'self_service'):
                # Fecha de despacho según /sla de ML (expected_date = fecha límite de despacho)
                from datetime import timezone as _tz, timedelta as _td
                tz_ar = _tz(_td(hours=-3))
                try:
                    import requests as _req2
                    sla_r = _req2.get(
                        f'https://api.mercadolibre.com/shipments/{shipping_id}/sla',
                        headers={'Authorization': f'Bearer {access_token}'},
                        timeout=10
                    )
                    if sla_r.status_code == 200:
                        expected_raw = sla_r.json().get('expected_date', '')
                        if expected_raw:
                            fecha_sla = datetime.fromisoformat(expected_raw).astimezone(tz_ar).date()
                            shipping_data['fecha_entrega_ml'] = f"{fecha_sla.day:02d}/{fecha_sla.month:02d}"
                            print(f"📅 Fecha despacho (SLA): {shipping_data['fecha_entrega_ml']}")
                        else:
                            raise ValueError("expected_date vacío")
                    else:
                        raise ValueError(f"SLA status {sla_r.status_code}")
                except Exception as e_sla:
                    print(f"⚠️ SLA falló ({e_sla}), usando hora de corte como fallback")
                    # ── FALLBACK: lógica hora de corte (comentada como respaldo) ──
                    try:
                        hora_corte_str = session.get('hora_corte_colecta', '14:00')
                    except RuntimeError:
                        hora_corte_str = '14:00'
                    try:
                        hh, mm = map(int, hora_corte_str.split(':'))
                    except:
                        hh, mm = 14, 0
                    fecha_orden_raw = fecha_orden_iso
                    if fecha_orden_raw:
                        dt_orden = datetime.fromisoformat(fecha_orden_raw.replace('Z', '+00:00'))
                        dt_orden_ar = dt_orden.astimezone(tz_ar)
                        corte = dt_orden_ar.replace(hour=hh, minute=mm, second=0, microsecond=0)
                        def proximo_dia_habil(fecha):
                            while fecha.weekday() >= 5:
                                fecha = fecha + timedelta(days=1)
                            return fecha
                        dia_orden = dt_orden_ar.date()
                        dia_siguiente = dia_orden + timedelta(days=1)
                        if dia_orden.weekday() >= 5:
                            fecha_colecta = proximo_dia_habil(dia_orden + timedelta(days=1))
                        elif (corte - dt_orden_ar).total_seconds() >= 3600:
                            fecha_colecta = dia_orden
                        else:
                            fecha_colecta = proximo_dia_habil(dia_siguiente)
                        shipping_data['fecha_entrega_ml'] = f"{fecha_colecta.day:02d}/{fecha_colecta.month:02d}"
                        print(f"📅 Fecha colecta (fallback corte {hora_corte_str}): {shipping_data['fecha_entrega_ml']}")
            else:
                # Flex y otros: usar estimated_delivery_limit
                fecha_entrega_raw = (
                    shipment.get('shipping_option', {})
                    .get('estimated_delivery_limit', {})
                    .get('date', '')
                )
                if fecha_entrega_raw:
                    dt = datetime.fromisoformat(fecha_entrega_raw.replace('Z', '+00:00'))
                    shipping_data['fecha_entrega_ml'] = f"{dt.day:02d}/{dt.month:02d}"
                    print(f"📅 Fecha entrega ML: {shipping_data['fecha_entrega_ml']}")
        except Exception as e:
            print(f"⚠️ Error capturando fecha entrega: {e}")
        
        # 🔧 MAPEO CORREGIDO según logs reales
        if logistic_type == 'fulfillment':
            shipping_data['metodo_envio'] = 'Full'
            shipping_data['fecha_entrega_ml'] = ''  # Full no muestra fecha en observaciones
            print(f"✅ MAPEADO A: Full")
        
        elif logistic_type == 'self_service':
            # Detectar Turbo: tag 'turbo' en el shipment
            tags = shipment.get('tags', [])
            if 'turbo' in tags:
                shipping_data['metodo_envio'] = 'Turbo'
                # Capturar rango horario desde offset
                try:
                    edt = shipment.get('shipping_option', {}).get('estimated_delivery_time', {})
                    hora_desde_raw = edt.get('date', '')
                    hora_hasta_raw = edt.get('offset', {}).get('date', '')
                    if hora_desde_raw and hora_hasta_raw:
                        from datetime import datetime, timezone, timedelta
                        tz_ar = timezone(timedelta(hours=-3))
                        h_desde = datetime.fromisoformat(hora_desde_raw).astimezone(tz_ar)
                        h_hasta = datetime.fromisoformat(hora_hasta_raw).astimezone(tz_ar)
                        shipping_data['turbo_rango'] = f"{h_desde.strftime('%H:%M')}-{h_hasta.strftime('%H:%M')}"
                        shipping_data['fecha_entrega_ml'] = f"Turbo {h_desde.strftime('%H:%M')}-{h_hasta.strftime('%H:%M')}"
                except Exception as e:
                    print(f"⚠️ Error capturando rango Turbo: {e}")
                    shipping_data['turbo_rango'] = ''
                print(f"✅ MAPEADO A: Turbo")
            else:
                shipping_data['metodo_envio'] = 'Flex'
                print(f"✅ MAPEADO A: Flex")
        
        elif logistic_type == 'xd_drop_off':
            shipping_data['metodo_envio'] = 'Colecta'
            print(f"✅ MAPEADO A: Colecta")
        
        elif logistic_type == 'cross_docking':
            shipping_data['metodo_envio'] = 'Colecta'
            print(f"✅ MAPEADO A: Colecta")
        
        elif logistic_type == 'default':
            # Default depende de la zona (Flete propio o Zippin)
            # Por ahora dejamos como Flete Propio y luego se puede ajustar manualmente
            shipping_data['metodo_envio'] = 'Flete Propio'
            print(f"⚠️ MAPEADO A: Flete Propio (default - ajustar según zona)")
        
        elif 'mercadoenvios' in str(shipping_mode).lower():
            shipping_data['metodo_envio'] = 'Colecta'
            print(f"✅ MAPEADO A: Colecta")
        
        else:
            shipping_data['metodo_envio'] = 'Colecta'
            print(f"⚠️ MAPEADO A: Colecta (default)")
        
        # Dirección
        receiver_address = shipment.get('receiver_address', {})
        
        if receiver_address:
            address_line = receiver_address.get('address_line', '')
            street_name = receiver_address.get('street_name', '')
            street_number = receiver_address.get('street_number', '')
            floor = receiver_address.get('floor', '')
            apartment = receiver_address.get('apartment', '')
            
            def limpiar_direccion(d):
                """Quita la palabra 'calle' al inicio si está presente"""
                d = d.strip()
                if d.lower().startswith('calle '):
                    d = d[6:].strip()
                return d

            if address_line and 'X' * 3 not in address_line.upper():
                shipping_data['direccion'] = limpiar_direccion(address_line)
            elif street_name and street_number:
                direccion = f"{street_name} {street_number}"
                if floor:
                    direccion += f" Piso {floor}"
                if apartment:
                    direccion += f" Depto {apartment}"
                shipping_data['direccion'] = limpiar_direccion(direccion)
            
            # Ciudad y provincia
            city = receiver_address.get('city', {})
            state = receiver_address.get('state', {})

            # Extraer nombre de ciudad y agregarlo a la dirección
            ciudad_nombre = city.get('name', '') if isinstance(city, dict) else (str(city) if city else '')
            if ciudad_nombre and shipping_data.get('direccion'):
                shipping_data['direccion'] += f', {ciudad_nombre}'

            if isinstance(city, dict):
                shipping_data['ciudad'] = str(city.get('name', ''))
            else:
                shipping_data['ciudad'] = str(city) if city else ''
            
            if isinstance(state, dict):
                shipping_data['provincia'] = str(state.get('name', ''))
            else:
                shipping_data['provincia'] = str(state) if state else ''
            
            shipping_data['codigo_postal'] = str(receiver_address.get('zip_code', ''))

            # Reclasificar Flete Propio → Zippin si el CP no está en nuestra cobertura
            if shipping_data.get('metodo_envio') == 'Flete Propio':
                cp = shipping_data['codigo_postal'].strip()
                if cp and cp not in CPS_FLETE_PROPIO:
                    shipping_data['metodo_envio'] = 'Zippin'
                    print(f"📦 CP {cp} fuera de cobertura → reclasificado a Zippin")
            
            # Inferir zona
            if shipping_data['ciudad']:
                ciudad_lower = shipping_data['ciudad'].lower()
                provincia_lower = shipping_data['provincia'].lower()
                
                if 'capital federal' in ciudad_lower or 'ciudad' in ciudad_lower or 'caba' in ciudad_lower or 'autonoma' in provincia_lower:
                    shipping_data['zona'] = 'Capital'
                elif any(x in ciudad_lower for x in ['plata', 'quilmes', 'avellaneda', 'berazategui', 'florencio varela', 'lanus', 'lomas', 'banfield', 'temperley', 'adrog', 'monte grande', 'bernal', 'wilde', 'dock sud', 'ezeiza', 'canning', 'longchamps', 'san francisco solano', 'varela']):
                    shipping_data['zona'] = 'Sur'
                elif any(x in ciudad_lower for x in ['san isidro', 'tigre', 'pilar', 'escobar', 'san fernando', 'martinez', 'acassuso', 'beccar', 'olivos', 'vicente lopez', 'florida', 'munro', 'villa adelina', 'boulogne', 'jose c paz', 'malvinas', 'del viso', 'nordelta', 'pacheco', 'grand bourg']):
                    shipping_data['zona'] = 'Norte-Noroeste'
                elif any(x in ciudad_lower for x in ['moron', 'merlo', 'ituzaingo', 'hurlingham', 'moreno', 'haedo', 'castelar', 'ramos mejia', 'san martin', 'ciudadela', 'lomas del mirador', 'villa luzuriaga', 'tapiales', 'laferrere', 'gonzalez catan', 'tres de febrero', 'liniers']):
                    shipping_data['zona'] = 'Oeste'
    
    except Exception as e:
        print(f"⚠️ Error al procesar shipping {shipping_id}: {str(e)}")
        import traceback
        traceback.print_exc()

    # Capturar fecha prometida al comprador y demora (schedule) desde estimated_delivery_time
    try:
        edt = shipment.get('shipping_option', {}).get('estimated_delivery_time', {}) if shipment else {}
        fecha_prometida_raw = edt.get('date', '')
        schedule_hs = edt.get('schedule') or 0
        if fecha_prometida_raw:
            from datetime import datetime as _dt
            fecha_prometida = _dt.fromisoformat(fecha_prometida_raw[:10]).date()
            shipping_data['fecha_entrega_prometida'] = fecha_prometida.strftime('%Y-%m-%d')
            print(f"📅 Fecha prometida al cliente: {fecha_prometida.strftime('%d/%m/%Y')}")
        else:
            shipping_data['fecha_entrega_prometida'] = ''
        if schedule_hs and float(schedule_hs) > 0:
            shipping_data['demora_ml_dias'] = int(round(float(schedule_hs) / 24))
            print(f"⏳ Demora ML: {shipping_data['demora_ml_dias']} días ({schedule_hs}hs schedule)")
        else:
            shipping_data['demora_ml_dias'] = 0
    except Exception as e:
        print(f"⚠️ Error capturando fecha prometida: {e}")
        shipping_data['fecha_entrega_prometida'] = ''
        shipping_data['demora_ml_dias'] = 0

    return shipping_data


def verificar_sku_en_bd(sku):
    """
    Verificar si un SKU existe en la base de datos
    Retorna: (existe, tipo_producto, nombre)
    """
    # Buscar en productos_base
    prod = query_one('SELECT tipo, nombre FROM productos_base WHERE sku = %s', (sku,))
    if prod:
        return True, prod['tipo'], prod['nombre']
    
    # Buscar en productos_compuestos
    combo = query_one('SELECT nombre FROM productos_compuestos WHERE sku = %s', (sku,))
    if combo:
        return True, 'combo', combo['nombre']
    
    return False, None, None


# ─── FUNCIÓN 1: Actualizar handling_time (tiempo de demora) en ML ───
def actualizar_handling_time_ml(mla_id, dias, access_token):
    """
    Actualizar el tiempo de disponibilidad (handling_time) en ML
    
    Args:
        mla_id: ID de la publicación
        dias: Cantidad de días de demora
        access_token: Token de ML
    
    Returns:
        (success: bool, message: str)
    """
    try:
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        # IMPORTANTE: Primero traer la publicación actual
        response_get = requests.get(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers
        )
        
        if response_get.status_code != 200:
            return False, "Error obteniendo publicación"
        
        # Actualizar solo handling_time
        data = {
            "sale_terms": [
                {
                    "id": "MANUFACTURING_TIME",
                    "value_name": f"{dias} días"
                }
            ]
        }
        
        response = requests.put(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers,
            json=data
        )
        
        if response.status_code == 200:
            return True, f"Demora configurada a {dias} días"
        else:
            error_data = response.json()
            error_msg = error_data.get('message', 'Error desconocido')
            return False, f"Error ML: {error_msg}"
    
    except Exception as e:
        return False, f"Error: {str(e)}"

# ============================================================================
# FUNCIÓN HELPER PARA QUITAR DEMORA (basada en la que funciona)
# ============================================================================

def quitar_handling_time_ml(mla_id, access_token):
    """
    Quitar el tiempo de disponibilidad (handling_time) en ML
    CON LOGGING DETALLADO
    """
    try:
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        # GET de la publicación
        response_get = requests.get(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers
        )
        
        if response_get.status_code != 200:
            return False, f"Error obteniendo publicación: {response_get.status_code}"
        
        item_data = response_get.json()
        
        # VER QUÉ SALE_TERMS TIENE
        sale_terms_antes = item_data.get('sale_terms', [])
        print(f"\n=== DEBUG {mla_id} ===")
        print(f"sale_terms ANTES: {sale_terms_antes}")
        
        # Filtrar para quitar MANUFACTURING_TIME
        sale_terms_despues = [
            term for term in sale_terms_antes
            if term.get('id') != 'MANUFACTURING_TIME'
        ]
        
        print(f"sale_terms DESPUÉS: {sale_terms_despues}")
        
        # Enviar actualización
        data = {
            "sale_terms": sale_terms_despues
        }
        
        print(f"Enviando: {data}")
        
        response = requests.put(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers,
            json=data
        )
        
        print(f"Response status: {response.status_code}")
        print(f"Response body: {response.text[:500]}")  # Primeros 500 chars
        
        if response.status_code == 200:
            # VERIFICAR si realmente se quitó
            response_check = requests.get(
                f'https://api.mercadolibre.com/items/{mla_id}',
                headers=headers
            )
            
            if response_check.status_code == 200:
                item_actualizado = response_check.json()
                sale_terms_final = item_actualizado.get('sale_terms', [])
                print(f"sale_terms FINAL (después de actualizar): {sale_terms_final}")
                
                # Verificar si todavía tiene MANUFACTURING_TIME
                tiene_demora = any(
                    term.get('id') == 'MANUFACTURING_TIME' 
                    for term in sale_terms_final
                )
                
                if tiene_demora:
                    return False, f"ML aceptó el cambio pero la demora sigue ahí"
                else:
                    return True, f"Demora quitada correctamente de {mla_id}"
            
            return True, f"Actualizado (sin verificar)"
        else:
            error_data = response.json() if response.text else {}
            error_msg = error_data.get('message', 'Error desconocido')
            return False, f"Error ML: {error_msg}"
    
    except Exception as e:
        print(f"Exception: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, f"Error: {str(e)}"



# ============================================================================
# RUTAS: MERCADO LIBRE
# Agregar DESPUÉS de la ruta /ventas/activas en app.py
# ============================================================================

# ─── FUNCIÓN 2: Sincronizar variantes con Z (demora) ───
@app.route('/alertas/<int:alerta_id>/configurar-demora-ml', methods=['POST'])
@login_required
def configurar_demora_ml_desde_alerta(alerta_id):
    """
    Configurar días de demora en las publicaciones CON Z
    Procesa solo la parte CON Z de la alerta
    """
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay token de ML configurado', 'error')
        return redirect(url_for('alertas_ml'))
    
    # Obtener días de demora del formulario
    dias_demora = request.form.get('dias_demora', type=int)
    
    if not dias_demora or dias_demora < 1 or dias_demora > 90:
        flash('❌ Los días de demora deben estar entre 1 y 90', 'error')
        return redirect(url_for('alertas_ml'))
    
    try:
        # Obtener la alerta
        alerta = query_db('SELECT * FROM alertas_stock WHERE id = %s', (alerta_id,))
        if alerta:
            alerta = alerta[0]
        
        if not alerta:
            flash('❌ Alerta no encontrada', 'error')
            return redirect(url_for('alertas_ml'))
        
        sku = alerta['sku']
        sku_con_z = f"{sku}Z"
        
        # Obtener publicaciones CON Z
        publicaciones = query_db(
            'SELECT * FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE',
            (sku_con_z,)
        )
        
        if not publicaciones:
            flash(f'⚠️ No hay publicaciones con Z mapeadas para {sku_con_z}', 'warning')
            return redirect(url_for('alertas_ml'))
        
        # Actualizar demora en cada publicación
        resultados = []
        errores = []
        
        for pub in publicaciones:
            mla_id = pub['mla_id']
            success, message = actualizar_handling_time_ml(mla_id, dias_demora, access_token)
            
            if success:
                resultados.append(f"{mla_id}: {message}")
            else:
                errores.append(f"{mla_id}: {message}")
        
        # ✅ LÓGICA DE PROCESAMIENTO INDEPENDIENTE
        tiene_variante_normal = len(query_db(
            'SELECT 1 FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE LIMIT 1',
            (sku,)
        )) > 0
        
        tipo_procesado_actual = alerta.get('tipo_procesado')
        
        if tiene_variante_normal:
            # Tiene variante normal - solo marcar que procesamos la parte Z
            if tipo_procesado_actual == 'normal':
                # Ya se había procesado normal → ahora marcar como ambos y cerrar alerta
                execute_db(
                    "UPDATE alertas_stock SET tipo_procesado = 'ambos', estado = 'procesada', fecha_procesada = NOW() WHERE id = %s",
                    (alerta_id,)
                )
            else:
                # Solo marcar que se procesó la parte Z
                execute_db(
                    "UPDATE alertas_stock SET tipo_procesado = 'z' WHERE id = %s",
                    (alerta_id,)
                )
        else:
            # No tiene variante normal - cerrar la alerta directamente
            execute_db(
                "UPDATE alertas_stock SET tipo_procesado = 'ambos', estado = 'procesada', fecha_procesada = NOW() WHERE id = %s",
                (alerta_id,)
            )
        
        # Mostrar resultados
        if resultados:
            flash(f'✅ Demora configurada en ML: {", ".join(resultados)}', 'success')
        
        if errores:
            flash(f'❌ Errores: {", ".join(errores)}', 'error')
    
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('alertas_ml'))

@app.route('/ventas/ml/configurar_token', methods=['GET', 'POST'])
@login_required
def ml_configurar_token():
    """
    Página para configurar/actualizar el token de ML
    Maneja 3 acciones:
    - GET: Mostrar página con URL de autorización
    - POST action=canjear_code: Canjear el code por access_token
    - POST action=token_manual: Guardar token manual (como antes)
    """
    
    # Credenciales de la app ML
    CLIENT_ID = os.getenv('ML_APP_ID')
    CLIENT_SECRET = os.getenv('ML_SECRET_KEY')
    REDIRECT_URI = os.getenv('ML_REDIRECT_URI')
    
    # Generar URL de autorización
    url_autorizacion = (
        f"https://auth.mercadolibre.com.ar/authorization"
        f"?response_type=code"
        f"&client_id={CLIENT_ID}"
        f"&redirect_uri={REDIRECT_URI}"
        f"&scope=offline_access"
    )
    
    if request.method == 'POST':
        accion = request.form.get('action', 'token_manual')
        
        # ─── OPCIÓN 1: Canjear code por token automáticamente ───
        if accion == 'canjear_code':
            code = request.form.get('code', '').strip()
            
            if not code:
                flash('❌ Debes ingresar el CODE de la URL de redirección', 'error')
                return redirect(url_for('ml_configurar_token'))
            
            try:
                response = requests.post(
                    "https://api.mercadolibre.com/oauth/token",
                    data={
                        "grant_type": "authorization_code",
                        "client_id": CLIENT_ID,
                        "client_secret": CLIENT_SECRET,
                        "code": code,
                        "redirect_uri": REDIRECT_URI
                    }
                )
                
                if response.status_code == 200:
                    data = response.json()
                    
                    token_data = {
                        "access_token": data.get("access_token"),
                        "refresh_token": data.get("refresh_token"),  # None si ML no lo da
                        "expires_at": time.time() + data.get("expires_in", 21600) - 300,
                        "client_id": CLIENT_ID,
                        "client_secret": CLIENT_SECRET
                    }
                    
                    if guardar_ml_token(token_data):
                        if token_data.get("refresh_token"):
                            flash('✅ Token configurado con auto-renovación activada', 'success')
                        else:
                            flash('✅ Token configurado correctamente (válido por 6 horas)', 'success')
                        return redirect(url_for('ventas_activas'))
                    else:
                        flash('❌ Error al guardar el token', 'error')
                
                else:
                    error_msg = response.json().get('message', 'Error desconocido')
                    flash(f'❌ Error al canjear el code: {error_msg}', 'error')
                    
            except Exception as e:
                flash(f'❌ Error: {str(e)}', 'error')
            
            return redirect(url_for('ml_configurar_token'))
        
        # ─── OPCIÓN 2: Pegar token manual (como antes) ───
        else:
            access_token = request.form.get('access_token', '').strip()
            
            if not access_token:
                flash('❌ Debes ingresar un ACCESS_TOKEN', 'error')
                return redirect(url_for('ml_configurar_token'))
            
            token_data = {
                'access_token': access_token,
                'fecha_configuracion': datetime.now().isoformat()
            }
            
            if guardar_ml_token(token_data):
                flash('✅ Token configurado correctamente', 'success')
                return redirect(url_for('ventas_activas'))
            else:
                flash('❌ Error al guardar token', 'error')
    
    # Verificar estado actual del token
    token_actual = cargar_ml_token()
    
    # Ver si tiene auto-refresh
    tiene_refresh = False
    horas_restantes = None
    try:
        token_path = 'config/ml_token.json'
        if os.path.exists(token_path):
            with open(token_path, 'r') as f:
                data = json.load(f)
                tiene_refresh = bool(data.get('refresh_token'))
                expires_at = data.get('expires_at')
                if expires_at:
                    segundos = expires_at - time.time()
                    if segundos > 0:
                        horas_restantes = round(segundos / 3600, 1)
    except:
        pass
    
    return render_template('ml_configurar_token.html',
                          token_actual=token_actual,
                          url_autorizacion=url_autorizacion,
                          tiene_refresh=tiene_refresh,
                          horas_restantes=horas_restantes)



@app.route('/ventas/ml/importar')
@login_required
def ml_importar_ordenes():
    """
    Traer órdenes de ML - FILTRO ARREGLADO
    """
    # Guardar hora de corte si viene en el request
    hora_corte = request.args.get('hora_corte', '').strip()
    if hora_corte:
        session['hora_corte_colecta'] = hora_corte

    # Si solo_guardar=1, guardar en session y volver sin importar
    if request.args.get('solo_guardar') == '1':
        return ('', 204)

    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay ACCESS_TOKEN configurado. Configuralo primero.', 'error')
        return redirect(url_for('ml_configurar_token'))
    
    # Obtener órdenes de ML
    success, result = obtener_ordenes_ml(access_token, limit=50)
    
    if not success:
        flash(f'❌ Error al obtener órdenes de ML: {result}', 'error')
        return redirect(url_for('ventas_activas'))
    
    # 🔧 FILTRO ARREGLADO
    ordenes_importadas = set()
    try:
        # Buscar en TODAS las ventas (no solo las que empiezan con ML-)
        # porque el numero_venta puede estar en cualquier formato
        todas_ventas = query_db('''
            SELECT id, numero_venta, fecha_venta, canal, mla_code 
            FROM ventas 
            ORDER BY fecha_venta DESC 
            LIMIT 200
        ''')
        
        print(f"\n📊 DEBUG FILTRO DE DUPLICADOS:")
        print(f"Total ventas en BD (últimas 200): {len(todas_ventas)}")
        
        # Revisar cada venta y ver si el numero_venta contiene un ID de ML
        for venta in todas_ventas:
            numero = venta['numero_venta']
            
            # Extraer números del numero_venta
            # Puede ser: "ML-2000015174126338", "2000015174126338", etc.
            import re
            numeros = re.findall(r'\d+', numero)
            
            for num in numeros:
                # Los IDs de ML tienen 16 dígitos y empiezan con 2000
                if len(num) == 16 and num.startswith('2000'):
                    ordenes_importadas.add(num)
                    print(f"   ✅ Encontrado: {numero} → ID ML: {num}")
        
        print(f"\n✅ Total IDs de ML ya importados: {len(ordenes_importadas)}\n")
        
    except Exception as e:
        print(f"⚠️ Error al obtener órdenes importadas: {e}")
        import traceback
        traceback.print_exc()
    
    # Procesar órdenes
    ordenes_procesadas = []
    ordenes_filtradas = 0
    
    for orden in result:
        orden_id = str(orden['id'])
        
        print(f"🔍 Orden ML: {orden_id}")
        
        if orden_id in ordenes_importadas:
            print(f"   ⏭️ YA IMPORTADA - Saltando")
            ordenes_filtradas += 1
            continue
        else:
            print(f"   ✅ NUEVA - Agregando")
        
        # Solo mostrar órdenes pagadas
        if orden['status'] in ['paid']:
            orden_data = procesar_orden_ml(orden)
            
            # Verificar SKU
            for item in orden_data['items']:
                sku = item['sku']
                if sku:
                    existe, tipo, nombre = verificar_sku_en_bd(sku)
                    item['existe_en_bd'] = existe
                    item['tipo_producto'] = tipo
                    item['nombre_bd'] = nombre
                else:
                    item['existe_en_bd'] = False
                    item['tipo_producto'] = None
                    item['nombre_bd'] = None
            
            ordenes_procesadas.append(orden_data)
    
    print(f"\n📊 RESUMEN:")
    print(f"   Órdenes en ML: {len(result)}")
    print(f"   Órdenes filtradas: {ordenes_filtradas}")
    print(f"   Órdenes a mostrar: {len(ordenes_procesadas)}\n")
    
    return render_template('ml_importar_ordenes.html', ordenes=ordenes_procesadas)


@app.route('/ventas/ml/seleccionar/<orden_id>')
@login_required
def ml_seleccionar_orden(orden_id):
    """
    Seleccionar orden - Con normalización automática de SKU (quita Z)
    GUARDA FECHA REAL DE VENTA Y DATOS DE FACTURACIÓN EN SESIÓN
    ✅ CORREGIDO: No sobrescribe costo_envio (ya viene del shipment)
    """
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay ACCESS_TOKEN configurado', 'error')
        return redirect(url_for('ml_configurar_token'))
    
    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        response = requests.get(f'https://api.mercadolibre.com/orders/{orden_id}', headers=headers)
        
        if response.status_code != 200:
            flash('❌ Error al obtener orden de ML', 'error')
            return redirect(url_for('ventas_activas'))
        
        orden = response.json()
        orden_data = procesar_orden_ml(orden)
        
        # OBTENER SHIPPING COMPLETO
        if orden_data['shipping']['shipping_id']:
            shipping_completo = obtener_shipping_completo(
                orden_data['shipping']['shipping_id'],
                access_token,
                orden_data.get('fecha_iso', '')
            )
            # ✅ CORREGIDO: Ya no sobrescribimos costo_envio
            # porque obtener_shipping_completo() ya lo captura del shipment
            orden_data['shipping'] = shipping_completo
        
        # ✅ NUEVO: OBTENER BILLING INFO (FACTURACIÓN)
        billing_data = None
        try:
            billing_response = requests.get(
                f'https://api.mercadolibre.com/orders/{orden_id}/billing_info',
                headers=headers
            )
            
            if billing_response.status_code == 200:
                billing_data = billing_response.json()
                print(f"✅ Billing info obtenida para orden {orden_id}")
            else:
                print(f"ℹ️ No hay billing info para orden {orden_id} (Consumidor Final)")
        
        except Exception as e:
            print(f"⚠️ Error al obtener billing info: {str(e)}")
        
        # Verificar mapeo de SKU CON NORMALIZACIÓN AUTOMÁTICA
        items_sin_mapear = []
        items_mapeados = []
        
        # logistic_type para mapeo automático de Compac (_DEP vs _FULL)
        logistic_type_actual = orden_data['shipping'].get('logistic_type_ml', '')
        es_full_ml = logistic_type_actual == 'fulfillment'

        for item in orden_data['items']:
            sku_ml_original = item['sku']
            if sku_ml_original:
                sku_a_usar, cant_override = normalizar_sku_ml(sku_ml_original)
                cantidad_final = cant_override if cant_override > 0 else item['cantidad']
                existe, tipo, nombre = verificar_sku_en_bd(sku_a_usar)

                # Auto-mapeo Compac: CCO{medida} → CCO{medida}_FULL o CCO{medida}_DEP
                if not existe and sku_a_usar.upper().startswith('CCO') and '_' not in sku_a_usar:
                    sufijo = '_FULL' if es_full_ml else '_DEP'
                    sku_compac = sku_a_usar.upper() + sufijo
                    existe, tipo, nombre = verificar_sku_en_bd(sku_compac)
                    if existe:
                        sku_a_usar = sku_compac
                        print(f"✅ Mapeo automático Compac: {sku_ml_original} → {sku_compac} (logistic: {logistic_type_actual})")

                if existe:
                    items_mapeados.append({
                        'sku_ml': sku_ml_original,
                        'sku_bd': sku_a_usar,
                        'titulo': item['titulo'],
                        'cantidad': cantidad_final,
                        'precio': item['precio'],
                        'nombre_bd': nombre
                    })
                else:
                    items_sin_mapear.append(item)
            else:
                items_sin_mapear.append(item)
        
        # Mapeo manual si hay productos sin mapear
        if items_sin_mapear:
            productos_bd = query_db('SELECT sku, nombre, tipo FROM productos_base ORDER BY nombre')
            combos_bd = query_db('SELECT sku, nombre FROM productos_compuestos ORDER BY nombre')
            
            return render_template('ml_mapear_productos.html',
                                 orden_id=orden_id,
                                 items_sin_mapear=items_sin_mapear,
                                 items_mapeados=items_mapeados,
                                 productos_bd=productos_bd,
                                 combos_bd=combos_bd,
                                 orden_data=orden_data)
        
        # ✅ GUARDAR EN SESIÓN CON FECHA REAL DE ML Y BILLING INFO
        from flask import session
        session['ml_orden_id'] = orden_id
        session['ml_items'] = items_mapeados
        session['ml_comprador_nombre'] = orden_data['comprador_nombre']
        session['ml_comprador_nickname'] = orden_data['comprador_nickname']
        session['ml_shipping'] = orden_data['shipping']
        session['ml_fecha_venta'] = orden_data['fecha'].isoformat()
        
        # ✅ NUEVO: Guardar billing info en sesión
        if billing_data:
            session['ml_billing_data'] = billing_data
        else:
            session['ml_billing_data'] = None
        
        flash('✅ Productos mapeados correctamente', 'success')
        return redirect(url_for('nueva_venta_desde_ml'))
        
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_activas'))


@app.route('/ventas/ml/mapear', methods=['POST'])
@login_required
def ml_guardar_mapeo():
    """
    Guardar mapeo - Obtiene shipping completo y billing info
    Con normalización automática de SKU (quita Z)
    GUARDA FECHA REAL DE VENTA Y DATOS DE FACTURACIÓN EN SESIÓN
    ✅ CORREGIDO: No sobrescribe costo_envio
    """
    orden_id = request.form.get('orden_id')
    items_mapeados = json.loads(request.form.get('items_mapeados', '[]'))
    items_form = request.form.getlist('item_sku_ml')
    
    for i, sku_ml in enumerate(items_form):
        sku_bd = request.form.get(f'mapeo_{i}')
        titulo = request.form.get(f'titulo_{i}')
        cantidad = int(request.form.get(f'cantidad_{i}'))
        precio = float(request.form.get(f'precio_{i}'))
        
        if sku_bd:
            existe, tipo, nombre = verificar_sku_en_bd(sku_bd)
            if existe:
                items_mapeados.append({
                    'sku_ml': sku_ml,
                    'sku_bd': sku_bd,
                    'titulo': titulo,
                    'cantidad': cantidad,
                    'precio': precio,
                    'nombre_bd': nombre
                })
    
    access_token = cargar_ml_token()
    
    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        response = requests.get(f'https://api.mercadolibre.com/orders/{orden_id}', headers=headers)
        
        if response.status_code == 200:
            orden = response.json()
            orden_data = procesar_orden_ml(orden)
            
            # OBTENER SHIPPING COMPLETO
            if orden_data['shipping']['shipping_id']:
                shipping_completo = obtener_shipping_completo(
                    orden_data['shipping']['shipping_id'],
                    access_token,
                    orden_data.get('fecha_iso', '')
                )
                # ✅ CORREGIDO: Ya no sobrescribimos costo_envio
                orden_data['shipping'] = shipping_completo
            
            # ✅ NUEVO: OBTENER BILLING INFO
            billing_data = None
            try:
                billing_response = requests.get(
                    f'https://api.mercadolibre.com/orders/{orden_id}/billing_info',
                    headers=headers
                )
                
                if billing_response.status_code == 200:
                    billing_data = billing_response.json()
            except:
                pass
            
            # ✅ GUARDAR EN SESIÓN CON FECHA REAL Y BILLING
            from flask import session
            session['ml_orden_id'] = orden_id
            session['ml_items'] = items_mapeados
            session['ml_comprador_nombre'] = orden_data['comprador_nombre']
            session['ml_comprador_nickname'] = orden_data['comprador_nickname']
            session['ml_shipping'] = orden_data['shipping']
            session['ml_fecha_venta'] = orden_data['fecha'].isoformat()
            session['ml_billing_data'] = billing_data
        else:
            from flask import session
            session['ml_orden_id'] = orden_id
            session['ml_items'] = items_mapeados
            session['ml_comprador_nombre'] = ''
            session['ml_comprador_nickname'] = ''
            session['ml_shipping'] = {}
            session['ml_fecha_venta'] = datetime.now().isoformat()
            session['ml_billing_data'] = None
    
    except Exception as e:
        from flask import session
        session['ml_orden_id'] = orden_id
        session['ml_items'] = items_mapeados
        session['ml_comprador_nombre'] = ''
        session['ml_comprador_nickname'] = ''
        session['ml_shipping'] = {}
        session['ml_fecha_venta'] = datetime.now().isoformat()
        session['ml_billing_data'] = None
        import traceback
        traceback.print_exc()
    
    flash('✅ Productos mapeados correctamente', 'success')
    return redirect(url_for('nueva_venta_desde_ml'))



@app.route('/ventas/nueva/ml')
@login_required
def nueva_venta_desde_ml():
    """
    Crear nueva venta con datos precargados desde ML
    ✅ CORREGIDO: Pasa ml_shipping al template
    """
    from flask import session
    
    print("\n" + "="*70)
    print("🔍 DEBUG: Cargando nueva venta desde ML")
    print("="*70)
    
    if 'ml_items' not in session:
        flash('❌ No hay datos de ML para importar', 'error')
        return redirect(url_for('ventas_activas'))
    
    ml_items = session.get('ml_items', [])
    ml_orden_id = session.get('ml_orden_id', '')
    ml_shipping = session.get('ml_shipping', {})
    
    print(f"📦 Orden ID: {ml_orden_id}")
    print(f"🛍️ Items: {len(ml_items)} productos")
    print(f"🚚 Shipping data en sesión: {ml_shipping}")
    
    # Obtener datos necesarios para el formulario
    productos = query_db('SELECT * FROM productos_base ORDER BY nombre')
    combos = query_db('SELECT * FROM productos_compuestos ORDER BY nombre')
    
    print("="*70 + "\n")
    
    return render_template('nueva_venta_ml.html',
                         productos=productos,
                         combos=combos,
                         ml_items=ml_items,
                         ml_orden_id=ml_orden_id,
                         ml_shipping=ml_shipping)  # ✅ NUEVO: Pasar ml_shipping



import requests
import json
from pprint import pprint

def debug_orden_ml_completa(orden_id, access_token):
    """
    Ver TODOS los datos que trae ML de una orden
    Incluyendo billing_info, buyer, shipping, payments, etc.
    """
    headers = {'Authorization': f'Bearer {access_token}'}
    
    print("\n" + "="*80)
    print(f"🔍 DEBUG COMPLETO - ORDEN ML: {orden_id}")
    print("="*80 + "\n")
    
    # ============================================
    # 1. ORDEN COMPLETA
    # ============================================
    print("📦 1. INFORMACIÓN COMPLETA DE LA ORDEN:")
    print("-" * 80)
    
    try:
        response = requests.get(
            f'https://api.mercadolibre.com/orders/{orden_id}',
            headers=headers
        )
        
        if response.status_code != 200:
            print(f"❌ Error {response.status_code}: {response.text}")
            return
        
        orden = response.json()
        
        # Guardar JSON completo en archivo
        with open(f'debug_orden_{orden_id}.json', 'w', encoding='utf-8') as f:
            json.dump(orden, f, indent=2, ensure_ascii=False)
        
        print(f"✅ JSON completo guardado en: debug_orden_{orden_id}.json")
        print()
        
        # Mostrar estructura principal
        print("📋 ESTRUCTURA PRINCIPAL:")
        for key in orden.keys():
            valor = orden[key]
            tipo = type(valor).__name__
            
            if isinstance(valor, (list, dict)):
                if isinstance(valor, list):
                    longitud = f"[{len(valor)} items]"
                else:
                    longitud = f"{{{len(valor)} keys}}"
                print(f"   • {key}: {tipo} {longitud}")
            else:
                print(f"   • {key}: {tipo} = {valor}")
        
        print()
        
    except Exception as e:
        print(f"❌ Error al obtener orden: {str(e)}")
        return
    
    # ============================================
    # 2. DATOS DEL COMPRADOR (BUYER)
    # ============================================
    print("\n👤 2. DATOS DEL COMPRADOR (buyer):")
    print("-" * 80)
    
    buyer = orden.get('buyer', {})
    
    if buyer:
        print("📋 Campos disponibles en 'buyer':")
        pprint(buyer, width=120)
    else:
        print("⚠️ No hay datos de buyer")
    
    print()
    
    # ============================================
    # 3. BILLING INFO (FACTURACIÓN)
    # ============================================
    print("\n🧾 3. INFORMACIÓN DE FACTURACIÓN (billing_info):")
    print("-" * 80)
    
    try:
        billing_response = requests.get(
            f'https://api.mercadolibre.com/orders/{orden_id}/billing_info',
            headers=headers
        )
        
        if billing_response.status_code == 200:
            billing = billing_response.json()
            
            print("✅ BILLING INFO DISPONIBLE:")
            print()
            
            # Guardar JSON de billing
            with open(f'debug_billing_{orden_id}.json', 'w', encoding='utf-8') as f:
                json.dump(billing, f, indent=2, ensure_ascii=False)
            
            print(f"✅ JSON guardado en: debug_billing_{orden_id}.json")
            print()
            
            print("📋 DATOS DE FACTURACIÓN:")
            pprint(billing, width=120)
            
        elif billing_response.status_code == 404:
            print("⚠️ NO HAY BILLING INFO (404 - Not Found)")
            print()
            print("Esto es normal si:")
            print("   • No sos Responsable Inscripto")
            print("   • La venta no requiere factura")
            print("   • El comprador es Consumidor Final sin CUIT")
        
        else:
            print(f"⚠️ Error {billing_response.status_code}: {billing_response.text}")
    
    except Exception as e:
        print(f"❌ Error al obtener billing_info: {str(e)}")
    
    print()
    
    # ============================================
    # 4. SHIPPING (ENVÍO)
    # ============================================
    print("\n🚚 4. DATOS DE ENVÍO (shipping):")
    print("-" * 80)
    
    shipping = orden.get('shipping', {})
    
    if shipping:
        print("📋 Campos disponibles en 'shipping':")
        pprint(shipping, width=120)
        
        # Si hay shipping_id, obtener detalles completos
        shipping_id = shipping.get('id')
        if shipping_id:
            print()
            print(f"📦 Obteniendo detalles completos del envío {shipping_id}...")
            
            try:
                ship_response = requests.get(
                    f'https://api.mercadolibre.com/shipments/{shipping_id}',
                    headers=headers
                )
                
                if ship_response.status_code == 200:
                    shipment = ship_response.json()
                    
                    with open(f'debug_shipment_{orden_id}.json', 'w', encoding='utf-8') as f:
                        json.dump(shipment, f, indent=2, ensure_ascii=False)
                    
                    print(f"✅ Shipment guardado en: debug_shipment_{orden_id}.json")
                    print()
                    
                    print("📍 DIRECCIÓN DE ENVÍO:")
                    receiver = shipment.get('receiver_address', {})
                    pprint(receiver, width=120)
                
            except Exception as e:
                print(f"⚠️ Error al obtener shipment: {str(e)}")
    else:
        print("⚠️ No hay datos de shipping")
    
    print()
    
    # ============================================
    # 5. PAYMENTS (PAGOS)
    # ============================================
    print("\n💳 5. INFORMACIÓN DE PAGOS (payments):")
    print("-" * 80)
    
    payments = orden.get('payments', [])
    
    if payments:
        for i, pago in enumerate(payments, 1):
            print(f"\n💰 Pago #{i}:")
            pprint(pago, width=120)
    else:
        print("⚠️ No hay datos de pagos")
    
    print()
    
    # ============================================
    # 6. ORDER ITEMS (PRODUCTOS)
    # ============================================
    print("\n📦 6. PRODUCTOS DE LA ORDEN (order_items):")
    print("-" * 80)
    
    items = orden.get('order_items', [])
    
    if items:
        for i, item in enumerate(items, 1):
            print(f"\n🛒 Item #{i}:")
            pprint(item, width=120)
    else:
        print("⚠️ No hay items")
    
    print()
    
    # ============================================
    # 7. RESUMEN
    # ============================================
    print("\n" + "="*80)
    print("📊 RESUMEN - DATOS ÚTILES PARA FACTURACIÓN:")
    print("="*80 + "\n")
    
    print("✅ DATOS DISPONIBLES:")
    print()
    
    # Buyer
    print("👤 COMPRADOR (buyer):")
    print(f"   • ID: {buyer.get('id', 'N/A')}")
    print(f"   • Nickname: {buyer.get('nickname', 'N/A')}")
    print(f"   • Email: {buyer.get('email', 'N/A')}")
    print(f"   • First name: {buyer.get('first_name', 'N/A')}")
    print(f"   • Last name: {buyer.get('last_name', 'N/A')}")
    print(f"   • Phone: {buyer.get('phone', {}).get('number', 'N/A')}")
    print()
    
    # Billing (si existe)
    try:
        if billing_response.status_code == 200:
            billing = billing_response.json()
            print("🧾 FACTURACIÓN (billing_info):")
            print(f"   • Doc type: {billing.get('doc_type', 'N/A')}")
            print(f"   • Doc number: {billing.get('doc_number', 'N/A')}")
            
            # Buscar más datos
            if 'additional_info' in billing:
                print(f"   • Additional info: {billing.get('additional_info')}")
            
            print()
        else:
            print("🧾 FACTURACIÓN: No disponible (consumidor final)")
            print()
    except:
        pass
    
    # Totales
    print("💰 TOTALES:")
    print(f"   • Total amount: ${orden.get('total_amount', 0)}")
    print(f"   • Paid amount: ${orden.get('paid_amount', 0)}")
    
    if payments:
        fee = payments[0].get('marketplace_fee', 0)
        print(f"   • Marketplace fee: ${fee}")
        print(f"   • Neto vendedor: ${orden.get('total_amount', 0) - fee}")
    
    print()
    
    # Dirección
    if shipping and 'receiver_address' in shipment:
        addr = shipment.get('receiver_address', {})
        print("📍 DIRECCIÓN:")
        print(f"   • Address line: {addr.get('address_line', 'N/A')}")
        print(f"   • City: {addr.get('city', {}).get('name', 'N/A')}")
        print(f"   • State: {addr.get('state', {}).get('name', 'N/A')}")
        print(f"   • Zip code: {addr.get('zip_code', 'N/A')}")
        print()
    
    print("="*80)
    print("✅ DEBUG COMPLETO - Revisá los archivos JSON generados")
    print("="*80 + "\n")

def extraer_billing_info_ml(billing_data):
    """
    Extraer datos de facturación del formato de ML
    
    ML devuelve los datos en formato:
    {
      "billing_info": {
        "additional_info": [
          {"type": "BUSINESS_NAME", "value": "FERNANDEZ GONZALO JAVIER"},
          {"type": "DOC_TYPE", "value": "CUIT"},
          ...
        ]
      }
    }
    
    Returns:
        dict con los datos extraídos
    """
    datos = {
        'business_name': None,
        'doc_type': None,
        'doc_number': None,
        'taxpayer_type': None,
        'city': None,
        'street': None,
        'state': None,
        'zip_code': None
    }
    
    try:
        # Obtener billing_info
        billing_info = billing_data.get('billing_info', {})
        
        # Doc type y number están en el nivel principal
        datos['doc_type'] = billing_info.get('doc_type')
        datos['doc_number'] = billing_info.get('doc_number')
        
        # Los demás datos están en additional_info
        additional_info = billing_info.get('additional_info', [])
        
        # Crear diccionario para búsqueda fácil
        info_dict = {}
        for item in additional_info:
            tipo = item.get('type')
            valor = item.get('value')
            if tipo and valor:
                info_dict[tipo] = valor
        
        # Extraer campos
        datos['business_name'] = info_dict.get('BUSINESS_NAME')
        datos['taxpayer_type'] = info_dict.get('TAXPAYER_TYPE_ID')
        datos['city'] = info_dict.get('CITY_NAME')
        datos['street'] = info_dict.get('STREET_NAME')
        datos['state'] = info_dict.get('STATE_NAME')
        datos['zip_code'] = info_dict.get('ZIP_CODE')
        
        # Si no hay doc_type en el nivel principal, intentar de additional_info
        if not datos['doc_type']:
            datos['doc_type'] = info_dict.get('DOC_TYPE')
        if not datos['doc_number']:
            datos['doc_number'] = info_dict.get('DOC_NUMBER')
        
        print(f"✅ Datos de facturación extraídos:")
        print(f"   • Razón social: {datos['business_name']}")
        print(f"   • {datos['doc_type']}: {datos['doc_number']}")
        print(f"   • Condición IVA: {datos['taxpayer_type']}")
        
    except Exception as e:
        print(f"⚠️ Error al extraer billing_info: {str(e)}")
    
    return datos



# ============================================================================
# RUTA TEMPORAL PARA USAR EN FLASK
# ============================================================================

@app.route('/debug/ml/<orden_id>')
@login_required
def debug_ml_orden(orden_id):
    """Ver qué datos trae ML de una orden"""
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('No hay token de ML', 'error')
        return redirect(url_for('ventas_activas'))
    
    headers = {'Authorization': f'Bearer {access_token}'}
    
    try:
        # Obtener orden
        response = requests.get(f'https://api.mercadolibre.com/orders/{orden_id}', headers=headers)
        
        if response.status_code != 200:
            return f"<h2>Error {response.status_code}</h2>"
        
        orden = response.json()
        buyer = orden.get('buyer', {})
        
        # Intentar obtener billing
        billing_html = ""
        try:
            billing_response = requests.get(
                f'https://api.mercadolibre.com/orders/{orden_id}/billing_info',
                headers=headers
            )
            
            if billing_response.status_code == 200:
                billing = billing_response.json()
                import json
                billing_json = json.dumps(billing, indent=2, ensure_ascii=False)
                billing_html = f"<h3>TIENE DATOS DE FACTURACION:</h3><pre>{billing_json}</pre>"
            else:
                billing_html = "<h3>NO TIENE DATOS DE FACTURACION (Consumidor Final)</h3>"
        except:
            billing_html = "<h3>Error al obtener billing</h3>"
        
        # Construir HTML de respuesta
        html_response = """
        <html>
        <head>
            <title>Debug ML Orden</title>
            <style>
                body {{ font-family: monospace; padding: 20px; background: #1e1e1e; color: #d4d4d4; }}
                pre {{ background: #2d2d2d; padding: 15px; border-radius: 5px; overflow-x: auto; }}
                h2 {{ color: #4fc3f7; }}
                h3 {{ color: #ce9178; }}
            </style>
        </head>
        <body>
            <h2>DEBUG ORDEN ML: {orden_id}</h2>
            
            <h3>COMPRADOR:</h3>
            <p>Nombre: {nombre}</p>
            <p>Nickname: {nickname}</p>
            <p>Email: {email}</p>
            
            {billing_info}
            
            <br><br>
            <a href="/ventas/activas" style="color: #4fc3f7;">Volver a Ventas Activas</a>
        </body>
        </html>
        """.format(
            orden_id=orden_id,
            nombre=f"{buyer.get('first_name', '')} {buyer.get('last_name', '')}",
            nickname=buyer.get('nickname', ''),
            email=buyer.get('email', 'No disponible'),
            billing_info=billing_html
        )
        
        return html_response
        
    except Exception as e:
        import traceback
        error_completo = traceback.format_exc()
        return f"<pre>Error: {error_completo}</pre>"

# ============================================================================
# FUNCIÓN OBTENER DATOS ML - CON STATUS REAL
# ============================================================================

CAMPAÑAS_CUOTAS = {
    'pcj-co-funded': 'Cuota Simple',
    '3x_campaign':  '3 cuotas s/interés',
    '6x_campaign':  '6 cuotas s/interés',
    '9x_campaign':  '9 cuotas s/interés',
    '12x_campaign': '12 cuotas s/interés',
}

def obtener_datos_ml(mla_id, access_token):
    """
    Consulta datos actuales de una publicación ML.
    Devuelve: titulo, stock, status, demora, precio, listing_type
    """
    CAMPAÑAS_CUOTAS = {
        'pcj-co-funded': 'Cuota Simple',
        '3x_campaign':   '3 cuotas s/interés',
        '6x_campaign':   '6 cuotas s/interés',
        '9x_campaign':   '9 cuotas s/interés',
        '12x_campaign':  '12 cuotas s/interés',
    }

    try:
        r = ml_request('get', f'https://api.mercadolibre.com/items/{mla_id}', access_token)

        if r.status_code != 200:
            return {
                'titulo': mla_id, 'stock': 0, 'status': 'unknown',
                'demora': None, 'precio': None, 'listing_type': None
            }

        data = r.json()

        demora = None
        campaign = None
        for term in data.get('sale_terms', []):
            if term.get('id') == 'MANUFACTURING_TIME':
                demora = term.get('value_name')
            if term.get('id') == 'INSTALLMENTS_CAMPAIGN':
                campaign = (term.get('value_name') or '').split('|')[0].strip()

        listing_type_id = data.get('listing_type_id', '')
        if listing_type_id == 'gold_special':
            financiacion = CAMPAÑAS_CUOTAS.get(campaign, 'Sin cuotas propias') if campaign else 'Sin cuotas propias'
        elif listing_type_id == 'gold_pro':
            financiacion = CAMPAÑAS_CUOTAS.get(campaign, '6 cuotas s/interés')
        else:
            financiacion = listing_type_id or '-'

        return {
            'titulo':       data.get('title', mla_id),
            'stock':        data.get('available_quantity', 0),
            'status':       data.get('status', 'unknown'),
            'demora':       demora,
            'precio':       data.get('price'),
            'listing_type': financiacion,
        }

    except Exception as e:
        print(f"Error obteniendo datos ML de {mla_id}: {e}")
        return {
            'titulo': mla_id, 'stock': 0, 'status': 'unknown',
            'demora': None, 'precio': None, 'listing_type': None
        }
# ============================================================================
# RUTAS CARGAR STOCK ML - CORREGIDAS CON COLUMNAS REALES
# ============================================================================

@app.route('/cargar-stock-ml', methods=['GET'])
@login_required
def cargar_stock_ml():
    """Mostrar página para cargar stock en ML"""
    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'porcentajes_ml'")
        porcentajes = json.loads(row['valor']) if row else PORCENTAJES_ML_DEFAULT
    except:
        porcentajes = PORCENTAJES_ML_DEFAULT

    return render_template('cargar_stock_ml.html',
                           sku_buscado=None,
                           publicaciones=[],
                           es_sku_con_z=False,
                           mensaje=None,
                           mensaje_tipo=None,
                           porcentajes=porcentajes,
                           precio_costos=None)

# ============================================================================
# RUTA BUSCAR SKU - ACTUALIZADA CON STATUS REAL DE ML
# ============================================================================


def ml_request(method, url, access_token, json_data=None, params=None, max_retries=4):
    """
    Helper para requests a ML con rate limiting global + retry exponencial.
    Toda llamada a ML debe pasar por acá.
    """
    global _ml_last_request

    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}

    for attempt in range(max_retries):
        # Rate limiting global: esperar si la última request fue muy reciente
        with _ml_rate_lock:
            now = time.time()
            elapsed = now - _ml_last_request
            if elapsed < _ML_MIN_INTERVAL:
                time.sleep(_ML_MIN_INTERVAL - elapsed)
            _ml_last_request = time.time()

        try:
            if method == 'get':
                r = requests.get(url, headers=headers, params=params, timeout=15)
            elif method == 'post':
                r = requests.post(url, headers=headers, json=json_data, timeout=15)
            else:
                r = requests.put(url, headers=headers, json=json_data, timeout=15)

            if r.status_code == 429:
                wait = min(5 * (2 ** attempt), 60)  # 5s, 10s, 20s, 40s
                print(f"⚠️ 429 ML - esperando {wait}s (intento {attempt+1}/{max_retries})")
                time.sleep(wait)
                continue

            return r

        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(3)
                continue
            raise

    return r

def obtener_datos_ml_batch(mla_ids, access_token):
    """
    Consulta datos de múltiples publicaciones ML en chunks de 20 (límite de la API).
    Devuelve dict: { mla_id: { titulo, stock, status, demora, precio, listing_type } }
    """
    if not mla_ids:
        return {}

    CAMPAÑAS_CUOTAS = {
        'pcj-co-funded': 'Cuota Simple',
        '3x_campaign':   '3 cuotas s/interés',
        '6x_campaign':   '6 cuotas s/interés',
        '9x_campaign':   '9 cuotas s/interés',
        '12x_campaign':  '12 cuotas s/interés',
    }

    resultado = {}

    # Cortar en chunks de 20 (límite de la API de ML)
    chunks = [mla_ids[i:i+20] for i in range(0, len(mla_ids), 20)]

    for chunk in chunks:
        try:
            r = ml_request('get', 'https://api.mercadolibre.com/items', access_token,
                           params={'ids': ','.join(chunk)})

            if r.status_code != 200:
                for mla_id in chunk:
                    resultado[mla_id] = {
                        'titulo': mla_id, 'stock': 0, 'status': 'unknown',
                        'demora': None, 'precio': None, 'listing_type': None, 'status_raw': 'unknown'
                    }
                continue

            items = r.json()
            for item in items:
                if item.get('code') != 200:
                    mla_id = item.get('body', {}).get('id', '')
                    resultado[mla_id] = {
                        'titulo': mla_id, 'stock': 0, 'status': 'unknown',
                        'demora': None, 'precio': None, 'listing_type': None, 'status_raw': 'unknown'
                    }
                    continue

                data = item['body']
                mla_id = data.get('id', '')

                demora = None
                campaign = None
                for term in data.get('sale_terms', []):
                    if term.get('id') == 'MANUFACTURING_TIME':
                        demora = term.get('value_name')
                    if term.get('id') == 'INSTALLMENTS_CAMPAIGN':
                        campaign = (term.get('value_name') or '').split('|')[0].strip()

                listing_type_id = data.get('listing_type_id', '')
                if listing_type_id == 'gold_special':
                    financiacion = CAMPAÑAS_CUOTAS.get(campaign, 'Sin cuotas propias') if campaign else 'Sin cuotas propias'
                elif listing_type_id == 'gold_pro':
                    financiacion = CAMPAÑAS_CUOTAS.get(campaign, '6 cuotas s/interés')
                else:
                    financiacion = listing_type_id

                # Extraer seller_sku de attributes o seller_custom_field
                seller_sku = data.get('seller_custom_field') or ''
                if not seller_sku:
                    for attr in data.get('attributes', []):
                        if attr.get('id') == 'SELLER_SKU':
                            seller_sku = attr.get('value_name', '')
                            break

                resultado[mla_id] = {
                    'titulo':              data.get('title', mla_id),
                    'stock':               data.get('available_quantity', 0),
                    'status':              data.get('status', 'unknown'),
                    'status_raw':          data.get('status', 'unknown'),
                    'demora':              demora,
                    'precio':              data.get('price'),
                    'listing_type':        financiacion,
                    'permalink':           data.get('permalink', ''),
                    'catalog_listing':     data.get('catalog_listing', False),
                    'catalog_product_id':  data.get('catalog_product_id'),
                    'category_id':         data.get('category_id'),
                    'domain_id':           data.get('domain_id'),
                    'seller_sku':          seller_sku,
                }

        except Exception as e:
            print(f"Error en obtener_datos_ml_batch (chunk {chunk}): {e}")
            for mla_id in chunk:
                resultado[mla_id] = {
                    'titulo': mla_id, 'stock': 0, 'status': 'unknown',
                    'demora': None, 'precio': None, 'listing_type': None, 'status_raw': 'unknown'
                }

    return resultado


def obtener_permalinks_ml(mla_ids, access_token):
    """
    Devuelve permalinks para los MLAs dados.
    Usa caché en sku_mla_mapeo.permalink — solo llama a la API si está vacío.
    """
    # Agregar columna permalink si no existe (idempotente)
    try:
        execute_db("ALTER TABLE sku_mla_mapeo ADD COLUMN permalink VARCHAR(500) DEFAULT NULL")
    except Exception:
        pass  # ya existe

    permalinks = {}
    sin_cache = []

    # Primero buscar en caché
    if mla_ids:
        filas = query_db(
            "SELECT mla_id, permalink FROM sku_mla_mapeo WHERE mla_id IN ({})".format(
                ','.join(['%s'] * len(mla_ids))), tuple(mla_ids))
        for f in filas:
            if f['permalink']:
                permalinks[f['mla_id']] = f['permalink']
            else:
                sin_cache.append(f['mla_id'])
        # MLAs que no están en la tabla
        en_tabla = {f['mla_id'] for f in filas}
        sin_cache += [m for m in mla_ids if m not in en_tabla]

    # Consultar API solo para los que no tienen caché
    for mla_id in sin_cache:
        try:
            r = ml_request('get', f'https://api.mercadolibre.com/items/{mla_id}',
                           access_token, params={'attributes': 'id,permalink'})
            if r.status_code == 200:
                url = r.json().get('permalink', '')
                permalinks[mla_id] = url
                if url:
                    execute_db(
                        "UPDATE sku_mla_mapeo SET permalink=%s WHERE mla_id=%s",
                        (url, mla_id))
        except Exception as e:
            print(f"Error permalink {mla_id}: {e}")
            permalinks[mla_id] = ''

    return permalinks

@app.route('/buscar-sku-ml', methods=['POST'])
@login_required
def buscar_sku_ml():
    sku_buscado = request.form.get('sku_buscar', '').strip().upper()

    if not sku_buscado:
        flash('Debes ingresar un SKU', 'warning')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()

    # Cargar porcentajes
    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'porcentajes_ml'")
        porcentajes = json.loads(row['valor']) if row else PORCENTAJES_ML_DEFAULT
    except:
        porcentajes = PORCENTAJES_ML_DEFAULT

    if not access_token:
        flash('❌ No hay token de ML configurado', 'warning')
        return render_template('cargar_stock_ml.html',
                               sku_buscado=sku_buscado,
                               publicaciones=[],
                               es_sku_con_z=sku_buscado.endswith('Z'),
                               porcentajes=porcentajes)

    mla_directo = None
    if sku_buscado.startswith('MLA'):
        mla_directo = sku_buscado
    elif sku_buscado.isdigit():
        mla_directo = f'MLA{sku_buscado}'

    if mla_directo:
        mla_ids = [mla_directo]
        es_sku_con_z = False
    else:
        es_sku_con_z = sku_buscado.endswith('Z')
        r = ml_request('get',
            f'https://api.mercadolibre.com/users/{ML_SELLER_ID}/items/search',
            access_token, params={'seller_sku': sku_buscado})
        if r.status_code != 200:
            flash(f'❌ Error consultando ML: {r.status_code}', 'danger')
            return render_template('cargar_stock_ml.html',
                                   sku_buscado=sku_buscado,
                                   publicaciones=[],
                                   es_sku_con_z=es_sku_con_z,
                                   porcentajes=porcentajes)
        mla_ids = r.json().get('results', [])

    if not mla_ids:
        flash(f'No se encontraron publicaciones para "{sku_buscado}"', 'warning')
        return render_template('cargar_stock_ml.html',
                               sku_buscado=sku_buscado,
                               publicaciones=[],
                               es_sku_con_z=es_sku_con_z,
                               porcentajes=porcentajes)

    estado_map = {
        'active': 'Activa', 'paused': 'Pausada', 'closed': 'Cerrada',
        'under_review': 'En revisión', 'inactive': 'Inactiva'
    }

    # Orden de tipos de publicación
    ORDEN_TIPOS = {
        'Sin cuotas propias': 0,
        'Cuota Simple':       1,
        '3 cuotas s/interés': 2,
        '6 cuotas s/interés': 3,
        '9 cuotas s/interés': 4,
        '12 cuotas s/interés': 5,
    }

    datos_batch = obtener_datos_ml_batch(mla_ids, access_token)
    permalinks = obtener_permalinks_ml(mla_ids, access_token)
    publicaciones = []
    for mla_id in mla_ids:
        datos_ml = datos_batch.get(mla_id, {
            'titulo': mla_id, 'stock': 0, 'status': 'unknown',
            'demora': None, 'precio': None, 'listing_type': None, 'status_raw': 'unknown'
        })
        status_ml = datos_ml.get('status', 'unknown')
        publicaciones.append({
            'mla':                 mla_id,
            'titulo':              datos_ml['titulo'],
            'stock_actual':        datos_ml['stock'],
            'demora':              datos_ml.get('demora'),
            'precio':              datos_ml.get('precio'),
            'listing_type':        datos_ml.get('listing_type'),
            'estado':              estado_map.get(status_ml, status_ml.capitalize()),
            'status_raw':          status_ml,
            'permalink':           permalinks.get(mla_id, ''),
            'catalog_listing':     datos_ml.get('catalog_listing', False),
            'catalog_product_id':  datos_ml.get('catalog_product_id'),
            'category_id':         datos_ml.get('category_id'),
        })

    # Ordenar por tipo de publicación
    publicaciones.sort(key=lambda p: ORDEN_TIPOS.get(p.get('listing_type', ''), 99))

    # Buscar precio calculado por costos para este SKU
    precio_costos = _get_precio_costos_sku(sku_buscado, porcentajes)

    # ── Catálogo: detectar cuotas faltantes ──────────────────────────────────
    TODOS_LOS_TIPOS = [
        'Sin cuotas propias', 'Cuota Simple',
        '3 cuotas s/interés', '6 cuotas s/interés',
        '9 cuotas s/interés', '12 cuotas s/interés',
    ]
    tipos_catalogo_existentes = set(
        p['listing_type'] for p in publicaciones
        if p.get('catalog_listing') and p.get('status_raw') in ('active', 'paused')
    )
    catalog_meta = next(
        (p for p in publicaciones if p.get('catalog_listing') and p.get('catalog_product_id')),
        None
    )
    cuotas_faltantes = [t for t in TODOS_LOS_TIPOS if t not in tipos_catalogo_existentes] if catalog_meta else []

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku_buscado,
                           publicaciones=publicaciones,
                           es_sku_con_z=es_sku_con_z,
                           mensaje=None,
                           mensaje_tipo=None,
                           porcentajes=porcentajes,
                           precio_costos=precio_costos,
                           cuotas_faltantes=cuotas_faltantes,
                           catalog_meta=catalog_meta)
# ============================================================================
# RUTA: Faltantes de catálogo ML (colchones)
# ============================================================================
@app.route('/faltantes-catalogo-ml')
@login_required
def faltantes_catalogo_ml():
    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'warning')
        return redirect(url_for('dashboard'))

    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'porcentajes_ml'")
        porcentajes = json.loads(row['valor']) if row else PORCENTAJES_ML_DEFAULT
    except:
        porcentajes = PORCENTAJES_ML_DEFAULT

    # 1. Obtener todos los IDs (activos + pausados)
    all_ids = []
    for status in ('active', 'paused'):
        offset = 0
        while True:
            r = ml_request('get',
                f'https://api.mercadolibre.com/users/{ML_SELLER_ID}/items/search',
                access_token,
                params={'limit': 100, 'offset': offset, 'status': status})
            if r.status_code != 200:
                break
            data = r.json()
            results = data.get('results', [])
            all_ids.extend(results)
            total = data.get('paging', {}).get('total', 0)
            offset += 100
            if offset >= total:
                break

    if not all_ids:
        return render_template('faltantes_catalogo_ml.html', resultados=[], total_skus=0)

    # 2. Fetch en batch (reutiliza obtener_datos_ml_batch)
    datos_batch = obtener_datos_ml_batch(all_ids, access_token)

    # 3. Filtrar catálogo + colchones, agrupar por SKU
    TODOS_LOS_TIPOS = [
        'Sin cuotas propias', 'Cuota Simple',
        '3 cuotas s/interés', '6 cuotas s/interés',
        '9 cuotas s/interés', '12 cuotas s/interés',
    ]
    TIPO_A_PRECIO = {
        'Sin cuotas propias': 'precio_sin_cuotas',
        'Cuota Simple':       'precio_1c',
        '3 cuotas s/interés': 'precio_3c',
        '6 cuotas s/interés': 'precio_6c',
        '9 cuotas s/interés': 'precio_9c',
        '12 cuotas s/interés':'precio_12c',
    }

    skus_data = {}
    for mla_id, datos in datos_batch.items():
        if not datos.get('catalog_listing'):
            continue
        if datos.get('domain_id') != 'MLA-MATTRESSES':
            continue
        sku = datos.get('seller_sku') or ''
        if not sku:
            continue
        if sku not in skus_data:
            skus_data[sku] = {
                'tipos': {},
                'catalog_product_id': datos.get('catalog_product_id'),
                'category_id':        datos.get('category_id'),
            }
        lt = datos.get('listing_type', '')
        if lt and lt not in skus_data[sku]['tipos']:
            skus_data[sku]['tipos'][lt] = mla_id

    # 4. Calcular faltantes con precio sugerido
    resultados = []
    for sku in sorted(skus_data.keys()):
        info = skus_data[sku]
        tipos_existentes = set(info['tipos'].keys())
        faltantes_tipos = [t for t in TODOS_LOS_TIPOS if t not in tipos_existentes]
        if not faltantes_tipos:
            continue
        pc = _get_precio_costos_sku(sku, porcentajes)
        mla_ref = next(iter(info['tipos'].values())) if info['tipos'] else None
        faltantes_con_precio = []
        for tipo in faltantes_tipos:
            precio_key = TIPO_A_PRECIO.get(tipo)
            precio_sug = pc.get(precio_key) if pc and precio_key else None
            faltantes_con_precio.append({'tipo': tipo, 'precio_sugerido': precio_sug})
        resultados.append({
            'sku':                 sku,
            'faltantes':           faltantes_con_precio,
            'existentes':          sorted(tipos_existentes, key=lambda t: TODOS_LOS_TIPOS.index(t) if t in TODOS_LOS_TIPOS else 99),
            'mla_ref':             mla_ref,
            'catalog_product_id':  info['catalog_product_id'],
        })

    return render_template('faltantes_catalogo_ml.html',
                           resultados=resultados,
                           total_skus=len(skus_data))


# ============================================================================
# RUTA: Publicar nueva publicación de catálogo (cuota faltante)
# ============================================================================
TIPO_A_PARAMS_ML = {
    'Sin cuotas propias':  {'listing_type_id': 'gold_special', 'campaign': None},
    'Cuota Simple':        {'listing_type_id': 'gold_special', 'campaign': 'pcj-co-funded'},
    '3 cuotas s/interés':  {'listing_type_id': 'gold_special', 'campaign': '3x_campaign'},
    '6 cuotas s/interés':  {'listing_type_id': 'gold_pro',     'campaign': None},
    '9 cuotas s/interés':  {'listing_type_id': 'gold_special', 'campaign': '9x_campaign'},
    '12 cuotas s/interés': {'listing_type_id': 'gold_special', 'campaign': '12x_campaign'},
}

@app.route('/debug-token-temp')
@login_required
def debug_token_temp():
    token = cargar_ml_token()
    return f"<pre>{token}</pre>"


@app.route('/publicar-catalogo-cuota', methods=['POST'])
@login_required
def publicar_catalogo_cuota():
    sku                = request.form.get('sku', '').strip().upper()
    tipo               = request.form.get('tipo', '').strip()
    precio_str         = request.form.get('precio', '').strip()
    catalog_product_id = request.form.get('catalog_product_id', '').strip()
    category_id        = request.form.get('category_id', '').strip()

    if not all([sku, tipo, precio_str, catalog_product_id, category_id]):
        flash('❌ Faltan datos para publicar en catálogo', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    try:
        precio = int(float(precio_str))
    except ValueError:
        flash('❌ Precio inválido', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    params = TIPO_A_PARAMS_ML.get(tipo)
    if not params:
        flash(f'❌ Tipo de publicación desconocido: {tipo}', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'warning')
        return redirect(url_for('cargar_stock_ml'))

    payload = {
        'site_id':             'MLA',
        'category_id':         category_id,
        'currency_id':         'ARS',
        'buying_mode':         'buy_it_now',
        'listing_type_id':     params['listing_type_id'],
        'price':               precio,
        'available_quantity':  1,
        'catalog_product_id':  catalog_product_id,
        'catalog_listing':     True,
        'seller_custom_field': sku,
    }
    if params['campaign']:
        payload['sale_terms'] = [{'id': 'INSTALLMENTS_CAMPAIGN', 'value_name': params['campaign']}]

    try:
        r = ml_request('post', 'https://api.mercadolibre.com/items', access_token, json_data=payload)
        resp = r.json()
        if r.status_code in (200, 201):
            nuevo_mla = resp.get('id', '?')
            flash(f'✅ Publicación creada: {nuevo_mla} — {tipo}', 'success')
        else:
            causa = resp.get('cause', [])
            detalle = causa[0].get('message', str(resp)) if causa else str(resp)
            flash(f'❌ Error ML: {detalle}', 'danger')
    except Exception as e:
        flash(f'❌ Excepción: {e}', 'danger')

    session['ultimo_sku_ml'] = sku
    return redirect(url_for('cargar_stock_ml'))

# ============================================================================
# 4. RUTA: Cambiar precio — INDIVIDUAL
# ============================================================================

@app.route('/cambiar-precio-mla', methods=['POST'])
@login_required
def cambiar_precio_mla():
    """Cambiar el precio de una publicación específica"""
    mla    = request.form.get('mla')
    sku    = request.form.get('sku')
    precio = request.form.get('precio', '').strip()

    if not mla or not sku or not precio:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    try:
        precio_float = float(precio)
        if precio_float <= 0:
            raise ValueError()
    except ValueError:
        flash('❌ Precio inválido', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    import requests as req
    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
    payload = {'price': precio_float}

    r = ml_request('put', f'https://api.mercadolibre.com/items/{mla}', access_token, json_data=payload)

    if r.status_code == 200:
        flash(f'✅ Precio actualizado a ${precio_float:,.0f} en {mla}', 'success')
    else:
        try:
            err = r.json()
        except:
            err = r.text
        flash(f'❌ Error ML {r.status_code} en {mla}: {err}', 'danger')

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token,
                               pubs_actuales=request.form.get('pubs_json'),
                               actualizar_mla=mla, campo='precio', valor=precio_float),
                           es_sku_con_z=sku.endswith('Z'))


# ============================================================================
# 5. RUTA: Cambiar precio — MASIVO
# ============================================================================

@app.route('/cambiar-precio-masivo', methods=['POST'])
@login_required
def cambiar_precio_masivo():
    """Cambiar el precio de todas las publicaciones de un SKU"""
    sku    = request.form.get('sku')
    precio = request.form.get('precio', '').strip()

    if not sku or not precio:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    try:
        precio_float = float(precio)
        if precio_float <= 0:
            raise ValueError()
    except ValueError:
        flash('❌ Precio inválido', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    mlas = query_db(
        "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,)
    )

    import requests as req
    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
    payload = {'price': precio_float}

    exitos, errores = 0, []
    for row in mlas:
        r = ml_request('put', f'https://api.mercadolibre.com/items/{row["mla_id"]}', access_token, json_data=payload)
        if r.status_code == 200:
            exitos += 1
        else:
            errores.append(f'{row["mla_id"]}: {r.status_code}')
        time.sleep(2)

    if exitos:
        flash(f'✅ Precio actualizado a ${precio_float:,.0f} en {exitos} publicación{"es" if exitos > 1 else ""}', 'success')
    for msg in errores[:3]:
        flash(f'❌ {msg}', 'danger')

    # Actualizar precio en todas las pubs localmente
    pubs_json = request.form.get('pubs_json')
    import json as _j
    try:
        pubs = _j.loads(pubs_json) if pubs_json else None
        if pubs:
            for p in pubs: p['precio'] = precio_float
    except: pubs = None
    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token, pubs_actuales=pubs),
                           es_sku_con_z=sku.endswith('Z'))

@app.route('/cambiar-precios-individuales', methods=['POST'])
@login_required
def cambiar_precios_individuales():
    """Actualizar precios individuales de múltiples MLAs de una vez"""
    import requests as req
    import json

    sku = request.form.get('sku')
    precios_json = request.form.get('precios_json', '[]')

    try:
        precios = json.loads(precios_json)
    except:
        flash('❌ Error al procesar los precios', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    if not precios:
        flash('❌ No se recibieron precios', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}

    exitos, errores = 0, []
    for item in precios:
        mla = item.get('mla')
        precio = item.get('precio')
        try:
            precio_float = float(precio)
            if precio_float <= 0:
                raise ValueError()
        except:
            errores.append(f'{mla}: precio inválido')
            continue

        r = ml_request('put', f'https://api.mercadolibre.com/items/{mla}', access_token, json_data={'price': precio_float})
        if r.status_code == 200:
            exitos += 1
        else:
            errores.append(f'{mla}: error {r.status_code}')
        time.sleep(2)

    if exitos:
        flash(f'✅ {exitos} precio{"s" if exitos > 1 else ""} actualizado{"s" if exitos > 1 else ""} correctamente', 'success')
    for msg in errores[:3]:
        flash(f'❌ {msg}', 'danger')

    # Actualizar precios individuales localmente
    pubs_json = request.form.get('pubs_json')
    import json as _j
    try:
        pubs = _j.loads(pubs_json) if pubs_json else None
        if pubs:
            precios_dict = {item['mla']: float(item['precio']) for item in precios}
            for p in pubs:
                if p['mla'] in precios_dict:
                    p['precio'] = precios_dict[p['mla']]
    except: pubs = None
    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token, pubs_actuales=pubs),
                           es_sku_con_z=sku.endswith('Z') if sku else False)



# ============================================================================
# RUTAS NUEVAS: Bajar stock a 0 y Cargar demora
# Agregar en app.py junto a las rutas de cargar_stock_ml
# ============================================================================

def _recargar_publicaciones(sku, access_token, pubs_actuales=None, actualizar_mla=None, campo=None, valor=None):
    """
    Helper: devuelve lista de publicaciones.
    Si se pasan pubs_actuales, las usa directamente y solo actualiza el campo indicado.
    Si no, consulta ML (solo al buscar por primera vez).
    """
    import json as _json

    estado_map = {
        'active': 'Activa', 'paused': 'Pausada', 'closed': 'Cerrada',
        'under_review': 'En revisión', 'inactive': 'Inactiva'
    }

    # Si tenemos datos actuales del form, usarlos sin llamar a ML
    if pubs_actuales:
        try:
            if isinstance(pubs_actuales, str):
                pubs_lista = _json.loads(pubs_actuales)
            else:
                pubs_lista = pubs_actuales
            # Aplicar el cambio puntual si se indicó
            if actualizar_mla and campo:
                for pub in pubs_lista:
                    if pub.get('mla') == actualizar_mla:
                        pub[campo] = valor
            return pubs_lista
        except:
            pass  # Si falla el parse, caer al batch

    # Sin datos actuales: consultar ML (solo pasa al buscar por primera vez)
    publicaciones = query_db(
        "SELECT mla_id, titulo_ml, activo FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
        (sku,)
    )
    pubs_lista = []
    if access_token and publicaciones:
        mla_ids = [row['mla_id'] for row in publicaciones]
        datos_batch = obtener_datos_ml_batch(mla_ids, access_token)
        permalinks = obtener_permalinks_ml(mla_ids, access_token)
        for row in publicaciones:
            datos_ml = datos_batch.get(row['mla_id'], {
                'titulo': row['mla_id'], 'stock': 0, 'status': 'unknown',
                'demora': None, 'precio': None, 'listing_type': None, 'status_raw': 'unknown'
            })
            status_ml = datos_ml.get('status', 'unknown')
            pubs_lista.append({
                'mla':          row['mla_id'],
                'titulo':       datos_ml['titulo'],
                'stock_actual': datos_ml['stock'],
                'demora':       datos_ml.get('demora'),
                'precio':       datos_ml.get('precio'),
                'listing_type': datos_ml.get('listing_type'),
                'estado':       estado_map.get(status_ml, status_ml.capitalize()),
                'status_raw':   status_ml,
                'permalink':    permalinks.get(row['mla_id'], ''),
            })
    else:
        for row in publicaciones:
            pubs_lista.append({
                'mla':          row['mla_id'],
                'titulo':       row['titulo_ml'] or 'Sin título',
                'stock_actual': '-', 'demora': None,
                'precio':       None, 'listing_type': None,
                'estado':       'Activa' if row['activo'] else 'Pausada',
                'status_raw':   'active' if row['activo'] else 'paused',
                'permalink':    '',
            })
    return pubs_lista




# ─── Bajar stock a 0 — INDIVIDUAL ────────────────────────────────────────────

@app.route('/bajar-stock-mla-cero', methods=['POST'])
@login_required
def bajar_stock_mla_cero():
    """Poner stock en 0 en una publicación específica"""
    mla = request.form.get('mla')
    sku = request.form.get('sku')

    if not mla or not sku:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    success, message = actualizar_stock_ml(mla, 0, access_token)

    if success:
        flash(f'✅ Stock bajado a 0 en {mla}', 'success')
    else:
        flash(f'❌ {message}', 'danger')

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token,
                               pubs_actuales=request.form.get('pubs_json'),
                               actualizar_mla=mla, campo='stock_actual', valor=0),
                           es_sku_con_z=sku.endswith('Z'))


# ─── Bajar stock a 0 — MASIVO ────────────────────────────────────────────────

@app.route('/bajar-stock-cero-masivo', methods=['POST'])
@login_required
def bajar_stock_cero_masivo():
    """Poner stock en 0 en todas las publicaciones de un SKU"""
    sku = request.form.get('sku')
    if not sku:
        flash('Falta el SKU', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    mlas = query_db(
        "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,)
    )

    exitos, errores = 0, []
    for row in mlas:
        ok, msg = actualizar_stock_ml(row['mla_id'], 0, access_token)
        if ok:
            exitos += 1
        else:
            errores.append(msg)
        time.sleep(2)

    if exitos:
        flash(f'✅ Stock bajado a 0 en {exitos} publicación{"es" if exitos > 1 else ""}', 'success')
    for msg in errores[:3]:
        flash(f'❌ {msg}', 'danger')

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token, pubs_actuales=request.form.get('pubs_json')),
                           es_sku_con_z=sku.endswith('Z'))


# ─── Cargar demora — INDIVIDUAL ──────────────────────────────────────────────

@app.route('/cargar-demora-mla', methods=['POST'])
@login_required
def cargar_demora_mla():
    """Poner X días de MANUFACTURING_TIME en una publicación"""
    mla  = request.form.get('mla')
    sku  = request.form.get('sku')
    dias = request.form.get('dias', '').strip()

    if not mla or not sku or not dias:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    import requests as req
    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
    payload = {"sale_terms": [{"id": "MANUFACTURING_TIME", "value_name": f"{dias} días"}]}

    r = ml_request('put', f'https://api.mercadolibre.com/items/{mla}', access_token, json_data=payload)

    if r.status_code == 200:
        flash(f'✅ Demora de {dias} días cargada en {mla}', 'success')
    else:
        try:
            err = r.json()
        except:
            err = r.text
        flash(f'❌ Error ML {r.status_code} en {mla}: {err}', 'danger')

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token, pubs_actuales=request.form.get('pubs_json')),
                           es_sku_con_z=sku.endswith('Z'))


# ─── Cargar demora — MASIVO ───────────────────────────────────────────────────

@app.route('/cargar-demora-masivo', methods=['POST'])
@login_required
def cargar_demora_masivo():
    """Poner X días de MANUFACTURING_TIME en todas las publicaciones de un SKU"""
    sku  = request.form.get('sku')
    dias = request.form.get('dias', '').strip()

    if not sku or not dias:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    mlas = query_db(
        "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,)
    )

    import requests as req
    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
    payload = {"sale_terms": [{"id": "MANUFACTURING_TIME", "value_name": f"{dias} días"}]}

    exitos, errores = 0, []
    for row in mlas:
        r = ml_request('put', f'https://api.mercadolibre.com/items/{row["mla_id"]}', access_token, json_data=payload)
        if r.status_code == 200:
            exitos += 1
        else:
            errores.append(f'{row["mla_id"]}: {r.status_code}')
        time.sleep(2)

    if exitos:
        flash(f'✅ {dias} días de demora cargados en {exitos} publicación{"es" if exitos > 1 else ""}', 'success')
    for msg in errores[:3]:
        flash(f'❌ {msg}', 'danger')

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token, pubs_actuales=request.form.get('pubs_json')),
                           es_sku_con_z=sku.endswith('Z'))



# ============================================================================
# RUTAS QUITAR DEMORA - CON STATUS REAL DE ML
# ============================================================================

# ============================================================================
# FUNCIÓN HELPER - Quitar MANUFACTURING_TIME completamente (envía null a ML)
# Reemplaza a quitar_handling_time_ml y actualizar_handling_time_ml
# ============================================================================

def quitar_manufacturing_time_ml(mla_id, access_token):
    """
    Elimina completamente el MANUFACTURING_TIME de una publicación ML.
    Envía value_id: null y value_name: null — ambos requeridos por la API.
    Retorna (True, mensaje) o (False, error)
    """
    import requests as req

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    payload = {
        "sale_terms": [
            {
                "id": "MANUFACTURING_TIME",
                "value_id": None,
                "value_name": None
            }
        ]
    }

    try:
        r = req.put(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers,
            json=payload
        )

        if r.status_code == 200:
            data = r.json()
            # Verificar que quedó en None
            mt_nuevo = None
            for term in data.get('sale_terms', []):
                if term.get('id') == 'MANUFACTURING_TIME':
                    mt_nuevo = term.get('value_name')
                    break

            if mt_nuevo is None:
                return True, f'Demora eliminada en {mla_id}'
            else:
                return True, f'Demora actualizada a {mt_nuevo} en {mla_id}'
        else:
            try:
                err = r.json()
            except:
                err = r.text
            return False, f'Error ML {r.status_code} en {mla_id}: {err}'

    except Exception as e:
        return False, f'Excepción en {mla_id}: {str(e)}'


# ============================================================================
# RUTA: Quitar demora de UNA publicación
# ============================================================================

@app.route('/quitar-demora-mla', methods=['POST'])
@login_required
def quitar_demora_mla():
    """Eliminar MANUFACTURING_TIME de una publicación específica"""

    mla = request.form.get('mla')
    sku = request.form.get('sku')

    if not mla or not sku:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    success, message = quitar_manufacturing_time_ml(mla, access_token)

    if success:
        flash(f'✅ {message}', 'success')
    else:
        flash(f'❌ {message}', 'danger')

    # Recargar publicaciones con datos actualizados de ML
    publicaciones = query_db(
        "SELECT mla_id, titulo_ml, activo FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
        (sku,)
    )

    pubs_lista = []
    access_token_refresh = cargar_ml_token()

    for row in publicaciones:
        if access_token_refresh:
            datos_ml = obtener_datos_ml(row['mla_id'], access_token_refresh)
            status_ml = datos_ml.get('status', 'unknown')
            estado_map = {
                'active': 'Activa', 'paused': 'Pausada', 'closed': 'Cerrada',
                'under_review': 'En revisión', 'inactive': 'Inactiva'
            }
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': datos_ml['titulo'],
                'stock_actual': datos_ml['stock'],
                'demora': datos_ml.get('demora'),
                'estado': estado_map.get(status_ml, status_ml.capitalize()),
                'status_raw': status_ml
            })
        else:
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': row['titulo_ml'] or 'Sin título',
                'stock_actual': '-',
                'demora': None,
                'estado': 'Activa' if row['activo'] else 'Pausada',
                'status_raw': 'active' if row['activo'] else 'paused'
            })

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=pubs_lista,
                           es_sku_con_z=sku.endswith('Z'))


# ============================================================================
# RUTA: Quitar demora de TODAS las publicaciones de un SKU
# ============================================================================

@app.route('/quitar-demora-masivo', methods=['POST'])
@login_required
def quitar_demora_masivo():
    """Eliminar MANUFACTURING_TIME de todas las publicaciones de un SKU"""
    sku = request.form.get('sku')
    if not sku:
        flash('Falta el SKU', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    mlas = query_db(
        "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,)
    )

    exitos, errores, mensajes_error = 0, 0, []

    for row in mlas:
        success, message = quitar_manufacturing_time_ml(row['mla_id'], access_token)
        if success:
            exitos += 1
        else:
            errores += 1
            mensajes_error.append(message)
        time.sleep(2)

    if exitos > 0:
        flash(f'✅ Demora eliminada en {exitos} publicación{"es" if exitos > 1 else ""}', 'success')
    if errores > 0:
        flash(f'⚠️ {errores} publicación{"es" if errores > 1 else ""} con errores', 'warning')
        for msg in mensajes_error[:3]:
            flash(msg, 'warning')

    # Recargar con UNA sola llamada batch
    publicaciones = query_db(
        "SELECT mla_id, titulo_ml, activo FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,)
    )
    estado_map = {
        'active': 'Activa', 'paused': 'Pausada', 'closed': 'Cerrada',
        'under_review': 'En revisión', 'inactive': 'Inactiva'
    }
    pubs_lista = []
    if publicaciones:
        mla_ids = [row['mla_id'] for row in publicaciones]
        datos_batch = obtener_datos_ml_batch(mla_ids, access_token)
        permalinks = obtener_permalinks_ml(mla_ids, access_token)
        for row in publicaciones:
            datos_ml = datos_batch.get(row['mla_id'], {
                'titulo': row['titulo_ml'] or row['mla_id'],
                'stock': '-', 'status': 'unknown', 'demora': None,
                'precio': None, 'listing_type': None
            })
            status_ml = datos_ml.get('status', 'unknown')
            pubs_lista.append({
                'mla':          row['mla_id'],
                'titulo':       datos_ml.get('titulo', row['titulo_ml'] or row['mla_id']),
                'stock_actual': datos_ml.get('stock', '-'),
                'demora':       datos_ml.get('demora'),
                'precio':       datos_ml.get('precio'),
                'listing_type': datos_ml.get('listing_type'),
                'estado':       estado_map.get(status_ml, status_ml.capitalize()),
                'status_raw':   status_ml,
                'permalink':    permalinks.get(row['mla_id'], ''),
            })

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=pubs_lista,
                           es_sku_con_z=sku.endswith('Z'))


# ============================================================================
# RUTAS CARGAR STOCK - CON STATUS REAL DE ML
# ============================================================================

@app.route('/cargar-stock-mla', methods=['POST'])
@login_required
def cargar_stock_mla():
    """Cargar stock en una publicación específica"""
    
    mla = request.form.get('mla')
    sku = request.form.get('sku')
    stock_nuevo = request.form.get('stock_nuevo')
    
    if not mla or not sku or not stock_nuevo:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))
    
    # Obtener token dinámico
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))
    
    try:
        stock_nuevo = int(stock_nuevo)
        
        # Usar la función helper actualizar_stock_ml
        success, message = actualizar_stock_ml(mla, stock_nuevo, access_token)
        
        if success:
            flash(f'✅ {message}', 'success')
        else:
            flash(f'❌ {message}', 'danger')
    
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'danger')
    
    # Volver a mostrar resultados CON DATOS ACTUALIZADOS DE ML
    publicaciones = query_db(
        "SELECT mla_id, titulo_ml, activo FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
        (sku,)
    )
    
    pubs_lista = []
    access_token_refresh = cargar_ml_token()
    
    for row in publicaciones:
        if access_token_refresh:
            datos_ml = obtener_datos_ml(row['mla_id'], access_token_refresh)
            
            # Mapear status de ML a español
            status_ml = datos_ml.get('status', 'unknown')
            
            if status_ml == 'active':
                estado_texto = 'Activa'
            elif status_ml == 'paused':
                estado_texto = 'Pausada'
            elif status_ml == 'closed':
                estado_texto = 'Cerrada'
            elif status_ml == 'under_review':
                estado_texto = 'En revisión'
            elif status_ml == 'inactive':
                estado_texto = 'Inactiva'
            else:
                estado_texto = status_ml.capitalize()
            
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': datos_ml['titulo'],
                'stock_actual': datos_ml['stock'],
                'demora': datos_ml.get('demora'),
                'estado': estado_texto,
                'status_raw': status_ml
            })
        else:
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': row['titulo_ml'] or 'Sin título',
                'stock_actual': '-',
                'demora': None,
                'estado': 'Activa' if row['activo'] else 'Pausada',
                'status_raw': 'active' if row['activo'] else 'paused'
            })
    
    return render_template('cargar_stock_ml.html',
                         sku_buscado=sku,
                         publicaciones=pubs_lista,
                         es_sku_con_z=sku.endswith('Z'))


@app.route('/cargar-stock-masivo', methods=['POST'])
@login_required
def cargar_stock_masivo():
    """Cargar el mismo stock en todas las publicaciones de un SKU"""
    
    sku = request.form.get('sku')
    stock_nuevo = request.form.get('stock_nuevo')
    
    if not sku or not stock_nuevo:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))
    
    # Obtener token dinámico
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))
    
    try:
        stock_nuevo = int(stock_nuevo)
        
        # Obtener todas las publicaciones del SKU
        mlas = query_db(
            "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
            (sku,)
        )
        
        exitos = 0
        errores = 0
        mensajes_error = []
        
        for row in mlas:
            mla = row['mla_id']
            
            # Usar la función helper actualizar_stock_ml
            success, message = actualizar_stock_ml(mla, stock_nuevo, access_token)
            
            if success:
                exitos += 1
            else:
                errores += 1
                mensajes_error.append(f"{mla}: {message}")
            time.sleep(2)
        
        if exitos > 0:
            flash(f'✅ Stock cargado en {exitos} publicaciones: {stock_nuevo} unidades', 'success')
        if errores > 0:
            flash(f'⚠️ {errores} publicaciones con errores', 'warning')
            if mensajes_error:
                for msg in mensajes_error[:3]:  # Mostrar solo los primeros 3 errores
                    flash(msg, 'warning')
    
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'danger')
    
    # Volver a mostrar resultados CON DATOS ACTUALIZADOS DE ML
    publicaciones = query_db(
        "SELECT mla_id, titulo_ml, activo FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
        (sku,)
    )
    
    pubs_lista = []
    access_token_refresh = cargar_ml_token()
    
    for row in publicaciones:
        if access_token_refresh:
            datos_ml = obtener_datos_ml(row['mla_id'], access_token_refresh)
            
            # Mapear status de ML a español
            status_ml = datos_ml.get('status', 'unknown')
            
            if status_ml == 'active':
                estado_texto = 'Activa'
            elif status_ml == 'paused':
                estado_texto = 'Pausada'
            elif status_ml == 'closed':
                estado_texto = 'Cerrada'
            elif status_ml == 'under_review':
                estado_texto = 'En revisión'
            elif status_ml == 'inactive':
                estado_texto = 'Inactiva'
            else:
                estado_texto = status_ml.capitalize()
            
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': datos_ml['titulo'],
                'stock_actual': datos_ml['stock'],
                'demora': datos_ml.get('demora'),
                'estado': estado_texto,
                'status_raw': status_ml
            })
        else:
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': row['titulo_ml'] or 'Sin título',
                'stock_actual': '-',
                'demora': None,
                'estado': 'Activa' if row['activo'] else 'Pausada',
                'status_raw': 'active' if row['activo'] else 'paused'
            })
    
    return render_template('cargar_stock_ml.html',
                         sku_buscado=sku,
                         publicaciones=pubs_lista,
                         es_sku_con_z=sku.endswith('Z'))


# ============================================================================
# AUDITORÍA ML - SECCIONES INDEPENDIENTES
# Reemplazar el bloque completo de auditoría en app.py
# ============================================================================

# ============================================================================
# HELPER: CALCULAR STOCK DISPONIBLE (compartido por los 3 tipos de auditoría)
# ============================================================================

def calcular_stock_por_sku():
    """
    Calcula stock disponible para todos los SKUs (base + combos).
    Devuelve dict: { sku: { nombre, stock_fisico, stock_disponible } }
    """
    # 1. Obtener stock físico de productos base
    productos_base_query = query_db('''
        SELECT 
            sku, 
            nombre, 
            tipo, 
            stock_actual,
            COALESCE(stock_full, 0) as stock_full
        FROM productos_base 
        ORDER BY tipo, nombre
    ''')
    
    # 2. Ventas activas descomponiendo combos
    ventas_activas = query_db('''
        SELECT 
            COALESCE(pb_comp.sku, iv.sku) as sku,
            SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
        FROM items_venta iv
        JOIN ventas v ON iv.venta_id = v.id
        LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
        LEFT JOIN componentes c ON c.producto_compuesto_id = pc.id
        LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
        WHERE v.estado_entrega = 'pendiente'
        GROUP BY sku
    ''')
    
    ventas_dict = {v['sku']: int(v['vendido']) for v in ventas_activas}
    
    # 3. Calcular stock disponible por SKU base
    stock_por_sku = {}
    
    for prod in productos_base_query:
        sku = prod['sku']
        vendido = ventas_dict.get(sku, 0)
        
        if '_DEP' in sku or '_FULL' in sku:
            stock_fisico = int(prod['stock_actual'])
            stock_disponible = stock_fisico - vendido
        elif prod['tipo'] == 'almohada':
            stock_dep = int(prod['stock_actual'])
            stock_full = int(prod['stock_full'])
            stock_fisico = stock_dep + stock_full
            stock_disponible = stock_fisico - vendido
        else:
            stock_fisico = int(prod['stock_actual'])
            stock_disponible = stock_fisico - vendido
        
        stock_por_sku[sku] = {
            'nombre': prod['nombre'],
            'stock_fisico': stock_fisico,
            'stock_disponible': stock_disponible
        }
    
    # 4. Calcular stock disponible de combos
    try:
        productos_combos = query_db('''
            SELECT sku, nombre
            FROM productos_compuestos
            WHERE activo = 1
            ORDER BY nombre
        ''')
        
        if productos_combos:
            for combo in productos_combos:
                sku_combo = combo['sku']
                componentes = query_db('''
                    SELECT pb.sku, c.cantidad_necesaria 
                    FROM componentes c
                    JOIN productos_base pb ON c.producto_base_id = pb.id
                    JOIN productos_compuestos pc ON c.producto_compuesto_id = pc.id
                    WHERE pc.sku = %s
                ''', (sku_combo,))
                
                stock_disponible_combo = 999999
                for comp in componentes:
                    sku_comp = comp['sku']
                    cant_necesaria = int(comp['cantidad_necesaria'])
                    prod_comp = stock_por_sku.get(sku_comp)
                    if prod_comp:
                        combos_posibles = prod_comp['stock_disponible'] // cant_necesaria if cant_necesaria > 0 else 0
                        stock_disponible_combo = min(stock_disponible_combo, combos_posibles)
                    else:
                        stock_disponible_combo = 0
                        break
                
                if stock_disponible_combo == 999999 or stock_disponible_combo < 0:
                    stock_disponible_combo = 0
                
                stock_por_sku[sku_combo] = {
                    'nombre': combo['nombre'],
                    'stock_fisico': 0,
                    'stock_disponible': stock_disponible_combo
                }
    except Exception as e:
        print(f"Error calculando combos: {str(e)}")
    
    return stock_por_sku


# ============================================================================
# PÁGINA PRINCIPAL DE AUDITORÍA (sin barrido, solo estructura)
# ============================================================================

@app.route('/auditoria-ml', methods=['GET'])
@login_required
def auditoria_ml():
    """Renderiza la página de auditoría. Los datos se cargan vía AJAX por sección."""
    return render_template('auditoria_ml.html')


# ============================================================================
# ENDPOINT AJAX: EJECUTAR AUDITORÍA POR TIPO
# GET /auditoria-ml/run/<tipo>
# tipo: 'pausadas_sin_stock' | 'pausadas_con_stock' | 'demoras'
# ============================================================================

@app.route('/auditoria-ml/run/<tipo>', methods=['GET'])
@login_required
def auditoria_ml_run(tipo):
    """
    Ejecuta un tipo específico de auditoría y devuelve JSON con los resultados.
    Usa batch de ML (20 MLAs por request) en lugar de llamadas individuales.
    """
    if tipo not in ['pausadas_sin_stock', 'pausadas_con_stock', 'demoras', 'stock_en_ml']:
        return jsonify({'error': 'Tipo de auditoría inválido'}), 400

    access_token = cargar_ml_token()
    if not access_token:
        return jsonify({'error': 'No hay token de ML configurado'}), 401

    try:
        stock_por_sku = calcular_stock_por_sku()

        if tipo == 'demoras':
            publicaciones_db = query_db("""
                SELECT mla_id, sku, titulo_ml
                FROM sku_mla_mapeo
                WHERE activo = TRUE AND sku LIKE '%%Z'
                ORDER BY sku
            """)
        else:
            publicaciones_db = query_db("""
                SELECT mla_id, sku, titulo_ml
                FROM sku_mla_mapeo
                WHERE activo = TRUE
                ORDER BY sku
            """)

        # PASO 1: filtrar por stock local primero (sin tocar la API de ML)
        pubs_a_consultar = []
        for pub in publicaciones_db:
            sku = pub['sku']
            stock_info = stock_por_sku.get(sku)
            if not stock_info and sku.endswith('Z'):
                stock_info = stock_por_sku.get(sku[:-1])
            if not stock_info:
                continue
            stock_disponible = stock_info['stock_disponible']

            if tipo == 'stock_en_ml':
                # Para esta auditoría queremos los que NO tienen stock disponible
                if stock_disponible > 0:
                    continue
            else:
                if stock_disponible <= 0:
                    continue

            pubs_a_consultar.append((pub, stock_disponible))

        # PASO 2: batch a ML — 20 MLAs por request en lugar de 1 x 1
        mla_ids = [pub['mla_id'] for pub, _ in pubs_a_consultar]
        datos_batch = obtener_datos_ml_batch(mla_ids, access_token)

        # PASO 3: clasificar resultados
        import re
        resultados = []
        for pub, stock_disponible in pubs_a_consultar:
            mla_id    = pub['mla_id']
            sku       = pub['sku']
            datos_ml  = datos_batch.get(mla_id, {})
            status_ml = datos_ml.get('status', 'unknown')
            stock_ml  = datos_ml.get('stock', 0)
            demora_ml = datos_ml.get('demora')

            item_base = {
                'mla':              mla_id,
                'sku':              sku,
                'titulo':           datos_ml.get('titulo', pub.get('titulo_ml', '')),
                'stock_disponible': stock_disponible,
                'stock_ml':         stock_ml,
                'status':           status_ml
            }

            if tipo == 'pausadas_sin_stock':
                if status_ml == 'paused' and stock_ml == 0:
                    resultados.append(item_base)

            elif tipo == 'pausadas_con_stock':
                if status_ml == 'paused' and stock_ml > 0:
                    resultados.append(item_base)

            elif tipo == 'demoras':
                if demora_ml and demora_ml != 'Sin especificar':
                    try:
                        numeros = re.findall(r'\d+', str(demora_ml))
                        if numeros and int(numeros[0]) > 0:
                            item_base['demora'] = demora_ml
                            resultados.append(item_base)
                    except Exception as e:
                        print(f"Error parseando demora '{demora_ml}': {e}")

            elif tipo == 'stock_en_ml':
                if stock_ml > 0:
                    es_z = sku.endswith('Z')
                    if es_z:
                        # Para SKUs Z: solo incluir si NO tiene demora cargada
                        tiene_demora = False
                        if demora_ml and demora_ml != 'Sin especificar':
                            try:
                                numeros = re.findall(r'\d+', str(demora_ml))
                                if numeros and int(numeros[0]) > 0:
                                    tiene_demora = True
                            except:
                                pass
                        if not tiene_demora:
                            resultados.append(item_base)
                    else:
                        # Para SKUs normales: siempre incluir si tiene stock en ML
                        resultados.append(item_base)

        print(f"✅ Auditoría '{tipo}': {len(resultados)} resultados (consultados {len(mla_ids)} MLAs en {-(-len(mla_ids)//20)} requests)")
        return jsonify({'tipo': tipo, 'resultados': resultados, 'total': len(resultados)})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# ACCIONES - Devuelven JSON para que el frontend refresque solo la sección
# ============================================================================

@app.route('/auditoria-ml/activar', methods=['POST'])
@login_required
def auditoria_activar_publicaciones():
    """Activar (despausar) publicaciones seleccionadas. Devuelve JSON."""
    
    mlas_seleccionadas = request.json.get('mlas', []) if request.is_json else request.form.getlist('mlas[]')
    
    if not mlas_seleccionadas:
        return jsonify({'error': 'No se seleccionaron publicaciones'}), 400
    
    access_token = cargar_ml_token()
    if not access_token:
        return jsonify({'error': 'No hay token de ML configurado'}), 401
    
    exitos = 0
    errores = []
    
    for mla in mlas_seleccionadas:
        try:
            url = f'https://api.mercadolibre.com/items/{mla}'
            headers = {'Authorization': f'Bearer {access_token}'}
            data = {'status': 'active'}
            
            response = ml_request('put', url, access_token, json_data=data)
            
            if response.status_code == 200:
                exitos += 1
            else:
                errores.append(f'{mla}: {response.status_code}')
            time.sleep(2)
        except Exception as e:
            errores.append(f'{mla}: {str(e)}')
    
    return jsonify({'exitos': exitos, 'errores': errores, 'total': len(mlas_seleccionadas)})


@app.route('/auditoria-ml/cargar-stock', methods=['POST'])
@login_required
def auditoria_cargar_stock():
    """Cargar stock en publicaciones seleccionadas. Devuelve JSON."""
    
    # Acepta JSON: { mla_stock: ["MLA123:5", "MLA456:3"] }
    # O form: mla_stock[] = "MLA123:5"
    if request.is_json:
        mlas_data = request.json.get('mla_stock', [])
    else:
        mlas_data = request.form.getlist('mla_stock')
    
    if not mlas_data:
        return jsonify({'error': 'No se seleccionaron publicaciones'}), 400
    
    access_token = cargar_ml_token()
    if not access_token:
        return jsonify({'error': 'No hay token de ML configurado'}), 401
    
    exitos = 0
    errores = []
    
    for item in mlas_data:
        try:
            mla, stock_str = item.split(':')
            stock = int(stock_str)
            success, message = actualizar_stock_ml(mla, stock, access_token)
            if success:
                exitos += 1
            else:
                errores.append(f'{mla}: {message}')
            time.sleep(2)
        except Exception as e:
            errores.append(f'{item}: {str(e)}')
    
    return jsonify({'exitos': exitos, 'errores': errores, 'total': len(mlas_data)})


@app.route('/auditoria-ml/reducir-demora', methods=['POST'])
@login_required
def auditoria_reducir_demora():
    """Quitar demora completamente en publicaciones seleccionadas. Devuelve JSON."""
    
    mlas_seleccionadas = request.json.get('mlas', []) if request.is_json else request.form.getlist('mlas_demora[]')
    
    if not mlas_seleccionadas:
        return jsonify({'error': 'No se seleccionaron publicaciones'}), 400
    
    access_token = cargar_ml_token()
    if not access_token:
        return jsonify({'error': 'No hay token de ML configurado'}), 401
    
    exitos = 0
    errores = []
    
    for mla in mlas_seleccionadas:
        try:
            success, message = quitar_manufacturing_time_ml(mla, access_token)
            if success:
                exitos += 1
            else:
                errores.append(f'{mla}: {message}')
            time.sleep(2)
        except Exception as e:
            errores.append(f'{mla}: {str(e)}')
    
    return jsonify({'exitos': exitos, 'errores': errores, 'total': len(mlas_seleccionadas)})

@app.route('/auditoria-ml/bajar-cero', methods=['POST'])
@login_required
def auditoria_bajar_cero():
    """Bajar stock a 0 en publicaciones seleccionadas. Para sección stock_en_ml."""
    if request.is_json:
        mlas_data = request.json.get('mla_stock', [])
    else:
        mlas_data = request.form.getlist('mla_stock')
    if not mlas_data:
        return jsonify({'error': 'No se seleccionaron publicaciones'}), 400
    access_token = cargar_ml_token()
    if not access_token:
        return jsonify({'error': 'No hay token de ML configurado'}), 401
    exitos = 0
    errores = []
    for item in mlas_data:
        try:
            mla, stock_str = item.split(':')
            stock = int(stock_str)  # siempre 0
            success, message = actualizar_stock_ml(mla, stock, access_token)
            if success:
                exitos += 1
            else:
                errores.append(f'{mla}: {message}')
            time.sleep(2)
        except Exception as e:
            errores.append(f'{item}: {str(e)}')
    return jsonify({'exitos': exitos, 'errores': errores, 'total': len(mlas_data)})


@app.route('/auditoria-ml/poner-demora', methods=['POST'])
@login_required
def auditoria_poner_demora():
    """Poner X días de demora en publicaciones Z seleccionadas. Para sección stock_en_ml."""
    if request.is_json:
        mlas_dias = request.json.get('mlas_dias', [])  # lista de "MLA123:15"
    else:
        mlas_dias = request.form.getlist('mlas_dias')
    if not mlas_dias:
        return jsonify({'error': 'No se seleccionaron publicaciones'}), 400
    access_token = cargar_ml_token()
    if not access_token:
        return jsonify({'error': 'No hay token de ML configurado'}), 401
    exitos = 0
    errores = []
    for item in mlas_dias:
        try:
            mla, dias_str = item.split(':')
            dias = int(dias_str)
            payload = {"sale_terms": [{"id": "MANUFACTURING_TIME", "value_name": f"{dias} días"}]}
            r = ml_request('put', f'https://api.mercadolibre.com/items/{mla}', access_token, json_data=payload)
            if r.status_code == 200:
                exitos += 1
            else:
                try:
                    err = r.json()
                except:
                    err = r.text
                errores.append(f'{mla}: {err}')
            time.sleep(2)
        except Exception as e:
            errores.append(f'{item}: {str(e)}')
    return jsonify({'exitos': exitos, 'errores': errores, 'total': len(mlas_dias)})



# ============================================================================
# ESTADÍSTICAS DE VENTAS
# ============================================================================

@app.route('/estadisticas')
@login_required
def estadisticas():
    from datetime import datetime, timedelta

    # ========================================
    # FILTROS
    # ========================================
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    filtro_canal = request.args.get('canal', '')
    filtro_metodo = request.args.get('metodo_envio', '')
    filtro_zona = request.args.get('zona', '')

    # Default: último mes
    if not fecha_desde:
        fecha_desde = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_hasta:
        fecha_hasta = datetime.now().strftime('%Y-%m-%d')

    # ========================================
    # BASE DE FILTROS (reutilizable)
    # ========================================
    base_where = """
        WHERE v.estado_entrega != 'cancelada'
        AND DATE(v.fecha_venta) BETWEEN %s AND %s
    """
    base_params = [fecha_desde, fecha_hasta]

    if filtro_canal == 'ML':
        base_where += " AND v.canal = 'Mercado Libre'"
    elif filtro_canal == 'no_ml':
        base_where += " AND v.canal != 'Mercado Libre'"

    if filtro_metodo:
        base_where += " AND v.metodo_envio = %s"
        base_params.append(filtro_metodo)

    if filtro_zona:
        base_where += " AND v.zona_envio = %s"
        base_params.append(filtro_zona)

    # ========================================
    # 1. MÉTRICAS RESUMEN
    # ========================================
    resumen = query_one(f"""
        SELECT
            COUNT(*) as total_ventas,
            COALESCE(SUM(v.importe_total), 0) as total_facturado,
            COALESCE(AVG(v.importe_total), 0) as ticket_promedio,
            COALESCE(SUM(iv.cantidad), 0) as total_unidades
        FROM ventas v
        LEFT JOIN items_venta iv ON iv.venta_id = v.id
        {base_where}
    """, tuple(base_params))

    # ========================================
    # 2. VENTAS POR DÍA (para gráfico de línea)
    # ========================================
    ventas_por_dia = query_db(f"""
        SELECT
            DATE(v.fecha_venta) as dia,
            COUNT(*) as cantidad,
            SUM(v.importe_total) as total
        FROM ventas v
        {base_where}
        GROUP BY DATE(v.fecha_venta)
        ORDER BY dia
    """, tuple(base_params))

    # ========================================
    # 3. DESGLOSE POR CANAL (para torta)
    # ========================================
    por_canal = query_db(f"""
        SELECT
            CASE WHEN v.canal = 'Mercado Libre' THEN 'MercadoLibre' ELSE 'Venta Directa' END as canal_label,
            COUNT(*) as cantidad,
            SUM(v.importe_total) as total
        FROM ventas v
        {base_where}
        GROUP BY canal_label
    """, tuple(base_params))

    # ========================================
    # 4. DESGLOSE POR MÉTODO DE ENVÍO (para torta)
    # ========================================
    por_metodo = query_db(f"""
        SELECT
            COALESCE(v.metodo_envio, 'Sin especificar') as metodo,
            COUNT(*) as cantidad,
            SUM(v.importe_total) as total
        FROM ventas v
        {base_where}
        GROUP BY v.metodo_envio
        ORDER BY cantidad DESC
    """, tuple(base_params))

    # ========================================
    # 5. TOP PRODUCTOS MÁS VENDIDOS
    # ========================================
    top_productos = query_db(f"""
        SELECT
            iv.sku,
            COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre,
            SUM(iv.cantidad) as total_unidades,
            SUM(iv.cantidad * iv.precio_unitario) as total_facturado
        FROM ventas v
        JOIN items_venta iv ON iv.venta_id = v.id
        LEFT JOIN productos_base pb ON iv.sku = pb.sku
        LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
        {base_where}
        GROUP BY iv.sku, nombre
        ORDER BY total_unidades DESC
        LIMIT 15
    """, tuple(base_params))

    # ========================================
    # 6. FILTROS DISPONIBLES (para los selects)
    # ========================================
    metodos_disponibles = query_db("""
        SELECT DISTINCT metodo_envio FROM ventas
        WHERE metodo_envio IS NOT NULL AND metodo_envio != ''
        ORDER BY metodo_envio
    """)

    zonas_disponibles = query_db("""
        SELECT DISTINCT zona_envio FROM ventas
        WHERE zona_envio IS NOT NULL AND zona_envio != ''
        ORDER BY zona_envio
    """)

    return render_template('estadisticas.html',
        resumen=resumen,
        ventas_por_dia=ventas_por_dia,
        por_canal=por_canal,
        por_metodo=por_metodo,
        top_productos=top_productos,
        metodos_disponibles=metodos_disponibles,
        zonas_disponibles=zonas_disponibles,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        filtro_canal=filtro_canal,
        filtro_metodo=filtro_metodo,
        filtro_zona=filtro_zona,
    )


# ============================================================================
# EXPORTAR REPOSICIÓN A EXCEL
# ============================================================================

@app.route('/estadisticas/exportar-reposicion')
@login_required
def exportar_reposicion():
    from flask import make_response
    from datetime import datetime, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    fecha_desde = request.args.get('fecha_desde', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_hasta = request.args.get('fecha_hasta', datetime.now().strftime('%Y-%m-%d'))
    filtro_canal = request.args.get('canal', '')
    filtro_metodo = request.args.get('metodo_envio', '')
    filtro_zona = request.args.get('zona', '')

    base_where = """
        WHERE v.estado_entrega != 'cancelada'
        AND DATE(v.fecha_venta) BETWEEN %s AND %s
    """
    base_params = [fecha_desde, fecha_hasta]

    if filtro_canal == 'ML':
        base_where += " AND v.canal = 'Mercado Libre'"
    elif filtro_canal == 'no_ml':
        base_where += " AND v.canal != 'Mercado Libre'"
    if filtro_metodo:
        base_where += " AND v.metodo_envio = %s"
        base_params.append(filtro_metodo)
    if filtro_zona:
        base_where += " AND v.zona_envio = %s"
        base_params.append(filtro_zona)

    # Stock disponible = stock_actual - pendiente de entrega (descomponiendo combos igual que nueva_venta)
    vendido_rows = query_db("""
        SELECT
            COALESCE(pb_comp.sku, iv.sku) as sku,
            SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
        FROM items_venta iv
        JOIN ventas v ON iv.venta_id = v.id
        LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
        LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
        LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
        WHERE v.estado_entrega NOT IN ('entregada', 'cancelada')
        GROUP BY COALESCE(pb_comp.sku, iv.sku)
    """)
    vendido_map = {r['sku']: int(r['vendido']) for r in vendido_rows}

    stock_fisico_rows = query_db("SELECT sku, stock_actual FROM productos_base")
    stock_disponible_map = {
        r['sku']: int(r['stock_actual']) - vendido_map.get(r['sku'], 0)
        for r in stock_fisico_rows
    }

    # Productos base vendidos directamente (no son combos)
    base_directos = query_db(f"""
        SELECT
            pb.sku,
            pb.nombre,
            SUM(iv.cantidad) as cantidad_vendida
        FROM ventas v
        JOIN items_venta iv ON iv.venta_id = v.id
        JOIN productos_base pb ON iv.sku = pb.sku
        {base_where}
        GROUP BY pb.sku, pb.nombre
    """, tuple(base_params))

    # Combos → explotar en componentes base
    combos = query_db(f"""
        SELECT
            pb.sku,
            pb.nombre,
            SUM(iv.cantidad * c.cantidad_necesaria) as cantidad_vendida
        FROM ventas v
        JOIN items_venta iv ON iv.venta_id = v.id
        JOIN productos_compuestos pc ON iv.sku = pc.sku
        JOIN componentes c ON c.producto_compuesto_id = pc.id
        JOIN productos_base pb ON pb.id = c.producto_base_id
        {base_where}
        GROUP BY pb.sku, pb.nombre
    """, tuple(base_params))

    # Consolidar sumando directos + combos
    totales = {}
    for row in list(base_directos) + list(combos):
        sku = row['sku']
        if sku in totales:
            totales[sku]['cantidad_vendida'] += int(row['cantidad_vendida'])
        else:
            totales[sku] = {
                'sku': sku,
                'nombre': row['nombre'],
                'cantidad_vendida': int(row['cantidad_vendida']),
                'stock_disponible': stock_disponible_map.get(sku, 0)
            }

    productos = sorted(totales.values(), key=lambda x: x['cantidad_vendida'], reverse=True)

    # CREAR EXCEL
    wb = Workbook()
    ws = wb.active
    ws.title = "Reposición"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Título
    ws.merge_cells('A1:D1')
    titulo = ws['A1']
    titulo.value = f"Reposición de Stock — {fecha_desde} al {fecha_hasta}"
    titulo.font = Font(bold=True, color='FFFFFF', name='Arial', size=12)
    titulo.fill = header_fill
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Headers
    headers = ['SKU', 'Descripción', 'Unidades Vendidas', 'Stock Disponible']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Datos
    for i, p in enumerate(productos):
        row = i + 3
        fill = alt_fill if i % 2 == 0 else PatternFill(fill_type=None)
        valores = [p['sku'], p['nombre'], int(p['cantidad_vendida']), int(p['stock_disponible'])]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name='Arial', size=10)
            cell.fill = fill
            cell.border = border
            if col in (3, 4):
                cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.freeze_panes = 'A3'

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = make_response(excel_file.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    fecha_str = datetime.now().strftime('%Y%m%d')
    response.headers['Content-Disposition'] = f'attachment; filename=reposicion_{fecha_str}.xlsx'
    return response

# ============================================================================
# TEST: GESTIÓN DE MANUFACTURING_TIME EN PUBLICACIONES ML
# Ruta segura y aislada - no toca nada del sistema existente
# ============================================================================

@app.route('/test/manufacturing-time', methods=['GET'])
@login_required
def test_manufacturing_time():
    """
    Página de prueba para ver y modificar el MANUFACTURING_TIME
    de publicaciones ML con sufijo Z
    """
    return render_template('test_manufacturing_time.html')


@app.route('/test/manufacturing-time/ver', methods=['POST'])
@login_required
def ver_manufacturing_time():
    """
    Consulta el estado actual de una publicación:
    - Muestra el MANUFACTURING_TIME actual
    - Muestra available_quantity
    - Solo lectura, no modifica nada
    """
    import requests as req

    item_id = request.form.get('item_id', '').strip().upper()
    if not item_id:
        return {'error': 'Falta item_id'}, 400

    access_token = cargar_ml_token()
    if not access_token:
        return {'error': 'No hay token ML activo'}, 400

    headers = {'Authorization': f'Bearer {access_token}'}

    try:
        r = req.get(f'https://api.mercadolibre.com/items/{item_id}', headers=headers)
        if r.status_code != 200:
            return {'error': f'ML devolvió {r.status_code}: {r.text}'}, 400

        data = r.json()

        # Extraer MANUFACTURING_TIME de sale_terms
        manufacturing_time = None
        for term in data.get('sale_terms', []):
            if term.get('id') == 'MANUFACTURING_TIME':
                manufacturing_time = term.get('value_name')
                break

        return {
            'ok': True,
            'item_id': item_id,
            'title': data.get('title'),
            'status': data.get('status'),
            'available_quantity': data.get('available_quantity'),
            'manufacturing_time': manufacturing_time,  # None = sin demora
            'permalink': data.get('permalink'),
        }

    except Exception as e:
        return {'error': str(e)}, 500


@app.route('/test/manufacturing-time/quitar', methods=['POST'])
@login_required
def quitar_manufacturing_time():
    """
    Elimina el MANUFACTURING_TIME de una publicación enviando null.
    Según la API de ML:
    PUT /items/{item_id}
    { "sale_terms": [{ "id": "MANUFACTURING_TIME", "value_id": null, "value_name": null }] }
    """
    import requests as req

    item_id = request.form.get('item_id', '').strip().upper()
    if not item_id:
        return {'error': 'Falta item_id'}, 400

    access_token = cargar_ml_token()
    if not access_token:
        return {'error': 'No hay token ML activo'}, 400

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    payload = {
        "sale_terms": [
            {
                "id": "MANUFACTURING_TIME",
                "value_id": None,
                "value_name": None
            }
        ]
    }

    try:
        r = req.put(
            f'https://api.mercadolibre.com/items/{item_id}',
            headers=headers,
            json=payload
        )

        if r.status_code == 200:
            data = r.json()
            # Verificar que quedó en None
            manufacturing_time_nuevo = None
            for term in data.get('sale_terms', []):
                if term.get('id') == 'MANUFACTURING_TIME':
                    manufacturing_time_nuevo = term.get('value_name')
                    break

            return {
                'ok': True,
                'item_id': item_id,
                'title': data.get('title'),
                'manufacturing_time_anterior': request.form.get('mt_anterior'),
                'manufacturing_time_nuevo': manufacturing_time_nuevo,
                'mensaje': '✅ Demora eliminada correctamente' if not manufacturing_time_nuevo else '⚠️ No se pudo eliminar'
            }
        else:
            return {
                'ok': False,
                'error': f'ML devolvió {r.status_code}',
                'detalle': r.json()
            }, 400

    except Exception as e:
        return {'error': str(e)}, 500


@app.route('/test/manufacturing-time/poner', methods=['POST'])
@login_required
def poner_manufacturing_time():
    """
    Pone o restaura el MANUFACTURING_TIME a un valor específico.
    Útil para revertir si algo sale mal.
    """
    import requests as req

    item_id = request.form.get('item_id', '').strip().upper()
    dias = request.form.get('dias', '').strip()

    if not item_id or not dias:
        return {'error': 'Faltan datos'}, 400

    access_token = cargar_ml_token()
    if not access_token:
        return {'error': 'No hay token ML activo'}, 400

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    payload = {
        "sale_terms": [
            {
                "id": "MANUFACTURING_TIME",
                "value_name": f"{dias} días"
            }
        ]
    }

    try:
        r = req.put(
            f'https://api.mercadolibre.com/items/{item_id}',
            headers=headers,
            json=payload
        )

        if r.status_code == 200:
            data = r.json()
            manufacturing_time_nuevo = None
            for term in data.get('sale_terms', []):
                if term.get('id') == 'MANUFACTURING_TIME':
                    manufacturing_time_nuevo = term.get('value_name')
                    break

            return {
                'ok': True,
                'item_id': item_id,
                'manufacturing_time_nuevo': manufacturing_time_nuevo,
                'mensaje': f'✅ Demora restaurada a {dias} días'
            }
        else:
            return {
                'ok': False,
                'error': f'ML devolvió {r.status_code}',
                'detalle': r.json()
            }, 400

    except Exception as e:
        return {'error': str(e)}, 500


@app.route('/debug-mla')
@login_required
def debug_mla():
    import requests
    access_token = cargar_ml_token()
    mla_id = 'MLA1538119895'
    headers = {'Authorization': f'Bearer {access_token}'}
    r = requests.get(f'https://api.mercadolibre.com/items/{mla_id}', headers=headers)
    data = r.json()
    
    listing_type_id = data.get('listing_type_id')
    campaign = None
    for term in data.get('sale_terms', []):
        if term.get('id') == 'INSTALLMENTS_CAMPAIGN':
            campaign = term
            break
    
    return {
        'listing_type_id': listing_type_id,
        'installments_campaign': campaign
    }


@app.route('/ml/callback')
@login_required
def ml_callback():
    """Recibe el code de ML y lo canjea automáticamente por el token"""
    code = request.args.get('code', '').strip()
    if not code:
        flash('❌ No se recibió el code de MercadoLibre', 'error')
        return redirect(url_for('ml_configurar_token'))
    
    CLIENT_ID = os.getenv('ML_APP_ID')
    CLIENT_SECRET = os.getenv('ML_SECRET_KEY')
    REDIRECT_URI = os.getenv('ML_REDIRECT_URI')
    
    try:
        response = requests.post(
            "https://api.mercadolibre.com/oauth/token",
            data={
                "grant_type": "authorization_code",
                "client_id": CLIENT_ID,
                "client_secret": CLIENT_SECRET,
                "code": code,
                "redirect_uri": REDIRECT_URI
            }
        )
        if response.status_code == 200:
            data = response.json()
            token_data = {
                "access_token": data.get("access_token"),
                "refresh_token": data.get("refresh_token"),
                "expires_at": time.time() + data.get("expires_in", 21600) - 300,
                "client_id": CLIENT_ID,
                "client_secret": CLIENT_SECRET
            }
            if guardar_ml_token(token_data):
                flash('✅ Token configurado con auto-renovación activada', 'success')
                return redirect(url_for('ventas_activas'))
            else:
                flash('❌ Error al guardar el token', 'error')
        else:
            error_msg = response.json().get('message', 'Error desconocido')
            flash(f'❌ Error al canjear el code: {error_msg}', 'error')
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
    
    return redirect(url_for('ml_configurar_token'))


# ============================================================================
# TIENDANUBE Y TIENDA PROPIA
# ============================================================================
from tiendanube_bp import tiendanube_bp
app.register_blueprint(tiendanube_bp)

@app.route('/ventas/guardar-trid', methods=['POST'])
@login_required
def guardar_trid():
    """Guardar o actualizar el TRID (código de tracking Correo Argentino) en notas"""
    try:
        data = request.get_json()
        venta_id = data.get('venta_id')
        trid = data.get('trid', '').strip().upper()

        if not venta_id or not trid:
            return jsonify({'ok': False, 'error': 'Datos incompletos'}), 400

        venta = query_one('SELECT notas FROM ventas WHERE id = %s', (venta_id,))
        if not venta:
            return jsonify({'ok': False, 'error': 'Venta no encontrada'}), 404

        notas = venta['notas'] or ''
        # Normalizar separadores (\n literal → \n real)
        notas = notas.replace('\\n', '\n')

        import re
        if re.search(r'TRID:', notas):
            # Reemplazar línea existente
            notas = re.sub(r'TRID:[^\n]*', f'TRID: {trid}', notas)
        else:
            # Agregar al final
            notas = notas.rstrip('\n') + f'\nTRID: {trid}'

        execute_db('UPDATE ventas SET notas = %s WHERE id = %s', (notas, venta_id))
        return jsonify({'ok': True, 'trid': trid})

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


from tienda_bp import tienda_bp
app.register_blueprint(tienda_bp)



# ============================================================================
# FLETES
# ============================================================================

def _crear_tablas_fletes():
    execute_db("""
        CREATE TABLE IF NOT EXISTS fleteros (
            id        INT AUTO_INCREMENT PRIMARY KEY,
            nombre    VARCHAR(100) NOT NULL UNIQUE,
            activo    TINYINT DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    execute_db("""
        CREATE TABLE IF NOT EXISTS fletes_registros (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            fletero_id  INT NOT NULL,
            fecha       DATE NOT NULL,
            descripcion VARCHAR(200) DEFAULT '',
            monto       DECIMAL(10,2) NOT NULL,
            pagado      TINYINT DEFAULT 0,
            created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (fletero_id) REFERENCES fleteros(id)
        )
    """)

@app.route('/fletes', methods=['GET'])
@login_required
def fletes():
    _crear_tablas_fletes()
    from datetime import date
    mes  = request.args.get('mes', date.today().strftime('%Y-%m'))
    try:
        anio, nmes = int(mes.split('-')[0]), int(mes.split('-')[1])
    except Exception:
        anio, nmes = date.today().year, date.today().month

    fleteros = query_db("SELECT * FROM fleteros ORDER BY nombre")

    registros = query_db("""
        SELECT r.id, r.fletero_id, r.fecha, r.descripcion, r.monto, r.pagado,
               f.nombre as fletero_nombre
        FROM fletes_registros r
        JOIN fleteros f ON f.id = r.fletero_id
        WHERE YEAR(r.fecha) = %s AND MONTH(r.fecha) = %s
        ORDER BY r.fecha, f.nombre, r.id
    """, (anio, nmes))

    # Totales por fletero del mes
    totales_fletero = query_db("""
        SELECT f.nombre, f.id,
               SUM(r.monto) as total,
               SUM(CASE WHEN r.pagado=1 THEN r.monto ELSE 0 END) as pagado,
               SUM(CASE WHEN r.pagado=0 THEN r.monto ELSE 0 END) as pendiente
        FROM fletes_registros r
        JOIN fleteros f ON f.id = r.fletero_id
        WHERE YEAR(r.fecha) = %s AND MONTH(r.fecha) = %s
        GROUP BY f.id, f.nombre
        ORDER BY f.nombre
    """, (anio, nmes))

    return render_template('fletes.html',
        fleteros        = fleteros,
        registros       = registros,
        totales_fletero = totales_fletero,
        mes             = mes,
        anio            = anio,
        nmes            = nmes,
    )


@app.route('/fletes/guardar', methods=['POST'])
@login_required
def fletes_guardar():
    _crear_tablas_fletes()
    data   = request.get_json()
    accion = data.get('accion')
    try:
        if accion == 'agregar_fletero':
            nombre = data.get('nombre', '').strip()
            if not nombre:
                return jsonify({'ok': False, 'error': 'Nombre requerido'})
            existe = query_db("SELECT id FROM fleteros WHERE nombre=%s", (nombre,))
            if existe:
                return jsonify({'ok': False, 'error': 'Ya existe ese fletero'})
            execute_db("INSERT INTO fleteros (nombre) VALUES (%s)", (nombre,))
        elif accion == 'toggle_fletero':
            execute_db("UPDATE fleteros SET activo = NOT activo WHERE id=%s", (data.get('id'),))
        elif accion == 'agregar_registro':
            execute_db("""
                INSERT INTO fletes_registros (fletero_id, fecha, descripcion, monto, pagado)
                VALUES (%s, %s, %s, %s, 0)
            """, (data['fletero_id'], data['fecha'], data.get('descripcion',''), float(data['monto'])))
        elif accion == 'eliminar_registro':
            execute_db("DELETE FROM fletes_registros WHERE id=%s", (data.get('id'),))
        elif accion == 'toggle_pagado_grupo':
            # Marcar/desmarcar todos los registros de un día+fletero
            nuevo = int(data.get('pagado', 0))
            execute_db("""
                UPDATE fletes_registros SET pagado=%s
                WHERE fletero_id=%s AND fecha=%s
            """, (nuevo, data['fletero_id'], data['fecha']))
        elif accion == 'editar_monto':
            execute_db("UPDATE fletes_registros SET monto=%s WHERE id=%s",
                       (float(data['monto']), data['id']))
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    return jsonify({'ok': True})


# ============================================================================
# MÓDULO VIAJES — Planificador de carga
# ============================================================================

def _pack_zona(zona, items):
    """
    Algoritmo de carga para objetos planos (colchones, bases, sommiers).

    Estrategia principal: parar todo de canto como libros en estantería.
    - Cada pieza ocupa su dimensión más chica (espesor) en el eje del ancho del camión
    - La cara de la pieza debe entrar en el largo × alto del camión
    - Se agrupa por parada: última parada va al fondo, primera a la boca
    - Suma de espesores ≤ ancho disponible de la zona

    Si no entran todos parados, el sobrante se intenta acomodar acostado encima.
    """
    Z_L = zona['largo_cm']   # fondo → boca
    Z_A = zona['ancho_cm']   # ancho (dimensión donde se acumulan los espesores)
    Z_H = zona['alto_cm']    # alto

    packed = []
    remaining = []
    ancho_usado = 0

    def _mejor_orientacion_parado(item):
        """
        Para una pieza parada de canto, su espesor ocupa el ancho del camión.
        La cara (las otras dos dimensiones) debe entrar en Z_L × Z_H.
        Retorna (espesor, cara_largo, cara_alto, label) o None si no entra parado.
        """
        dims = sorted([
            item.get('largo_cm') or 0,
            item.get('ancho_cm') or 0,
            item.get('alto_cm') or 0,
        ])
        if not all(dims):
            return None

        d_min, d_mid, d_max = dims  # d_min = espesor (va al ancho del camión)

        # La cara es d_mid × d_max, debe entrar en Z_L × Z_H
        if d_max <= Z_L and d_mid <= Z_H:
            return (d_min, d_max, d_mid, 'parado')
        if d_max <= Z_H and d_mid <= Z_L:
            return (d_min, d_mid, d_max, 'parado (girado)')
        return None

    def _mejor_orientacion_acostado(item):
        """
        Para una pieza acostada: la dimensión más chica es el alto.
        La cara (largo × ancho) debe entrar en Z_L × Z_A.
        """
        dims = sorted([
            item.get('largo_cm') or 0,
            item.get('ancho_cm') or 0,
            item.get('alto_cm') or 0,
        ])
        if not all(dims):
            return None
        d_min, d_mid, d_max = dims
        if d_max <= Z_L and d_mid <= Z_A:
            return (d_max, d_mid, d_min, 'acostado')
        if d_max <= Z_A and d_mid <= Z_L:
            return (d_mid, d_max, d_min, 'acostado (girado)')
        return None

    # Paso 1: intentar parar todo de canto
    for item in items:
        if not (item.get('largo_cm') and item.get('ancho_cm') and item.get('alto_cm')):
            packed.append({**item, 'zona_nombre': zona['nombre_zona'],
                           'orientacion': '—', 'sin_medidas': True,
                           'espesor': 0})
            continue

        orient = _mejor_orientacion_parado(item)
        if orient and ancho_usado + orient[0] <= Z_A:
            espesor, il, ih, lbl = orient
            ancho_usado += espesor
            packed.append({**item, 'zona_nombre': zona['nombre_zona'],
                           'orientacion': lbl,
                           'il': il, 'ia': espesor, 'ih': ih,
                           'espesor': espesor})
        else:
            remaining.append(item)

    # Paso 2: los que no entran parados, intentar acostarlos encima
    # (esto es marginal para el caso normal, pero cubre colchones de una plaza extras, etc.)
    alto_usado_acostado = 0
    segunda_vuelta = []
    for item in remaining:
        orient = _mejor_orientacion_acostado(item)
        if orient and alto_usado_acostado + orient[2] <= Z_H:
            il, ia, ih, lbl = orient
            alto_usado_acostado += ih
            packed.append({**item, 'zona_nombre': zona['nombre_zona'],
                           'orientacion': lbl,
                           'il': il, 'ia': ia, 'ih': ih,
                           'espesor': ih})
        else:
            segunda_vuelta.append(item)

    return {
        'zona': zona,
        'productos': packed,
        'ancho_usado': ancho_usado,
        'ancho_total': Z_A,
        'alto_usado': alto_usado_acostado,
        'alto_total': Z_H,
        'largo_total': Z_L,
    }, segunda_vuelta


def _calcular_carga_viaje(viaje_id):
    """
    Calcula si los productos del viaje entran en las zonas del vehículo.
    Retorna dict con resultado detallado, o dict con 'error' si hay problema.
    """
    viaje = query_one("""
        SELECT v.*, f.nombre AS fletero_nombre
        FROM viajes v JOIN fleteros f ON f.id = v.fletero_id
        WHERE v.id = %s
    """, (viaje_id,))
    if not viaje:
        return None

    zonas = query_db("""
        SELECT * FROM fletero_zonas
        WHERE fletero_id = %s ORDER BY orden
    """, (viaje['fletero_id'],))

    if not zonas:
        return {'error': 'El fletero no tiene zonas configuradas. '
                         'Agregá las dimensiones del vehículo primero.'}

    # Items expandidos por cantidad; último en entregar → primero en cargar (DESC)
    # Primero traemos los items que son productos_base directamente
    raw_base = query_db("""
        SELECT
            p.orden_entrega, p.cliente, p.direccion,
            i.sku, i.cantidad,
            pb.nombre AS producto_nombre,
            pb.largo_cm, pb.ancho_cm, pb.alto_cm, pb.peso_gramos, pb.tipo
        FROM viaje_paradas p
        JOIN items_venta i ON i.venta_id = p.venta_id
        JOIN productos_base pb ON pb.sku = i.sku
        WHERE p.viaje_id = %s
        ORDER BY p.orden_entrega DESC, (pb.largo_cm * pb.ancho_cm) DESC
    """, (viaje_id,))

    # Luego los que son productos_compuestos → expandir en componentes (cada componente = pieza física)
    raw_comp = query_db("""
        SELECT
            p.orden_entrega, p.cliente, p.direccion,
            i.sku, i.cantidad AS cantidad_venta,
            pc.nombre AS nombre_compuesto,
            pb.nombre AS producto_nombre,
            pb.largo_cm, pb.ancho_cm, pb.alto_cm, pb.peso_gramos, pb.tipo,
            c.cantidad_necesaria
        FROM viaje_paradas p
        JOIN items_venta i ON i.venta_id = p.venta_id
        JOIN productos_compuestos pc ON pc.sku = i.sku
        JOIN componentes c ON c.producto_compuesto_id = pc.id
        JOIN productos_base pb ON pb.id = c.producto_base_id
        WHERE p.viaje_id = %s
        ORDER BY p.orden_entrega DESC, (pb.largo_cm * pb.ancho_cm) DESC
    """, (viaje_id,))

    if not raw_base and not raw_comp:
        return {'error': 'El viaje no tiene paradas con ventas asignadas.'}

    items = []
    for row in raw_base:
        for _ in range(int(row['cantidad'] or 1)):
            items.append(dict(row))

    # Cada unidad vendida del compuesto genera N piezas físicas (una por componente × cantidad_necesaria)
    for row in raw_comp:
        total_piezas = int(row['cantidad_venta'] or 1) * int(row['cantidad_necesaria'] or 1)
        for _ in range(total_piezas):
            items.append({
                'orden_entrega': row['orden_entrega'],
                'cliente': row['cliente'],
                'direccion': row['direccion'],
                'sku': row['sku'],
                'cantidad': 1,
                'producto_nombre': f"{row['producto_nombre']} (de {row['nombre_compuesto']})",
                'largo_cm': row['largo_cm'],
                'ancho_cm': row['ancho_cm'],
                'alto_cm': row['alto_cm'],
                'peso_gramos': row['peso_gramos'],
                'tipo': row['tipo'],
            })

    # Ordenar: última parada primero (va al fondo del camión), piezas más grandes primero
    items.sort(key=lambda x: (-x['orden_entrega'],
                               -((x['largo_cm'] or 0) * (x['ancho_cm'] or 0))))

    zona_results = []
    sobrantes = items[:]
    for zona in zonas:
        if not sobrantes:
            break
        res, sobrantes = _pack_zona(zona, sobrantes)
        zona_results.append(res)

    return {
        'viaje': viaje,
        'zonas': zona_results,
        'sobrantes': sobrantes,
        'entra_todo': len(sobrantes) == 0,
        'total_items': len(items),
    }


@app.route('/viajes', methods=['GET'])
@login_required
def viajes_lista():
    viajes = query_db("""
        SELECT v.*, f.nombre AS fletero_nombre,
               COUNT(p.id) AS total_paradas
        FROM viajes v
        JOIN fleteros f ON f.id = v.fletero_id
        LEFT JOIN viaje_paradas p ON p.viaje_id = v.id
        GROUP BY v.id
        ORDER BY v.fecha DESC, v.id DESC
    """)
    fleteros = query_db("SELECT * FROM fleteros WHERE activo = 1 ORDER BY nombre")
    return render_template('viajes.html', viajes=viajes, fleteros=fleteros)


@app.route('/viajes/nuevo', methods=['GET'])
@login_required
def viaje_nuevo():
    fleteros = query_db("""
        SELECT f.*, COUNT(z.id) AS total_zonas
        FROM fleteros f
        LEFT JOIN fletero_zonas z ON z.fletero_id = f.id
        WHERE f.activo = 1
        GROUP BY f.id
        ORDER BY f.nombre
    """)
    ventas_pendientes = query_db("""
        SELECT v.id, v.nombre_cliente AS cliente, v.direccion_entrega, v.metodo_envio,
               GROUP_CONCAT(CONCAT(i.cantidad, 'x ', COALESCE(pb.nombre, pc.nombre, i.sku))
                            ORDER BY COALESCE(pb.nombre, pc.nombre) SEPARATOR ', ') AS detalle
        FROM ventas v
        JOIN items_venta i ON i.venta_id = v.id
        LEFT JOIN productos_base pb ON pb.sku = i.sku
        LEFT JOIN productos_compuestos pc ON pc.sku = i.sku
        WHERE v.estado_entrega = 'pendiente'
          AND v.metodo_envio IN ('Turbo', 'Flex', 'Flete Propio')
        GROUP BY v.id
        ORDER BY v.fecha_venta DESC
        LIMIT 200
    """)
    return render_template('viaje_form.html',
                           fleteros=fleteros,
                           ventas_pendientes=ventas_pendientes)


@app.route('/viajes/<int:viaje_id>', methods=['GET'])
@login_required
def viaje_detalle(viaje_id):
    viaje = query_one("""
        SELECT v.*, f.nombre AS fletero_nombre
        FROM viajes v JOIN fleteros f ON f.id = v.fletero_id
        WHERE v.id = %s
    """, (viaje_id,))
    if not viaje:
        flash('Viaje no encontrado', 'warning')
        return redirect(url_for('viajes_lista'))

    paradas = query_db("""
        SELECT p.*,
               vt.nombre_cliente AS venta_cliente, vt.direccion_entrega,
               GROUP_CONCAT(
                   CONCAT(i.cantidad, 'x ', COALESCE(pb.nombre, pc.nombre, i.sku))
                   ORDER BY COALESCE(pb.nombre, pc.nombre) SEPARATOR ' | '
               ) AS items_texto
        FROM viaje_paradas p
        LEFT JOIN ventas vt ON vt.id = p.venta_id
        LEFT JOIN items_venta i ON i.venta_id = p.venta_id
        LEFT JOIN productos_base pb ON pb.sku = i.sku
        LEFT JOIN productos_compuestos pc ON pc.sku = i.sku
        WHERE p.viaje_id = %s
        GROUP BY p.id
        ORDER BY p.orden_entrega
    """, (viaje_id,))

    zonas_fletero = query_db("""
        SELECT * FROM fletero_zonas
        WHERE fletero_id = (SELECT fletero_id FROM viajes WHERE id = %s)
        ORDER BY orden
    """, (viaje_id,))

    ventas_pendientes = query_db("""
        SELECT v.id, v.nombre_cliente AS cliente, v.direccion_entrega, v.metodo_envio,
               GROUP_CONCAT(CONCAT(i.cantidad, 'x ', COALESCE(pb.nombre, pc.nombre, i.sku))
                            ORDER BY COALESCE(pb.nombre, pc.nombre) SEPARATOR ', ') AS detalle
        FROM ventas v
        JOIN items_venta i ON i.venta_id = v.id
        LEFT JOIN productos_base pb ON pb.sku = i.sku
        LEFT JOIN productos_compuestos pc ON pc.sku = i.sku
        WHERE v.estado_entrega = 'pendiente'
          AND v.metodo_envio IN ('Turbo', 'Flex', 'Flete Propio')
        GROUP BY v.id
        ORDER BY v.fecha_venta DESC
        LIMIT 200
    """)

    resultado = _calcular_carga_viaje(viaje_id)

    return render_template('viaje_detalle.html',
                           viaje=viaje,
                           paradas=paradas,
                           zonas_fletero=zonas_fletero,
                           ventas_pendientes=ventas_pendientes,
                           resultado=resultado)


@app.route('/viajes/guardar', methods=['POST'])
@login_required
def viajes_guardar():
    data = request.get_json()
    accion = data.get('accion')
    try:
        if accion == 'crear_viaje':
            vid = execute_db("""
                INSERT INTO viajes (fletero_id, fecha, estado, notas)
                VALUES (%s, %s, 'borrador', %s)
            """, (data['fletero_id'], data['fecha'], data.get('notas', '')))
            return jsonify({'ok': True, 'viaje_id': vid})

        elif accion == 'eliminar_viaje':
            execute_db("DELETE FROM viaje_paradas WHERE viaje_id = %s", (data['viaje_id'],))
            execute_db("DELETE FROM viajes WHERE id = %s", (data['viaje_id'],))

        elif accion == 'cambiar_estado':
            execute_db("UPDATE viajes SET estado = %s WHERE id = %s",
                       (data['estado'], data['viaje_id']))

        elif accion == 'agregar_parada':
            execute_db("""
                INSERT INTO viaje_paradas
                    (viaje_id, orden_entrega, venta_id, cliente, direccion, notas)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (data['viaje_id'], int(data['orden_entrega']),
                  data.get('venta_id') or None,
                  data.get('cliente', ''), data.get('direccion', ''),
                  data.get('notas', '')))

        elif accion == 'eliminar_parada':
            execute_db("DELETE FROM viaje_paradas WHERE id = %s AND viaje_id = %s",
                       (data['parada_id'], data['viaje_id']))

        elif accion == 'reordenar_parada':
            execute_db("UPDATE viaje_paradas SET orden_entrega = %s WHERE id = %s",
                       (int(data['orden_entrega']), data['parada_id']))

        elif accion == 'guardar_zona':
            if data.get('zona_id'):
                execute_db("""
                    UPDATE fletero_zonas
                    SET nombre_zona=%s, largo_cm=%s, ancho_cm=%s, alto_cm=%s, orden=%s
                    WHERE id=%s AND fletero_id=%s
                """, (data['nombre_zona'], int(data['largo_cm']), int(data['ancho_cm']),
                      int(data['alto_cm']), int(data.get('orden', 1)),
                      data['zona_id'], data['fletero_id']))
            else:
                execute_db("""
                    INSERT INTO fletero_zonas
                        (fletero_id, nombre_zona, largo_cm, ancho_cm, alto_cm, orden)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (data['fletero_id'], data['nombre_zona'],
                      int(data['largo_cm']), int(data['ancho_cm']),
                      int(data['alto_cm']), int(data.get('orden', 1))))

        elif accion == 'eliminar_zona':
            execute_db("DELETE FROM fletero_zonas WHERE id=%s AND fletero_id=%s",
                       (data['zona_id'], data['fletero_id']))

        elif accion == 'get_zonas':
            zonas = query_db(
                "SELECT * FROM fletero_zonas WHERE fletero_id=%s ORDER BY orden",
                (data['fletero_id'],))
            return jsonify({'ok': True, 'zonas': [dict(z) for z in zonas]})

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    return jsonify({'ok': True})


# ============================================================================
# PANEL TIENDA WEB — Precios y Ofertas
# ============================================================================

@app.route('/tienda-admin/precios', methods=['GET'])
@admin_required
def tienda_precios():
    productos_base = query_db("""
        SELECT sku, nombre, tipo, linea, modelo, medida, precio_base, descuento_catalogo, 'base' as origen
        FROM productos_base
        WHERE tipo IN ('colchon','almohada')
        ORDER BY tipo, linea, modelo, medida
    """)
    bases = query_db("""
        SELECT sku, nombre, precio_base, descuento_catalogo
        FROM productos_base
        WHERE tipo = 'base'
        ORDER BY nombre
    """)
    # Calcular precio de conjuntos dinámicamente = colchon + base * cantidad
    colchones_map = {p['sku']: p for p in productos_base if p['tipo'] == 'colchon'}
    bases_map     = {b['sku']: b for b in bases}
    conjuntos_cfg = query_db("SELECT colchon_sku, base_sku_default, cantidad_bases FROM conjunto_configuracion WHERE activo = 1")
    # Descuentos actuales de productos_compuestos (para poder editarlos)
    desc_comp = {r['sku']: r for r in query_db("SELECT sku, descuento_catalogo FROM productos_compuestos WHERE activo = 1")}

    conjuntos_calc = []
    for cfg in conjuntos_cfg:
        col = colchones_map.get(cfg['colchon_sku'])
        base = bases_map.get(cfg['base_sku_default'])
        if not col or not base:
            continue
        precio_col  = float(col['precio_base'] or 0)
        precio_base = float(base['precio_base'] or 0)
        cant        = int(cfg['cantidad_bases'] or 1)
        precio_conj = precio_col + precio_base * cant
        # SKU conjunto = S + resto del SKU colchon (ej: CTR80 -> STR80)
        sku_col  = cfg['colchon_sku']
        sku_conj = 'S' + sku_col[1:] if sku_col.startswith('C') else 'S' + sku_col
        desc_actual = desc_comp.get(sku_conj, {}).get('descuento_catalogo', None)
        conjuntos_calc.append({
            'sku':                sku_conj,
            'sku_colchon':        sku_col,
            'base_sku':           cfg['base_sku_default'],
            'nombre':             f"Sommier y Colchón {col['modelo']} {col['medida']}cm",
            'precio_colchon':     precio_col,
            'precio_base_unit':   precio_base,
            'cantidad_bases':     cant,
            'precio_conjunto':    precio_conj,
            'descuento_catalogo': desc_actual,
        })
    conjuntos_calc.sort(key=lambda x: (x['nombre']))

    return render_template('tienda_precios.html',
        productos_base=productos_base,
        bases=bases,
        conjuntos_calc=conjuntos_calc,
        precio_costos_map=_build_precio_costos_map(),
    )


@app.route('/tienda-admin/precios/guardar', methods=['POST'])
@admin_required
def tienda_precios_guardar():
    data = request.get_json()
    cambios = data.get('cambios', [])
    if not cambios:
        return jsonify({'ok': False, 'error': 'Sin cambios'})
    actualizados = 0
    try:
        for c in cambios:
            sku    = c.get('sku', '').strip()
            precio = c.get('precio')
            origen = c.get('origen', 'base')
            if not sku or precio is None:
                continue
            try:
                precio = float(str(precio).replace(',', '.'))
            except ValueError:
                continue
            if origen == 'compuesto':
                execute_db("UPDATE productos_compuestos SET precio_base=%s WHERE sku=%s", (precio, sku))
            else:
                execute_db("UPDATE productos_base SET precio_base=%s WHERE sku=%s", (precio, sku))
            actualizados += 1
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    return jsonify({'ok': True, 'actualizados': actualizados})


@app.route('/tienda-admin/precios/descuento', methods=['POST'])
@admin_required
def tienda_precios_descuento():
    data = request.get_json()
    accion = data.get('accion')  # 'set', 'quitar', 'set_todos', 'quitar_todos'
    try:
        if accion == 'set':
            sku    = data.get('sku', '').strip()
            pct    = float(data.get('pct', 0))
            origen = data.get('origen', 'base')
            if origen == 'compuesto':
                execute_db("UPDATE productos_compuestos SET descuento_catalogo=%s WHERE sku=%s", (pct, sku))
            else:
                execute_db("UPDATE productos_base SET descuento_catalogo=%s WHERE sku=%s", (pct, sku))
        elif accion == 'quitar':
            sku    = data.get('sku', '').strip()
            origen = data.get('origen', 'base')
            if origen == 'compuesto':
                execute_db("UPDATE productos_compuestos SET descuento_catalogo=NULL WHERE sku=%s", (sku,))
            else:
                execute_db("UPDATE productos_base SET descuento_catalogo=NULL WHERE sku=%s", (sku,))
        elif accion == 'set_todos':
            pct = float(data.get('pct', 0))
            execute_db("UPDATE productos_base SET descuento_catalogo=%s WHERE tipo IN ('colchon','almohada')", (pct,))
            execute_db("UPDATE productos_compuestos SET descuento_catalogo=%s WHERE activo=1", (pct,))
        elif accion == 'quitar_todos':
            execute_db("UPDATE productos_base SET descuento_catalogo=NULL WHERE tipo IN ('colchon','almohada')", ())
            execute_db("UPDATE productos_compuestos SET descuento_catalogo=NULL WHERE activo=1", ())
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    return jsonify({'ok': True})


@app.route('/tienda-admin/ofertas', methods=['GET'])
@login_required
def tienda_ofertas():
    ofertas = query_db("""
        SELECT o.id, o.sku, o.descuento_pct, o.orden, o.activo,
               COALESCE(pb.nombre, pc.nombre, o.sku) as nombre
        FROM ofertas_home o
        LEFT JOIN productos_base pb ON pb.sku = o.sku
        LEFT JOIN productos_compuestos pc ON pc.sku = o.sku
        ORDER BY o.orden, o.id
    """)
    skus_base = query_db("SELECT sku, nombre, 'base' as origen FROM productos_base WHERE tipo IN ('colchon','almohada') ORDER BY nombre")
    skus_comp = query_db("SELECT sku, nombre, 'compuesto' as origen FROM productos_compuestos WHERE activo=1 ORDER BY nombre")
    todos_skus = list(skus_base) + list(skus_comp)
    return render_template('tienda_ofertas.html',
        ofertas=ofertas,
        todos_skus=todos_skus,
    )


@app.route('/tienda-admin/ofertas/guardar', methods=['POST'])
@login_required
def tienda_ofertas_guardar():
    data = request.get_json()
    accion = data.get('accion')
    try:
        if accion == 'agregar':
            sku  = data.get('sku', '').strip()
            pct  = float(data.get('descuento_pct', 8))
            existe = query_db("SELECT id FROM ofertas_home WHERE sku=%s", (sku,))
            if existe:
                return jsonify({'ok': False, 'error': 'SKU ya existe en ofertas'})
            max_orden = query_db("SELECT COALESCE(MAX(orden),0)+1 as orden FROM ofertas_home")
            orden = max_orden[0]['orden'] if max_orden else 1
            execute_db("INSERT INTO ofertas_home (sku, descuento_pct, orden, activo) VALUES (%s,%s,%s,1)", (sku, pct, orden))
        elif accion == 'eliminar':
            execute_db("DELETE FROM ofertas_home WHERE id=%s", (data.get('id'),))
        elif accion == 'toggle':
            execute_db("UPDATE ofertas_home SET activo = NOT activo WHERE id=%s", (data.get('id'),))
        elif accion == 'pct':
            execute_db("UPDATE ofertas_home SET descuento_pct=%s WHERE id=%s", (float(data.get('descuento_pct', 8)), data.get('id')))
        elif accion == 'reordenar':
            for item in data.get('items', []):
                execute_db("UPDATE ofertas_home SET orden=%s WHERE id=%s", (item['orden'], item['id']))
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    return jsonify({'ok': True})



# ── NOTA DE PEDIDO PDF ────────────────────────────────────────────────────────

@app.route('/ventas/<int:venta_id>/nota-pedido')
@login_required
def nota_pedido_pdf(venta_id):
    from flask import make_response
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
    venta = cursor.fetchone()
    if not venta:
        cursor.close(); conn.close()
        return 'Venta no encontrada', 404

    cursor.execute('''
        SELECT iv.sku, iv.cantidad, iv.precio_unitario,
               COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre
        FROM items_venta iv
        LEFT JOIN productos_base pb ON pb.sku = iv.sku
        LEFT JOIN productos_compuestos pc ON pc.sku = iv.sku
        WHERE iv.venta_id = %s
    ''', (venta_id,))
    items = cursor.fetchall()
    cursor.close(); conn.close()

    # Parsear notas
    notas_raw = (venta.get('notas') or '').replace('\\n', '\n')
    notas_dict = {}
    for linea in notas_raw.split('\n'):
        if ':' in linea:
            k, v = linea.split(':', 1)
            notas_dict[k.strip()] = v.strip()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    st_title  = ParagraphStyle('title',  fontSize=16, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=4)
    st_sub    = ParagraphStyle('sub',    fontSize=9,  fontName='Helvetica',      alignment=TA_CENTER, spaceAfter=12, textColor=colors.grey)
    st_h      = ParagraphStyle('h',      fontSize=10, fontName='Helvetica-Bold', spaceBefore=10, spaceAfter=4)
    st_normal = ParagraphStyle('n',      fontSize=9,  fontName='Helvetica',      spaceAfter=3)
    st_right  = ParagraphStyle('r',      fontSize=9,  fontName='Helvetica',      alignment=TA_RIGHT)

    story = []

    # Encabezado
    story.append(Paragraph('MERCADOMUEBLES', st_title))
    story.append(Paragraph('Distribuidores oficiales Cannon · Bahía Blanca 1777, Floresta, CABA', st_sub))
    story.append(HRFlowable(width='100%', thickness=1.5, color=colors.HexColor('#1a1a2e')))
    story.append(Spacer(1, 8))

    # Número y fecha
    fecha_str = venta['fecha_venta'].strftime('%d/%m/%Y %H:%M') if venta.get('fecha_venta') else ''
    story.append(Paragraph(f'<b>NOTA DE PEDIDO</b> · {venta["numero_venta"]}', st_h))
    story.append(Paragraph(f'Fecha: {fecha_str}', st_normal))
    story.append(Spacer(1, 8))

    # Datos del cliente
    story.append(Paragraph('DATOS DEL CLIENTE', st_h))
    story.append(Paragraph(f'<b>Nombre:</b> {venta.get("nombre_cliente") or "-"}', st_normal))
    if venta.get('dni_cliente'):
        story.append(Paragraph(f'<b>DNI/CUIT:</b> {venta["dni_cliente"]}', st_normal))
    if venta.get('telefono_cliente'):
        story.append(Paragraph(f'<b>Teléfono:</b> {venta["telefono_cliente"]}', st_normal))
    story.append(Spacer(1, 8))

    # Entrega
    story.append(Paragraph('ENTREGA', st_h))
    tipo = venta.get('tipo_entrega', '')
    metodo = venta.get('metodo_envio', '')
    if tipo == 'retiro':
        story.append(Paragraph('<b>Tipo:</b> Retiro en local', st_normal))
        story.append(Paragraph('Bahía Blanca 1777, Floresta, CABA · Lunes a Viernes 8-12hs y 14-16:30hs', st_normal))
    else:
        story.append(Paragraph(f'<b>Tipo:</b> Envío a domicilio · <b>Método:</b> {metodo}', st_normal))
        if venta.get('direccion_entrega'):
            story.append(Paragraph(f'<b>Dirección:</b> {venta["direccion_entrega"]}', st_normal))
        # Datos según método
        if metodo == 'ME2':
            for campo in ['MPID', 'VEID', 'SHID', 'TRID']:
                if notas_dict.get(campo):
                    story.append(Paragraph(f'<b>{campo}:</b> {notas_dict[campo]}', st_normal))
        elif metodo in ('Zippin', 'Flete Propio'):
            for campo in ['MPID', 'ZNID', 'ZN_URL']:
                if notas_dict.get(campo):
                    story.append(Paragraph(f'<b>{campo}:</b> {notas_dict[campo]}', st_normal))
            if venta.get('fecha_entrega_estimada'):
                story.append(Paragraph(f'<b>Entrega estimada:</b> {venta["fecha_entrega_estimada"]}', st_normal))
    story.append(Spacer(1, 8))

    # Productos
    story.append(Paragraph('PRODUCTOS', st_h))
    tabla_data = [['SKU', 'Descripción', 'Cant.', 'Precio Unit.', 'Subtotal']]
    total_productos = 0
    for it in items:
        sub = float(it['precio_unitario'] or 0) * int(it['cantidad'] or 1)
        total_productos += sub
        tabla_data.append([
            it['sku'],
            it['nombre'] or it['sku'],
            str(it['cantidad']),
            f'${float(it["precio_unitario"]):,.0f}'.replace(',', '.'),
            f'${sub:,.0f}'.replace(',', '.'),
        ])
    tabla = Table(tabla_data, colWidths=[3*cm, 8*cm, 1.5*cm, 3*cm, 3*cm])
    tabla.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,-1), 8),
        ('ALIGN',       (2,0), (-1,-1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('TOPPADDING',  (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
    ]))
    story.append(tabla)
    story.append(Spacer(1, 6))

    # Totales
    costo_flete = float(venta.get('costo_flete') or 0)
    total_final = total_productos + costo_flete
    totales = [[f'Subtotal productos:', f'${total_productos:,.0f}'.replace(',', '.')]]
    if costo_flete > 0:
        totales.append([f'Costo de envío:', f'${costo_flete:,.0f}'.replace(',', '.')])
    totales.append(['TOTAL:', f'${total_final:,.0f}'.replace(',', '.')])
    t_totales = Table(totales, colWidths=[14*cm, 4.5*cm])
    t_totales.setStyle(TableStyle([
        ('ALIGN',    (1,0), (1,-1), 'RIGHT'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('LINEABOVE',(0,-1),(-1,-1), 1, colors.HexColor('#1a1a2e')),
        ('TOPPADDING',(0,0),(-1,-1), 3),
    ]))
    story.append(t_totales)
    story.append(Spacer(1, 8))

    # Pago
    story.append(Paragraph('PAGO', st_h))
    metodo_pago = venta.get('metodo_pago') or 'MercadoPago'
    abonado = float(venta.get('importe_abonado') or 0)
    mp = float(venta.get('pago_mercadopago') or 0)
    ef = float(venta.get('pago_efectivo') or 0)
    story.append(Paragraph(f'<b>Método:</b> {metodo_pago}', st_normal))
    story.append(Paragraph(f'<b>Total abonado:</b> ${abonado:,.0f}'.replace(',', '.'), st_normal))
    if mp > 0:
        story.append(Paragraph(f'  · MercadoPago: ${mp:,.0f}'.replace(',', '.'), st_normal))
    if ef > 0:
        story.append(Paragraph(f'  · Efectivo: ${ef:,.0f}'.replace(',', '.'), st_normal))

    # Demora sin stock
    demora_val = notas_dict.get('DEMORA', '')
    if demora_val:
        story.append(Spacer(1, 10))
        st_demora = ParagraphStyle('demora', fontSize=10, fontName='Helvetica-Bold',
                                   textColor=colors.HexColor('#856404'),
                                   backColor=colors.HexColor('#fff3cd'),
                                   borderPad=6, spaceBefore=4, spaceAfter=4)
        story.append(Paragraph(f'⚠ PRODUCTO CON DEMORA: {demora_val}', st_demora))

    story.append(Spacer(1, 16))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.grey))
    story.append(Paragraph('ventas@mercadomuebles.com.ar · www.mercadomuebles.com.ar · WhatsApp 11 2627-5185', st_sub))

    doc.build(story)
    buf.seek(0)
    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=nota-pedido-{venta["numero_venta"]}.pdf'
    return response


@app.route('/ventas/<int:venta_id>/papel-azul')
@login_required
def papel_azul_pdf(venta_id):
    """Genera el papel azul de despacho para ventas Flex o Flete Propio."""
    from flask import make_response
    import io
    from datetime import datetime
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import simpleSplit

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
    venta = cursor.fetchone()
    if not venta:
        cursor.close(); conn.close()
        return 'Venta no encontrada', 404

    cursor.execute('''
        SELECT iv.sku, iv.cantidad, iv.precio_unitario,
               COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre,
               pb.tipo as tipo_producto
        FROM items_venta iv
        LEFT JOIN productos_base pb ON pb.sku = iv.sku
        LEFT JOIN productos_compuestos pc ON pc.sku = iv.sku
        WHERE iv.venta_id = %s
    ''', (venta_id,))
    items = cursor.fetchall()
    cursor.close(); conn.close()

    es_flex  = (venta.get('metodo_envio') or '').lower() == 'flex'
    es_turbo = (venta.get('metodo_envio') or '').lower() == 'turbo'

    # Calcular saldo
    importe_total   = float(venta.get('importe_total') or 0)
    importe_abonado = float(venta.get('importe_abonado') or 0)
    costo_flete     = float(venta.get('costo_flete') or 0)
    saldo = 0 if (es_flex or es_turbo) else max(0, importe_total + costo_flete - importe_abonado)

    # Desglosar items
    def desglosar_item(sku, nombre, cantidad):
        sku_up = sku.upper()
        if sku_up.startswith('S') and len(sku_up) > 1 and sku_up[1].isalpha() and not sku_up.startswith('SU'):
            sku_col = 'C' + sku_up[1:]
            conn2 = get_db_connection()
            cur2 = conn2.cursor()
            cur2.execute("""
                SELECT cc.cantidad_bases, cc.base_sku_default,
                       pb_col.nombre as nombre_col,
                       pb_base.medida as medida_base
                FROM conjunto_configuracion cc
                JOIN productos_base pb_col ON pb_col.sku = cc.colchon_sku
                LEFT JOIN productos_base pb_base ON pb_base.sku = cc.base_sku_default
                WHERE cc.colchon_sku = %s AND cc.activo = 1
            """, (sku_col,))
            cfg = cur2.fetchone()
            cur2.close(); conn2.close()
            if cfg:
                cant_bases = int(cfg['cantidad_bases'] or 1)
                nombre_col = cfg['nombre_col'] or nombre
                medida_base = cfg['medida_base'] or ''
                try:
                    ancho_base = int(medida_base.split('x')[0])
                except:
                    ancho_base = 0
                tipo_pata = 'x7' if ancho_base > 100 else 'x6'
                bases_str = f"{cant_bases} base{'s' if cant_bases > 1 else ''}"
                patas_str = f"{cant_bases} pata {tipo_pata}" if cant_bases == 1 else f"{cant_bases} patas {tipo_pata}"
                linea1 = f"{cantidad} Sommier {nombre_col}"
                linea2 = f"({cantidad} colchón + {bases_str} + {patas_str})"
                return [linea1, linea2]
        return [f"{cantidad} {nombre}"]

    descripcion_items = []
    for item in items:
        descripcion_items.extend(desglosar_item(item['sku'], item['nombre'], item['cantidad']))

    # Extraer rango horario de notas para Turbo (formato: "Turbo HH:MM-HH:MM")
    turbo_rango = ''
    if es_turbo:
        notas_venta = venta.get('notas') or ''
        import re as _re
        m = _re.search(r'(\d{1,2}:\d{2}-\d{1,2}:\d{2})', notas_venta)
        if m:
            turbo_rango = m.group(1)

    # ── Canvas PDF ───────────────────────────────────────────────────────────
    PAGE_W = 165 * mm
    PAGE_H = 215 * mm
    MARGIN_L = 7 * mm
    MARGIN_R = 7 * mm
    MARGIN_TOP = 70 * mm     # área imprimible empieza en y = PAGE_H - 70mm
    MARGIN_BOT = 27 * mm
    TXT_W = PAGE_W - MARGIN_L - MARGIN_R  # ancho útil

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))

    # Posición Y inicial (desde arriba del área imprimible)
    y = PAGE_H - MARGIN_TOP

    def draw_text(text, x, y, font, size, align='left', max_width=None):
        """Dibuja texto y retorna nueva y."""
        c.setFont(font, size)
        if align == 'right':
            c.drawRightString(x, y, text)
        elif align == 'center':
            c.drawCentredString(x, y, text)
        else:
            if max_width:
                lines = simpleSplit(text, font, size, max_width)
                for line in lines:
                    c.drawString(x, y, line)
                    y -= size * 1.3
                return y
            c.drawString(x, y, text)
        return y - size * 1.3

    # ── Fecha (línea 1) y Nombre (línea 2), centrados ──
    fecha_hoy = datetime.now().strftime('%d-%m-%Y')
    nombre_cliente = venta.get('nombre_cliente') or '-'
    cx = PAGE_W / 2
    c.setFont('Helvetica', 13)
    c.drawCentredString(cx, y, fecha_hoy)
    y -= 8 * mm
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(cx, y, nombre_cliente)
    y -= 8 * mm

    # Teléfono — solo para Flete Propio
    if not es_flex and not es_turbo:
        telefono = venta.get('telefono_cliente') or ''
        if telefono:
            c.setFont('Helvetica', 13)
            c.drawCentredString(cx, y, telefono)
            y -= 8 * mm

    y -= 5 * mm

    # ── Producto ── (bold, grande, centrado, 2 líneas)
    for i, linea in enumerate(descripcion_items):
        font = 'Helvetica-Bold'
        size = 17 if i == 0 else 14
        c.setFont(font, size)
        lines = simpleSplit(linea, font, size, TXT_W)
        for ln in lines:
            c.drawCentredString(cx, y, ln)
            y -= size * 1.4
    y -= 8 * mm

    # ── Total ──
    saldo_str = f'Total: ${"{:,.0f}".format(saldo).replace(",",".")}.-'
    c.setFont('Helvetica-Bold', 17)
    c.drawCentredString(cx, y, saldo_str)
    y -= 16 * mm

    # ── Dirección ── (22pt)
    direccion = venta.get('direccion_entrega') or ''
    if direccion:
        c.setFont('Helvetica-Bold', 22)
        lines = simpleSplit(direccion, 'Helvetica-Bold', 22, TXT_W)
        for ln in lines:
            c.drawCentredString(cx, y, ln)
            y -= 22 * 1.4
    y -= 16 * mm

    # ── FLEX / TURBO ── (49pt)
    if es_flex:
        c.setFont('Helvetica-Bold', 49)
        c.drawCentredString(cx, y, 'FLEX')
    elif es_turbo:
        c.setFont('Helvetica-Bold', 49)
        c.drawCentredString(cx, y, 'TURBO')
        if turbo_rango:
            y -= 49 * 1.3
            c.setFont('Helvetica-Bold', 28)
            c.drawCentredString(cx, y, turbo_rango)

    c.save()
    buf.seek(0)
    response = make_response(buf.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=despacho-{venta_id}.pdf'
    return response

@app.route('/tienda-admin/cupones', methods=['GET'])
@login_required
def tienda_cupones():
    execute_db("""
        CREATE TABLE IF NOT EXISTS cupones (
            id               INT AUTO_INCREMENT PRIMARY KEY,
            codigo           VARCHAR(50) NOT NULL UNIQUE,
            tipo             ENUM('pct','fijo') NOT NULL DEFAULT 'pct',
            valor            DECIMAL(10,2) NOT NULL,
            minimo_compra    DECIMAL(10,2) DEFAULT 0,
            usos_maximos     INT DEFAULT NULL,
            usos_actuales    INT DEFAULT 0,
            fecha_vencimiento DATE DEFAULT NULL,
            solo_un_uso      TINYINT DEFAULT 0,
            activo           TINYINT DEFAULT 1,
            created_at       TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    execute_db("""
        CREATE TABLE IF NOT EXISTS cupones_uso (
            id           INT AUTO_INCREMENT PRIMARY KEY,
            cupon_id     INT NOT NULL,
            email        VARCHAR(255),
            telefono     VARCHAR(50),
            venta_numero VARCHAR(100),
            fecha_uso    TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cupones = query_db("""
        SELECT c.*, COUNT(cu.id) as usos_reales
        FROM cupones c
        LEFT JOIN cupones_uso cu ON cu.cupon_id = c.id
        GROUP BY c.id
        ORDER BY c.created_at DESC
    """)
    execute_db("""
        CREATE TABLE IF NOT EXISTS suscriptores (
            id       INT AUTO_INCREMENT PRIMARY KEY,
            email    VARCHAR(255) NOT NULL UNIQUE,
            cupon_id INT DEFAULT NULL,
            fecha    TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    suscriptores = query_db("""
        SELECT s.id, s.email, s.fecha, c.codigo, c.usos_actuales, c.activo as cupon_activo
        FROM suscriptores s
        LEFT JOIN cupones c ON c.id = s.cupon_id
        ORDER BY s.fecha DESC
    """)
    return render_template('tienda_cupones.html', cupones=cupones, suscriptores=suscriptores)


@app.route('/tienda-admin/cupones/guardar', methods=['POST'])
@login_required
def tienda_cupones_guardar():
    data   = request.get_json()
    accion = data.get('accion')
    try:
        if accion == 'crear':
            codigo   = data.get('codigo', '').strip().upper()
            tipo     = data.get('tipo', 'pct')
            valor    = float(data.get('valor', 0))
            minimo   = float(data.get('minimo_compra', 0) or 0)
            usos_max = int(data.get('usos_maximos') or 0) or None
            venc     = data.get('fecha_vencimiento') or None
            solo_uno = int(data.get('solo_un_uso', 0))
            if not codigo or valor <= 0:
                return jsonify({'ok': False, 'error': 'Código y valor son obligatorios'})
            existe = query_db("SELECT id FROM cupones WHERE codigo=%s", (codigo,))
            if existe:
                return jsonify({'ok': False, 'error': 'Código ya existe'})
            execute_db("""
                INSERT INTO cupones (codigo, tipo, valor, minimo_compra, usos_maximos, fecha_vencimiento, solo_un_uso, activo)
                VALUES (%s, %s, %s, %s, %s, %s, %s, 1)
            """, (codigo, tipo, valor, minimo, usos_max, venc, solo_uno))
        elif accion == 'toggle':
            execute_db("UPDATE cupones SET activo = NOT activo WHERE id=%s", (data.get('id'),))
        elif accion == 'eliminar':
            execute_db("DELETE FROM cupones WHERE id=%s", (data.get('id'),))
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    return jsonify({'ok': True})


@app.route('/tienda-admin/suscriptores/eliminar', methods=['POST'])
@login_required
def tienda_suscriptores_eliminar():
    data = request.get_json()
    try:
        execute_db("DELETE FROM suscriptores WHERE id=%s", (data.get('id'),))
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    return jsonify({'ok': True})
# ============================================================================

# SKUs de almohadas a excluir de actualización ML
SKUS_ALMOHADA = {
    'CERVICAL', 'CLASICA', 'DORAL', 'DUAL', 'EXCLUSIVE', 'PLATINO',
    'PRUEBA', 'RENOVATION', 'SUBLIME', 'DORALX2', 'DUALX2',
    'EXCLUSIVEX2', 'PLATINOX2', 'PLATINOX4',
    # combos con almohada como ÚNICO componente de sommier (CEX140+2 SÍ se actualiza)
}

# SKUs compac a excluir
def _es_compac(sku):
    return sku.upper().startswith('CCO')

# SKU de almohada pura (no combos mixtos)
def _es_almohada(sku):
    return sku.upper() in SKUS_ALMOHADA

# ¿Aplica lógica Z a este SKU? Solo sommiers (S*) todas las medidas
# y colchones (C* no CCO) a partir de 140
def _aplica_logica_z(sku):
    """
    Determina si aplica lógica Z (demora) para un SKU.
    Sommiers (S*): siempre.
    Colchones (C*, no CCO): solo si el ANCHO >= 140.
    El ancho está en los primeros dígitos del SKU antes del modelo.
    Ej: CPR8020 → ancho=80, CPR14020 → ancho=140, CEX140 → ancho=140
    """
    sku = sku.upper()
    if sku.startswith('S'):
        return True
    if sku.startswith('C') and not sku.startswith('CCO'):
        import re
        # Buscar el primer grupo de dígitos que sea el ancho (80, 90, 100, 140, 150, 160, 180, 200)
        nums = re.findall(r'\d+', sku)
        if nums:
            # El ancho es el primer número del SKU
            # Para CPR8020: nums=['8020'] → tomar los primeros 2-3 dígitos
            # Para CEX140: nums=['140'] → 140
            # Para CPR14020: nums=['14020'] → primeros 3 dígitos = 140
            primer_num = nums[0]
            # Si tiene 4+ dígitos, los primeros 2-3 son el ancho
            if len(primer_num) >= 4:
                # Probar con 3 dígitos primero (140, 150, 160, 180, 200)
                ancho = int(primer_num[:3])
                if ancho not in (80, 90, 100, 140, 150, 160, 180, 200):
                    # Probar con 2 dígitos (80, 90)
                    ancho = int(primer_num[:2])
            else:
                ancho = int(primer_num)
            return ancho >= 140
    return False


def _poner_demora_ml(mla_id, access_token, dias=7):
    """Poner demora de manufacturing_time en una publi ML."""
    payload = {"sale_terms": [{"id": "MANUFACTURING_TIME", "value_name": f"{dias} días"}]}
    try:
        r = ml_request('put', f'https://api.mercadolibre.com/items/{mla_id}',
                       access_token, json_data=payload)
        return r.status_code == 200
    except Exception as e:
        print(f"[AUTO-ML] Error poniendo demora en {mla_id}: {e}")
        return False


def _quitar_demora_ml(mla_id, access_token):
    """Quitar demora de manufacturing_time en una publi ML."""
    ok, msg = quitar_manufacturing_time_ml(mla_id, access_token)
    return ok


def actualizar_publicaciones_ml(skus_base_afectados):
    """
    Dado un set de SKUs base que cambiaron disponible, actualiza en ML
    todas las publicaciones relacionadas (sin Z y con Z).

    Exclusiones: almohadas puras, compac (CCO*).
    Lógica Z: aplica solo a sommiers y colchones >= 140.
    """
    if not skus_base_afectados:
        return

    access_token = cargar_ml_token()
    if not access_token:
        print("[AUTO-ML] Sin access_token, abortando actualización.")
        return

    # Calcular stock disponible de todos los SKUs
    try:
        stock_todos = calcular_stock_por_sku()
    except Exception as e:
        print(f"[AUTO-ML] Error calculando stock: {e}")
        return

    # Para cada SKU base afectado, encontrar todos los combos/productos que lo usan
    skus_a_actualizar = set()

    for sku_base in skus_base_afectados:
        sku_base = sku_base.upper()
        # El propio SKU base (si tiene publis directas)
        skus_a_actualizar.add(sku_base)
        # Combos que lo contienen
        try:
            combos = query_db('''
                SELECT pc.sku FROM productos_compuestos pc
                JOIN componentes c ON c.producto_compuesto_id = pc.id
                JOIN productos_base pb ON c.producto_base_id = pb.id
                WHERE pb.sku = %s AND pc.activo = 1
            ''', (sku_base,))
            for combo in combos:
                skus_a_actualizar.add(combo['sku'].upper())
        except Exception as e:
            print(f"[AUTO-ML] Error buscando combos de {sku_base}: {e}")

    print(f"[AUTO-ML] SKUs a actualizar en ML: {skus_a_actualizar}")

    for sku in skus_a_actualizar:
        # Excluir almohadas puras y compac
        if _es_almohada(sku) or _es_compac(sku):
            print(f"[AUTO-ML] Excluido: {sku}")
            continue

        disponible = 0
        if sku in stock_todos:
            disponible = max(0, stock_todos[sku]['stock_disponible'])
        else:
            print(f"[AUTO-ML] SKU {sku} no encontrado en stock_todos, disponible=0")

        # ── Publicaciones SIN Z ──────────────────────────────────────
        try:
            pubs_sin_z = query_db(
                "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
                (sku,)
            )
            for pub in pubs_sin_z:
                mla = pub['mla_id']
                ok, msg = actualizar_stock_ml(mla, disponible, access_token)
                print(f"[AUTO-ML] {sku} sin Z → {mla} stock={disponible}: {'✅' if ok else '❌'} {msg}")
                time.sleep(0.5)
        except Exception as e:
            print(f"[AUTO-ML] Error actualizando sin Z de {sku}: {e}")

        # ── Publicaciones CON Z (solo si aplica lógica Z) ────────────
        if not _aplica_logica_z(sku):
            continue

        sku_z = sku + 'Z'
        try:
            pubs_z = query_db(
                "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
                (sku_z,)
            )
            if not pubs_z:
                continue

            for pub in pubs_z:
                mla = pub['mla_id']
                if disponible > 0:
                    # Hay stock: subir disponible y quitar demora
                    ok, msg = actualizar_stock_ml(mla, disponible, access_token)
                    print(f"[AUTO-ML] {sku_z} → {mla} stock={disponible}: {'✅' if ok else '❌'} {msg}")
                    time.sleep(0.5)
                    ok2 = _quitar_demora_ml(mla, access_token)
                    print(f"[AUTO-ML] {sku_z} → {mla} quitar demora: {'✅' if ok2 else '❌'}")
                    time.sleep(0.5)
                else:
                    # Sin stock: stock=1 + 7 días demora
                    ok, msg = actualizar_stock_ml(mla, 1, access_token)
                    print(f"[AUTO-ML] {sku_z} → {mla} stock=1: {'✅' if ok else '❌'} {msg}")
                    time.sleep(0.5)
                    ok2 = _poner_demora_ml(mla, access_token, dias=7)
                    print(f"[AUTO-ML] {sku_z} → {mla} poner demora 7d: {'✅' if ok2 else '❌'}")
                    time.sleep(0.5)
        except Exception as e:
            print(f"[AUTO-ML] Error actualizando con Z de {sku}: {e}")


def _extraer_skus_base_de_items(items):
    """
    Dado una lista de items [{sku, cantidad}], devuelve el set de SKUs base
    involucrados (expandiendo combos a sus componentes).
    """
    skus_base = set()
    for item in items:
        sku = item['sku'].upper()
        # ¿Es combo?
        combo = query_db("SELECT id FROM productos_compuestos WHERE sku = %s", (sku,))
        if combo:
            comps = query_db('''
                SELECT pb.sku FROM componentes c
                JOIN productos_base pb ON c.producto_base_id = pb.id
                WHERE c.producto_compuesto_id = %s
            ''', (combo[0]['id'],))
            for c in comps:
                skus_base.add(c['sku'].upper())
        else:
            skus_base.add(sku)
    return skus_base


# ── Tabla para tracking de auto-import ──────────────────────────────────────
def _init_auto_import_table():
    try:
        execute_db("""
            CREATE TABLE IF NOT EXISTS auto_import_log (
                id INT AUTO_INCREMENT PRIMARY KEY,
                ventas_nuevas INT DEFAULT 0,
                ultima_ejecucion TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                errores TEXT DEFAULT NULL
            )
        """)
        # Asegurar que solo existe 1 fila con id=1
        execute_db("DELETE FROM auto_import_log WHERE id > 1")
        existe = query_db("SELECT id FROM auto_import_log WHERE id = 1 LIMIT 1")
        if not existe:
            execute_db("INSERT INTO auto_import_log (id, ventas_nuevas) VALUES (1, 0)")
        # Agregar columna cancelada_en_ml si no existe
        try:
            execute_db("ALTER TABLE ventas ADD COLUMN cancelada_en_ml TINYINT DEFAULT 0")
            print("[AUTO-ML] Columna cancelada_en_ml agregada.")
        except Exception:
            pass  # Ya existe, ignorar
    except Exception as e:
        print(f"[AUTO-ML] Error init tabla: {e}")


def _importar_orden_automatica(orden, access_token):
    """
    Importa automáticamente una orden de ML sin intervención del usuario.
    Retorna True si se importó, False si falló o requiere mapeo manual.
    """
    import re as _re
    from datetime import timezone as _tz, timedelta as _td

    try:
        orden_id = str(orden['id'])
        orden_data = procesar_orden_ml(orden)

        # ML oculta first_name/last_name en búsqueda masiva — consultar orden individual
        try:
            headers_ml = {'Authorization': f'Bearer {access_token}'}
            r_individual = requests.get(
                f'https://api.mercadolibre.com/orders/{orden_id}',
                headers=headers_ml
            )
            if r_individual.status_code == 200:
                buyer_full = r_individual.json().get('buyer', {})
                fn = (buyer_full.get('first_name') or '').strip()
                ln = (buyer_full.get('last_name') or '').strip()
                nombre_completo = f"{fn} {ln}".strip()
                if nombre_completo:
                    orden_data['comprador_nombre'] = nombre_completo
        except Exception as e_buyer:
            print(f"[AUTO-ML] No se pudo obtener nombre completo: {e_buyer}")

        # Verificar que todos los items tienen SKU mapeado en BD
        items_bd = []
        for item in orden_data['items']:
            sku_ml = item.get('sku', '')
            sku_norm, cant_override = normalizar_sku_ml(sku_ml) if sku_ml else ('', 0)
            if not sku_norm:
                print(f"[AUTO-ML] Orden {orden_id}: item sin SKU, requiere mapeo manual")
                return False, []
            cantidad_final = cant_override if cant_override > 0 else item['cantidad']
            # Mapear compac según ubicacion_despacho (se determina después del shipping,
            # pero podemos pre-calcular basándonos en el logistic_type del shipment)
            # El mapeo final se hace más abajo cuando ya tenemos ubicacion_despacho
            existe, tipo, nombre = verificar_sku_en_bd(sku_norm)
            # Para compac, buscar con sufijo _DEP o _FULL si el SKU base no existe
            if not existe and sku_norm.upper().startswith('CCO'):
                existe_dep, _, _ = verificar_sku_en_bd(sku_norm + '_DEP')
                existe_full, _, _ = verificar_sku_en_bd(sku_norm + '_FULL')
                if existe_dep or existe_full:
                    existe = True  # se mapea más abajo
                else:
                    print(f"[AUTO-ML] Orden {orden_id}: SKU compac {sku_norm} sin _DEP/_FULL en BD")
                    return False, []
            elif not existe:
                print(f"[AUTO-ML] Orden {orden_id}: SKU {sku_norm} no existe en BD, requiere mapeo manual")
                return False, []
            items_bd.append({'sku': sku_norm, 'cantidad': cantidad_final, 'precio': item['precio']})

        # Auto-agregar PLATINO si el título contiene "almohada" y el SKU no es almohada
        SKUS_ALMOHADA_PURAS = {'CERVICAL','CLASICA','DORAL','DUAL','EXCLUSIVE','PLATINO','RENOVATION','SUBLIME'}
        platino_a_agregar = 0
        for item in orden_data['items']:
            sku_ml = item.get('sku', '').upper()
            titulo = item.get('titulo', '').lower()
            es_almohada_sku = any(a in sku_ml for a in SKUS_ALMOHADA_PURAS)
            if 'almohada' in titulo and not es_almohada_sku:
                import re as _re2
                ancho = 0
                m = _re2.search(r'(\d{2,3})\s*x\s*(\d{2,3})', titulo)
                if m:
                    ancho = min(int(m.group(1)), int(m.group(2)))
                if not ancho:
                    cm = _re2.search(r'(\d{2,3})\s*cm', titulo)
                    if cm: ancho = int(cm.group(1))
                if not ancho:
                    if '2 plaza' in titulo or 'doble' in titulo: ancho = 140
                    elif 'queen' in titulo: ancho = 160
                    elif 'king' in titulo: ancho = 180
                    elif 'plaza y media' in titulo or '1.5' in titulo: ancho = 100
                    elif '1 plaza' in titulo or 'individual' in titulo: ancho = 80
                cant = 0
                if ancho >= 140: cant = 2
                elif 0 < ancho <= 100: cant = 1
                platino_a_agregar = max(platino_a_agregar, cant)

        if platino_a_agregar > 0:
            existe_plat, _, _ = verificar_sku_en_bd('PLATINO')
            if existe_plat:
                items_bd.append({'sku': 'PLATINO', 'cantidad': platino_a_agregar, 'precio': 0})
                print(f"[AUTO-ML] ✅ Auto-agregado PLATINO x{platino_a_agregar}")
        shipping = {}
        if orden_data['shipping'].get('shipping_id'):
            shipping = obtener_shipping_completo(
                orden_data['shipping']['shipping_id'],
                access_token,
                orden_data.get('fecha_iso', '')
            )
        else:
            shipping = orden_data['shipping']

        # Billing info
        billing_info = {
            'business_name': None, 'doc_type': None, 'doc_number': None,
            'taxpayer_type': None, 'city': None, 'street': None,
            'state': None, 'zip_code': None
        }
        try:
            headers_ml = {'Authorization': f'Bearer {access_token}'}
            br = requests.get(
                f'https://api.mercadolibre.com/orders/{orden_id}/billing_info',
                headers=headers_ml
            )
            if br.status_code == 200:
                bd = br.json()
                iva_map = {
                    'IVA Exento': 'Exento',
                    'IVA Responsable Inscripto': 'Responsable Inscripto',
                    'Monotributo': 'Responsable Monotributo',
                }
                # Estructura nueva: billing_info.additional_info array [{type, value}]
                bi_root = bd.get('billing_info', {})
                add_info = bi_root.get('additional_info', [])
                if add_info:
                    ai = {item['type']: item['value'] for item in add_info}
                    taxpayer_raw = ai.get('TAXPAYER_TYPE_ID', '')
                    # Nombre de razón social (CUIT) o nombre físico
                    business = ai.get('BUSINESS_NAME') or ''
                    billing_info = {
                        'business_name': business or None,
                        'doc_type':      bi_root.get('doc_type') or ai.get('DOC_TYPE'),
                        'doc_number':    bi_root.get('doc_number') or ai.get('DOC_NUMBER'),
                        'taxpayer_type': iva_map.get(taxpayer_raw, taxpayer_raw) or None,
                        'city':          ai.get('CITY_NAME'),
                        'street':        ai.get('STREET_NAME'),
                        'state':         ai.get('STATE_NAME'),
                        'zip_code':      ai.get('ZIP_CODE'),
                    }
                else:
                    # Estructura vieja: buyer.billing_info
                    buyer_info = bd.get('buyer', {})
                    billing_doc = buyer_info.get('billing_info', {})
                    doc_type_raw = billing_doc.get('doc_type', '')
                    taxpayer_raw = billing_doc.get('taxpayer_type', {})
                    if isinstance(taxpayer_raw, dict):
                        taxpayer_raw = taxpayer_raw.get('description', '')
                    billing_info = {
                        'business_name': billing_doc.get('business_name'),
                        'doc_type':      doc_type_raw,
                        'doc_number':    billing_doc.get('doc_number'),
                        'taxpayer_type': iva_map.get(taxpayer_raw, taxpayer_raw),
                        'city':          billing_doc.get('city'),
                        'street':        billing_doc.get('street'),
                        'state':         billing_doc.get('state'),
                        'zip_code':      billing_doc.get('zip_code'),
                    }
        except Exception as e:
            print(f"[AUTO-ML] Error obteniendo billing de {orden_id}: {e}")

        # Datos de la venta
        fecha_venta = orden_data['fecha']
        canal = 'Mercado Libre'
        # mla_code guarda el APODO (nickname) — se muestra en negrita en la tabla
        # nombre_cliente guarda el nombre real — si ML no lo provee, queda vacío
        mla_code = orden_data.get('comprador_nickname', '') or f"ML-{orden_id}"
        nombre_real = orden_data.get('comprador_nombre', '').strip()
        # Si hay nombre real distinto al nickname, usarlo; sino usar el nickname (igual que flujo manual)
        nombre_cliente = nombre_real if nombre_real else mla_code
        numero_venta = f"ML-{orden_id}"
        telefono_cliente = ''
        tipo_entrega = 'envio' if shipping.get('tiene_envio') else 'retiro'
        metodo_envio = shipping.get('metodo_envio', '')
        ubicacion_despacho = 'FULL' if metodo_envio == 'Full' else 'DEP'
        # Flex → Delega si todos los productos son almohadas o compac (igual que flujo manual)
        if metodo_envio == 'Flex':
            SKUS_ALM = {'CERVICAL','CLASICA','DORAL','DUAL','EXCLUSIVE','PLATINO','RENOVATION','SUBLIME'}
            es_todo_alm_o_compac = all(
                item['sku'].upper().startswith('CCO') or
                '_DEP' in item['sku'].upper() or
                '_FULL' in item['sku'].upper() or
                any(a in item['sku'].upper() for a in SKUS_ALM)
                for item in items_bd
            )
            if es_todo_alm_o_compac:
                metodo_envio = 'Delega'
                zona_envio = ''
        # Zona solo para Flete Propio
        zona_envio = shipping.get('zona', '') if metodo_envio == 'Flete Propio' else zona_envio if metodo_envio == 'Delega' else ''
        direccion_entrega = shipping.get('direccion', '')

        # Mapear SKUs compac a _DEP o _FULL ahora que conocemos ubicacion_despacho
        sufijo = '_FULL' if ubicacion_despacho == 'FULL' else '_DEP'
        items_bd = [
            dict(item, sku=item['sku'] + sufijo) if item['sku'].upper().startswith('CCO') and '_DEP' not in item['sku'] and '_FULL' not in item['sku'] else item
            for item in items_bd
        ]
        costo_flete = float(shipping.get('costo_envio', 0) or 0)
        metodo_pago = 'Mercadopago'
        importe_total = float(orden_data['total'])          # solo productos
        importe_abonado = float(orden_data.get('paid_amount') or orden_data['total'])  # productos + flete
        pago_mp = importe_abonado
        pago_efectivo = 0.0
        estado_entrega = 'pendiente'
        estado_pago = 'pagado'

        # Notas: si hay fecha_entrega_ml solo esa fecha, sino el numero de orden
        fecha_entrega_ml = shipping.get('fecha_entrega_ml', '')
        if fecha_entrega_ml:
            notas = fecha_entrega_ml
        else:
            notas = f"Importado desde ML - Orden: {orden_id}"

        # Agregar demora ML a notas si aplica
        demora_ml_dias = shipping.get('demora_ml_dias', 0)
        if demora_ml_dias:
            notas += f"\nDEMORA_ML: {demora_ml_dias} días"

        # Fecha prometida al cliente (para fecha_entrega_estimada)
        fecha_entrega_prometida = shipping.get('fecha_entrega_prometida') or None

        # Insertar
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO ventas (
                    numero_venta, fecha_venta, canal, mla_code,
                    nombre_cliente, telefono_cliente,
                    tipo_entrega, metodo_envio, ubicacion_despacho,
                    zona_envio, direccion_entrega,
                    costo_flete, metodo_pago, importe_total, importe_abonado,
                    pago_mercadopago, pago_efectivo,
                    estado_entrega, estado_pago, notas,
                    fecha_entrega_estimada,
                    factura_business_name, factura_doc_type, factura_doc_number,
                    factura_taxpayer_type, factura_city, factura_street,
                    factura_state, factura_zip_code
                ) VALUES (
                    %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,%s,%s,%s
                )
            ''', (
                numero_venta, fecha_venta, canal, mla_code,
                nombre_cliente, telefono_cliente,
                tipo_entrega, metodo_envio, ubicacion_despacho,
                zona_envio, direccion_entrega,
                costo_flete, metodo_pago, importe_total, importe_abonado,
                pago_mp, pago_efectivo,
                estado_entrega, estado_pago, notas,
                fecha_entrega_prometida,
                billing_info['business_name'], billing_info['doc_type'],
                billing_info['doc_number'], billing_info['taxpayer_type'],
                billing_info['city'], billing_info['street'],
                billing_info['state'], billing_info['zip_code'],
            ))
            venta_id = cursor.lastrowid

            for item in items_bd:
                cursor.execute('''
                    INSERT INTO items_venta (venta_id, sku, cantidad, precio_unitario)
                    VALUES (%s, %s, %s, %s)
                ''', (venta_id, item['sku'], item['cantidad'], item['precio']))

            conn.commit()

            # Detectar alertas (por compatibilidad con flujo existente)
            try:
                detectar_alertas_stock_bajo(cursor, items_bd, venta_id)
                conn.commit()
            except Exception as e:
                print(f"[AUTO-ML] Error detectando alertas: {e}")

            print(f"[AUTO-ML] ✅ Venta importada: {mla_code}")
            return True, items_bd

        except Exception as e:
            conn.rollback()
            print(f"[AUTO-ML] Error insertando venta {orden_id}: {e}")
            return False, []
        finally:
            cursor.close()
            conn.close()

    except Exception as e:
        import traceback
        print(f"[AUTO-ML] Error procesando orden: {e}")
        traceback.print_exc()
        return False, []


def job_auto_importar_ml():
    """
    Job que corre cada 60 segundos.
    Importa órdenes nuevas de ML y actualiza publicaciones.
    """
    with app.app_context():
        try:
            # Chequear si el auto-import está activo
            try:
                row = query_db("SELECT valor FROM configuracion WHERE clave = 'auto_import_activo' LIMIT 1")
                if row and row[0]['valor'] == '0':
                    print("[AUTO-ML] Auto-import desactivado, saltando.")
                    return
            except Exception:
                pass  # Si falla la consulta, continuar igual

            print("[AUTO-ML] 🔄 Iniciando auto-import...")
            access_token = cargar_ml_token()
            if not access_token:
                print("[AUTO-ML] Sin token, saltando.")
                return

            success, result = obtener_ordenes_ml(access_token, limit=50)
            if not success:
                print(f"[AUTO-ML] Error obteniendo órdenes: {result}")
                return

            # Obtener IDs ya importados
            import re as _re
            ordenes_importadas = set()
            try:
                todas_ventas = query_db('''
                    SELECT numero_venta FROM ventas
                    ORDER BY fecha_venta DESC LIMIT 300
                ''')
                for v in todas_ventas:
                    nums = _re.findall(r'\d+', v['numero_venta'] or '')
                    for n in nums:
                        if len(n) == 16 and n.startswith('2000'):
                            ordenes_importadas.add(n)
            except Exception as e:
                print(f"[AUTO-ML] Error cargando importadas: {e}")

            ventas_nuevas = 0
            skus_base_afectados = set()

            for orden in result:
                orden_id = str(orden['id'])
                if orden_id in ordenes_importadas:
                    continue
                if orden.get('status') not in ['paid']:
                    continue

                ok, items_bd = _importar_orden_automatica(orden, access_token)
                if ok:
                    ventas_nuevas += 1
                    skus_afectados = _extraer_skus_base_de_items(items_bd)
                    skus_base_afectados.update(skus_afectados)
                    time.sleep(1)

            # Actualizar log
            try:
                execute_db(
                    "UPDATE auto_import_log SET ventas_nuevas = %s, ultima_ejecucion = NOW() WHERE id = 1",
                    (ventas_nuevas,)
                )
            except Exception as e:
                print(f"[AUTO-ML] Error actualizando log: {e}")

            print(f"[AUTO-ML] ✅ Import completo. Ventas nuevas: {ventas_nuevas}")

            # Actualizar publicaciones ML si hubo cambios
            if skus_base_afectados:
                print(f"[AUTO-ML] Actualizando ML para SKUs: {skus_base_afectados}")
                actualizar_publicaciones_ml_con_progreso(skus_base_afectados)

        except Exception as e:
            import traceback
            print(f"[AUTO-ML] Error en job: {e}")
            traceback.print_exc()


def job_verificar_cancelaciones_ml():
    """
    Job que corre cada 10 minutos.
    Verifica si ventas activas importadas desde ML fueron canceladas en ML.
    Si status == 'cancelled' → marca cancelada_en_ml = 1.
    """
    with app.app_context():
        try:
            access_token = cargar_ml_token()
            if not access_token:
                return

            import re as _re
            # Ventas activas de ML de los últimos 30 días
            ventas_ml = query_db("""
                SELECT id, numero_venta FROM ventas
                WHERE estado_entrega = 'pendiente'
                  AND canal = 'Mercado Libre'
                  AND fecha_venta >= DATE_SUB(NOW(), INTERVAL 30 DAY)
            """)
            if not ventas_ml:
                return

            marcadas = 0
            for v in ventas_ml:
                nums = _re.findall(r'\d{16}', v['numero_venta'] or '')
                if not nums:
                    continue
                orden_id = nums[0]
                try:
                    import requests as _req
                    r = _req.get(
                        f'https://api.mercadolibre.com/orders/{orden_id}',
                        headers={'Authorization': f'Bearer {access_token}'},
                        timeout=5
                    )
                    if r.status_code != 200:
                        continue
                    data = r.json()
                    if data.get('status') == 'cancelled':
                        execute_db(
                            "UPDATE ventas SET cancelada_en_ml = 1 WHERE id = %s AND cancelada_en_ml = 0",
                            (v['id'],)
                        )
                        marcadas += 1
                    time.sleep(0.3)
                except Exception:
                    continue

            if marcadas:
                print(f"[CANCEL-ML] ⚠️ {marcadas} venta(s) marcadas como canceladas en ML.")

        except Exception as e:
            print(f"[CANCEL-ML] Error: {e}")


@app.route('/ventas/auto-import-toggle', methods=['POST'])
@login_required
def auto_import_toggle():
    data = request.get_json()
    activo = bool(data.get('activo', True))
    try:
        existe = query_db("SELECT valor FROM configuracion WHERE clave = 'auto_import_activo' LIMIT 1")
        if existe:
            execute_db("UPDATE configuracion SET valor = %s WHERE clave = 'auto_import_activo'",
                       ('1' if activo else '0',))
        else:
            execute_db("INSERT INTO configuracion (clave, valor) VALUES ('auto_import_activo', %s)",
                       ('1' if activo else '0',))
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    return jsonify({'ok': True, 'activo': activo})


# ── Progreso de actualización ML ────────────────────────────────────────────
def actualizar_stock_compac_dep_ml(sku_dep, cantidad_disponible, access_token):
    """
    Actualiza el stock selling_address (DEP) en ML para todas las publicaciones
    del SKU compac base (ej: CCO140_DEP → busca CCO140 en ML).
    Usa el endpoint user-products/stock/type/selling_address.
    """
    import requests as _req
    sku_base = sku_dep.replace('_DEP', '').replace('_FULL', '')
    headers = {'Authorization': f'Bearer {access_token}'}

    try:
        me = _req.get('https://api.mercadolibre.com/users/me', headers=headers, timeout=10).json()
        user_id = me.get('id')
        if not user_id:
            print(f"[COMPAC-ML] No se pudo obtener user_id")
            return
    except Exception as e:
        print(f"[COMPAC-ML] Error obteniendo user_id: {e}")
        return

    # Buscar MLAs activas con ese seller_sku
    try:
        r = _req.get(
            f'https://api.mercadolibre.com/users/{user_id}/items/search?seller_sku={sku_base}&status=active',
            headers=headers, timeout=10
        ).json()
        mlas = r.get('results', [])
    except Exception as e:
        print(f"[COMPAC-ML] Error buscando MLAs de {sku_base}: {e}")
        return

    print(f"[COMPAC-ML] {sku_dep} → disponible={cantidad_disponible} | MLAs: {mlas}")

    for mla in mlas:
        try:
            # Obtener user_product_id y x-version
            item_r = _req.get(f'https://api.mercadolibre.com/items/{mla}', headers=headers, timeout=10).json()
            up_id = item_r.get('user_product_id')
            if not up_id:
                print(f"[COMPAC-ML] {mla}: sin user_product_id, saltando")
                continue

            stock_r = _req.get(f'https://api.mercadolibre.com/user-products/{up_id}/stock',
                               headers=headers, timeout=10)
            x_version = stock_r.headers.get('x-version', '1')
            stock_data = stock_r.json()

            # Verificar si tiene selling_address
            locations = stock_data.get('locations', [])
            tiene_dep = any(loc['type'] == 'selling_address' for loc in locations)
            solo_full = all(loc['type'] == 'meli_facility' for loc in locations) if locations else False
            if not tiene_dep or solo_full:
                print(f"[COMPAC-ML] {mla} ({up_id}): solo FULL, saltando")
                continue

            # Actualizar stock DEP
            put_r = _req.put(
                f'https://api.mercadolibre.com/user-products/{up_id}/stock/type/selling_address',
                headers={**headers, 'Content-Type': 'application/json', 'x-version': str(x_version)},
                json={'quantity': int(cantidad_disponible)},
                timeout=10
            )
            if put_r.status_code in (200, 204):
                print(f"[COMPAC-ML] {mla} ({up_id}): ✅ DEP stock={cantidad_disponible}")
            elif put_r.status_code == 409:
                # x-version desactualizada, reintentar una vez
                stock_r2 = _req.get(f'https://api.mercadolibre.com/user-products/{up_id}/stock',
                                    headers=headers, timeout=10)
                x_version2 = stock_r2.headers.get('x-version', '1')
                put_r2 = _req.put(
                    f'https://api.mercadolibre.com/user-products/{up_id}/stock/type/selling_address',
                    headers={**headers, 'Content-Type': 'application/json', 'x-version': str(x_version2)},
                    json={'quantity': int(cantidad_disponible)},
                    timeout=10
                )
                if put_r2.status_code in (200, 204):
                    print(f"[COMPAC-ML] {mla} ({up_id}): ✅ DEP stock={cantidad_disponible} (retry)")
                else:
                    print(f"[COMPAC-ML] {mla} ({up_id}): ❌ {put_r2.status_code} {put_r2.text[:100]}")
            else:
                print(f"[COMPAC-ML] {mla} ({up_id}): ❌ {put_r.status_code} {put_r.text[:100]}")
            time.sleep(0.3)
        except Exception as e:
            print(f"[COMPAC-ML] Error procesando {mla}: {e}")


def _ml_progress_save(data):
    """Guarda el progreso en la BD para que sea compartido entre workers."""
    try:
        import json as _json
        valor = _json.dumps(data)
        execute_db(
            "INSERT INTO configuracion (clave, valor) VALUES ('ml_progress', %s) "
            "ON DUPLICATE KEY UPDATE valor = %s",
            (valor, valor)
        )
    except Exception as e:
        print(f"[AUTO-ML] Error guardando progreso: {e}")

def _ml_progress_get():
    """Lee el progreso desde la BD."""
    try:
        import json as _json
        row = query_db("SELECT valor FROM configuracion WHERE clave = 'ml_progress' LIMIT 1")
        if row:
            return _json.loads(row[0]['valor'])
    except Exception:
        pass
    return {'running': False, 'total': 0, 'done': 0, 'ok': [], 'errors': [], 'skus': []}


def actualizar_publicaciones_ml_con_progreso(skus_base_afectados):
    """Corre la actualización de ML guardando progreso en BD (compartido entre workers)."""
    resultados_ok = []
    resultados_err = []

    # Calcular total de publicaciones
    pubs_count = 0
    for sku in skus_base_afectados:
        try:
            pubs_count += len(query_db("SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,)))
            pubs_count += len(query_db("SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku + 'Z',)))
            combos = query_db('''SELECT pc.sku FROM productos_compuestos pc
                JOIN componentes c ON c.producto_compuesto_id = pc.id
                JOIN productos_base pb ON c.producto_base_id = pb.id
                WHERE pb.sku = %s AND pc.activo = 1''', (sku,))
            for combo in combos:
                pubs_count += len(query_db("SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (combo['sku'],)))
                pubs_count += len(query_db("SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (combo['sku'] + 'Z',)))
        except Exception:
            pass

    _ml_progress_save({'running': True, 'total': pubs_count or 1, 'done': 0,
                       'ok': [], 'errors': [], 'skus': list(skus_base_afectados)})

    access_token = cargar_ml_token()
    if not access_token:
        _ml_progress_save({'running': False, 'total': 0, 'done': 0,
                           'ok': [], 'errors': ['Sin access token ML'], 'skus': []})
        return

    try:
        stock_todos = calcular_stock_por_sku()
    except Exception as e:
        _ml_progress_save({'running': False, 'total': 0, 'done': 0,
                           'ok': [], 'errors': [f'Error calculando stock: {e}'], 'skus': []})
        return

    skus_a_actualizar = set()
    for sku_base in skus_base_afectados:
        sku_base = sku_base.upper()
        skus_a_actualizar.add(sku_base)
        try:
            combos = query_db('''SELECT pc.sku FROM productos_compuestos pc
                JOIN componentes c ON c.producto_compuesto_id = pc.id
                JOIN productos_base pb ON c.producto_base_id = pb.id
                WHERE pb.sku = %s AND pc.activo = 1''', (sku_base,))
            for combo in combos:
                skus_a_actualizar.add(combo['sku'].upper())
        except Exception:
            pass

    done = 0
    for sku in skus_a_actualizar:
        if _es_almohada(sku):
            continue
        # Compac _DEP → actualizar selling_address en ML
        if '_DEP' in sku.upper() and sku.upper().startswith('CCO'):
            disponible = max(0, stock_todos.get(sku, {}).get('stock_disponible', 0))
            try:
                actualizar_stock_compac_dep_ml(sku, disponible, access_token)
                resultados_ok.append(f"{sku} DEP stock={disponible}")
            except Exception as e:
                resultados_err.append(f"{sku}: {e}")
            done += 1
            _ml_progress_save({'running': True, 'total': pubs_count or 1, 'done': done,
                               'ok': resultados_ok[:], 'errors': resultados_err[:], 'skus': list(skus_base_afectados)})
            continue
        # Compac _FULL → ML lo gestiona solo
        if '_FULL' in sku.upper() and sku.upper().startswith('CCO'):
            print(f"[AUTO-ML] {sku} es FULL, ML gestiona el stock solo")
            continue
        # Compac base sin sufijo → también saltar
        if _es_compac(sku):
            continue
        disponible = max(0, stock_todos.get(sku, {}).get('stock_disponible', 0))

        # Sin Z
        try:
            pubs_sin_z = query_db("SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,))
            for pub in pubs_sin_z:
                mla = pub['mla_id']
                ok, msg = actualizar_stock_ml(mla, disponible, access_token)
                if ok:
                    resultados_ok.append(f"{sku} \u2192 {mla} stock={disponible}")
                else:
                    resultados_err.append(f"{mla}: {msg}")
                done += 1
                _ml_progress_save({'running': True, 'total': pubs_count or 1, 'done': done,
                                   'ok': resultados_ok[:], 'errors': resultados_err[:], 'skus': list(skus_base_afectados)})
                time.sleep(0.1)
        except Exception as e:
            resultados_err.append(f"Error sin Z {sku}: {e}")

        # Con Z
        if not _aplica_logica_z(sku):
            continue
        sku_z = sku + 'Z'
        try:
            pubs_z = query_db("SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku_z,))
            for pub in pubs_z:
                mla = pub['mla_id']
                if disponible > 0:
                    ok, msg = actualizar_stock_ml(mla, disponible, access_token)
                    label = f"{sku_z} \u2192 {mla} stock={disponible}"
                    time.sleep(0.1)
                    _quitar_demora_ml(mla, access_token)
                    label += ' + quitar demora'
                else:
                    ok, msg = actualizar_stock_ml(mla, 1, access_token)
                    label = f"{sku_z} \u2192 {mla} stock=1+demora"
                    time.sleep(0.1)
                    _poner_demora_ml(mla, access_token, dias=7)
                if ok:
                    resultados_ok.append(label)
                else:
                    resultados_err.append(f"{mla}: {msg}")
                done += 1
                _ml_progress_save({'running': True, 'total': pubs_count or 1, 'done': done,
                                   'ok': resultados_ok[:], 'errors': resultados_err[:], 'skus': list(skus_base_afectados)})
                time.sleep(0.1)
        except Exception as e:
            resultados_err.append(f"Error con Z {sku_z}: {e}")

    _ml_progress_save({'running': False, 'total': pubs_count or done, 'done': done,
                       'ok': resultados_ok, 'errors': resultados_err, 'skus': list(skus_base_afectados)})


@app.route('/ventas/ml-progress')
@login_required
def ml_progress():
    return jsonify(_ml_progress_get())



# ── Endpoint: cuántas ventas nuevas hay (para popup) ────────────────────────
@app.route('/ventas/nuevas-count')
@login_required
def ventas_nuevas_count():
    try:
        row = query_db("SELECT ventas_nuevas FROM auto_import_log WHERE id = 1 LIMIT 1")
        count = row[0]['ventas_nuevas'] if row else 0
        return jsonify({'count': count})
    except Exception:
        return jsonify({'count': 0})


@app.route('/ventas/nuevas-reset', methods=['POST'])
@login_required
def ventas_nuevas_reset():
    try:
        execute_db("UPDATE auto_import_log SET ventas_nuevas = 0 WHERE id = 1")
    except Exception:
        pass
    return ('', 204)


# ── Iniciar APScheduler ──────────────────────────────────────────────────────
def job_completar_notas_mp():
    """
    Job que corre cada 10 minutos.
    Busca ventas de tienda web (canal='tienda_web') de las últimas 2 horas
    que no tengan MPID en notas, consulta la API de MP y completa MPID y VEID.
    """
    with app.app_context():
        try:
            import requests as _req
            mp_token = os.getenv('MP_ACCESS_TOKEN', '')
            if not mp_token:
                return

            ventas = query_db("""
                SELECT id, numero_venta, notas
                FROM ventas
                WHERE canal = 'tienda_web'
                  AND fecha_registro >= DATE_SUB(NOW(), INTERVAL 2 HOUR)
                  AND (notas IS NULL OR notas NOT LIKE '%MPID:%')
            """)
            if not ventas:
                return

            for v in ventas:
                # Extraer payment_id del numero_venta (MP-XXXXXXXXX)
                numero = v['numero_venta'] or ''
                if not numero.startswith('MP-'):
                    continue
                payment_id = numero[3:]

                try:
                    # Consultar payment
                    rp = _req.get(
                        f'https://api.mercadopago.com/v1/payments/{payment_id}',
                        headers={'Authorization': f'Bearer {mp_token}'},
                        timeout=5
                    )
                    if rp.status_code != 200:
                        continue
                    payment = rp.json()
                    if payment.get('status') != 'approved':
                        continue

                    order_id = (payment.get('order') or {}).get('id')

                    notas_parts = [f"MPID: {payment_id}"]
                    if order_id:
                        notas_parts.append(f"VEID: {order_id}")
                        # Intentar obtener SHID del merchant_order
                        try:
                            rm = _req.get(
                                f'https://api.mercadopago.com/merchant_orders/{order_id}',
                                headers={'Authorization': f'Bearer {mp_token}'},
                                timeout=5
                            )
                            if rm.status_code == 200:
                                mo = rm.json()
                                shipments = mo.get('shipments', [])
                                if shipments:
                                    shid = shipments[0].get('id')
                                    if shid:
                                        notas_parts.append(f"SHID: {shid}")
                        except Exception:
                            pass

                    notas_nuevas = "\n".join(notas_parts)
                    notas_actuales = v['notas'] or ''
                    # Agregar al final sin pisar notas manuales existentes
                    if notas_actuales.strip():
                        notas_finales = notas_actuales.rstrip('\n') + '\n' + notas_nuevas
                    else:
                        notas_finales = notas_nuevas

                    execute_db(
                        "UPDATE ventas SET notas = %s WHERE id = %s",
                        (notas_finales, v['id'])
                    )
                    print(f"[MP-NOTAS] ✅ Completadas notas para {numero}: {notas_nuevas.replace(chr(10), ' | ')}")

                except Exception as e:
                    print(f"[MP-NOTAS] Error procesando {numero}: {e}")

        except Exception as e:
            print(f"[MP-NOTAS] Error general: {e}")


# ============================================================================
# DASHBOARD AGENCIA
# ============================================================================

def _ga4_query(property_id, date_ranges, dimensions, metrics):
    """Consulta la GA4 Data API y retorna rows."""
    try:
        import json as _json
        from google.oauth2 import service_account
        from google.analytics.data_v1beta import BetaAnalyticsDataClient
        from google.analytics.data_v1beta.types import (
            RunReportRequest, DateRange, Dimension, Metric
        )
        sa_path = os.path.join(os.path.dirname(__file__), 'config', 'ga4_service_account.json')
        credentials = service_account.Credentials.from_service_account_file(
            sa_path,
            scopes=['https://www.googleapis.com/auth/analytics.readonly']
        )
        client = BetaAnalyticsDataClient(credentials=credentials)
        request = RunReportRequest(
            property=f'properties/{property_id}',
            date_ranges=[DateRange(**dr) for dr in date_ranges],
            dimensions=[Dimension(name=d) for d in dimensions],
            metrics=[Metric(name=m) for m in metrics],
        )
        response = client.run_report(request)
        return response
    except Exception as e:
        print(f'[GA4] Error: {e}')
        return None


@app.route('/agencia')
@login_required
def agencia_dashboard():
    if current_user.rol not in ('admin', 'vendedor', 'viewer', 'agencia'):
        return redirect(url_for('login'))

    GA4_PROPERTY_ID = '528968739'

    # ── Período ──────────────────────────────────────────────────────────────
    periodo = request.args.get('periodo', '30d')
    fecha_desde_custom = request.args.get('desde', '')
    fecha_hasta_custom = request.args.get('hasta', '')

    from datetime import date, timedelta
    hoy = date.today()
    if periodo == '7d':
        fecha_desde = (hoy - timedelta(days=6)).strftime('%Y-%m-%d')
        fecha_hasta = hoy.strftime('%Y-%m-%d')
        label_periodo = 'Últimos 7 días'
    elif periodo == '15d':
        fecha_desde = (hoy - timedelta(days=14)).strftime('%Y-%m-%d')
        fecha_hasta = hoy.strftime('%Y-%m-%d')
        label_periodo = 'Últimos 15 días'
    elif periodo == 'mes':
        fecha_desde = hoy.replace(day=1).strftime('%Y-%m-%d')
        fecha_hasta = hoy.strftime('%Y-%m-%d')
        label_periodo = 'Mes en curso'
    elif periodo == 'custom' and fecha_desde_custom and fecha_hasta_custom:
        fecha_desde = fecha_desde_custom
        fecha_hasta = fecha_hasta_custom
        label_periodo = f'{fecha_desde_custom} → {fecha_hasta_custom}'
    else:  # 30d default
        fecha_desde = (hoy - timedelta(days=29)).strftime('%Y-%m-%d')
        fecha_hasta = hoy.strftime('%Y-%m-%d')
        label_periodo = 'Últimos 30 días'

    # ── Datos de la DB (tienda web) ───────────────────────────────────────────
    ventas_db = query_db("""
        SELECT v.id, v.importe_total, v.importe_abonado, v.fecha_venta
        FROM ventas v
        WHERE v.canal = 'tienda_web'
          AND v.estado_entrega != 'cancelada'
          AND DATE(v.fecha_venta) BETWEEN %s AND %s
    """, (fecha_desde, fecha_hasta))

    total_ventas = len(ventas_db)
    total_facturado = sum(float(v['importe_total'] or 0) for v in ventas_db)
    ticket_promedio = total_facturado / total_ventas if total_ventas else 0

    # Productos más vendidos
    top_productos = query_db("""
        SELECT COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre, SUM(iv.cantidad) as total_unidades
        FROM items_venta iv
        JOIN ventas v ON v.id = iv.venta_id
        LEFT JOIN productos_base pb ON pb.sku = iv.sku
        LEFT JOIN productos_compuestos pc ON pc.sku = iv.sku
        WHERE v.canal = 'tienda_web'
          AND v.estado_entrega != 'cancelada'
          AND DATE(v.fecha_venta) BETWEEN %s AND %s
        GROUP BY iv.sku
        ORDER BY total_unidades DESC
        LIMIT 8
    """, (fecha_desde, fecha_hasta))

    # ── Datos de GA4 ──────────────────────────────────────────────────────────
    ga4_visitas = 0
    ga4_sesiones = 0
    tasa_conversion = 0.0
    fuentes = []

    try:
        dr = [{'start_date': fecha_desde, 'end_date': fecha_hasta}]

        # Visitas y sesiones
        r1 = _ga4_query(GA4_PROPERTY_ID, dr, [], ['activeUsers', 'sessions'])
        if r1 and r1.rows:
            row = r1.rows[0]
            ga4_visitas  = int(row.metric_values[0].value)
            ga4_sesiones = int(row.metric_values[1].value)

        # Tasa de conversión (purchase events / sesiones)
        tasa_conversion = round((total_ventas / ga4_sesiones * 100), 2) if ga4_sesiones else 0.0

        # Fuentes de tráfico
        r2 = _ga4_query(GA4_PROPERTY_ID, dr, ['sessionDefaultChannelGroup'], ['sessions'])
        if r2 and r2.rows:
            fuentes = sorted([
                {'fuente': row.dimension_values[0].value, 'sesiones': int(row.metric_values[0].value)}
                for row in r2.rows
            ], key=lambda x: x['sesiones'], reverse=True)

    except Exception as e:
        print(f'[AGENCIA] Error GA4: {e}')

    return render_template('agencia.html',
        periodo=periodo,
        label_periodo=label_periodo,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        fecha_desde_custom=fecha_desde_custom,
        fecha_hasta_custom=fecha_hasta_custom,
        total_ventas=total_ventas,
        total_facturado=total_facturado,
        ticket_promedio=ticket_promedio,
        top_productos=top_productos,
        ga4_visitas=ga4_visitas,
        ga4_sesiones=ga4_sesiones,
        tasa_conversion=tasa_conversion,
        fuentes=fuentes,
    )


def iniciar_scheduler():
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        _init_auto_import_table()
        scheduler = BackgroundScheduler(daemon=True)
        scheduler.add_job(
            job_auto_importar_ml,
            'interval',
            seconds=120,
            id='auto_import_ml',
            replace_existing=True,
            max_instances=1
        )
        scheduler.add_job(
            job_verificar_cancelaciones_ml,
            'interval',
            minutes=10,
            id='verificar_cancelaciones_ml',
            replace_existing=True,
            max_instances=1
        )
        scheduler.add_job(
            job_completar_notas_mp,
            'interval',
            minutes=10,
            id='completar_notas_mp',
            replace_existing=True,
            max_instances=1
        )
        scheduler.start()
        print("[AUTO-ML] ✅ Scheduler iniciado — auto-import cada 120s, cancelaciones cada 10min, notas MP cada 10min")
        return scheduler
    except Exception as e:
        print(f"[AUTO-ML] Error iniciando scheduler: {e}")
        return None


# Iniciar al cargar el módulo (funciona con gunicorn y Flask dev)
_scheduler = iniciar_scheduler()

# Crear tabla de logs si no existe
crear_tabla_sistema_logs()


# ============================================================================
# SISTEMA DE LOGS
# ============================================================================

@app.route('/admin/logs')
@login_required
@admin_required
def admin_logs():
    from datetime import timedelta
    from flask import make_response
    import csv, io

    nivel    = request.args.get('nivel', '')
    modulo   = request.args.get('modulo', '')
    sku      = request.args.get('sku', '').strip()
    desde    = request.args.get('desde', '')
    hasta    = request.args.get('hasta', '')
    exportar = request.args.get('exportar', '')

    conditions = []
    params = []
    if nivel:
        conditions.append('nivel = %s'); params.append(nivel)
    if modulo:
        conditions.append('modulo = %s'); params.append(modulo)
    if sku:
        conditions.append('sku LIKE %s'); params.append(f'%{sku}%')
    if desde:
        conditions.append('timestamp >= %s'); params.append(desde)
    if hasta:
        conditions.append('timestamp <= %s'); params.append(hasta + ' 23:59:59')

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

    if exportar == 'csv':
        logs = query_db(f"SELECT * FROM sistema_logs {where} ORDER BY timestamp DESC LIMIT 5000", params)
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(['id','timestamp','nivel','modulo','accion','detalle','sku','venta_id','usuario','ip'])
        for row in logs:
            writer.writerow([row.get(k,'') for k in ['id','timestamp','nivel','modulo','accion','detalle','sku','venta_id','usuario','ip']])
        output = make_response(si.getvalue())
        output.headers['Content-Disposition'] = 'attachment; filename=sistema_logs.csv'
        output.headers['Content-type'] = 'text/csv'
        return output

    logs   = query_db(f"SELECT * FROM sistema_logs {where} ORDER BY timestamp DESC LIMIT 500", params)
    modulos = query_db("SELECT DISTINCT modulo FROM sistema_logs WHERE modulo IS NOT NULL ORDER BY modulo")

    return render_template('admin_logs.html',
        logs=logs, modulos=modulos,
        filtros={'nivel': nivel, 'modulo': modulo, 'sku': sku, 'desde': desde, 'hasta': hasta})


# ============================================================================
# INICIAR APLICACIÓN
# ============================================================================

if __name__ == '__main__':
    # Configuración para red local
    app.run(
        host='0.0.0.0',      # Escuchar en todas las interfaces de red
        port=5000,            # Puerto del servidor
        debug=False,          # Desactivar debug en producción
        threaded=True         # Permitir múltiples usuarios simultáneos
    )

# ============================================================================
# MÓDULO COSTOS — CALCULADORA DE PRECIOS CANNON
# ============================================================================

def _get_precio_costos_sku(sku, porcentajes_ml=None):
    """
    Retorna el precio calculado por costos para un SKU dado.
    Útil para mostrar en cargar_stock_ml.
    Retorna dict con precio_sin_cuotas, 1c, 3c, 6c, 9c, 12c o None si no hay datos.
    """
    try:
        if porcentajes_ml is None:
            row = query_one("SELECT valor FROM configuracion WHERE clave = 'porcentajes_ml'")
            porcentajes_ml = json.loads(row['valor']) if row else PORCENTAJES_ML_DEFAULT

        cfg = _get_config_costos()
        multiplicador    = cfg.get('multiplicador', 1.85)
        prontopago_pct   = cfg.get('prontopago', 5.0)
        desc_cliente_pct = cfg.get('cliente', 0.0)

        # Buscar SKU en cannon_productos
        sku_buscar = sku[:-1] if sku.endswith('Z') else sku
        # Para sommiers, buscar el colchon base
        sku_col = ('C' + sku_buscar[1:]) if sku_buscar.startswith('S') and len(sku_buscar) > 1 and sku_buscar[1].isalpha() else sku_buscar

        cp = query_one("""
            SELECT cp.sku, cp.descripcion, clp.precio_lista,
                   cd_adi.valor as desc_adicional,
                   ce_col.costo as costo_colecta,
                   ce_flex.costo as costo_flex
            FROM cannon_productos cp
            JOIN cannon_lista_precios clp ON clp.codigo_material = cp.codigo_material
            LEFT JOIN cannon_descuentos cd_adi ON cd_adi.clave = CONCAT('adicional_', cp.sku)
            LEFT JOIN cannon_costos_envio ce_col ON ce_col.sku = %s AND ce_col.tipo = 'colecta'
            LEFT JOIN cannon_costos_envio ce_flex ON ce_flex.sku = %s AND ce_flex.tipo = 'flex'
            WHERE cp.sku = %s
        """, (sku_buscar, sku_buscar, sku_col))

        if not cp or not cp['precio_lista']:
            return None

        descuentos = {r['clave']: {'valor': float(r['valor']), 'desc_adicional': float(r['desc_adicional'] or 0)}
                      for r in query_db("SELECT clave, valor, desc_adicional FROM cannon_descuentos WHERE tipo = 'descuento_linea'")}

        def _detectar_clave_simple(desc, sku_up):
            desc = (desc or '').upper()
            if sku_up in ('CLASICA','SUBLIME','CERVICAL','RENOVATION','PLATINO','DORAL','DUAL','EXCLUSIVE'): return 'almohadas'
            if desc.startswith('ALM'): return 'almohadas'
            if 'EUROPILLOW' in desc:
                if 'SUBLIME' in desc: return 'sublime_europillow'
                if 'RENOVATION' in desc: return 'renovation_europillow'
            if 'PILLOW' in desc or 'PIL' in desc:
                if 'EXCLUSIVE' in desc: return 'exclusive_pillow'
                if 'DORAL' in desc: return 'doral_pillow'
            if 'PRINCESS' in desc: return 'princess_23' if '23' in desc else 'princess_20'
            if 'ESPECIAL DE LUJO' in desc: return 'especial_de_lujo'
            if 'EXCLUSIVE' in desc: return 'exclusive'
            if 'RENOVATION' in desc: return 'renovation'
            if 'TROPICAL' in desc: return 'tropical'
            if 'SONAR' in desc or 'SOÑAR' in desc: return 'sonar'
            if 'PLATINO' in desc: return 'platino'
            if 'DORAL' in desc: return 'doral'
            if 'SUBLIME' in desc: return 'sublime'
            if 'BASE' in desc or sku_up.startswith('BASE_') or desc.startswith('SOM '): return 'bases'
            return None

        clave = _detectar_clave_simple(cp['descripcion'], sku_col.upper())
        desc_entry = descuentos.get(clave, {'valor': 0, 'desc_adicional': 0}) if clave else {'valor': 0, 'desc_adicional': 0}
        desc_linea = desc_entry['valor']
        desc_adi = desc_entry['desc_adicional'] + float(cp['desc_adicional'] or 0)

        precio_lista = round(_calcular_precio_lista(
            float(cp['precio_lista']), desc_linea, desc_cliente_pct, desc_adi, prontopago_pct, multiplicador
        ) / 1000) * 1000

        # Si es sommier, sumar base
        es_sommier = sku_buscar.startswith('S') and len(sku_buscar) > 1 and sku_buscar[1].isalpha()
        if es_sommier:
            cfg_conj = query_one("SELECT base_sku_default, cantidad_bases FROM conjunto_configuracion WHERE colchon_sku = %s AND activo=1", (sku_col,))
            if cfg_conj:
                base_sku = cfg_conj['base_sku_default']
                cant = int(cfg_conj['cantidad_bases'] or 1)
                cp_base = query_one("""
                    SELECT clp.precio_lista FROM cannon_productos cp
                    JOIN cannon_lista_precios clp ON clp.codigo_material = cp.codigo_material
                    WHERE cp.sku = %s
                """, (base_sku,))
                if cp_base:
                    desc_base = descuentos.get('bases', {'valor': 40})['valor']
                    precio_base_calc = round(_calcular_precio_lista(
                        float(cp_base['precio_lista']), desc_base, desc_cliente_pct, 0, prontopago_pct, multiplicador
                    ) / 1000) * 1000
                    precio_lista = round((precio_lista + precio_base_calc * cant) / 1000) * 1000

        # Costo envío
        es_z = sku.endswith('Z')
        sku_base_num = sku_buscar[:-1] if sku_buscar.endswith('Z') else sku_buscar
        ancho = int(sku_base_num[-3:]) if sku_base_num[-3:].isdigit() else 0
        if not es_z and clave not in ('bases', 'almohadas'):
            if ancho <= 100:
                costo_envio = float(cp['costo_colecta'] or 0)
            else:
                costo_envio = float(cp['costo_flex'] or 0)
        else:
            costo_envio = 0

        precio_sc = round((precio_lista + costo_envio) / 1000) * 1000

        def _pc(base, pct):
            return round(base * 0.76 / (0.76 - pct / 100) / 1000) * 1000

        return {
            'precio_lista':     precio_lista,
            'precio_sin_cuotas': precio_sc,
            'precio_1c':  _pc(precio_sc, porcentajes_ml.get('cuota_simple', 5.0)),
            'precio_3c':  _pc(precio_sc, porcentajes_ml.get('cuotas_3', 9.4)),
            'precio_6c':  _pc(precio_sc, porcentajes_ml.get('cuotas_6', 15.1)),
            'precio_9c':  _pc(precio_sc, porcentajes_ml.get('cuotas_9', 20.7)),
            'precio_12c': _pc(precio_sc, porcentajes_ml.get('cuotas_12', 25.9)),
        }
    except Exception as e:
        print(f"[_get_precio_costos_sku] Error: {e}")
        return None

COSTOS_MODELOS_CLAVES = [
    'tropical','princess_20','princess_23','especial_de_lujo',
    'exclusive','exclusive_pillow','renovation','renovation_europillow',
    'sonar','platino','doral','doral_pillow','sublime','sublime_europillow',
    'bases','almohadas',
]

def _get_config_costos():
    """Retorna dict con todos los descuentos y multiplicador desde DB."""
    rows = query_db("SELECT clave, valor, tipo FROM cannon_descuentos")
    return {r['clave']: float(r['valor']) for r in rows}

def _calcular_precio_lista(precio_cannon, desc_linea_pct, desc_cliente_pct, desc_adicional_pct, prontopago_pct, multiplicador):
    """
    precio_lista = precio_cannon
                   × (1 - desc_linea/100)
                   × (1 - desc_cliente/100)
                   × (1 - desc_adicional/100)
                   × 1/(1 + prontopago/100)   ← prontopago correcto
                   × multiplicador
    """
    costo = precio_cannon
    costo *= (1 - desc_linea_pct / 100)
    if desc_cliente_pct:
        costo *= (1 - desc_cliente_pct / 100)
    if desc_adicional_pct:
        costo *= (1 - desc_adicional_pct / 100)
    costo *= 1 / (1 + prontopago_pct / 100)
    return round(costo * multiplicador)


def _build_precio_costos_map():
    """Construye un mapa sku → precio_lista_costos para mostrar en tienda_precios."""
    try:
        cfg = _get_config_costos()
        multiplicador    = cfg.get('multiplicador', 1.85)
        prontopago_pct   = cfg.get('prontopago', 5.0)
        desc_cliente_pct = cfg.get('cliente', 0.0)
        descuentos = {r['clave']: {'valor': float(r['valor']), 'desc_adicional': float(r['desc_adicional'] or 0)}
                      for r in query_db("SELECT clave, valor, desc_adicional FROM cannon_descuentos WHERE tipo = 'descuento_linea'")}
        rows = query_db("""
            SELECT cp.sku, clp.precio_lista, cp.descripcion,
                   cd_adi.valor as desc_adicional
            FROM cannon_productos cp
            JOIN cannon_lista_precios clp ON clp.codigo_material = cp.codigo_material
            LEFT JOIN cannon_descuentos cd_adi ON cd_adi.clave = CONCAT('adicional_', cp.sku)
            WHERE cp.sku IS NOT NULL
        """)
        mapa = {}
        for r in rows:
            sku = r['sku']
            desc = (r['descripcion'] or '').upper()
            clave = None
            if sku in ('CLASICA','SUBLIME','CERVICAL','RENOVATION','PLATINO','DORAL','DUAL','EXCLUSIVE') or desc.startswith('ALM'):
                clave = 'almohadas'
            elif 'EUROPILLOW' in desc:
                clave = 'sublime_europillow' if 'SUBLIME' in desc else 'renovation_europillow'
            elif 'PILLOW' in desc or 'PIL' in desc:
                clave = 'exclusive_pillow' if 'EXCLUSIVE' in desc else 'doral_pillow'
            elif 'PRINCESS' in desc: clave = 'princess_23' if '23' in desc else 'princess_20'
            elif 'ESPECIAL DE LUJO' in desc: clave = 'especial_de_lujo'
            elif 'EXCLUSIVE' in desc: clave = 'exclusive'
            elif 'RENOVATION' in desc: clave = 'renovation'
            elif 'TROPICAL' in desc: clave = 'tropical'
            elif 'SONAR' in desc or 'SOÑAR' in desc: clave = 'sonar'
            elif 'PLATINO' in desc: clave = 'platino'
            elif 'DORAL' in desc: clave = 'doral'
            elif 'BASE' in desc or sku.startswith('BASE_') or desc.startswith('SOM '): clave = 'bases'
            elif 'SUBLIME' in desc: clave = 'sublime'
            desc_entry = descuentos.get(clave, {'valor': 0, 'desc_adicional': 0}) if clave else {'valor': 0, 'desc_adicional': 0}
            desc_adi = desc_entry['desc_adicional'] + float(r['desc_adicional'] or 0)
            precio = round(_calcular_precio_lista(
                float(r['precio_lista']), desc_entry['valor'], desc_cliente_pct, desc_adi, prontopago_pct, multiplicador
            ) / 1000) * 1000
            mapa[sku] = precio
        # Agregar sommiers
        conjuntos = query_db("SELECT colchon_sku, base_sku_default, cantidad_bases FROM conjunto_configuracion WHERE activo=1")
        for c in conjuntos:
            sku_col = c['colchon_sku']
            base_sku = c['base_sku_default']
            cant = int(c['cantidad_bases'] or 1)
            precio_col = mapa.get(sku_col, 0)
            precio_base = mapa.get(base_sku, 0)
            if precio_col and precio_base:
                sku_conj = 'S' + sku_col[1:] if sku_col.startswith('C') else 'S' + sku_col
                mapa[sku_conj] = round((precio_col + precio_base * cant) / 1000) * 1000
        return mapa
    except Exception as e:
        print(f"[_build_precio_costos_map] Error: {e}")
        return {}


@app.route('/costos')
@admin_required
def costos_index():
    """Calculadora de precios — vista principal."""
    cfg = _get_config_costos()
    multiplicador  = cfg.get('multiplicador', 1.85)
    prontopago_pct = cfg.get('prontopago', 5.0)

    # Porcentajes de cuotas ML desde configuracion
    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'porcentajes_ml'")
        porcentajes_ml = json.loads(row['valor']) if row else PORCENTAJES_ML_DEFAULT
    except Exception:
        porcentajes_ml = PORCENTAJES_ML_DEFAULT

    # JOIN cannon_productos + cannon_lista_precios
    productos = query_db("""
        SELECT cp.id, cp.codigo_material, cp.descripcion, cp.sku, cp.ean,
               clp.precio_lista as precio_cannon, clp.vigencia,
               ce.tipo as envio_tipo, ce.costo as envio_costo,
               cd_adi.valor as desc_adicional
        FROM cannon_productos cp
        LEFT JOIN cannon_lista_precios clp ON clp.codigo_material = cp.codigo_material
        LEFT JOIN cannon_costos_envio ce ON ce.sku = cp.sku
        LEFT JOIN cannon_descuentos cd_adi ON cd_adi.clave = CONCAT('adicional_', cp.sku)
        ORDER BY cp.descripcion
    """)

    # Descuentos por modelo
    descuentos = {r['clave']: float(r['valor']) for r in
                  query_db("SELECT clave, valor FROM cannon_descuentos WHERE tipo = 'descuento_linea'")}

    return render_template('costos.html',
        productos=productos,
        cfg=cfg,
        descuentos=descuentos,
        multiplicador=multiplicador,
        prontopago_pct=prontopago_pct,
        porcentajes_ml=porcentajes_ml,
        COSTOS_MODELOS_CLAVES=COSTOS_MODELOS_CLAVES,
    )


@app.route('/costos/descuentos', methods=['GET', 'POST'])
@admin_required
def costos_descuentos():
    """Configurar descuentos por modelo, prontopago y multiplicador."""
    if request.method == 'POST':
        data = request.get_json() or {}
        cambios = data.get('cambios', [])
        try:
            for c in cambios:
                campo = c.get('campo', 'valor')
                if campo == 'desc_adicional':
                    execute_db(
                        "UPDATE cannon_descuentos SET desc_adicional = %s WHERE clave = %s",
                        (float(c['valor']), c['clave'])
                    )
                else:
                    execute_db(
                        "UPDATE cannon_descuentos SET valor = %s WHERE clave = %s",
                        (float(c['valor']), c['clave'])
                    )
            return jsonify(ok=True, actualizados=len(cambios))
        except Exception as e:
            return jsonify(ok=False, error=str(e))
    rows = query_db("SELECT clave, descripcion, valor, desc_adicional, tipo FROM cannon_descuentos ORDER BY tipo, descripcion")
    return render_template('costos_descuentos.html', descuentos=rows)


@app.route('/costos/importar', methods=['GET', 'POST'])
@admin_required
def costos_importar():
    """Subir Excel de lista de precios Cannon."""
    if request.method == 'POST':
        tipo = request.form.get('tipo', 'lista')  # 'lista' o 'almohadas'
        archivo = request.files.get('archivo')
        if not archivo or not archivo.filename.endswith('.xlsx'):
            flash('❌ Debe ser un archivo .xlsx', 'danger')
            return redirect(url_for('costos_importar'))
        try:
            import io
            from openpyxl import load_workbook
            from datetime import date as _date
            wb = load_workbook(io.BytesIO(archivo.read()), read_only=True)
            ws = wb.active
            vigencia_str = request.form.get('vigencia', '')
            try:
                vigencia = _date.fromisoformat(vigencia_str)
            except Exception:
                vigencia = _date.today()

            insertados = 0
            for row in ws.iter_rows(values_only=True):
                codigo = row[0]
                precio = row[2]  # Columna C = Importe
                if not codigo or not isinstance(codigo, (int, float)):
                    continue
                if not precio or not isinstance(precio, (int, float)):
                    continue
                execute_db("""
                    INSERT INTO cannon_lista_precios (codigo_material, precio_lista, vigencia)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE precio_lista = VALUES(precio_lista), vigencia = VALUES(vigencia)
                """, (int(codigo), float(precio), vigencia))
                insertados += 1
            wb.close()
            flash(f'✅ {insertados} precios importados correctamente', 'success')
        except Exception as e:
            flash(f'❌ Error: {e}', 'danger')
        return redirect(url_for('costos_importar'))

    # Última vigencia cargada
    ultima = query_one("SELECT MAX(vigencia) as v FROM cannon_lista_precios")
    return render_template('costos_importar.html', ultima_vigencia=ultima['v'] if ultima else None)


@app.route('/costos/productos', methods=['GET', 'POST'])
@admin_required
def costos_productos():
    """Ver/editar SKU de productos Cannon y descuentos adicionales."""
    if request.method == 'POST':
        data = request.get_json() or {}
        pid  = data.get('id')
        sku  = (data.get('sku') or '').strip().upper() or None
        desc_adi = data.get('desc_adicional')
        try:
            execute_db("UPDATE cannon_productos SET sku = %s WHERE id = %s", (sku, pid))
            if desc_adi is not None:
                clave = f"adicional_{sku}" if sku else None
                if clave:
                    execute_db("""
                        INSERT INTO cannon_descuentos (clave, descripcion, valor, tipo)
                        VALUES (%s, %s, %s, 'descuento_adicional')
                        ON DUPLICATE KEY UPDATE valor = VALUES(valor)
                    """, (clave, f'Desc. adicional {sku}', float(desc_adi)))
            return jsonify(ok=True)
        except Exception as e:
            return jsonify(ok=False, error=str(e))

    q = request.args.get('q', '').strip()
    where = "WHERE (cp.descripcion LIKE %s OR cp.sku LIKE %s)" if q else ""
    params = (f'%{q}%', f'%{q}%') if q else ()
    productos = query_db(f"""
        SELECT cp.*, clp.precio_lista,
               cd.valor as desc_adicional
        FROM cannon_productos cp
        LEFT JOIN cannon_lista_precios clp ON clp.codigo_material = cp.codigo_material
        LEFT JOIN cannon_descuentos cd ON cd.clave = CONCAT('adicional_', cp.sku)
        {where}
        ORDER BY cp.descripcion
    """, params)
    return render_template('costos_productos.html', productos=productos, q=q)


@app.route('/costos/envio', methods=['GET', 'POST'])
@admin_required
def costos_envio():
    """Gestionar costos de envío colecta/flex por SKU."""
    if request.method == 'POST':
        data = request.get_json() or {}
        cambios = data.get('cambios', [])
        try:
            for c in cambios:
                execute_db("""
                    INSERT INTO cannon_costos_envio (sku, tipo, costo)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE tipo = VALUES(tipo), costo = VALUES(costo)
                """, (c['sku'].strip().upper(), c['tipo'], float(c['costo'])))
            return jsonify(ok=True, actualizados=len(cambios))
        except Exception as e:
            return jsonify(ok=False, error=str(e))

    skus = query_db("""
        SELECT pb.sku, pb.nombre, pb.tipo, pb.medida,
               ce.tipo as envio_tipo, ce.costo as envio_costo
        FROM productos_base pb
        LEFT JOIN cannon_costos_envio ce ON ce.sku = pb.sku
        WHERE pb.tipo IN ('colchon','almohada')
        ORDER BY pb.tipo, pb.nombre
    """)
    return render_template('costos_envio.html', skus=skus)


# SKUs para barrido de costos colecta ML
SKUS_COLECTA_BARRIDO = [
    'CTR80','CPR8020','CPR9020','CPR10020','CPR8023','CPR9023','CPR10023',
    'CEX80','CEX100','CEXP80','CEXP90','CEXP100',
    'CRE80','CRE100','CREP80','CREP90','CREP100',
    'CSO80','CSO100','CDO80','CDO90','CDO100',
]

@app.route('/costos/envio/barrido-ml', methods=['GET'])
@admin_required
def costos_envio_barrido_ml():
    """Consulta costos de colecta ML para los SKUs de barrido y compara con lo guardado."""
    import requests as _requests

    access_token = cargar_ml_token()
    if not access_token:
        return jsonify(ok=False, error='Sin token ML'), 400

    resultados = []
    for sku in SKUS_COLECTA_BARRIDO:
        # Buscar MLA gold_special para este SKU
        r = ml_request('get',
            f'https://api.mercadolibre.com/users/{ML_SELLER_ID}/items/search',
            access_token, params={'seller_sku': sku})
        if r.status_code != 200:
            resultados.append({'sku': sku, 'error': f'ML error {r.status_code}'})
            continue

        mla_ids = r.json().get('results', [])
        if not mla_ids:
            resultados.append({'sku': sku, 'error': 'Sin publicaciones'})
            continue

        # Tomar el primer MLA disponible
        mla = mla_ids[0]

        # Consultar costo de colecta
        r2 = ml_request('get',
            f'https://api.mercadolibre.com/users/{ML_SELLER_ID}/shipping_options/free',
            access_token,
            params={'item_id': mla, 'verbose': 'true', 'free_shipping': 'False',
                    'mode': 'me2', 'logistic_type': 'cross_docking'})

        if r2.status_code != 200:
            resultados.append({'sku': sku, 'mla': mla, 'error': f'Shipping error {r2.status_code}'})
            continue

        costo_ml = r2.json().get('coverage', {}).get('all_country', {}).get('list_cost', 0)
        costo_ml = round(float(costo_ml) / 0.86)

        # Costo actual guardado
        ce = query_one("SELECT costo FROM cannon_costos_envio WHERE sku = %s AND tipo = 'colecta'", (sku,))
        costo_actual = float(ce['costo']) if ce else None
        diferencia = round(costo_ml - costo_actual) if costo_actual is not None else None

        resultados.append({
            'sku':          sku,
            'mla':          mla,
            'costo_ml':     costo_ml,
            'costo_actual': costo_actual,
            'diferencia':   diferencia,
            'cambio':       diferencia != 0 if diferencia is not None else True,
        })

    return jsonify(ok=True, resultados=resultados)


@app.route('/costos/calcular')
@admin_required
def costos_calcular():
    """Tabla de precios calculados con opción de aplicar a productos_base."""
    cfg = _get_config_costos()
    multiplicador   = cfg.get('multiplicador', 1.85)
    prontopago_pct  = cfg.get('prontopago', 5.0)
    desc_cliente_pct = cfg.get('cliente', 0.0)

    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'porcentajes_ml'")
        porcentajes_ml = json.loads(row['valor']) if row else PORCENTAJES_ML_DEFAULT
    except Exception:
        porcentajes_ml = PORCENTAJES_ML_DEFAULT

    # Productos base (colchones y almohadas) con SKU y precio cargado
    productos_raw = query_db("""
        SELECT cp.id, cp.codigo_material, cp.descripcion, cp.sku,
               clp.precio_lista as precio_cannon, clp.vigencia,
               ce_col.costo as costo_colecta,
               ce_flex.costo as costo_flex,
               cd_adi.valor as desc_adicional,
               pb.precio_base as precio_actual
        FROM cannon_productos cp
        JOIN cannon_lista_precios clp ON clp.codigo_material = cp.codigo_material
        LEFT JOIN cannon_costos_envio ce_col ON ce_col.sku = cp.sku AND ce_col.tipo = 'colecta'
        LEFT JOIN cannon_costos_envio ce_flex ON ce_flex.sku = cp.sku AND ce_flex.tipo = 'flex'
        LEFT JOIN cannon_descuentos cd_adi ON cd_adi.clave = CONCAT('adicional_', cp.sku)
        LEFT JOIN productos_base pb ON pb.sku = cp.sku
        WHERE cp.sku IS NOT NULL
        ORDER BY cp.descripcion
    """)

    descuentos = {r['clave']: {'valor': float(r['valor']), 'desc_adicional': float(r['desc_adicional'] or 0)} for r in
                  query_db("SELECT clave, valor, desc_adicional FROM cannon_descuentos WHERE tipo = 'descuento_linea'")}

    # Configuración de conjuntos para calcular precio sommier
    conjuntos_cfg = {r['colchon_sku']: r for r in
                     query_db("SELECT colchon_sku, base_sku_default, cantidad_bases FROM conjunto_configuracion WHERE activo=1")}
    # Precios de bases desde cannon_lista_precios + cannon_productos
    bases_precio = {r['sku']: float(r['precio_lista'] or 0) for r in query_db("""
        SELECT cp.sku, clp.precio_lista
        FROM cannon_productos cp
        JOIN cannon_lista_precios clp ON clp.codigo_material = cp.codigo_material
        WHERE cp.sku IS NOT NULL AND cp.sku LIKE %s
    """, ('BASE%',))}

    def _detectar_clave(descripcion, sku=''):
        desc = (descripcion or '').upper()
        sku_up = (sku or '').upper()
        # Almohadas primero — SKU o descripción
        if sku_up in ('CLASICA','SUBLIME','CERVICAL','RENOVATION','PLATINO','DORAL','DUAL','EXCLUSIVE'):
            return 'almohadas'
        if desc.startswith('ALM'):
            return 'almohadas'
        # Europillow antes que pillow
        if 'EUROPILLOW' in desc or 'EURO PILLOW' in desc:
            if 'SUBLIME' in desc: return 'sublime_europillow'
            if 'RENOVATION' in desc: return 'renovation_europillow'
        if 'PILLOW' in desc or 'PIL' in desc:
            if 'EXCLUSIVE' in desc: return 'exclusive_pillow'
            if 'DORAL' in desc: return 'doral_pillow'
        if 'PRINCESS' in desc:
            return 'princess_23' if '23' in desc else 'princess_20'
        if 'ESPECIAL DE LUJO' in desc or 'ESP DE LUJO' in desc or 'ESPECIAL LUJ' in desc: return 'especial_de_lujo'
        if 'EXCLUSIVE' in desc: return 'exclusive'
        if 'RENOVATION' in desc: return 'renovation'
        if 'TROPICAL' in desc: return 'tropical'
        if 'SONAR' in desc or 'SOÑAR' in desc: return 'sonar'
        if 'PLATINO' in desc: return 'platino'
        if 'DORAL' in desc: return 'doral'
        if 'SUBLIME' in desc: return 'sublime'
        if 'BASE' in desc or desc.startswith('SOM'): return 'bases'
        return None

    def _precio_cuotas(base, pct):
        """precio_Xc = base × 0.76 / (0.76 - coef/100)"""
        return round(base * 0.76 / (0.76 - pct / 100) / 1000) * 1000

    def _mil(n):
        """Redondea al millar más cercano."""
        return round(n / 1000) * 1000

    # Mapa colchon_sku → precio_lista calculado (para sommiers)
    colchones_precio_lista = {}

    productos = []
    for p in productos_raw:
        precio_cannon = float(p['precio_cannon'] or 0)
        if not precio_cannon:
            continue
        sku = p['sku'] or ''
        clave = _detectar_clave(p['descripcion'], sku)
        desc_entry = descuentos.get(clave, {'valor': 0, 'desc_adicional': 0}) if clave else {'valor': 0, 'desc_adicional': 0}
        desc_linea = desc_entry['valor']
        desc_adi_linea = desc_entry['desc_adicional']
        desc_adi_sku = float(p['desc_adicional'] or 0)
        desc_adi   = desc_adi_linea + desc_adi_sku
        precio_lista = _mil(_calcular_precio_lista(
            precio_cannon, desc_linea, desc_cliente_pct, desc_adi, prontopago_pct, multiplicador
        ))

        # Guardar precio colchon para calcular sommiers después
        if clave != 'bases' and clave != 'almohadas':
            colchones_precio_lista[sku] = precio_lista

        es_z = sku.endswith('Z')
        sku_base = sku[:-1] if es_z else sku
        ancho = int(sku_base[-3:]) if sku_base[-3:].isdigit() else 0
        es_colecta = (not es_z) and (ancho <= 100) and (clave not in ('bases','almohadas'))
        es_almohada = clave == 'almohadas'

        costo_envio_ml = 0
        tipo_pub = None
        if not es_z and not clave == 'bases':
            if es_colecta or es_almohada:
                costo_envio_ml = float(p['costo_colecta'] or 0)
                tipo_pub = 'colecta'
            else:
                costo_envio_ml = float(p['costo_flex'] or 0)
                tipo_pub = 'flex'

        precio_ml_sin_cuotas = _mil(precio_lista + costo_envio_ml) if not es_z else precio_lista

        productos.append({
            'id':              p['id'],
            'sku':             sku,
            'descripcion':     p['descripcion'],
            'precio_cannon':   precio_cannon,
            'clave_descuento': clave,
            'desc_linea':      desc_linea,
            'desc_cliente':    desc_cliente_pct,
            'desc_adi':        desc_adi,
            'precio_lista':    precio_lista,
            'precio_actual':   float(p['precio_actual'] or 0),
            'costo_envio_ml':  costo_envio_ml,
            'tipo_pub':        tipo_pub,
            'es_z':            es_z,
            'es_conjunto':     False,
            'precio_ml_sin_cuotas': precio_ml_sin_cuotas,
            'precio_ml_1c':    _precio_cuotas(precio_ml_sin_cuotas, porcentajes_ml.get('cuota_simple', 5.0)),
            'precio_ml_3c':    _precio_cuotas(precio_ml_sin_cuotas, porcentajes_ml.get('cuotas_3', 9.4)),
            'precio_ml_6c':    _precio_cuotas(precio_ml_sin_cuotas, porcentajes_ml.get('cuotas_6', 15.1)),
            'precio_ml_9c':    _precio_cuotas(precio_ml_sin_cuotas, porcentajes_ml.get('cuotas_9', 20.7)),
            'precio_ml_12c':   _precio_cuotas(precio_ml_sin_cuotas, porcentajes_ml.get('cuotas_12', 25.9)),
        })

        # Colchones >100cm sin Z: agregar entrada Z (ME1/Flex Propio = mismo precio web)
        if not es_z and ancho > 100 and clave not in ('bases', 'almohadas'):
            sku_z = sku + 'Z'
            pb_z = query_one("SELECT precio_base FROM productos_base WHERE sku = %s", (sku_z,))
            precio_actual_z = float(pb_z['precio_base'] or 0) if pb_z else 0
            productos.append({
                'id':              None,
                'sku':             sku_z,
                'descripcion':     p['descripcion'] + ' (ME1/Flex propio)',
                'precio_cannon':   precio_cannon,
                'clave_descuento': clave,
                'desc_linea':      desc_linea,
                'desc_cliente':    desc_cliente_pct,
                'desc_adi':        desc_adi,
                'precio_lista':    precio_lista,
                'precio_actual':   precio_actual_z,
                'costo_envio_ml':  0,
                'tipo_pub':        'me1',
                'es_z':            True,
                'es_conjunto':     False,
                'precio_ml_sin_cuotas': precio_lista,
                'precio_ml_1c':    _precio_cuotas(precio_lista, porcentajes_ml.get('cuota_simple', 5.0)),
                'precio_ml_3c':    _precio_cuotas(precio_lista, porcentajes_ml.get('cuotas_3', 9.4)),
                'precio_ml_6c':    _precio_cuotas(precio_lista, porcentajes_ml.get('cuotas_6', 15.1)),
                'precio_ml_9c':    _precio_cuotas(precio_lista, porcentajes_ml.get('cuotas_9', 20.7)),
                'precio_ml_12c':   _precio_cuotas(precio_lista, porcentajes_ml.get('cuotas_12', 25.9)),
            })

    # ── Agregar sommiers ──────────────────────────────────────────────────────
    for sku_col, cfg_conj in conjuntos_cfg.items():
        precio_col = colchones_precio_lista.get(sku_col)
        if not precio_col:
            continue
        base_sku  = cfg_conj['base_sku_default']
        cant      = int(cfg_conj['cantidad_bases'] or 1)

        precio_cannon_base = bases_precio.get(base_sku, 0)
        if precio_cannon_base:
            precio_base_calc = _mil(_calcular_precio_lista(
                precio_cannon_base,
                descuentos.get('bases', {'valor': 40})['valor'],
                desc_cliente_pct, 0, prontopago_pct, multiplicador
            ))
        else:
            precio_base_calc = 0

        precio_conjunto = _mil(precio_col + precio_base_calc * cant)

        sku_conj   = 'S' + sku_col[1:] if sku_col.startswith('C') else 'S' + sku_col
        sku_conj_z = sku_conj + 'Z'

        # Precio actual del conjunto = colchon + base en productos_base
        pb_col  = query_one("SELECT precio_base FROM productos_base WHERE sku = %s", (sku_col,))
        pb_base = query_one("SELECT precio_base FROM productos_base WHERE sku = %s", (base_sku,))
        precio_col_actual  = float(pb_col['precio_base'] or 0) if pb_col else 0
        precio_base_actual = float(pb_base['precio_base'] or 0) if pb_base else 0
        precio_actual_conj = precio_col_actual + precio_base_actual * cant

        sku_base_col = sku_col
        ancho = int(sku_base_col[-3:]) if sku_base_col[-3:].isdigit() else 999

        ce_flex = query_one("SELECT costo FROM cannon_costos_envio WHERE sku = %s AND tipo = 'flex'", (sku_conj,))
        costo_flex = float(ce_flex['costo'] or 0) if ce_flex else 0

        precio_ml_flex = _mil(precio_conjunto + costo_flex)

        productos.append({
            'id':              None,
            'sku':             sku_conj,
            'descripcion':     f'SOMMIER + {sku_col} ({base_sku} ×{cant})',
            'precio_cannon':   None,
            'clave_descuento': 'conjunto',
            'desc_linea':      None,
            'desc_cliente':    None,
            'desc_adi':        None,
            'precio_lista':    precio_conjunto,
            'precio_actual':   precio_actual_conj,
            'costo_envio_ml':  costo_flex,
            'tipo_pub':        'flex',
            'es_z':            False,
            'es_conjunto':     True,
            'precio_ml_sin_cuotas': precio_ml_flex,
            'precio_ml_1c':    _precio_cuotas(precio_ml_flex, porcentajes_ml.get('cuota_simple', 5.0)),
            'precio_ml_3c':    _precio_cuotas(precio_ml_flex, porcentajes_ml.get('cuotas_3', 9.4)),
            'precio_ml_6c':    _precio_cuotas(precio_ml_flex, porcentajes_ml.get('cuotas_6', 15.1)),
            'precio_ml_9c':    _precio_cuotas(precio_ml_flex, porcentajes_ml.get('cuotas_9', 20.7)),
            'precio_ml_12c':   _precio_cuotas(precio_ml_flex, porcentajes_ml.get('cuotas_12', 25.9)),
        })
        # SKU con Z — mismo precio web
        pb_conj_z = query_one("SELECT precio_base FROM productos_base WHERE sku = %s", (sku_conj_z,))
        precio_actual_z = float(pb_conj_z['precio_base'] or 0) if pb_conj_z else precio_actual_conj
        productos.append({
            'id':              None,
            'sku':             sku_conj_z,
            'descripcion':     f'SOMMIER Z + {sku_col} ({base_sku} ×{cant})',
            'precio_cannon':   None,
            'clave_descuento': 'conjunto_z',
            'desc_linea':      None,
            'desc_cliente':    None,
            'desc_adi':        None,
            'precio_lista':    precio_conjunto,
            'precio_actual':   precio_actual_z,
            'costo_envio_ml':  0,
            'tipo_pub':        'me1',
            'es_z':            True,
            'es_conjunto':     True,
            'precio_ml_sin_cuotas': precio_conjunto,
            'precio_ml_1c':    _precio_cuotas(precio_conjunto, porcentajes_ml.get('cuota_simple', 5.0)),
            'precio_ml_3c':    _precio_cuotas(precio_conjunto, porcentajes_ml.get('cuotas_3', 9.4)),
            'precio_ml_6c':    _precio_cuotas(precio_conjunto, porcentajes_ml.get('cuotas_6', 15.1)),
            'precio_ml_9c':    _precio_cuotas(precio_conjunto, porcentajes_ml.get('cuotas_9', 20.7)),
            'precio_ml_12c':   _precio_cuotas(precio_conjunto, porcentajes_ml.get('cuotas_12', 25.9)),
        })

    productos.sort(key=lambda x: x['sku'])

    return render_template('costos_calcular.html',
        productos=productos,
        cfg=cfg,
        porcentajes_ml=porcentajes_ml,
        multiplicador=multiplicador,
        prontopago_pct=prontopago_pct,
        desc_cliente_pct=desc_cliente_pct,
    )


@app.route('/costos/aplicar', methods=['POST'])
@admin_required
def costos_aplicar():
    """Aplica precios calculados a productos_base."""
    data = request.get_json() or {}
    cambios = data.get('cambios', [])
    if not cambios:
        return jsonify(ok=False, error='Sin cambios')
    try:
        aplicados = 0
        for c in cambios:
            sku    = c.get('sku', '').strip()
            precio = c.get('precio')
            if not sku or precio is None:
                continue
            execute_db(
                "UPDATE productos_base SET precio_base = %s WHERE sku = %s",
                (float(precio), sku)
            )
            aplicados += 1
        return jsonify(ok=True, aplicados=aplicados)
    except Exception as e:
        return jsonify(ok=False, error=str(e))



# ============================================================================
# CATÁLOGO DE PRODUCTOS
# ============================================================================

_tablas_productos_ok = False

def _crear_tablas_productos():
    """Tabla de fotos y columnas opcionales en productos_base. Solo corre una vez por proceso."""
    global _tablas_productos_ok
    if _tablas_productos_ok:
        return
    execute_db("""
        CREATE TABLE IF NOT EXISTS productos_fotos (
            id         INT AUTO_INCREMENT PRIMARY KEY,
            sku        VARCHAR(50) NOT NULL,
            filename   VARCHAR(255) NOT NULL,
            orden      INT DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_sku_fotos (sku),
            UNIQUE KEY uq_sku_file (sku, filename)
        )
    """)
    # Columnas opcionales — ignorar si ya existen
    _cols = [
        ('activo',                 'TINYINT DEFAULT 1'),
        ('tipo_base',              'VARCHAR(50) DEFAULT NULL'),
        ('modelo_almohada',        'VARCHAR(100) DEFAULT NULL'),
        ('peso_gramos',            'INT DEFAULT NULL'),
        ('alto_cm',                'DECIMAL(6,1) DEFAULT NULL'),
        ('ancho_cm',               'DECIMAL(6,1) DEFAULT NULL'),
        ('largo_cm',               'DECIMAL(6,1) DEFAULT NULL'),
        ('stock_minimo_pausar',    'INT DEFAULT 0'),
        ('stock_minimo_reactivar', 'INT DEFAULT 1'),
    ]
    for col, defn in _cols:
        try:
            execute_db(f"ALTER TABLE productos_base ADD COLUMN {col} {defn}")
        except Exception:
            pass  # columna ya existe
    _tablas_productos_ok = True


def _get_fotos(sku):
    """Devuelve lista de filenames ordenados para un SKU."""
    rows = query_db(
        "SELECT filename FROM productos_fotos WHERE sku=%s ORDER BY orden, id",
        (sku,)
    )
    return [r['filename'] for r in rows]


@app.route('/productos')
@admin_required
def productos_lista():
    _crear_tablas_productos()

    busq      = request.args.get('q', '').strip()
    linea_sel = request.args.get('linea', '').strip()
    filtro    = request.args.get('filtro', 'todos')

    # ── Productos base ────────────────────────────────────────────────────────
    where_b  = "WHERE 1=1"
    params_b = []
    if busq:
        where_b += " AND (pb.nombre LIKE %s OR pb.sku LIKE %s OR pb.modelo LIKE %s)"
        params_b += [f'%{busq}%', f'%{busq}%', f'%{busq}%']
    if linea_sel:
        where_b += " AND pb.linea = %s"
        params_b.append(linea_sel)
    if filtro == 'activos':
        where_b += " AND COALESCE(pb.activo,1) = 1"
    elif filtro == 'inactivos':
        where_b += " AND COALESCE(pb.activo,1) = 0"

    productos_b = query_db(f"""
        SELECT pb.id, pb.sku, pb.nombre, pb.tipo,
               COALESCE(pb.linea,'')              AS linea,
               COALESCE(pb.modelo,'')             AS modelo,
               COALESCE(pb.medida,'')             AS medida,
               COALESCE(pb.precio_base,0)         AS precio_base,
               COALESCE(pb.descuento_catalogo,0)  AS descuento_catalogo,
               COALESCE(pb.stock_actual,0)        AS stock_actual,
               COALESCE(pb.activo,1)              AS activo,
               COUNT(pf.id)                       AS cant_fotos
        FROM productos_base pb
        LEFT JOIN productos_fotos pf ON pf.sku = pb.sku
        {where_b}
        GROUP BY pb.id
        ORDER BY pb.tipo, pb.linea, pb.nombre
    """, params_b)

    # ── Sommiers ──────────────────────────────────────────────────────────────
    where_c  = "WHERE 1=1"
    params_c = []
    if busq:
        where_c += " AND (pc.nombre LIKE %s OR pc.sku LIKE %s)"
        params_c += [f'%{busq}%', f'%{busq}%']
    if filtro == 'activos':
        where_c += " AND pc.activo = 1"
    elif filtro == 'inactivos':
        where_c += " AND pc.activo = 0"

    productos_c = [] if linea_sel else query_db(f"""
        SELECT pc.id, pc.sku, pc.nombre, 'sommier' AS tipo,
               'sommier' AS linea, '' AS modelo, '' AS medida,
               0 AS precio_base, 0 AS descuento_catalogo,
               0 AS stock_actual, pc.activo,
               COUNT(pf.id) AS cant_fotos
        FROM productos_compuestos pc
        LEFT JOIN productos_fotos pf ON pf.sku = pc.sku
        {where_c}
        GROUP BY pc.id
        ORDER BY pc.nombre
    """, params_c)

    # ── Precio y stock real de sommiers ──────────────────────────────────────
    if productos_c:
        conjuntos_cfg_cat = {r['colchon_sku']: r for r in query_db(
            "SELECT colchon_sku, base_sku_default, cantidad_bases FROM conjunto_configuracion WHERE activo=1"
        )}
        pb_map_cat = {r['sku']: r for r in query_db(
            "SELECT sku, precio_base, stock_actual FROM productos_base"
        )}
        for p in productos_c:
            sku_conj = p['sku']
            sku_col = ('C' + sku_conj[1:]) if (sku_conj.startswith('S') and len(sku_conj) > 1 and sku_conj[1].isalpha()) else None
            if sku_col and sku_col in conjuntos_cfg_cat:
                cfg_c = conjuntos_cfg_cat[sku_col]
                base_sku = cfg_c['base_sku_default']
                cant = int(cfg_c['cantidad_bases'] or 1)
                pb_col  = pb_map_cat.get(sku_col, {})
                pb_base = pb_map_cat.get(base_sku, {})
                precio_col  = float(pb_col.get('precio_base') or 0)
                precio_b_u  = float(pb_base.get('precio_base') or 0)
                p['precio_base']  = precio_col + precio_b_u * cant
                p['stock_actual'] = min(
                    int(pb_col.get('stock_actual') or 0),
                    int(pb_base.get('stock_actual') or 0)
                )

    productos = list(productos_b) + list(productos_c)

    # Sin miniatura — se ve desde la sección de fotos de cada SKU
    for p in productos:
        p['foto_thumb'] = None

    # Líneas disponibles para el filtro
    lineas_rows = query_db("""
        SELECT DISTINCT linea FROM productos_base
        WHERE linea IS NOT NULL AND linea != ''
        ORDER BY linea
    """)
    lineas = [r['linea'] for r in lineas_rows]

    # Modo JSON para búsqueda en tiempo real
    if request.args.get('json'):
        def p_json(p):
            return {
                'id':               p['id'],
                'sku':              p['sku'],
                'nombre':           p['nombre'],
                'tipo':             p['tipo'],
                'linea':            p['linea'],
                'modelo':           p['modelo'],
                'medida':           p['medida'],
                'precio_base':      float(p['precio_base'] or 0),
                'descuento_catalogo': float(p['descuento_catalogo'] or 0),
                'stock_actual':     int(p['stock_actual'] or 0),
                'activo':           bool(p['activo']),
                'cant_fotos':       int(p['cant_fotos'] or 0),
            }
        return jsonify(productos=[p_json(p) for p in productos], total=len(productos))

    demora_row = query_one("SELECT valor FROM configuracion WHERE clave='demora_sin_stock'")
    demora_dias = int(demora_row['valor']) if demora_row and demora_row['valor'] else 0

    nl_monto_row  = query_one("SELECT valor FROM configuracion WHERE clave='nl_monto'")
    nl_minimo_row = query_one("SELECT valor FROM configuracion WHERE clave='nl_minimo'")
    nl_monto  = int(nl_monto_row['valor'])  if nl_monto_row  and nl_monto_row['valor']  else 0
    nl_minimo = int(nl_minimo_row['valor']) if nl_minimo_row and nl_minimo_row['valor'] else 0

    coef_3_row = query_one("SELECT valor FROM configuracion WHERE clave='cuotas_3_coef'")
    coef_6_row = query_one("SELECT valor FROM configuracion WHERE clave='cuotas_6_coef'")
    cuotas_3_coef = float(coef_3_row['valor']) if coef_3_row and coef_3_row['valor'] else 1.11
    cuotas_6_coef = float(coef_6_row['valor']) if coef_6_row and coef_6_row['valor'] else 1.22

    return render_template('productos_lista.html',
        productos     = productos,
        busq          = busq,
        linea_sel     = linea_sel,
        lineas        = lineas,
        filtro        = filtro,
        demora_dias   = demora_dias,
        nl_monto      = nl_monto,
        nl_minimo     = nl_minimo,
        cuotas_3_coef = cuotas_3_coef,
        cuotas_6_coef = cuotas_6_coef,
    )


@app.route('/productos/demora', methods=['POST'])
@admin_required
def productos_demora_guardar():
    dias = request.form.get('dias', '').strip()
    try:
        if dias == '' or int(dias) == 0:
            execute_db("INSERT INTO configuracion (clave, valor) VALUES ('demora_sin_stock', '0') ON DUPLICATE KEY UPDATE valor='0'")
            return jsonify(ok=True, dias=0, msg='Demora desactivada')
        dias_int = max(1, int(dias))
        execute_db(
            "INSERT INTO configuracion (clave, valor) VALUES ('demora_sin_stock', %s) ON DUPLICATE KEY UPDATE valor=%s",
            (str(dias_int), str(dias_int))
        )
        return jsonify(ok=True, dias=dias_int, msg=f'Demora guardada: {dias_int} días')
    except Exception as e:
        return jsonify(ok=False, msg=str(e))


@app.route('/productos/newsletter-cupon', methods=['POST'])
@admin_required
def productos_newsletter_cupon():
    monto  = request.form.get('monto', '').strip()
    minimo = request.form.get('minimo', '').strip()
    try:
        monto_int  = max(0, int(monto))  if monto  else 0
        minimo_int = max(0, int(minimo)) if minimo else 0
        execute_db(
            "INSERT INTO configuracion (clave, valor) VALUES ('nl_monto', %s) ON DUPLICATE KEY UPDATE valor=%s",
            (str(monto_int), str(monto_int))
        )
        execute_db(
            "INSERT INTO configuracion (clave, valor) VALUES ('nl_minimo', %s) ON DUPLICATE KEY UPDATE valor=%s",
            (str(minimo_int), str(minimo_int))
        )
        fmt_monto  = f'${monto_int:,.0f}'.replace(',', '.')
        fmt_minimo = f'${minimo_int:,.0f}'.replace(',', '.')
        return jsonify(ok=True, msg=f'✅ {fmt_monto} OFF / mín. {fmt_minimo}')
    except Exception as e:
        return jsonify(ok=False, msg=str(e))


@app.route('/productos/cuotas-coeficientes', methods=['POST'])
@admin_required
def productos_cuotas_coeficientes():
    coef_3 = request.form.get('coef_3', '').strip()
    coef_6 = request.form.get('coef_6', '').strip()
    try:
        coef_3_f = round(max(1.0, float(coef_3)), 4) if coef_3 else 1.11
        coef_6_f = round(max(1.0, float(coef_6)), 4) if coef_6 else 1.22
        execute_db(
            "INSERT INTO configuracion (clave, valor) VALUES ('cuotas_3_coef', %s) ON DUPLICATE KEY UPDATE valor=%s",
            (str(coef_3_f), str(coef_3_f))
        )
        execute_db(
            "INSERT INTO configuracion (clave, valor) VALUES ('cuotas_6_coef', %s) ON DUPLICATE KEY UPDATE valor=%s",
            (str(coef_6_f), str(coef_6_f))
        )
        pct_3 = round((coef_3_f - 1) * 100, 1)
        pct_6 = round((coef_6_f - 1) * 100, 1)
        return jsonify(ok=True, msg=f'✅ 3 cuotas: +{pct_3}% / 6 cuotas: +{pct_6}%',
                       coef_3=coef_3_f, coef_6=coef_6_f)
    except Exception as e:
        return jsonify(ok=False, msg=str(e))


@app.route('/productos/toggle/<sku>', methods=['POST'])
@admin_required
def productos_toggle(sku):
    _crear_tablas_productos()
    try:
        # Buscar por SKU — evita colisión de IDs entre tablas
        pb = query_one("SELECT sku, COALESCE(activo,1) AS activo FROM productos_base WHERE sku=%s", (sku,))
        if pb:
            nuevo = 0 if pb['activo'] else 1
            execute_db("UPDATE productos_base SET activo=%s WHERE sku=%s", (nuevo, sku))
            verb = 'Activado' if nuevo else 'Desactivado'
            return jsonify(ok=True, activo=bool(nuevo), msg=f"{verb}: {sku}")

        pc = query_one("SELECT sku, activo FROM productos_compuestos WHERE sku=%s", (sku,))
        if pc:
            nuevo = 0 if pc['activo'] else 1
            execute_db("UPDATE productos_compuestos SET activo=%s WHERE sku=%s", (nuevo, sku))
            verb = 'Activado' if nuevo else 'Desactivado'
            return jsonify(ok=True, activo=bool(nuevo), msg=f"{verb}: {sku}")

        return jsonify(ok=False, msg='Producto no encontrado')
    except Exception as e:
        return jsonify(ok=False, msg=str(e))


@app.route('/productos/<sku>/fotos')
@admin_required
def productos_fotos(sku):
    _crear_tablas_productos()
    prod = query_one("SELECT id, sku, nombre, tipo FROM productos_base WHERE sku=%s", (sku,))
    if not prod:
        prod = query_one(
            "SELECT id, sku, nombre, 'sommier' AS tipo FROM productos_compuestos WHERE sku=%s", (sku,)
        )
    if not prod:
        flash('Producto no encontrado', 'error')
        return redirect(url_for('productos_lista'))

    fotos = _get_fotos(sku)
    return render_template('productos_fotos.html', producto=prod, fotos=fotos)


@app.route('/productos/<sku>/fotos/subir', methods=['POST'])
@admin_required
def productos_fotos_subir(sku):
    from werkzeug.utils import secure_filename
    import uuid
    _crear_tablas_productos()

    archivos = request.files.getlist('fotos')
    if not archivos or all(a.filename == '' for a in archivos):
        return jsonify(ok=False, msg='Sin archivos')

    carpeta = os.path.join(app.root_path, 'static', 'img', 'productos', sku)
    os.makedirs(carpeta, exist_ok=True)

    subidos = 0
    errores = []
    for archivo in archivos:
        if not archivo or archivo.filename == '':
            continue
        ext = os.path.splitext(secure_filename(archivo.filename))[1].lower()
        if ext not in ('.jpg', '.jpeg', '.png', '.webp'):
            errores.append(f"{archivo.filename}: formato no válido")
            continue
        if archivo.content_length and archivo.content_length > 5 * 1024 * 1024:
            errores.append(f"{archivo.filename}: supera 5 MB")
            continue

        nombre = f"{uuid.uuid4().hex}{ext}"
        archivo.save(os.path.join(carpeta, nombre))

        max_o = query_one(
            "SELECT COALESCE(MAX(orden),0)+1 AS o FROM productos_fotos WHERE sku=%s", (sku,)
        )
        execute_db(
            "INSERT IGNORE INTO productos_fotos (sku, filename, orden) VALUES (%s,%s,%s)",
            (sku, nombre, max_o['o'] if max_o else 1)
        )
        subidos += 1

    fotos = _get_fotos(sku)
    if subidos:
        return jsonify(ok=True, fotos=fotos, msg=f'{subidos} foto(s) subida(s)')
    else:
        return jsonify(ok=False, fotos=fotos, msg='; '.join(errores) or 'No se pudo subir ninguna foto')


@app.route('/productos/<sku>/fotos/eliminar', methods=['POST'])
@admin_required
def productos_fotos_eliminar(sku):
    _crear_tablas_productos()
    data     = request.get_json() or {}
    filename = (data.get('filename') or '').strip()
    if not filename:
        return jsonify(ok=False, msg='Sin filename')

    ruta = os.path.join(app.root_path, 'static', 'img', 'productos', sku, filename)
    try:
        os.remove(ruta)
    except Exception:
        pass

    execute_db(
        "DELETE FROM productos_fotos WHERE sku=%s AND filename=%s", (sku, filename)
    )
    fotos = _get_fotos(sku)
    return jsonify(ok=True, fotos=fotos, msg='Foto eliminada')


@app.route('/productos/<sku>/fotos/reordenar', methods=['POST'])
@admin_required
def productos_fotos_reordenar(sku):
    _crear_tablas_productos()
    data  = request.get_json() or {}
    orden = data.get('orden', [])  # lista de filenames en nuevo orden
    for i, filename in enumerate(orden):
        execute_db(
            "UPDATE productos_fotos SET orden=%s WHERE sku=%s AND filename=%s",
            (i + 1, sku, filename)
        )
    return jsonify(ok=True, fotos=_get_fotos(sku))


@app.route('/productos/nuevo', methods=['GET', 'POST'])
@admin_required
def productos_nuevo():
    _crear_tablas_productos()
    if request.method == 'POST':
        sku    = request.form.get('sku', '').strip().upper()
        nombre = request.form.get('nombre', '').strip()
        tipo   = request.form.get('tipo', 'colchon')
        linea  = request.form.get('linea', '').strip()
        modelo = request.form.get('modelo', '').strip()
        medida = request.form.get('medida', '').strip()
        tipo_base         = request.form.get('tipo_base', '').strip() or None
        modelo_almohada   = request.form.get('modelo_almohada', '').strip() or None
        precio_base       = float(request.form.get('precio_base', 0) or 0)
        descuento_catalogo = float(request.form.get('descuento_catalogo', 0) or 0) or None
        stock_inicial     = int(request.form.get('stock_actual', 0) or 0)
        stock_min_pausar  = int(request.form.get('stock_minimo_pausar', 0) or 0)
        stock_min_reactiv = int(request.form.get('stock_minimo_reactivar', 1) or 1)
        peso_gramos       = int(request.form.get('peso_gramos', 0) or 0) or None
        alto_cm           = float(request.form.get('alto_cm', 0) or 0) or None
        ancho_cm          = float(request.form.get('ancho_cm', 0) or 0) or None
        largo_cm          = float(request.form.get('largo_cm', 0) or 0) or None

        if not sku or not nombre:
            flash('SKU y nombre son obligatorios', 'error')
            return render_template('productos_form.html', producto=None, modo='nuevo')

        if query_one("SELECT sku FROM productos_base WHERE sku=%s", (sku,)):
            flash(f'Ya existe un producto con SKU {sku}', 'error')
            return render_template('productos_form.html', producto=None, modo='nuevo')

        execute_db("""
            INSERT INTO productos_base
                (sku, nombre, tipo, linea, modelo, medida,
                 tipo_base, modelo_almohada,
                 precio_base, descuento_catalogo,
                 stock_actual, activo,
                 stock_minimo_pausar, stock_minimo_reactivar,
                 peso_gramos, alto_cm, ancho_cm, largo_cm)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,1,%s,%s,%s,%s,%s,%s)
        """, (sku, nombre, tipo, linea, modelo, medida,
              tipo_base, modelo_almohada,
              precio_base, descuento_catalogo,
              stock_inicial,
              stock_min_pausar, stock_min_reactiv,
              peso_gramos, alto_cm, ancho_cm, largo_cm))
        flash(f'Producto {sku} creado correctamente', 'success')
        return redirect(url_for('productos_lista'))

    return render_template('productos_form.html', producto=None, modo='nuevo')


@app.route('/productos/editar/<sku>', methods=['GET', 'POST'])
@admin_required
def productos_editar(sku):
    _crear_tablas_productos()
    # Buscar en productos_base primero, luego en productos_compuestos
    producto = query_one("SELECT * FROM productos_base WHERE sku=%s", (sku,))
    es_sommier = False
    if not producto:
        pc = query_one("SELECT id, sku, nombre, activo FROM productos_compuestos WHERE sku=%s", (sku,))
        if not pc:
            flash('Producto no encontrado', 'error')
            return redirect(url_for('productos_lista'))
        # Normalizar como dict compatible con el form
        producto = {
            'id': pc['id'], 'sku': pc['sku'], 'nombre': pc['nombre'],
            'tipo': 'sommier', 'linea': None, 'modelo': None, 'medida': None,
            'tipo_base': None, 'modelo_almohada': None,
            'precio_base': 0, 'descuento_catalogo': None,
            'stock_actual': 0, 'activo': pc['activo'],
            'stock_minimo_pausar': 0, 'stock_minimo_reactivar': 1,
            'peso_gramos': None, 'alto_cm': None, 'ancho_cm': None, 'largo_cm': None,
        }
        es_sommier = True

    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        if es_sommier:
            execute_db("UPDATE productos_compuestos SET nombre=%s WHERE sku=%s", (nombre, sku))
        else:
            linea             = request.form.get('linea', '').strip()
            modelo            = request.form.get('modelo', '').strip()
            medida            = request.form.get('medida', '').strip()
            tipo_base         = request.form.get('tipo_base', '').strip() or None
            modelo_almohada   = request.form.get('modelo_almohada', '').strip() or None
            precio_base       = float(request.form.get('precio_base', 0) or 0)
            descuento_catalogo = float(request.form.get('descuento_catalogo', 0) or 0) or None
            stock_min_pausar  = int(request.form.get('stock_minimo_pausar', 0) or 0)
            stock_min_reactiv = int(request.form.get('stock_minimo_reactivar', 1) or 1)
            peso_gramos       = int(request.form.get('peso_gramos', 0) or 0) or None
            alto_cm           = float(request.form.get('alto_cm', 0) or 0) or None
            ancho_cm          = float(request.form.get('ancho_cm', 0) or 0) or None
            largo_cm          = float(request.form.get('largo_cm', 0) or 0) or None
            execute_db("""
                UPDATE productos_base SET
                    nombre=%s, linea=%s, modelo=%s, medida=%s,
                    tipo_base=%s, modelo_almohada=%s,
                    precio_base=%s, descuento_catalogo=%s,
                    stock_minimo_pausar=%s, stock_minimo_reactivar=%s,
                    peso_gramos=%s, alto_cm=%s, ancho_cm=%s, largo_cm=%s
                WHERE sku=%s
            """, (nombre, linea, modelo, medida,
                  tipo_base, modelo_almohada,
                  precio_base, descuento_catalogo,
                  stock_min_pausar, stock_min_reactiv,
                  peso_gramos, alto_cm, ancho_cm, largo_cm,
                  sku))
        flash('Producto actualizado correctamente', 'success')
        return redirect(url_for('productos_lista'))

    return render_template('productos_form.html', producto=producto, modo='editar', es_sommier=es_sommier)


# ============================================================================
# COTIZADOR DE ENVÍOS ZIPNOVA
# ============================================================================

ZIPNOVA_BASE_URL_ADM   = 'https://api.zipnova.com.ar/v2'
ZIPNOVA_ACCOUNT_ID_ADM = os.getenv('ZIPNOVA_ACCOUNT_ID', '5786')
ZIPNOVA_ORIGIN_ID_ADM  = os.getenv('ZIPNOVA_ORIGIN_ID',  '374397')
ZIPNOVA_API_KEY_ADM    = os.getenv('ZIPNOVA_API_KEY', '')
ZIPNOVA_API_SECRET_ADM = os.getenv('ZIPNOVA_API_SECRET', '')
ZIPNOVA_PATAS_PESO_ADM = 1000  # gramos


def _zipnova_auth_adm():
    return (ZIPNOVA_API_KEY_ADM, ZIPNOVA_API_SECRET_ADM)


def _armar_bultos_cotizador(skus_lista):
    """
    Dado una lista de SKUs, arma los bultos para cotizar en Zipnova.
    Misma lógica que armar_bultos_zipnova de tienda_bp.
    skus_lista: lista de dicts {sku, cantidad}
    """
    bultos = []
    peso_patas_acum = 0
    hay_patas = False

    conn = get_db_connection()
    cur = conn.cursor()

    skus = [s['sku'] for s in skus_lista]
    if not skus:
        cur.close(); conn.close()
        return []

    placeholders = ','.join(['%s'] * len(skus))
    cur.execute(f"SELECT id, sku FROM productos_compuestos WHERE sku IN ({placeholders})", skus)
    compuestos_map = {r['sku']: r['id'] for r in cur.fetchall()}

    SKUS_ALMOHADA = {'CLASICA','SUBLIME','CERVICAL','RENOVATION','PLATINO','DORAL','DUAL','EXCLUSIVE'}

    for item in skus_lista:
        sku      = item['sku'].strip().upper()
        cantidad = int(item.get('cantidad', 1))

        for _ in range(cantidad):
            if sku in compuestos_map:
                # Sommier compuesto
                comp_id = compuestos_map[sku]
                cur.execute("""
                    SELECT pb.sku, pb.nombre, pb.alto_cm, pb.ancho_cm, pb.largo_cm,
                           pb.peso_gramos, c.cantidad_necesaria
                    FROM componentes c
                    JOIN productos_base pb ON c.producto_base_id = pb.id
                    WHERE c.producto_compuesto_id = %s
                """, (comp_id,))
                componentes = cur.fetchall()
                for comp in componentes:
                    csku = comp['sku']
                    cant = comp['cantidad_necesaria']
                    if csku in SKUS_ALMOHADA:
                        peso_patas_acum += (comp['peso_gramos'] or 0) * cant
                    elif csku.startswith('BASE_'):
                        for __ in range(cant):
                            bultos.append({
                                'sku':         csku,
                                'description': comp['nombre'],
                                'weight':      max(10, comp['peso_gramos'] or 20000),
                                'height':      comp['alto_cm']  or 21,
                                'width':       comp['ancho_cm'] or 100,
                                'length':      comp['largo_cm'] or 190,
                            })
                    else:
                        for __ in range(cant):
                            bultos.append({
                                'sku':         csku,
                                'description': comp['nombre'],
                                'weight':      max(10, comp['peso_gramos'] or 20000),
                                'height':      comp['alto_cm']  or 27,
                                'width':       comp['ancho_cm'] or 100,
                                'length':      comp['largo_cm'] or 190,
                            })
                peso_patas_acum += ZIPNOVA_PATAS_PESO_ADM
                hay_patas = True
            else:
                # Colchon o producto simple
                cur.execute(
                    "SELECT nombre, alto_cm, ancho_cm, largo_cm, peso_gramos FROM productos_base WHERE sku = %s",
                    (sku,)
                )
                pb = cur.fetchone()
                if pb:
                    bultos.append({
                        'sku':         sku,
                        'description': pb['nombre'],
                        'weight':      max(10, pb['peso_gramos'] or 20000),
                        'height':      pb['alto_cm']  or 27,
                        'width':       pb['ancho_cm'] or 100,
                        'length':      pb['largo_cm'] or 190,
                    })

    if hay_patas:
        bultos.append({
            'sku':         'PATAS',
            'description': 'Patas y accesorios',
            'weight':      max(10, int(peso_patas_acum)),
            'height':      30,
            'width':       20,
            'length':      10,
        })

    cur.close()
    conn.close()
    return bultos


@app.route('/cotizador-envio')
@login_required
def cotizador_envio():
    """Página del cotizador de envíos Zipnova."""
    productos = query_db("""
        SELECT sku, nombre, tipo FROM productos_base
        WHERE tipo IN ('colchon','almohada')
        ORDER BY tipo, nombre
    """)
    sommiers = query_db("""
        SELECT sku, nombre FROM productos_compuestos WHERE activo = 1 ORDER BY nombre
    """)
    todos = list(productos) + list(sommiers)
    return render_template('cotizador_envio.html', productos=todos)


@app.route('/cotizador-envio/localidades')
@login_required
def cotizador_localidades():
    """Devuelve localidades para un CP dado."""
    from tienda_bp import CP_LOCALIDADES
    cp = request.args.get('cp', '').strip()
    nombres = CP_LOCALIDADES.get(cp, []) or CP_LOCALIDADES.get(cp.zfill(4), [])
    return jsonify(sorted(nombres))


@app.route('/cotizador-envio/cotizar', methods=['POST'])
@login_required
def cotizador_cotizar():
    """AJAX: cotiza envío Zipnova para los SKUs indicados."""
    data = request.get_json() or {}
    skus_lista = data.get('skus', [])
    cp         = (data.get('cp') or '').strip()
    ciudad     = (data.get('ciudad') or '').strip()
    provincia  = (data.get('provincia') or 'Buenos Aires').strip()
    declared   = int(data.get('declared_value') or 0)

    if not skus_lista or not cp or not ciudad:
        return jsonify({'ok': False, 'error': 'Faltan datos (SKUs, CP o ciudad)'})

    try:
        bultos = _armar_bultos_cotizador(skus_lista)
        if not bultos:
            return jsonify({'ok': False, 'error': 'No se pudieron armar los bultos para los SKUs indicados'})

        payload = {
            'account_id':     ZIPNOVA_ACCOUNT_ID_ADM,
            'origin_id':      ZIPNOVA_ORIGIN_ID_ADM,
            'declared_value': declared,
            'destination': {
                'zipcode': cp,
                'city':    ciudad,
                'state':   provincia,
            },
            'items': bultos,
        }
        resp = requests.post(
            f"{ZIPNOVA_BASE_URL_ADM}/shipments/quote",
            json=payload,
            auth=_zipnova_auth_adm(),
            timeout=15,
        )
        resp.raise_for_status()
        resultados = resp.json().get('all_results') or resp.json().get('results') or []

        # Filtrar solo entrega a domicilio
        CODIGOS_DOMICILIO = {'standard_delivery', 'express_delivery', 'same_day', 'next_day'}
        resultados = [
            r for r in resultados
            if (r.get('service_type') or {}).get('code', 'standard_delivery') in CODIGOS_DOMICILIO
            or isinstance(r.get('service_type'), str) and 'pickup' not in r.get('service_type', '').lower()
        ] or resultados  # fallback: si no hay domicilio, mostrar todos

        opciones = []
        for r in resultados:
            amounts = r.get('amounts', {})
            price = (amounts.get('price_incl_tax')
                     or amounts.get('price')
                     or r.get('total_price')
                     or r.get('price') or 0)
            carrier = r.get('carrier', {})
            service = r.get('service_type', {})
            delivery = r.get('delivery_time', {})
            times = delivery.get('times', {}).get('total', {})
            dias_str = str(delivery.get('max') or delivery.get('min') or '—')
            opciones.append({
                'carrier':    carrier.get('name') or r.get('carrier_name') or '—',
                'servicio':   service.get('name') or r.get('service_name') or '',
                'precio':     float(price),
                'precio_fmt': f"${float(price):,.0f}".replace(',', '.'),
                'dias':       dias_str,
                'carrier_id': carrier.get('id') or r.get('carrier_id') or '',
            })
        opciones.sort(key=lambda x: x['precio'])

        return jsonify({
            'ok':      True,
            'bultos':  bultos,
            'opciones': opciones,
        })

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
