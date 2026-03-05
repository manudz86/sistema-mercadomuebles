# FUNCIÓN 1: FACTURA INDIVIDUAL

@app.route('/ventas/historicas/<int:venta_id>/generar-factura-excel')
def generar_factura_excel(venta_id):
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    incluir_flete = request.args.get('incluir_flete', 'false').lower() == 'true'
    metodos_con_flete = ['Flete Propio', 'Zippin']

    try:
        venta = query_one('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_historicas'))

        items = list(query_db('''
            SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
            FROM items_venta iv
            LEFT JOIN productos_base pb ON iv.sku = pb.sku
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            WHERE iv.venta_id = %s
        ''', (venta_id,)))

        tiene_flete_real = (
            venta.get('metodo_envio') in metodos_con_flete
            and venta.get('costo_flete')
            and venta['costo_flete'] > 0
            and incluir_flete
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturacion"

        total_cols = len(items) + (1 if tiene_flete_real else 0)
        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia']
        for idx in range(1, total_cols + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        headers.append('importe total')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row_data = []

        if venta.get('mla_code'):
            id_venta = venta['mla_code']
        elif venta.get('factura_business_name'):
            id_venta = venta['factura_business_name']
        else:
            id_venta = venta['nombre_cliente']
        row_data.append(id_venta)

        row_data.append(venta.get('factura_taxpayer_type') or 'Consumidor Final')
        row_data.append(venta.get('factura_business_name') or venta['nombre_cliente'])
        row_data.append(str(venta.get('factura_doc_number') or venta.get('dni_cliente') or '99999999'))

        if venta.get('factura_street'):
            direccion = str(venta['factura_street']) + ', ' + str(venta.get('factura_city', ''))
        else:
            direccion = venta.get('direccion_entrega') or ''
        row_data.append(direccion)

        row_data.append(
            venta.get('factura_state') or venta.get('provincia_cliente') or venta.get('zona_envio') or 'Capital Federal'
        )

        for item in items:
            row_data.append(item['sku'])
            row_data.append(int(item['cantidad']))
            row_data.append(float(item['precio_unitario']))

        if tiene_flete_real:
            row_data.append('FLETE')
            row_data.append(1)
            row_data.append(float(venta['costo_flete']))

        if tiene_flete_real:
            importe_total = float(venta['importe_total'])
        elif venta.get('metodo_envio') in metodos_con_flete and venta.get('costo_flete') and venta['costo_flete'] > 0:
            importe_total = float(venta['importe_total']) - float(venta['costo_flete'])
        else:
            importe_total = float(venta['importe_total'])
        row_data.append(importe_total)

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=2, column=col_idx, value=value)

        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        nombre_archivo = "factura_" + venta['numero_venta'].replace('/', '-') + ".xlsx"
        response.headers['Content-Disposition'] = 'attachment; filename=' + nombre_archivo

        execute_db('''
            UPDATE ventas SET factura_generada = TRUE, factura_fecha_generacion = NOW() WHERE id = %s
        ''', (venta_id,))

        return response

    except Exception as e:
        flash('Error al generar factura: ' + str(e), 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))


# FUNCION 2: FACTURA MULTIPLE

@app.route('/ventas/historicas/facturar-multiple-excel')
def facturar_multiple_excel():
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    incluir_flete = request.args.get('incluir_flete', 'false').lower() == 'true'
    metodos_con_flete = ['Flete Propio', 'Zippin']

    try:
        ids_str = request.args.get('ids', '')
        if not ids_str:
            flash('No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_historicas'))

        venta_ids = [int(id) for id in ids_str.split(',') if id.strip()]
        if not venta_ids:
            flash('No se seleccionaron ventas validas', 'error')
            return redirect(url_for('ventas_historicas'))

        placeholders = ', '.join(['%s'] * len(venta_ids))
        ventas = query_db('SELECT * FROM ventas WHERE id IN (' + placeholders + ') ORDER BY id DESC', tuple(venta_ids))

        if not ventas:
            flash('No se encontraron ventas', 'error')
            return redirect(url_for('ventas_historicas'))

        ventas_con_items = []
        max_columnas_items = 0

        for venta in ventas:
            items = list(query_db('''
                SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
            ''', (venta['id'],)))

            tiene_flete_real = (
                venta.get('metodo_envio') in metodos_con_flete
                and venta.get('costo_flete')
                and venta['costo_flete'] > 0
                and incluir_flete
            )

            columnas = len(items) + (1 if tiene_flete_real else 0)
            if columnas > max_columnas_items:
                max_columnas_items = columnas

            ventas_con_items.append({'venta': venta, 'items': items, 'tiene_flete_real': tiene_flete_real})

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturacion Multiple"

        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia']
        for idx in range(1, max_columnas_items + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        headers.append('importe total')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row = 2

        for venta_data in ventas_con_items:
            venta = venta_data['venta']
            items = venta_data['items']
            tiene_flete_real = venta_data['tiene_flete_real']

            row_data = []

            if venta.get('mla_code'):
                id_venta = venta['mla_code']
            elif venta.get('factura_business_name'):
                id_venta = venta['factura_business_name']
            else:
                id_venta = venta['nombre_cliente']
            row_data.append(id_venta)

            row_data.append(venta.get('factura_taxpayer_type') or 'Consumidor Final')
            row_data.append(venta.get('factura_business_name') or venta['nombre_cliente'])
            row_data.append(str(venta.get('factura_doc_number') or venta.get('dni_cliente') or '99999999'))

            if venta.get('factura_street'):
                direccion = str(venta['factura_street']) + ', ' + str(venta.get('factura_city', ''))
            else:
                direccion = venta.get('direccion_entrega') or ''
            row_data.append(direccion)

            row_data.append(
                venta.get('factura_state') or venta.get('provincia_cliente') or venta.get('zona_envio') or 'Capital Federal'
            )

            for item in items:
                row_data.append(item['sku'])
                row_data.append(int(item['cantidad']))
                row_data.append(float(item['precio_unitario']))

            if tiene_flete_real:
                row_data.append('FLETE')
                row_data.append(1)
                row_data.append(float(venta['costo_flete']))

            columnas_usadas = len(items) + (1 if tiene_flete_real else 0)
            for _ in range(max_columnas_items - columnas_usadas):
                row_data.extend(['', '', ''])

            if tiene_flete_real:
                importe_total = float(venta['importe_total'])
            elif venta.get('metodo_envio') in metodos_con_flete and venta.get('costo_flete') and venta['costo_flete'] > 0:
                importe_total = float(venta['importe_total']) - float(venta['costo_flete'])
            else:
                importe_total = float(venta['importe_total'])
            row_data.append(importe_total)

            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_idx, value=value)

            current_row += 1

        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = "facturas_multiple_" + fecha_str + ".xlsx"
        response.headers['Content-Disposition'] = 'attachment; filename=' + nombre_archivo

        placeholders = ', '.join(['%s'] * len(venta_ids))
        execute_db('UPDATE ventas SET factura_generada = TRUE, factura_fecha_generacion = NOW() WHERE id IN (' + placeholders + ')', tuple(venta_ids))

        return response

    except Exception as e:
        flash('Error al generar facturas: ' + str(e), 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))
