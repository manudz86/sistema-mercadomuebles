# ============================================================================
# ESTADÍSTICAS DE VENTAS
# ============================================================================

@app.route('/estadisticas')
def estadisticas():
    from datetime import datetime, timedelta

    # ========================================
    # FILTROS
    # ========================================
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    filtro_canal = request.args.get('canal', '')
    filtro_metodo = request.args.get('metodo_envio', '')
    filtro_zona = request.args.get('zona', '')

    # Default: último mes
    if not fecha_desde:
        fecha_desde = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_hasta:
        fecha_hasta = datetime.now().strftime('%Y-%m-%d')

    # ========================================
    # BASE DE FILTROS (reutilizable)
    # ========================================
    base_where = """
        WHERE v.estado_entrega != 'cancelada'
        AND DATE(v.fecha_venta) BETWEEN %s AND %s
    """
    base_params = [fecha_desde, fecha_hasta]

    if filtro_canal == 'ML':
        base_where += " AND v.canal = 'Mercado Libre'"
    elif filtro_canal == 'no_ml':
        base_where += " AND v.canal != 'Mercado Libre'"

    if filtro_metodo:
        base_where += " AND v.metodo_envio = %s"
        base_params.append(filtro_metodo)

    if filtro_zona:
        base_where += " AND v.zona_envio = %s"
        base_params.append(filtro_zona)

    # ========================================
    # 1. MÉTRICAS RESUMEN
    # ========================================
    resumen = query_one(f"""
        SELECT
            COUNT(*) as total_ventas,
            COALESCE(SUM(v.importe_total), 0) as total_facturado,
            COALESCE(AVG(v.importe_total), 0) as ticket_promedio,
            COALESCE(SUM(iv.cantidad), 0) as total_unidades
        FROM ventas v
        LEFT JOIN items_venta iv ON iv.venta_id = v.id
        {base_where}
    """, tuple(base_params))

    # ========================================
    # 2. VENTAS POR DÍA (para gráfico de línea)
    # ========================================
    ventas_por_dia = query_db(f"""
        SELECT
            DATE(v.fecha_venta) as dia,
            COUNT(*) as cantidad,
            SUM(v.importe_total) as total
        FROM ventas v
        {base_where}
        GROUP BY DATE(v.fecha_venta)
        ORDER BY dia
    """, tuple(base_params))

    # ========================================
    # 3. DESGLOSE POR CANAL (para torta)
    # ========================================
    por_canal = query_db(f"""
        SELECT
            CASE WHEN v.canal = 'Mercado Libre' THEN 'MercadoLibre' ELSE 'Venta Directa' END as canal_label,
            COUNT(*) as cantidad,
            SUM(v.importe_total) as total
        FROM ventas v
        {base_where}
        GROUP BY canal_label
    """, tuple(base_params))

    # ========================================
    # 4. DESGLOSE POR MÉTODO DE ENVÍO (para torta)
    # ========================================
    por_metodo = query_db(f"""
        SELECT
            COALESCE(v.metodo_envio, 'Sin especificar') as metodo,
            COUNT(*) as cantidad,
            SUM(v.importe_total) as total
        FROM ventas v
        {base_where}
        GROUP BY v.metodo_envio
        ORDER BY cantidad DESC
    """, tuple(base_params))

    # ========================================
    # 5. TOP PRODUCTOS MÁS VENDIDOS
    # ========================================
    top_productos = query_db(f"""
        SELECT
            iv.sku,
            COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre,
            SUM(iv.cantidad) as total_unidades,
            SUM(iv.cantidad * iv.precio_unitario) as total_facturado
        FROM ventas v
        JOIN items_venta iv ON iv.venta_id = v.id
        LEFT JOIN productos_base pb ON iv.sku = pb.sku
        LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
        {base_where}
        GROUP BY iv.sku, nombre
        ORDER BY total_unidades DESC
        LIMIT 15
    """, tuple(base_params))

    # ========================================
    # 6. FILTROS DISPONIBLES (para los selects)
    # ========================================
    metodos_disponibles = query_db("""
        SELECT DISTINCT metodo_envio FROM ventas
        WHERE metodo_envio IS NOT NULL AND metodo_envio != ''
        ORDER BY metodo_envio
    """)

    zonas_disponibles = query_db("""
        SELECT DISTINCT zona_envio FROM ventas
        WHERE zona_envio IS NOT NULL AND zona_envio != ''
        ORDER BY zona_envio
    """)

    return render_template('estadisticas.html',
        resumen=resumen,
        ventas_por_dia=ventas_por_dia,
        por_canal=por_canal,
        por_metodo=por_metodo,
        top_productos=top_productos,
        metodos_disponibles=metodos_disponibles,
        zonas_disponibles=zonas_disponibles,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        filtro_canal=filtro_canal,
        filtro_metodo=filtro_metodo,
        filtro_zona=filtro_zona,
    )


# ============================================================================
# EXPORTAR REPOSICIÓN A EXCEL
# ============================================================================

@app.route('/estadisticas/exportar-reposicion')
def exportar_reposicion():
    from flask import make_response
    from datetime import datetime, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    fecha_desde = request.args.get('fecha_desde', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_hasta = request.args.get('fecha_hasta', datetime.now().strftime('%Y-%m-%d'))
    filtro_canal = request.args.get('canal', '')
    filtro_metodo = request.args.get('metodo_envio', '')
    filtro_zona = request.args.get('zona', '')

    base_where = """
        WHERE v.estado_entrega != 'cancelada'
        AND DATE(v.fecha_venta) BETWEEN %s AND %s
    """
    base_params = [fecha_desde, fecha_hasta]

    if filtro_canal == 'ML':
        base_where += " AND v.canal = 'Mercado Libre'"
    elif filtro_canal == 'no_ml':
        base_where += " AND v.canal != 'Mercado Libre'"
    if filtro_metodo:
        base_where += " AND v.metodo_envio = %s"
        base_params.append(filtro_metodo)
    if filtro_zona:
        base_where += " AND v.zona_envio = %s"
        base_params.append(filtro_zona)

    # Productos base vendidos directamente (no son combos)
    base_directos = query_db(f"""
        SELECT
            pb.sku,
            pb.nombre,
            SUM(iv.cantidad) as cantidad_vendida,
            COALESCE(pb.stock_actual, 0) as stock_actual
        FROM ventas v
        JOIN items_venta iv ON iv.venta_id = v.id
        JOIN productos_base pb ON iv.sku = pb.sku
        {base_where}
        GROUP BY pb.sku, pb.nombre, pb.stock_actual
    """, tuple(base_params))

    # Combos → explotar en componentes base
    combos = query_db(f"""
        SELECT
            pb.sku,
            pb.nombre,
            SUM(iv.cantidad * c.cantidad_necesaria) as cantidad_vendida,
            COALESCE(pb.stock_actual, 0) as stock_actual
        FROM ventas v
        JOIN items_venta iv ON iv.venta_id = v.id
        JOIN productos_compuestos pc ON iv.sku = pc.sku
        JOIN componentes c ON c.producto_compuesto_id = pc.id
        JOIN productos_base pb ON pb.id = c.producto_base_id
        {base_where}
        GROUP BY pb.sku, pb.nombre, pb.stock_actual
    """, tuple(base_params))

    # Consolidar: sumar si el mismo SKU base aparece en ambas listas
    totales = {}
    for row in list(base_directos) + list(combos):
        sku = row['sku']
        if sku in totales:
            totales[sku]['cantidad_vendida'] += int(row['cantidad_vendida'])
        else:
            totales[sku] = {
                'sku': sku,
                'nombre': row['nombre'],
                'cantidad_vendida': int(row['cantidad_vendida']),
                'stock_actual': int(row['stock_actual'])
            }

    productos = sorted(totales.values(), key=lambda x: x['cantidad_vendida'], reverse=True)

    # CREAR EXCEL
    wb = Workbook()
    ws = wb.active
    ws.title = "Reposición"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Título
    ws.merge_cells('A1:D1')
    titulo = ws['A1']
    titulo.value = f"Reposición de Stock — {fecha_desde} al {fecha_hasta}"
    titulo.font = Font(bold=True, color='FFFFFF', name='Arial', size=12)
    titulo.fill = header_fill
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Headers
    headers = ['SKU', 'Descripción', 'Unidades Vendidas', 'Stock Actual']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Datos
    for i, p in enumerate(productos):
        row = i + 3
        fill = alt_fill if i % 2 == 0 else PatternFill(fill_type=None)
        valores = [p['sku'], p['nombre'], int(p['cantidad_vendida']), int(p['stock_actual'])]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name='Arial', size=10)
            cell.fill = fill
            cell.border = border
            if col in (3, 4):
                cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.freeze_panes = 'A3'

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = make_response(excel_file.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    fecha_str = datetime.now().strftime('%Y%m%d')
    response.headers['Content-Disposition'] = f'attachment; filename=reposicion_{fecha_str}.xlsx'
    return response
