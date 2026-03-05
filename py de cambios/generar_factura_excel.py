@app.route('/ventas/historicas/<int:venta_id>/generar-factura-excel')
def generar_factura_excel(venta_id):
    """
    Generar archivo .xlsx con datos de facturación de UNA venta
    Cada venta en una fila
    """
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    try:
        # Obtener venta
        venta = query_one('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        
        if not venta:
            flash('❌ Venta no encontrada', 'error')
            return redirect(url_for('ventas_historicas'))
        
        # Obtener items
        items = query_db('''
            SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
            FROM items_venta iv
            LEFT JOIN productos_base pb ON iv.sku = pb.sku
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            WHERE iv.venta_id = %s
        ''', (venta_id,))
        
        # Si tiene costo de envío, agregarlo como item FLETE
        if venta.get('costo_flete') and venta['costo_flete'] > 0:
            items = list(items)  # Convertir a lista si es necesario
            items.append({
                'sku': 'FLETE',
                'nombre_producto': 'Costo de Envío',
                'cantidad': 1,
                'precio_unitario': venta['costo_flete']
            })
        
        # ============================================
        # CREAR EXCEL
        # ============================================
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación"
        
        # ============================================
        # ENCABEZADOS
        # ============================================
        
        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia']
        
        # Agregar columnas dinámicas por cada SKU
        for idx, item in enumerate(items, 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        
        headers.append('importe total')
        
        # Escribir encabezados
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # ============================================
        # DATOS DE LA VENTA
        # ============================================
        
        row_data = []
        
        # ID Venta
        row_data.append(venta['numero_venta'])
        
        # Categoría IVA
        if venta.get('factura_taxpayer_type'):
            categoria_iva = venta['factura_taxpayer_type']
        else:
            categoria_iva = 'Consumidor Final'
        row_data.append(categoria_iva)
        
        # Nombre
        if venta.get('factura_business_name'):
            nombre = venta['factura_business_name']
        else:
            nombre = venta['nombre_cliente']
        row_data.append(nombre)
        
        # DNI/CUIT
        if venta.get('factura_doc_number'):
            dni = venta['factura_doc_number']
        else:
            dni = ''
        row_data.append(dni)
        
        # Dirección
        if venta.get('factura_street'):
            direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
        elif venta.get('direccion_entrega'):
            direccion = venta['direccion_entrega']
        else:
            direccion = ''
        row_data.append(direccion)
        
        # Provincia
        if venta.get('factura_state'):
            provincia = venta['factura_state']
        elif venta.get('zona_envio'):
            provincia = venta['zona_envio']
        else:
            provincia = ''
        row_data.append(provincia)
        
        # SKUs (cada producto)
        for item in items:
            row_data.append(item['sku'])
            row_data.append(item['cantidad'])
            row_data.append(item['cantidad'] * item['precio_unitario'])
        
        # Importe Total
        row_data.append(venta['importe_total'])
        
        # Escribir fila de datos
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=2, column=col_idx, value=value)
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # ============================================
        # GUARDAR Y ENVIAR
        # ============================================
        
        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Crear respuesta
        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Nombre del archivo
        nombre_archivo = f"factura_{venta['numero_venta'].replace('/', '-')}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        
        # ============================================
        # MARCAR COMO GENERADA
        # ============================================
        
        execute_db('''
            UPDATE ventas 
            SET factura_generada = TRUE,
                factura_fecha_generacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        return response
        
    except Exception as e:
        flash(f'❌ Error al generar factura: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))
