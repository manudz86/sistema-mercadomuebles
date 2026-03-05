# ============================================================================
# FUNCIÓN 1: FACTURA INDIVIDUAL
# ============================================================================

@app.route('/ventas/historicas/<int:venta_id>/generar-factura-excel')
def generar_factura_excel(venta_id):
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    try:
        venta = query_one('SELECT * FROM ventas WHERE id = %s', (venta_id,))

        if not venta:
            flash('❌ Venta no encontrada', 'error')
            return redirect(url_for('ventas_historicas'))

        items = list(query_db('''
            SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
            FROM items_venta iv
            LEFT JOIN productos_base pb ON iv.sku = pb.sku
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            WHERE iv.venta_id = %s
        ''', (venta_id,)))

        # Calcular precios normalizados
        costo_flete = float(venta.get('costo_flete') or 0)
        base_productos = float(venta['importe_total']) - costo_flete
        items_sin_flete = [i for i in items if i['sku'] != 'FLETE']
        total_bruto = sum(float(i['precio_unitario']) * int(i['cantidad']) for i in items_sin_flete)
        factor = (base_productos / total_bruto) if total_bruto > 0 else 1

        # Cantidad de columnas SKU: items + 1 si hay flete
        cant_slots = len(items_sin_flete) + (1 if costo_flete > 0 else 0)

        # CREAR EXCEL
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación"

        # ENCABEZADOS
        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia']
        for idx in range(1, cant_slots + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        headers.append('importe total')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row_data = []

        # ID Venta: apodo ML si es ML, nombre cliente si no
        if venta.get('mla_code'):
            id_venta = venta['mla_code']
        else:
            if venta.get('factura_business_name'):
                id_venta = venta['factura_business_name']
            else:
                id_venta = venta['nombre_cliente']
        row_data.append(id_venta)

        # Categoría IVA
        if venta.get('factura_taxpayer_type'):
            categoria_iva = venta['factura_taxpayer_type']
        else:
            categoria_iva = 'Consumidor Final'
        row_data.append(categoria_iva)

        # Nombre
        if venta.get('factura_business_name'):
            nombre = venta['factura_business_name']
        else:
            nombre = venta['nombre_cliente']
        row_data.append(nombre)

        # DNI/CUIT
        if venta.get('factura_doc_number'):
            dni = str(venta['factura_doc_number'])
        elif venta.get('dni_cliente'):
            dni = str(venta['dni_cliente'])
        else:
            dni = '99999999'
        row_data.append(dni)

        # Dirección
        if venta.get('factura_street'):
            direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
        elif venta.get('direccion_entrega'):
            direccion = venta['direccion_entrega']
        else:
            direccion = ''
        row_data.append(direccion)

        # Provincia
        if venta.get('factura_state'):
            provincia = venta['factura_state']
        elif venta.get('provincia_cliente'):
            provincia = venta['provincia_cliente']
        elif venta.get('zona_envio'):
            provincia = venta['zona_envio']
        else:
            provincia = 'Capital Federal'
        row_data.append(provincia)

        # SKUs con precio normalizado (sin flete embebido)
        for item in items_sin_flete:
            precio_normalizado = round(float(item['precio_unitario']) * factor, 2)
            row_data.append(item['sku'])
            row_data.append(int(item['cantidad']))
            row_data.append(precio_normalizado)

        # FLETE como línea separada
        if costo_flete > 0:
            row_data.append('FLETE')
            row_data.append(1)
            row_data.append(costo_flete)

        # Importe Total real
        row_data.append(float(venta['importe_total']))

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=2, column=col_idx, value=value)

        # Ajustar columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        nombre_archivo = f"factura_{venta['numero_venta'].replace('/', '-')}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        execute_db('''
            UPDATE ventas
            SET factura_generada = TRUE,
                factura_fecha_generacion = NOW()
            WHERE id = %s
        ''', (venta_id,))

        return response

    except Exception as e:
        flash(f'❌ Error al generar factura: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))


# ============================================================================
# FUNCIÓN 2: FACTURA MÚLTIPLE
# ============================================================================

@app.route('/ventas/historicas/facturar-multiple-excel')
def facturar_multiple_excel():
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    try:
        ids_str = request.args.get('ids', '')
        if not ids_str:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_historicas'))

        venta_ids = [int(id) for id in ids_str.split(',') if id.strip()]

        if not venta_ids:
            flash('❌ No se seleccionaron ventas válidas', 'error')
            return redirect(url_for('ventas_historicas'))

        placeholders = ', '.join(['%s'] * len(venta_ids))
        ventas = query_db(f'SELECT * FROM ventas WHERE id IN ({placeholders}) ORDER BY id DESC', tuple(venta_ids))

        if not ventas:
            flash('❌ No se encontraron ventas', 'error')
            return redirect(url_for('ventas_historicas'))

        # Preparar datos de cada venta con normalización
        ventas_preparadas = []
        max_slots = 0

        for venta in ventas:
            items = list(query_db('''
                SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
            ''', (venta['id'],)))

            costo_flete = float(venta.get('costo_flete') or 0)
            base_productos = float(venta['importe_total']) - costo_flete
            items_sin_flete = [i for i in items if i['sku'] != 'FLETE']
            total_bruto = sum(float(i['precio_unitario']) * int(i['cantidad']) for i in items_sin_flete)
            factor = (base_productos / total_bruto) if total_bruto > 0 else 1

            # Slots = items + 1 para flete si existe
            cant_slots = len(items_sin_flete) + (1 if costo_flete > 0 else 0)
            if cant_slots > max_slots:
                max_slots = cant_slots

            ventas_preparadas.append({
                'venta': venta,
                'items_sin_flete': items_sin_flete,
                'costo_flete': costo_flete,
                'factor': factor,
                'cant_slots': cant_slots
            })

        # CREAR EXCEL
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación Múltiple"

        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia']
        for idx in range(1, max_slots + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        headers.append('importe total')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row = 2

        for venta_prep in ventas_preparadas:
            venta = venta_prep['venta']
            items_sin_flete = venta_prep['items_sin_flete']
            costo_flete = venta_prep['costo_flete']
            factor = venta_prep['factor']
            cant_slots = venta_prep['cant_slots']

            row_data = []

            # ID Venta: apodo ML si es ML, nombre cliente si no
            if venta.get('mla_code'):
                id_venta = venta['mla_code']
            else:
                if venta.get('factura_business_name'):
                    id_venta = venta['factura_business_name']
                else:
                    id_venta = venta['nombre_cliente']
            row_data.append(id_venta)

            # Categoría IVA
            if venta.get('factura_taxpayer_type'):
                categoria_iva = venta['factura_taxpayer_type']
            else:
                categoria_iva = 'Consumidor Final'
            row_data.append(categoria_iva)

            # Nombre
            if venta.get('factura_business_name'):
                nombre = venta['factura_business_name']
            else:
                nombre = venta['nombre_cliente']
            row_data.append(nombre)

            # DNI
            if venta.get('factura_doc_number'):
                dni = str(venta['factura_doc_number'])
            elif venta.get('dni_cliente'):
                dni = str(venta['dni_cliente'])
            else:
                dni = '99999999'
            row_data.append(dni)

            # Dirección
            if venta.get('factura_street'):
                direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
            elif venta.get('direccion_entrega'):
                direccion = venta['direccion_entrega']
            else:
                direccion = ''
            row_data.append(direccion)

            # Provincia
            if venta.get('factura_state'):
                provincia = venta['factura_state']
            elif venta.get('provincia_cliente'):
                provincia = venta['provincia_cliente']
            elif venta.get('zona_envio'):
                provincia = venta['zona_envio']
            else:
                provincia = 'Capital Federal'
            row_data.append(provincia)

            # SKUs con precio normalizado
            for item in items_sin_flete:
                precio_normalizado = round(float(item['precio_unitario']) * factor, 2)
                row_data.append(item['sku'])
                row_data.append(int(item['cantidad']))
                row_data.append(precio_normalizado)

            # FLETE como línea separada
            if costo_flete > 0:
                row_data.append('FLETE')
                row_data.append(1)
                row_data.append(costo_flete)

            # Rellenar slots vacíos hasta max_slots
            slots_usados = cant_slots
            for _ in range(max_slots - slots_usados):
                row_data.extend(['', '', ''])

            # Importe Total real
            row_data.append(float(venta['importe_total']))

            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_idx, value=value)

            current_row += 1

        # Ajustar columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"facturas_multiple_{fecha_str}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        placeholders = ', '.join(['%s'] * len(venta_ids))
        execute_db(f'''
            UPDATE ventas
            SET factura_generada = TRUE,
                factura_fecha_generacion = NOW()
            WHERE id IN ({placeholders})
        ''', tuple(venta_ids))

        return response

    except Exception as e:
        flash(f'❌ Error al generar facturas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))
