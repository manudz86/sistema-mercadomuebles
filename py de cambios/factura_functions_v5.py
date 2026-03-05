@app.route('/ventas/historicas/<int:venta_id>/generar-factura-excel')
def generar_factura_excel(venta_id):
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    try:
        venta = query_one('SELECT * FROM ventas WHERE id = %s', (venta_id,))

        if not venta:
            flash('❌ Venta no encontrada', 'error')
            return redirect(url_for('ventas_historicas'))

        items = list(query_db('''
            SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
            FROM items_venta iv
            LEFT JOIN productos_base pb ON iv.sku = pb.sku
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            WHERE iv.venta_id = %s
        ''', (venta_id,)))

        costo_flete = float(venta.get('costo_flete') or 0)

        incluir_flete_param = request.args.get('incluir_flete', 'false').lower()
        incluir_flete = (incluir_flete_param == 'true' and costo_flete > 0)

        cant_slots = len(items) + (1 if incluir_flete else 0)

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación"

        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia', 'rubro']
        for idx in range(1, cant_slots + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        headers.append('importe total')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row_data = []

        if venta.get('mla_code'):
            id_venta = venta['mla_code']
        else:
            id_venta = venta.get('factura_business_name') or venta['nombre_cliente']
        row_data.append(id_venta)

        row_data.append(venta.get('factura_taxpayer_type') or 'Consumidor Final')

        row_data.append(venta.get('factura_business_name') or venta['nombre_cliente'])

        if venta.get('factura_doc_number'):
            dni = str(venta['factura_doc_number'])
        elif venta.get('dni_cliente'):
            dni = str(venta['dni_cliente'])
        else:
            dni = '99999999'
        row_data.append(dni)

        if venta.get('factura_street'):
            direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
        else:
            direccion = venta.get('direccion_entrega') or ''
        row_data.append(direccion)

        provincia_str = (venta.get('factura_state') or venta.get('provincia_cliente') or
                         venta.get('zona_envio') or 'Capital Federal')
        row_data.append(provincia_a_codigo(provincia_str))

        rubro = 'F' if len(dni.replace('.', '').replace('-', '').strip()) == 11 else 'R'
        row_data.append(rubro)

        for item in items:
            row_data.append(item['sku'])
            row_data.append(int(item['cantidad']))
            row_data.append(float(item['precio_unitario']))

        if incluir_flete:
            row_data.append('FLETE')
            row_data.append(1)
            row_data.append(costo_flete)

        total_excel = float(venta['importe_total']) + (costo_flete if incluir_flete else 0)
        row_data.append(total_excel)

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=2, column=col_idx, value=value)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        nombre_archivo = f"factura_{venta['numero_venta'].replace('/', '-')}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        execute_db('''
            UPDATE ventas SET factura_generada = TRUE, factura_fecha_generacion = NOW()
            WHERE id = %s
        ''', (venta_id,))

        return response

    except Exception as e:
        flash(f'❌ Error al generar factura: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))


# ============================================================================
# FUNCIÓN 2: FACTURA MÚLTIPLE
# ============================================================================

@app.route('/ventas/historicas/facturar-multiple-excel')
def facturar_multiple_excel():
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    try:
        ids_str = request.args.get('ids', '')
        if not ids_str:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_historicas'))

        venta_ids = [int(id) for id in ids_str.split(',') if id.strip()]

        if not venta_ids:
            flash('❌ No se seleccionaron ventas válidas', 'error')
            return redirect(url_for('ventas_historicas'))

        ids_con_flete_str = request.args.get('ids_con_flete', '')
        ids_con_flete = set()
        if ids_con_flete_str:
            ids_con_flete = {int(i) for i in ids_con_flete_str.split(',') if i.strip()}

        placeholders = ', '.join(['%s'] * len(venta_ids))
        ventas = query_db(f'SELECT * FROM ventas WHERE id IN ({placeholders}) ORDER BY id DESC', tuple(venta_ids))

        if not ventas:
            flash('❌ No se encontraron ventas', 'error')
            return redirect(url_for('ventas_historicas'))

        ventas_preparadas = []
        max_slots = 0

        for venta in ventas:
            items = list(query_db('''
                SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
            ''', (venta['id'],)))

            costo_flete = float(venta.get('costo_flete') or 0)
            incluir_flete = (venta['id'] in ids_con_flete and costo_flete > 0)

            cant_slots = len(items) + (1 if incluir_flete else 0)
            if cant_slots > max_slots:
                max_slots = cant_slots

            ventas_preparadas.append({
                'venta': venta,
                'items': items,
                'costo_flete': costo_flete,
                'incluir_flete': incluir_flete,
                'cant_slots': cant_slots
            })

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación Múltiple"

        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia', 'rubro']
        for idx in range(1, max_slots + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        headers.append('importe total')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row = 2

        for venta_prep in ventas_preparadas:
            venta = venta_prep['venta']
            items = venta_prep['items']
            costo_flete = venta_prep['costo_flete']
            incluir_flete = venta_prep['incluir_flete']
            cant_slots = venta_prep['cant_slots']

            row_data = []

            if venta.get('mla_code'):
                id_venta = venta['mla_code']
            else:
                id_venta = venta.get('factura_business_name') or venta['nombre_cliente']
            row_data.append(id_venta)

            row_data.append(venta.get('factura_taxpayer_type') or 'Consumidor Final')
            row_data.append(venta.get('factura_business_name') or venta['nombre_cliente'])

            if venta.get('factura_doc_number'):
                dni = str(venta['factura_doc_number'])
            elif venta.get('dni_cliente'):
                dni = str(venta['dni_cliente'])
            else:
                dni = '99999999'
            row_data.append(dni)

            if venta.get('factura_street'):
                direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
            else:
                direccion = venta.get('direccion_entrega') or ''
            row_data.append(direccion)

            provincia_str = (venta.get('factura_state') or venta.get('provincia_cliente') or
                             venta.get('zona_envio') or 'Capital Federal')
            row_data.append(provincia_a_codigo(provincia_str))

            rubro = 'F' if len(dni.replace('.', '').replace('-', '').strip()) == 11 else 'R'
            row_data.append(rubro)

            for item in items:
                row_data.append(item['sku'])
                row_data.append(int(item['cantidad']))
                row_data.append(float(item['precio_unitario']))

            if incluir_flete:
                row_data.append('FLETE')
                row_data.append(1)
                row_data.append(costo_flete)

            for _ in range(max_slots - cant_slots):
                row_data.extend(['', '', ''])

            total_excel = float(venta['importe_total']) + (costo_flete if incluir_flete else 0)
            row_data.append(total_excel)

            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_idx, value=value)

            current_row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"facturas_multiple_{fecha_str}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        placeholders = ', '.join(['%s'] * len(venta_ids))
        execute_db(f'''
            UPDATE ventas SET factura_generada = TRUE, factura_fecha_generacion = NOW()
            WHERE id IN ({placeholders})
        ''', tuple(venta_ids))

        return response

    except Exception as e:
        flash(f'❌ Error al generar facturas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))
