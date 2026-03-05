# ============================================================================
# FUNCIÓN 1: FACTURA INDIVIDUAL (ACTUALIZADA)
# ============================================================================

@app.route('/ventas/historicas/<int:venta_id>/generar-factura-excel')
def generar_factura_excel(venta_id):
    """
    Generar archivo .xlsx con datos de facturación de UNA venta
    Cada venta en una fila
    """
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    try:
        # Obtener venta
        venta = query_one('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        
        if not venta:
            flash('❌ Venta no encontrada', 'error')
            return redirect(url_for('ventas_historicas'))
        
        # Obtener items
        items = query_db('''
            SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
            FROM items_venta iv
            LEFT JOIN productos_base pb ON iv.sku = pb.sku
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            WHERE iv.venta_id = %s
        ''', (venta_id,))
        
        # Si tiene costo de envío, agregarlo como item FLETE
        if venta.get('costo_flete') and venta['costo_flete'] > 0:
            items = list(items)
            items.append({
                'sku': 'FLETE',
                'nombre_producto': 'Costo de Envío',
                'cantidad': 1,
                'precio_unitario': venta['costo_flete']
            })
        
        # ============================================
        # CREAR EXCEL
        # ============================================
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación"
        
        # ENCABEZADOS
        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia']
        
        for idx, item in enumerate(items, 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        
        headers.append('importe total')
        
        # Escribir encabezados
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # ============================================
        # DATOS DE LA VENTA
        # ============================================
        
        row_data = []
        
        # ID Venta
        row_data.append(venta['numero_venta'])
        
        # Categoría IVA
        if venta.get('factura_taxpayer_type'):
            categoria_iva = venta['factura_taxpayer_type']
        else:
            categoria_iva = 'Consumidor Final'
        row_data.append(categoria_iva)
        
        # Nombre
        if venta.get('factura_business_name'):
            nombre = venta['factura_business_name']
        else:
            nombre = venta['nombre_cliente']
        row_data.append(nombre)
        
        # ✅ DNI/CUIT - DEFAULT "99999999" si está vacío
        if venta.get('factura_doc_number'):
            dni = str(venta['factura_doc_number'])
        elif venta.get('dni_cliente'):
            dni = str(venta['dni_cliente'])
        else:
            dni = '99999999'
        row_data.append(dni)
        
        # Dirección
        if venta.get('factura_street'):
            direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
        elif venta.get('direccion_entrega'):
            direccion = venta['direccion_entrega']
        else:
            direccion = ''
        row_data.append(direccion)
        
        # ✅ Provincia - DEFAULT "Capital Federal" si está vacía
        if venta.get('factura_state'):
            provincia = venta['factura_state']
        elif venta.get('provincia_cliente'):
            provincia = venta['provincia_cliente']
        elif venta.get('zona_envio'):
            provincia = venta['zona_envio']
        else:
            provincia = 'Capital Federal'
        row_data.append(provincia)
        
        # SKUs (cada producto)
        for item in items:
            row_data.append(item['sku'])
            row_data.append(item['cantidad'])
            row_data.append(item['cantidad'] * item['precio_unitario'])
        
        # Importe Total
        row_data.append(venta['importe_total'])
        
        # Escribir fila de datos
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=2, column=col_idx, value=value)
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # ============================================
        # GUARDAR Y ENVIAR
        # ============================================
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        nombre_archivo = f"factura_{venta['numero_venta'].replace('/', '-')}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        
        # ============================================
        # MARCAR COMO GENERADA
        # ============================================
        
        execute_db('''
            UPDATE ventas 
            SET factura_generada = TRUE,
                factura_fecha_generacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        return response
        
    except Exception as e:
        flash(f'❌ Error al generar factura: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))


# ============================================================================
# FUNCIÓN 2: FACTURA MÚLTIPLE (ACTUALIZADA - SIN TOTAL GENERAL)
# ============================================================================

@app.route('/ventas/historicas/facturar-multiple-excel')
def facturar_multiple_excel():
    """
    Generar UN archivo .xlsx con TODAS las ventas seleccionadas
    Cada venta en UNA FILA
    """
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    try:
        # Obtener IDs
        ids_str = request.args.get('ids', '')
        if not ids_str:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        venta_ids = [int(id) for id in ids_str.split(',') if id.strip()]
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas válidas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        # Obtener todas las ventas
        placeholders = ', '.join(['%s'] * len(venta_ids))
        query = f'''
            SELECT * FROM ventas 
            WHERE id IN ({placeholders})
            ORDER BY id DESC
        '''
        ventas = query_db(query, tuple(venta_ids))
        
        if not ventas:
            flash('❌ No se encontraron ventas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        # Obtener items de cada venta
        ventas_con_items = []
        max_items = 0
        
        for venta in ventas:
            items = query_db('''
                SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
            ''', (venta['id'],))
            
            items = list(items)
            
            # Si tiene costo de envío, agregarlo como item FLETE
            if venta.get('costo_flete') and venta['costo_flete'] > 0:
                items.append({
                    'sku': 'FLETE',
                    'nombre_producto': 'Costo de Envío',
                    'cantidad': 1,
                    'precio_unitario': venta['costo_flete']
                })
            
            ventas_con_items.append({
                'venta': venta,
                'items': items
            })
            
            if len(items) > max_items:
                max_items = len(items)
        
        # ============================================
        # CREAR EXCEL
        # ============================================
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación Múltiple"
        
        # ENCABEZADOS
        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia']
        
        for idx in range(1, max_items + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        
        headers.append('importe total')
        
        # Escribir encabezados
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # ============================================
        # DATOS DE CADA VENTA
        # ============================================
        
        current_row = 2
        
        for venta_data in ventas_con_items:
            venta = venta_data['venta']
            items = venta_data['items']
            
            row_data = []
            
            # ID Venta
            row_data.append(venta['numero_venta'])
            
            # Categoría IVA
            if venta.get('factura_taxpayer_type'):
                categoria_iva = venta['factura_taxpayer_type']
            else:
                categoria_iva = 'Consumidor Final'
            row_data.append(categoria_iva)
            
            # Nombre
            if venta.get('factura_business_name'):
                nombre = venta['factura_business_name']
            else:
                nombre = venta['nombre_cliente']
            row_data.append(nombre)
            
            # ✅ DNI/CUIT - DEFAULT "99999999" si está vacío
            if venta.get('factura_doc_number'):
                dni = str(venta['factura_doc_number'])
            elif venta.get('dni_cliente'):
                dni = str(venta['dni_cliente'])
            else:
                dni = '99999999'
            row_data.append(dni)
            
            # Dirección
            if venta.get('factura_street'):
                direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
            elif venta.get('direccion_entrega'):
                direccion = venta['direccion_entrega']
            else:
                direccion = ''
            row_data.append(direccion)
            
            # ✅ Provincia - DEFAULT "Capital Federal" si está vacía
            if venta.get('factura_state'):
                provincia = venta['factura_state']
            elif venta.get('provincia_cliente'):
                provincia = venta['provincia_cliente']
            elif venta.get('zona_envio'):
                provincia = venta['zona_envio']
            else:
                provincia = 'Capital Federal'
            row_data.append(provincia)
            
            # SKUs
            for item in items:
                row_data.append(item['sku'])
                row_data.append(item['cantidad'])
                row_data.append(item['cantidad'] * item['precio_unitario'])
            
            # Rellenar vacíos si tiene menos items que el máximo
            items_restantes = max_items - len(items)
            for _ in range(items_restantes):
                row_data.extend(['', '', ''])
            
            # Importe Total
            row_data.append(venta['importe_total'])
            
            # Escribir fila
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_idx, value=value)
            
            current_row += 1
        
        # ✅ SIN FILA DE TOTAL GENERAL (eliminada)
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # ============================================
        # GUARDAR Y ENVIAR
        # ============================================
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"facturas_multiple_{fecha_str}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        
        # Marcar como generadas
        placeholders = ', '.join(['%s'] * len(venta_ids))
        execute_db(f'''
            UPDATE ventas 
            SET factura_generada = TRUE,
                factura_fecha_generacion = NOW()
            WHERE id IN ({placeholders})
        ''', tuple(venta_ids))
        
        return response
        
    except Exception as e:
        flash(f'❌ Error al generar facturas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))
