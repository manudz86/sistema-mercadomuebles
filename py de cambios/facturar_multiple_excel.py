@app.route('/ventas/historicas/facturar-multiple-excel')
def facturar_multiple_excel():
    """
    Generar UN archivo .xlsx con TODAS las ventas seleccionadas
    Cada venta en UNA FILA
    """
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    try:
        # Obtener IDs
        ids_str = request.args.get('ids', '')
        if not ids_str:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        venta_ids = [int(id) for id in ids_str.split(',') if id.strip()]
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas válidas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        # ============================================
        # OBTENER TODAS LAS VENTAS
        # ============================================
        placeholders = ', '.join(['%s'] * len(venta_ids))
        query = f'''
            SELECT * FROM ventas 
            WHERE id IN ({placeholders})
            ORDER BY id DESC
        '''
        ventas = query_db(query, tuple(venta_ids))
        
        if not ventas:
            flash('❌ No se encontraron ventas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        # ============================================
        # OBTENER ITEMS DE CADA VENTA
        # ============================================
        ventas_con_items = []
        max_items = 0  # Para saber cuántas columnas de SKU necesitamos
        
        for venta in ventas:
            items = query_db('''
                SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
            ''', (venta['id'],))
            
            items = list(items)
            
            # Si tiene costo de envío, agregarlo como item FLETE
            if venta.get('costo_flete') and venta['costo_flete'] > 0:
                items.append({
                    'sku': 'FLETE',
                    'nombre_producto': 'Costo de Envío',
                    'cantidad': 1,
                    'precio_unitario': venta['costo_flete']
                })
            
            ventas_con_items.append({
                'venta': venta,
                'items': items
            })
            
            # Actualizar máximo de items
            if len(items) > max_items:
                max_items = len(items)
        
        # ============================================
        # CREAR EXCEL
        # ============================================
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación Múltiple"
        
        # ============================================
        # ENCABEZADOS
        # ============================================
        
        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia']
        
        # Agregar columnas dinámicas según el máximo de SKUs
        for idx in range(1, max_items + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        
        headers.append('importe total')
        
        # Escribir encabezados
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # ============================================
        # DATOS DE CADA VENTA (cada venta en una fila)
        # ============================================
        
        current_row = 2
        
        for venta_data in ventas_con_items:
            venta = venta_data['venta']
            items = venta_data['items']
            
            row_data = []
            
            # ID Venta
            row_data.append(venta['numero_venta'])
            
            # Categoría IVA
            if venta.get('factura_taxpayer_type'):
                categoria_iva = venta['factura_taxpayer_type']
            else:
                categoria_iva = 'Consumidor Final'
            row_data.append(categoria_iva)
            
            # Nombre
            if venta.get('factura_business_name'):
                nombre = venta['factura_business_name']
            else:
                nombre = venta['nombre_cliente']
            row_data.append(nombre)
            
            # DNI/CUIT
            if venta.get('factura_doc_number'):
                dni = venta['factura_doc_number']
            else:
                dni = ''
            row_data.append(dni)
            
            # Dirección
            if venta.get('factura_street'):
                direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
            elif venta.get('direccion_entrega'):
                direccion = venta['direccion_entrega']
            else:
                direccion = ''
            row_data.append(direccion)
            
            # Provincia
            if venta.get('factura_state'):
                provincia = venta['factura_state']
            elif venta.get('zona_envio'):
                provincia = venta['zona_envio']
            else:
                provincia = ''
            row_data.append(provincia)
            
            # SKUs (cada producto)
            for item in items:
                row_data.append(item['sku'])
                row_data.append(item['cantidad'])
                row_data.append(item['cantidad'] * item['precio_unitario'])
            
            # Rellenar con vacíos si esta venta tiene menos items que el máximo
            items_restantes = max_items - len(items)
            for _ in range(items_restantes):
                row_data.extend(['', '', ''])
            
            # Importe Total
            row_data.append(venta['importe_total'])
            
            # Escribir fila de datos
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_idx, value=value)
            
            current_row += 1
        
        # ============================================
        # FILA DE TOTALES (opcional)
        # ============================================
        
        # Total general
        total_general = sum([v['venta']['importe_total'] for v in ventas_con_items])
        
        ws.cell(row=current_row, column=1, value="TOTAL GENERAL").font = Font(bold=True)
        ws.cell(row=current_row, column=len(headers), value=total_general).font = Font(bold=True)
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # ============================================
        # GUARDAR Y ENVIAR
        # ============================================
        
        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Crear respuesta
        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Nombre del archivo
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"facturas_multiple_{fecha_str}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        
        # ============================================
        # MARCAR TODAS COMO GENERADAS
        # ============================================
        
        placeholders = ', '.join(['%s'] * len(venta_ids))
        execute_db(f'''
            UPDATE ventas 
            SET factura_generada = TRUE,
                factura_fecha_generacion = NOW()
            WHERE id IN ({placeholders})
        ''', tuple(venta_ids))
        
        return response
        
    except Exception as e:
        flash(f'❌ Error al generar facturas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))
