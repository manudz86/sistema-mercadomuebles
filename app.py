"""
SISTEMA INVENTARIO CANNON - VERSIÓN SIMPLIFICADA
Todo el código en un solo archivo para evitar problemas de imports en Windows
"""

import os
import json
import requests  # pip install requests
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from dotenv import load_dotenv
import pymysql

import threading

_ml_rate_lock = threading.Lock()
_ml_last_request = 0.0
_ML_MIN_INTERVAL = 0.7  # ~1.4 requests/segundo máximo

ML_SELLER_ID = 29563319

# Cargar configuración
load_dotenv('config/.env')

# Configurar Flask
app = Flask(__name__, template_folder='templates')
app.secret_key = os.getenv('SECRET_KEY', 'cambiar-en-produccion-123456')

# Filtro personalizado para dashboard visual
@app.template_filter('zero_dash')
def zero_dash(value):
    """Convierte 0 en '-' para el dashboard visual"""
    return '-' if value == 0 or value is None else value

# ============================================================================
# FLASK-LOGIN SETUP
# ============================================================================
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '⚠️ Debés iniciar sesión para acceder'
login_manager.login_message_category = 'warning'

class User(UserMixin):
    def __init__(self, id, username, rol, activo):
        self.id = id
        self.username = username
        self.rol = rol
        self.activo = activo

@login_manager.user_loader
def load_user(user_id):
    row = query_one('SELECT * FROM usuarios WHERE id = %s AND activo = TRUE', (user_id,))
    if row:
        return User(row['id'], row['username'], row['rol'], row['activo'])
    return None

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.rol != 'admin':
            flash('❌ No tenés permisos para realizar esta acción', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

# Configuración de base de datos
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'port': int(os.getenv('DB_PORT', 3306)),
    'user': os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', ''),
    'database': os.getenv('DB_NAME', 'inventario_cannon'),
    'charset': 'utf8mb4',
    'cursorclass': pymysql.cursors.DictCursor
}

# ============================================================================
# FUNCIONES DE BASE DE DATOS
# ============================================================================

def get_db_connection():
    """Crear conexión a la base de datos"""
    return pymysql.connect(**DB_CONFIG)

def query_db(query, params=None):
    """Ejecutar query SELECT y retornar resultados"""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(query, params or ())
            return cursor.fetchall()
    finally:
        conn.close()

def query_one(query, params=None):
    """Ejecutar query SELECT y retornar un solo resultado"""
    results = query_db(query, params)
    return results[0] if results else None

def execute_db(query, params=None):
    """Ejecutar query INSERT/UPDATE/DELETE"""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(query, params or ())
            conn.commit()
            return cursor.lastrowid
    finally:
        conn.close()

# ============================================================================
# RUTAS - PÁGINAS
# ============================================================================

# ============================================================================
# RUTAS DE LOGIN / LOGOUT
# ============================================================================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user_row = query_one('SELECT * FROM usuarios WHERE username = %s AND activo = TRUE', (username,))
        if user_row and check_password_hash(user_row['password_hash'], password):
            user = User(user_row['id'], user_row['username'], user_row['rol'], user_row['activo'])
            login_user(user, remember=True)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        flash('❌ Usuario o contraseña incorrectos', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('✅ Sesión cerrada', 'success')
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    """Dashboard principal"""
    stats = {'ventas_activas': 0, 'ventas_en_proceso': 0, 'alertas_pendientes': 0}
    stock_critico = []
    
    try:
        # Contar ventas activas
        result = query_one("SELECT COUNT(*) as total FROM ventas WHERE estado_entrega = 'pendiente'")
        stats['ventas_activas'] = result['total'] if result else 0
        
        # Contar ventas en proceso
        result = query_one("SELECT COUNT(*) as total FROM ventas WHERE estado_entrega = 'en_proceso'")
        stats['ventas_en_proceso'] = result['total'] if result else 0
        
        # Contar alertas
        result = query_one("SELECT COUNT(*) as total FROM alertas_stock WHERE estado = 'pendiente'")
        stats['alertas_pendientes'] = result['total'] if result else 0
        
        # Stock crítico
        stock_critico = query_db("SELECT * FROM stock_disponible_ml WHERE estado_stock = 'SIN_STOCK' LIMIT 10")
        
    except Exception as e:
        flash(f'Error al cargar dashboard: {str(e)}', 'error')
    
    return render_template('dashboard.html', stats=stats, stock_critico=stock_critico, now=datetime.now)




# ============================================================================
# FUNCIONES AUXILIARES: VERIFICACIÓN DE STOCK
# Agregar ANTES de las rutas de /ventas/activas en app.py
# ============================================================================

def verificar_stock_disponible(cursor, items, ubicacion_despacho):
    """
    Verifica si hay stock suficiente para todos los items de una venta.
    Retorna (True, []) si hay stock suficiente
    Retorna (False, [lista de errores]) si falta stock
    """
    errores = []
    
    for item in items:
        sku = item['sku']
        cantidad = item['cantidad']
        
        # Verificar si es un combo
        cursor.execute('SELECT id FROM productos_compuestos WHERE sku = %s', (sku,))
        combo = cursor.fetchone()
        
        if combo:
            # Es combo: verificar componentes
            cursor.execute('''
                SELECT pb.sku, pb.nombre, pb.tipo, c.cantidad_necesaria
                FROM componentes c
                JOIN productos_base pb ON c.producto_base_id = pb.id
                WHERE c.producto_compuesto_id = %s
            ''', (combo['id'],))
            componentes = cursor.fetchall()
            
            for comp in componentes:
                sku_comp = comp['sku']
                nombre_comp = comp['nombre']
                cant_necesaria = comp['cantidad_necesaria'] * cantidad
                tipo_comp = comp['tipo']
                
                # Verificar stock del componente
                stock_disponible = obtener_stock_disponible(cursor, sku_comp, tipo_comp, ubicacion_despacho)
                
                if stock_disponible < cant_necesaria:
                    errores.append(f"{nombre_comp} (SKU: {sku_comp}): Necesitas {cant_necesaria}, disponible {stock_disponible}")
        
        else:
            # Es producto simple
            cursor.execute('SELECT nombre, tipo FROM productos_base WHERE sku = %s', (sku,))
            prod = cursor.fetchone()
            
            if not prod:
                errores.append(f"Producto {sku} no encontrado en base de datos")
                continue
            
            nombre = prod['nombre']
            tipo = prod['tipo']
            
            # Verificar stock
            stock_disponible = obtener_stock_disponible(cursor, sku, tipo, ubicacion_despacho)
            
            if stock_disponible < cantidad:
                errores.append(f"{nombre} (SKU: {sku}): Necesitas {cantidad}, disponible {stock_disponible}")
    
    if errores:
        return False, errores
    else:
        return True, []


def obtener_stock_disponible(cursor, sku, tipo, ubicacion_despacho):
    """
    Obtiene el stock disponible de un producto según su tipo y ubicación.
    """
    # COMPAC: tiene _DEP y _FULL
    if '_DEP' in sku or '_FULL' in sku:
        if ubicacion_despacho == 'FULL':
            sku_real = sku.replace('_DEP', '_FULL')
        else:
            sku_real = sku.replace('_FULL', '_DEP')
        
        cursor.execute('SELECT stock_actual FROM productos_base WHERE sku = %s', (sku_real,))
        prod = cursor.fetchone()
        return prod['stock_actual'] if prod else 0
    
    # ALMOHADAS: tienen stock_actual (DEP) y stock_full (FULL)
    elif tipo == 'almohada':
        cursor.execute('SELECT stock_actual, stock_full FROM productos_base WHERE sku = %s', (sku,))
        prod = cursor.fetchone()
        
        if not prod:
            return 0
        
        if ubicacion_despacho == 'FULL':
            return prod['stock_full']
        else:
            return prod['stock_actual']
    
    # BASES CHICAS (80200, 90200, 100200): usar stock directo
    elif tipo == 'base' and any(x in sku for x in ['80200', '90200', '100200']):
        cursor.execute('SELECT stock_actual FROM productos_base WHERE sku = %s', (sku,))
        prod = cursor.fetchone()
        return prod['stock_actual'] if prod else 0
    
    # BASES GRANDES (160, 180, 200): dividir stock de bases chicas entre 2
    elif tipo == 'base' and any(x in sku for x in ['160', '180', '200']):
        # Determinar SKU de bases chicas
        if '160' in sku:
            sku_chica = sku.replace('160', '80200')
        elif '180' in sku:
            sku_chica = sku.replace('180', '90200')
        elif '200' in sku:
            sku_chica = sku.replace('200', '100200')
        else:
            return 0
        
        cursor.execute('SELECT stock_actual FROM productos_base WHERE sku = %s', (sku_chica,))
        prod = cursor.fetchone()
        
        if not prod:
            return 0
        
        # Una base grande = 2 bases chicas
        # Stock disponible de bases grandes = stock_chicas / 2
        return prod['stock_actual'] // 2
    
    # OTROS: stock_actual normal
    else:
        cursor.execute('SELECT stock_actual FROM productos_base WHERE sku = %s', (sku,))
        prod = cursor.fetchone()
        return prod['stock_actual'] if prod else 0



# Porcentajes ML por defecto
PORCENTAJES_ML_DEFAULT = {
    'cuota_simple':  5.00,
    'cuotas_3':      9.40,
    'cuotas_6':     15.10,
    'cuotas_9':     20.70,
    'cuotas_12':    25.90,
}

@app.route('/configuracion/porcentajes-ml', methods=['GET'])
@login_required
def get_porcentajes_ml():
    """Retorna los porcentajes de ML guardados en DB"""
    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'porcentajes_ml'")
        if row:
            return jsonify({'ok': True, 'porcentajes': json.loads(row['valor'])})
        return jsonify({'ok': True, 'porcentajes': PORCENTAJES_ML_DEFAULT})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/configuracion/porcentajes-ml', methods=['POST'])
@login_required
def guardar_porcentajes_ml():
    """Guarda los porcentajes de ML en DB"""
    try:
        data = request.get_json()
        porcentajes = data.get('porcentajes', {})
        valor = json.dumps(porcentajes)
        execute_db(
            "INSERT INTO configuracion (clave, valor) VALUES ('porcentajes_ml', %s) "
            "ON DUPLICATE KEY UPDATE valor = %s, actualizado_at = NOW()",
            (valor, valor)
        )
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# ============================================================================
# FUNCIÓN CORREGIDA: EXCLUIR VENTA ACTUAL
# Reemplazar detectar_alertas_stock_bajo() en app.py
# ============================================================================

def detectar_alertas_stock_bajo(cursor, items_vendidos, venta_id_actual=None):
    """
    Detecta productos con stock disponible <= 0.
    
    Args:
        cursor: Cursor de BD
        items_vendidos: Lista de items de la venta actual
        venta_id_actual: ID de la venta actual para excluirla del cálculo
    
    Returns:
        Lista de productos sin stock
    """
    productos_sin_stock = []
    
    try:
        print("\n" + "="*60)
        print("🔍 DEBUG - DETECTAR ALERTAS")
        print("="*60)
        print(f"Items vendidos: {items_vendidos}")
        print(f"Venta ID actual (excluir): {venta_id_actual}")
        
        # ============================================
        # 1. SKUs A VERIFICAR
        # ============================================
        skus_a_verificar = set()
        cantidades_venta_actual = {}
        
        for item in items_vendidos:
            sku = item['sku']
            cantidad = item['cantidad']
            
            print(f"\n📦 Procesando item: {sku} x{cantidad}")
            
            cursor.execute('SELECT id FROM productos_compuestos WHERE sku = %s', (sku,))
            es_combo = cursor.fetchone()
            
            if es_combo:
                print(f"  → Es COMBO")
                cursor.execute('''
                    SELECT pb.sku, c.cantidad_necesaria
                    FROM componentes c
                    JOIN productos_base pb ON c.producto_base_id = pb.id
                    WHERE c.producto_compuesto_id = %s
                ''', (es_combo['id'],))
                
                componentes = cursor.fetchall()
                for comp in componentes:
                    comp_sku = comp['sku']
                    comp_cant = comp['cantidad_necesaria'] * cantidad
                    skus_a_verificar.add(comp_sku)
                    cantidades_venta_actual[comp_sku] = cantidades_venta_actual.get(comp_sku, 0) + comp_cant
            else:
                print(f"  → Es PRODUCTO BASE")
                skus_a_verificar.add(sku)
                cantidades_venta_actual[sku] = cantidades_venta_actual.get(sku, 0) + cantidad
        
        print(f"\n📋 SKUs a verificar: {skus_a_verificar}")
        print(f"📋 Cantidades venta actual: {cantidades_venta_actual}")
        
        # ============================================
        # 2. VENTAS ACTIVAS (EXCLUYENDO LA ACTUAL)
        # ============================================
        if venta_id_actual:
            cursor.execute('''
                SELECT 
                    COALESCE(pb_comp.sku, iv.sku) as sku,
                    SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
                FROM items_venta iv
                JOIN ventas v ON iv.venta_id = v.id
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
                LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
                WHERE v.estado_entrega = 'pendiente'
                AND v.id != %s
                GROUP BY sku
            ''', (venta_id_actual,))
        else:
            cursor.execute('''
                SELECT 
                    COALESCE(pb_comp.sku, iv.sku) as sku,
                    SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
                FROM items_venta iv
                JOIN ventas v ON iv.venta_id = v.id
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
                LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
                WHERE v.estado_entrega = 'pendiente'
                GROUP BY sku
            ''')
        
        ventas_activas = cursor.fetchall()
        ventas_dict = {v['sku']: int(v['vendido']) for v in ventas_activas}
        
        print(f"\n📊 Ventas activas en BD (sin venta actual): {ventas_dict}")
        
        # ============================================
        # 3. VERIFICAR CADA SKU
        # ============================================
        skus_sin_stock = []
        
        for sku in skus_a_verificar:
            print(f"\n🔎 Verificando: {sku}")
            
            cursor.execute('''
                SELECT sku, nombre, stock_actual, COALESCE(stock_full, 0) as stock_full, tipo
                FROM productos_base
                WHERE sku = %s
            ''', (sku,))
            
            prod = cursor.fetchone()
            
            if not prod:
                print(f"  ❌ NO EXISTE")
                continue
            
            print(f"  ✅ Encontrado: {prod['nombre']}")
            
            stock_fisico = prod['stock_actual'] + prod['stock_full']
            vendido_anterior = ventas_dict.get(sku, 0)
            vendido_actual = cantidades_venta_actual.get(sku, 0)
            vendido_total = vendido_anterior + vendido_actual
            stock_disponible = stock_fisico - vendido_total
            
            print(f"  📊 Stock físico: {stock_fisico}")
            print(f"  📊 Vendido anterior: {vendido_anterior}")
            print(f"  📊 Vendido actual: {vendido_actual}")
            print(f"  📊 Vendido TOTAL: {vendido_total}")
            print(f"  📊 Disponible: {stock_disponible}")
            
            if stock_disponible <= 0:
                print(f"  ⚠️ SIN STOCK - DEBE ALERTAR")
                
                skus_sin_stock.append(sku)
                
                productos_sin_stock.append({
                    'sku': sku,
                    'nombre': prod['nombre'],
                    'stock_fisico': stock_fisico,
                    'vendido': vendido_total,
                    'stock_disponible': stock_disponible,
                    'tipo_producto': 'base'
                })
                
                # Guardar alerta
                try:
                    cursor.execute('SELECT id FROM alertas_stock WHERE sku = %s AND estado = "pendiente"', (sku,))
                    
                    if not cursor.fetchone():
                        print(f"  💾 Guardando alerta...")
                        cursor.execute('''
                            INSERT INTO alertas_stock 
                            (sku, nombre_producto, stock_fisico, stock_vendido, stock_disponible, tipo_alerta, estado)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku, prod['nombre'], stock_fisico, vendido_total, stock_disponible, 'SIN_STOCK', 'pendiente'))
                        print(f"  ✅ Alerta guardada")
                    else:
                        print(f"  ℹ️ Alerta ya existe")
                except Exception as e:
                    print(f"  ❌ Error al guardar alerta: {str(e)}")
            else:
                print(f"  ✅ Stock OK")
        
        # ============================================
        # 4. BUSCAR COMBOS AFECTADOS
        # ============================================
        combos_afectados = []
        
        if skus_sin_stock:
            for sku_sin_stock in skus_sin_stock:
                try:
                    cursor.execute('''
                        SELECT DISTINCT pc.sku as combo_sku, pc.nombre as combo_nombre
                        FROM productos_compuestos pc
                        JOIN componentes c ON pc.id = c.producto_compuesto_id
                        JOIN productos_base pb ON c.producto_base_id = pb.id
                        WHERE pb.sku = %s AND pc.activo = 1
                    ''', (sku_sin_stock,))
                    
                    combos_que_usan = cursor.fetchall()
                    
                    for combo in combos_que_usan:
                        combo_sku = combo['combo_sku']
                        combo_nombre = combo['combo_nombre']
                        
                        if not any(p['sku'] == combo_sku for p in productos_sin_stock):
                            combos_afectados.append({
                                'sku': combo_sku,
                                'nombre': combo_nombre,
                                'componente_faltante': sku_sin_stock,
                                'tipo_producto': 'combo'
                            })
                            
                            try:
                                cursor.execute('SELECT id FROM alertas_stock WHERE sku = %s AND estado = "pendiente"', (combo_sku,))
                                
                                if not cursor.fetchone():
                                    cursor.execute('''
                                        INSERT INTO alertas_stock 
                                        (sku, nombre_producto, stock_fisico, stock_vendido, stock_disponible, tipo_alerta, estado, mlas_afectados)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                                    ''', (combo_sku, combo_nombre, 0, 0, 0, 'COMBO_SIN_COMPONENTE', 'pendiente', sku_sin_stock))
                            except Exception as e:
                                print(f"  ❌ Error al guardar alerta combo: {str(e)}")
                                
                except Exception as e:
                    print(f"⚠️ Error al buscar combos: {str(e)}")
        
        print(f"\n📋 Total productos sin stock: {len(productos_sin_stock) + len(combos_afectados)}")
        print("="*60 + "\n")
        
        return productos_sin_stock + combos_afectados
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return []



@app.route('/ventas/activas')
@login_required
def ventas_activas():
    """Lista de ventas activas con filtros de búsqueda"""
    try:
        # ========================================
        # OBTENER FILTROS
        # ========================================
        filtro_buscar = request.args.get('buscar', '').strip()
        filtro_tipo_entrega = request.args.get('tipo_entrega', '')
        filtro_metodo_envio = request.args.get('metodo_envio', '')
        filtro_zona = request.args.get('zona', '')
        filtro_canal = request.args.get('canal', '')
        filtro_estado_pago = request.args.get('estado_pago', '')
        
        # ========================================
        # CONSTRUIR QUERY CON FILTROS
        # ========================================
        query = '''
            SELECT 
                id, numero_venta, fecha_venta, canal, mla_code,
                nombre_cliente, telefono_cliente,
                tipo_entrega, metodo_envio, ubicacion_despacho,
                zona_envio, direccion_entrega, costo_flete,
metodo_pago, importe_total, importe_abonado,
                pago_mercadopago, pago_efectivo,
                estado_entrega, estado_pago, notas
            FROM ventas
            WHERE estado_entrega = 'pendiente'
        '''
        params = []
        
        # Filtro: Búsqueda de texto (apodo, nombre, productos)
        if filtro_buscar:
            query += '''
                AND (
                    mla_code LIKE %s 
                    OR nombre_cliente LIKE %s
                    OR id IN (
                        SELECT venta_id FROM items_venta WHERE sku LIKE %s
                    )
                )
            '''
            busqueda = f'%{filtro_buscar}%'
            params.extend([busqueda, busqueda, busqueda])
        
        # Filtro: Tipo de entrega
        if filtro_tipo_entrega:
            query += ' AND tipo_entrega = %s'
            params.append(filtro_tipo_entrega)
        
        # Filtro: Método de envío
        if filtro_metodo_envio:
            query += ' AND metodo_envio = %s'
            params.append(filtro_metodo_envio)
        
        # Filtro: Zona
        if filtro_zona:
            query += ' AND zona_envio = %s'
            params.append(filtro_zona)
        
        # Filtro: Canal
        if filtro_canal:
            query += ' AND canal = %s'
            params.append(filtro_canal)
        
        # Filtro: Estado de pago
        if filtro_estado_pago:
            if filtro_estado_pago == 'pagado':
                query += ' AND importe_abonado >= importe_total'
            elif filtro_estado_pago == 'pendiente':
                query += ' AND importe_abonado = 0'
            elif filtro_estado_pago == 'parcial':
                query += ' AND importe_abonado > 0 AND importe_abonado < importe_total'
        
        # Ordenar: más antiguas arriba
        query += ' ORDER BY id DESC'
        
        # Ejecutar query
        ventas = query_db(query, tuple(params) if params else None)
        
        # ========================================
        # OBTENER ITEMS DE CADA VENTA
        # ========================================
        for venta in ventas:
            items = query_db('''
                SELECT 
                    iv.sku, 
                    iv.cantidad, 
                    iv.precio_unitario,
                    COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
                ORDER BY iv.id
            ''', (venta['id'],))
            venta['items'] = items
        
        return render_template('ventas_activas.html', 
                             ventas=ventas,
                             filtro_buscar=filtro_buscar,
                             filtro_tipo_entrega=filtro_tipo_entrega,
                             filtro_metodo_envio=filtro_metodo_envio,
                             filtro_zona=filtro_zona,
                             filtro_canal=filtro_canal,
                             filtro_estado_pago=filtro_estado_pago)
        
    except Exception as e:
        flash(f'Error al cargar ventas activas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('index'))

@app.route('/ventas/activas/<int:venta_id>/proceso', methods=['POST'])
@login_required
def pasar_a_proceso(venta_id):
    """Pasar venta a proceso de envío (descuenta stock con verificación)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_activas'))
        
        # Obtener items
        cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
        items = cursor.fetchall()
        
        # ========================================
        # VERIFICAR STOCK ANTES DE DESCONTAR
        # ========================================
        hay_stock, errores = verificar_stock_disponible(cursor, items, venta['ubicacion_despacho'])
        
        if not hay_stock:
            # No hay stock suficiente - mostrar modal de error
            mensaje_html = f'''
                <p><strong>No se puede procesar la venta {venta["numero_venta"]}</strong></p>
                <p>Los siguientes productos no tienen stock suficiente:</p>
                <ul class="list-unstyled">
            '''
            for error in errores:
                mensaje_html += f'<li class="text-danger mb-2"><i class="bi bi-x-circle-fill"></i> {error}</li>'
            
            mensaje_html += '''
                </ul>
                <div class="alert alert-info mt-3">
                    <i class="bi bi-info-circle"></i> 
                    Por favor, <strong>carga más stock</strong> antes de procesar esta venta.
                </div>
            '''
            flash(mensaje_html, 'error_stock')
            return redirect(url_for('ventas_activas'))
        
        # ========================================
        # HAY STOCK - PROCEDER CON DESCUENTO
        # ========================================
        for item in items:
            descontar_stock_item(cursor, item, venta['ubicacion_despacho'])
        
        # Actualizar estado
        cursor.execute('''
            UPDATE ventas 
            SET estado_entrega = 'en_proceso',
                fecha_modificacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        conn.commit()
        flash(f'✅ Venta {venta["numero_venta"]} pasada a Proceso de Envío. Stock descontado.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al pasar a proceso: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_activas'))


@app.route('/ventas/activas/<int:venta_id>/entregada', methods=['POST'])
@login_required
def marcar_entregada(venta_id):
    """Marcar venta como entregada (descuenta stock si no se descontó, con verificación)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_activas'))
        
        # Si está pendiente, necesita descontar stock
        if venta['estado_entrega'] == 'pendiente':
            cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
            items = cursor.fetchall()
            
            # ========================================
            # VERIFICAR STOCK ANTES DE DESCONTAR
            # ========================================
            hay_stock, errores = verificar_stock_disponible(cursor, items, venta['ubicacion_despacho'])
            
            if not hay_stock:
                # No hay stock suficiente - mostrar modal de error
                mensaje_html = f'''
                    <p><strong>No se puede marcar como entregada la venta {venta["numero_venta"]}</strong></p>
                    <p>Los siguientes productos no tienen stock suficiente:</p>
                    <ul class="list-unstyled">
                '''
                for error in errores:
                    mensaje_html += f'<li class="text-danger mb-2"><i class="bi bi-x-circle-fill"></i> {error}</li>'
                
                mensaje_html += '''
                    </ul>
                    <div class="alert alert-info mt-3">
                        <i class="bi bi-info-circle"></i> 
                        Por favor, <strong>carga más stock</strong> antes de marcar como entregada.
                    </div>
                '''
                flash(mensaje_html, 'error_stock')
                return redirect(url_for('ventas_activas'))
            
            # HAY STOCK - DESCONTAR
            for item in items:
                descontar_stock_item(cursor, item, venta['ubicacion_despacho'])
        
        # Actualizar estado Y FECHA DE ENTREGA
        cursor.execute('''
            UPDATE ventas 
            SET estado_entrega = 'entregada',
                fecha_entrega = NOW(),
                fecha_modificacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        conn.commit()
        flash(f'✅ Venta {venta["numero_venta"]} marcada como Entregada.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al marcar como entregada: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_activas'))


@app.route('/ventas/activas/<int:venta_id>/cancelar', methods=['POST'])
@login_required
def cancelar_venta(venta_id):
    """Cancelar venta (NO descuenta stock)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT numero_venta FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_activas'))
        
        # Actualizar estado
        cursor.execute('''
            UPDATE ventas 
            SET estado_entrega = 'cancelada',
                fecha_modificacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        conn.commit()
        flash(f'✅ Venta {venta["numero_venta"]} cancelada. No se descontó stock.', 'info')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al cancelar venta: {str(e)}', 'error')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_activas'))

@app.route('/ventas/activas/<int:venta_id>/eliminar', methods=['POST'])
@login_required
def eliminar_venta(venta_id):
    """
    Eliminar venta completamente de la base de datos
    NO descuenta stock (igual que cancelar)
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener info de la venta antes de borrar
        cursor.execute('SELECT numero_venta FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('❌ Venta no encontrada', 'error')
            return redirect(url_for('ventas_activas'))
        
        numero_venta = venta['numero_venta']
        
        # 1. Eliminar items de venta
        cursor.execute('DELETE FROM items_venta WHERE venta_id = %s', (venta_id,))
        items_eliminados = cursor.rowcount
        
        # 2. Eliminar venta
        cursor.execute('DELETE FROM ventas WHERE id = %s', (venta_id,))
        
        conn.commit()
        
        flash(f'✅ Venta {numero_venta} eliminada correctamente ({items_eliminados} items borrados). No se descontó stock.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al eliminar venta: {str(e)}', 'error')
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_activas'))




# ============================================================================
# RUTAS: ACCIONES MÚLTIPLES EN VENTAS ACTIVAS
# Agregar en app.py después de las rutas individuales de activas
# ============================================================================

@app.route('/ventas/activas/pasar-proceso-multiple', methods=['POST'])
@login_required
def pasar_a_proceso_multiple():
    """
    Pasar múltiples ventas a proceso de envío
    Descuenta stock con verificación
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_activas'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        ventas_sin_stock = []
        
        for venta_id in venta_ids:
            try:
                # Obtener venta
                cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta or venta['estado_entrega'] != 'pendiente':
                    continue
                
                # Obtener items
                cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
                items = cursor.fetchall()
                
                # VERIFICAR STOCK ANTES DE DESCONTAR
                hay_stock, errores = verificar_stock_disponible(cursor, items, venta['ubicacion_despacho'])
                
                if not hay_stock:
                    # No hay stock - agregar a lista de errores
                    ventas_sin_stock.append({
                        'numero': venta['numero_venta'],
                        'errores': errores
                    })
                    continue
                
                # HAY STOCK - PROCEDER CON DESCUENTO
                for item in items:
                    descontar_stock_item(cursor, item, venta['ubicacion_despacho'])
                
                # Actualizar estado
                cursor.execute('''
                    UPDATE ventas 
                    SET estado_entrega = 'en_proceso',
                        fecha_modificacion = NOW()
                    WHERE id = %s
                ''', (venta_id,))
                
                ventas_procesadas += 1
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # MENSAJES
        if ventas_procesadas > 0:
            flash(f'✅ {ventas_procesadas} venta(s) pasadas a Proceso de Envío. Stock descontado.', 'success')
        
        if ventas_sin_stock:
            mensaje_html = '<div class="alert alert-warning"><strong>⚠️ Algunas ventas no se pudieron procesar por falta de stock:</strong><ul class="mt-2">'
            for v in ventas_sin_stock:
                mensaje_html += f'<li><strong>{v["numero"]}</strong><ul class="list-unstyled ms-3">'
                for error in v['errores']:
                    mensaje_html += f'<li class="text-danger"><small>{error}</small></li>'
                mensaje_html += '</ul></li>'
            mensaje_html += '</ul></div>'
            flash(mensaje_html, 'error_stock')
        
        if ventas_procesadas == 0 and not ventas_sin_stock:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        
        # Mantener filtros
        filtros = {}
        for key in ['buscar', 'zona', 'metodo_envio', 'tipo_entrega', 'estado_pago']:
            if request.form.get(key):
                filtros[key] = request.form.get(key)
        
        return redirect(url_for('ventas_activas', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_activas'))


@app.route('/ventas/activas/marcar-entregadas-multiple', methods=['POST'])
@login_required
def marcar_entregadas_multiple():
    """
    Marcar múltiples ventas como entregadas
    Descuenta stock si no se descontó, con verificación
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_activas'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        ventas_sin_stock = []
        
        for venta_id in venta_ids:
            try:
                # Obtener venta
                cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta or venta['estado_entrega'] != 'pendiente':
                    continue
                
                # Si está pendiente, necesita descontar stock
                cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
                items = cursor.fetchall()
                
                # VERIFICAR STOCK ANTES DE DESCONTAR
                hay_stock, errores = verificar_stock_disponible(cursor, items, venta['ubicacion_despacho'])
                
                if not hay_stock:
                    # No hay stock - agregar a lista de errores
                    ventas_sin_stock.append({
                        'numero': venta['numero_venta'],
                        'errores': errores
                    })
                    continue
                
                # HAY STOCK - DESCONTAR
                for item in items:
                    descontar_stock_item(cursor, item, venta['ubicacion_despacho'])
                
                # Actualizar estado Y FECHA DE ENTREGA
                cursor.execute('''
                    UPDATE ventas 
                    SET estado_entrega = 'entregada',
                        fecha_entrega = NOW(),
                        fecha_modificacion = NOW()
                    WHERE id = %s
                ''', (venta_id,))
                
                ventas_procesadas += 1
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # MENSAJES
        if ventas_procesadas > 0:
            flash(f'✅ {ventas_procesadas} venta(s) marcadas como Entregadas.', 'success')
        
        if ventas_sin_stock:
            mensaje_html = '<div class="alert alert-warning"><strong>⚠️ Algunas ventas no se pudieron marcar como entregadas por falta de stock:</strong><ul class="mt-2">'
            for v in ventas_sin_stock:
                mensaje_html += f'<li><strong>{v["numero"]}</strong><ul class="list-unstyled ms-3">'
                for error in v['errores']:
                    mensaje_html += f'<li class="text-danger"><small>{error}</small></li>'
                mensaje_html += '</ul></li>'
            mensaje_html += '</ul></div>'
            flash(mensaje_html, 'error_stock')
        
        if ventas_procesadas == 0 and not ventas_sin_stock:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        
        # Mantener filtros
        filtros = {}
        for key in ['buscar', 'zona', 'metodo_envio', 'tipo_entrega', 'estado_pago']:
            if request.form.get(key):
                filtros[key] = request.form.get(key)
        
        return redirect(url_for('ventas_activas', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_activas'))


@app.route('/ventas/activas/cancelar-multiple', methods=['POST'])
@login_required
def cancelar_ventas_multiple():
    """
    Cancelar múltiples ventas
    NO descuenta stock (porque son ventas pendientes)
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_activas'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        
        for venta_id in venta_ids:
            try:
                # Verificar que exista y esté pendiente
                cursor.execute('SELECT id FROM ventas WHERE id = %s AND estado_entrega = %s', (venta_id, 'pendiente'))
                venta = cursor.fetchone()
                
                if not venta:
                    continue
                
                # Actualizar estado
                cursor.execute('''
                    UPDATE ventas 
                    SET estado_entrega = 'cancelada',
                        fecha_modificacion = NOW()
                    WHERE id = %s
                ''', (venta_id,))
                
                ventas_procesadas += 1
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        if ventas_procesadas == 0:
            flash('❌ No se pudieron cancelar las ventas seleccionadas', 'error')
        else:
            flash(f'✅ {ventas_procesadas} venta(s) canceladas. No se descontó stock.', 'info')
        
        # Mantener filtros
        filtros = {}
        for key in ['buscar', 'zona', 'metodo_envio', 'tipo_entrega', 'estado_pago']:
            if request.form.get(key):
                filtros[key] = request.form.get(key)
        
        return redirect(url_for('ventas_activas', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_activas'))



# ============================================================================
# PROCESO DE ENVÍO - RUTAS ACTUALIZADAS
# Incluye: Volver a Activas + Motivo de Cancelación
# Reemplazar las rutas de /ventas/proceso en app.py
# ============================================================================

@app.route('/ventas/proceso')
@login_required
def ventas_proceso():
    """Lista de ventas en proceso de envío con filtros"""
    try:
        # ========================================
        # OBTENER FILTROS
        # ========================================
        filtro_buscar = request.args.get('buscar', '').strip()
        filtro_tipo_entrega = request.args.get('tipo_entrega', '')
        filtro_metodo_envio = request.args.get('metodo_envio', '')
        filtro_zona = request.args.get('zona', '')
        filtro_canal = request.args.get('canal', '')
        filtro_estado_pago = request.args.get('estado_pago', '')
        
        # ========================================
        # CONSTRUIR QUERY CON FILTROS
        # ========================================
        query = '''
            SELECT 
                id, numero_venta, fecha_venta, canal, mla_code,
                nombre_cliente, telefono_cliente,
                tipo_entrega, metodo_envio, ubicacion_despacho,
                zona_envio, direccion_entrega, costo_flete,
metodo_pago, importe_total, importe_abonado,
                pago_mercadopago, pago_efectivo,
                estado_entrega, estado_pago, notas
            FROM ventas
            WHERE estado_entrega = 'en_proceso'
        '''
        params = []
        
        # Filtro: Búsqueda de texto
        if filtro_buscar:
            query += '''
                AND (
                    mla_code LIKE %s 
                    OR nombre_cliente LIKE %s
                    OR id IN (
                        SELECT venta_id FROM items_venta WHERE sku LIKE %s
                    )
                )
            '''
            busqueda = f'%{filtro_buscar}%'
            params.extend([busqueda, busqueda, busqueda])
        
        # Filtro: Tipo de entrega
        if filtro_tipo_entrega:
            query += ' AND tipo_entrega = %s'
            params.append(filtro_tipo_entrega)
        
        # Filtro: Método de envío
        if filtro_metodo_envio:
            query += ' AND metodo_envio = %s'
            params.append(filtro_metodo_envio)
        
        # Filtro: Zona
        if filtro_zona:
            query += ' AND zona_envio = %s'
            params.append(filtro_zona)
        
        # Filtro: Canal
        if filtro_canal:
            query += ' AND canal = %s'
            params.append(filtro_canal)
        
        # Filtro: Estado de pago
        if filtro_estado_pago:
            if filtro_estado_pago == 'pagado':
                query += ' AND importe_abonado >= importe_total'
            elif filtro_estado_pago == 'pendiente':
                query += ' AND importe_abonado = 0'
            elif filtro_estado_pago == 'parcial':
                query += ' AND importe_abonado > 0 AND importe_abonado < importe_total'
        
        # Ordenar: más antiguas arriba
        query += ' ORDER BY id DESC'
        
        # Ejecutar query
        ventas = query_db(query, tuple(params) if params else None)
        
        # ========================================
        # OBTENER ITEMS DE CADA VENTA
        # ========================================
        for venta in ventas:
            items = query_db('''
                SELECT 
                    iv.sku, 
                    iv.cantidad, 
                    iv.precio_unitario,
                    COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
                ORDER BY iv.id
            ''', (venta['id'],))
            venta['items'] = items
        
        return render_template('proceso_envio.html', 
                             ventas=ventas,
                             filtro_buscar=filtro_buscar,
                             filtro_tipo_entrega=filtro_tipo_entrega,
                             filtro_metodo_envio=filtro_metodo_envio,
                             filtro_zona=filtro_zona,
                             filtro_canal=filtro_canal,
                             filtro_estado_pago=filtro_estado_pago)
        
    except Exception as e:
        flash(f'Error al cargar proceso de envío: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('index'))


@app.route('/ventas/proceso/<int:venta_id>/volver_activas', methods=['POST'])
@login_required
def proceso_volver_activas(venta_id):
    """Volver venta de proceso a activas (devuelve stock)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_proceso'))
        
        if venta['estado_entrega'] != 'en_proceso':
            flash('La venta no está en proceso', 'warning')
            return redirect(url_for('ventas_proceso'))
        
        # Obtener items de la venta
        cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
        items = cursor.fetchall()
        
        # DEVOLVER STOCK (porque ya se había descontado)
        for item in items:
            devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
        
        # Actualizar estado a 'pendiente' (volver a activas)
        cursor.execute('''
            UPDATE ventas 
            SET estado_entrega = 'pendiente',
                fecha_modificacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        conn.commit()
        flash(f'✅ Venta {venta["numero_venta"]} devuelta a Ventas Activas. Stock restaurado.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al volver a activas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_proceso'))


@app.route('/ventas/proceso/<int:venta_id>/entregada', methods=['POST'])
@login_required
def proceso_marcar_entregada(venta_id):
    """Marcar venta en proceso como entregada (stock ya descontado)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT numero_venta, estado_entrega FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_proceso'))
        
        if venta['estado_entrega'] != 'en_proceso':
            flash('La venta no está en proceso', 'warning')
            return redirect(url_for('ventas_proceso'))
        
        # Actualizar estado Y FECHA DE ENTREGA (NO descuenta stock, ya se descontó)
        cursor.execute('''
            UPDATE ventas 
            SET estado_entrega = 'entregada',
                fecha_entrega = NOW(),
                fecha_modificacion = NOW()
            WHERE id = %s
        ''', (venta_id,))
        
        conn.commit()
        flash(f'✅ Venta {venta["numero_venta"]} marcada como Entregada.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_proceso'))


@app.route('/ventas/proceso/<int:venta_id>/cancelar', methods=['POST'])
@login_required
def proceso_cancelar_devolver(venta_id):
    """Cancelar venta en proceso y DEVOLVER stock descontado (con motivo opcional)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_proceso'))
        
        if venta['estado_entrega'] != 'en_proceso':
            flash('La venta no está en proceso', 'warning')
            return redirect(url_for('ventas_proceso'))
        
        # Obtener motivo de cancelación (opcional)
        motivo_cancelacion = request.form.get('motivo_cancelacion', '').strip()
        
        # Obtener items de la venta
        cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
        items = cursor.fetchall()
        
        # DEVOLVER STOCK (lo opuesto a descontar)
        for item in items:
            devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
        
        # Actualizar estado y agregar motivo si existe
        if motivo_cancelacion:
            # Agregar motivo a las notas existentes
            notas_actuales = venta.get('notas', '') or ''
            notas_nuevas = f"{notas_actuales}\n[CANCELACIÓN] {motivo_cancelacion}".strip()
            
            cursor.execute('''
                UPDATE ventas 
                SET estado_entrega = 'cancelada',
                    notas = %s,
                    fecha_modificacion = NOW()
                WHERE id = %s
            ''', (notas_nuevas, venta_id))
        else:
            cursor.execute('''
                UPDATE ventas 
                SET estado_entrega = 'cancelada',
                    fecha_modificacion = NOW()
                WHERE id = %s
            ''', (venta_id,))
        
        conn.commit()
        
        mensaje = f'✅ Venta {venta["numero_venta"]} cancelada. Stock devuelto correctamente.'
        if motivo_cancelacion:
            mensaje += f' Motivo: {motivo_cancelacion}'
        
        flash(mensaje, 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al cancelar y devolver stock: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_proceso'))

# ============================================================================
# RUTAS: ACCIONES MÚLTIPLES EN VENTAS EN PROCESO
# Agregar en app.py después de las rutas individuales de proceso
# ============================================================================

@app.route('/ventas/proceso/volver-activas-multiple', methods=['POST'])
@login_required
def proceso_volver_activas_multiple():
    """
    Volver múltiples ventas en proceso a activas
    Devuelve stock de todas
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_proceso'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        
        for venta_id in venta_ids:
            try:
                # Obtener venta
                cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta or venta['estado_entrega'] != 'en_proceso':
                    continue
                
                # Obtener items
                cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
                items = cursor.fetchall()
                
                # DEVOLVER STOCK (porque ya se había descontado)
                for item in items:
                    devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
                
                # Actualizar estado a 'pendiente' (volver a activas)
                cursor.execute('''
                    UPDATE ventas 
                    SET estado_entrega = 'pendiente',
                        fecha_modificacion = NOW()
                    WHERE id = %s
                ''', (venta_id,))
                
                ventas_procesadas += 1
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        if ventas_procesadas == 0:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        else:
            flash(f'✅ {ventas_procesadas} venta(s) devueltas a Ventas Activas. Stock restaurado.', 'success')
        
        # Mantener filtros
        filtros = {}
        if request.form.get('buscar'):
            filtros['buscar'] = request.form.get('buscar')
        if request.form.get('zona'):
            filtros['zona'] = request.form.get('zona')
        if request.form.get('metodo_envio'):
            filtros['metodo_envio'] = request.form.get('metodo_envio')
        
        return redirect(url_for('ventas_proceso', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_proceso'))


@app.route('/ventas/proceso/marcar-entregadas-multiple', methods=['POST'])
@login_required
def proceso_marcar_entregadas_multiple():
    """
    Marcar múltiples ventas en proceso como entregadas
    NO toca stock (ya estaba descontado)
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_proceso'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        
        for venta_id in venta_ids:
            try:
                # Verificar que esté en proceso
                cursor.execute('SELECT estado_entrega FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta or venta['estado_entrega'] != 'en_proceso':
                    continue
                
                # Actualizar estado Y FECHA DE ENTREGA (NO descuenta stock)
                cursor.execute('''
                    UPDATE ventas 
                    SET estado_entrega = 'entregada',
                        fecha_entrega = NOW(),
                        fecha_modificacion = NOW()
                    WHERE id = %s
                ''', (venta_id,))
                
                ventas_procesadas += 1
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        if ventas_procesadas == 0:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        else:
            flash(f'✅ {ventas_procesadas} venta(s) marcadas como Entregadas.', 'success')
        
        # Mantener filtros
        filtros = {}
        if request.form.get('buscar'):
            filtros['buscar'] = request.form.get('buscar')
        if request.form.get('zona'):
            filtros['zona'] = request.form.get('zona')
        if request.form.get('metodo_envio'):
            filtros['metodo_envio'] = request.form.get('metodo_envio')
        
        return redirect(url_for('ventas_proceso', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_proceso'))


@app.route('/ventas/proceso/cancelar-multiple', methods=['POST'])
@login_required
def proceso_cancelar_multiple():
    """
    Cancelar múltiples ventas en proceso y DEVOLVER stock
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        motivo_cancelacion = request.form.get('motivo_cancelacion', '').strip()
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_proceso'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        
        for venta_id in venta_ids:
            try:
                # Obtener venta
                cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta or venta['estado_entrega'] != 'en_proceso':
                    continue
                
                # Obtener items
                cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
                items = cursor.fetchall()
                
                # DEVOLVER STOCK
                for item in items:
                    devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
                
                # Actualizar estado y agregar motivo si existe
                if motivo_cancelacion:
                    notas_actuales = venta.get('notas', '') or ''
                    notas_nuevas = f"{notas_actuales}\n[CANCELACIÓN MÚLTIPLE] {motivo_cancelacion}".strip()
                    
                    cursor.execute('''
                        UPDATE ventas 
                        SET estado_entrega = 'cancelada',
                            notas = %s,
                            fecha_modificacion = NOW()
                        WHERE id = %s
                    ''', (notas_nuevas, venta_id))
                else:
                    cursor.execute('''
                        UPDATE ventas 
                        SET estado_entrega = 'cancelada',
                            fecha_modificacion = NOW()
                        WHERE id = %s
                    ''', (venta_id,))
                
                ventas_procesadas += 1
            
            except Exception as e:
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        if ventas_procesadas == 0:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        else:
            mensaje = f'✅ {ventas_procesadas} venta(s) canceladas. Stock devuelto correctamente.'
            if motivo_cancelacion:
                mensaje += f' Motivo: {motivo_cancelacion}'
            flash(mensaje, 'success')
        
        # Mantener filtros
        filtros = {}
        if request.form.get('buscar'):
            filtros['buscar'] = request.form.get('buscar')
        if request.form.get('zona'):
            filtros['zona'] = request.form.get('zona')
        if request.form.get('metodo_envio'):
            filtros['metodo_envio'] = request.form.get('metodo_envio')
        
        return redirect(url_for('ventas_proceso', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_proceso'))


@app.route('/ventas/editar/<int:venta_id>', methods=['GET', 'POST'])
@login_required
def editar_venta(venta_id):
    """Editar una venta activa"""
    
    if request.method == 'GET':
        # Obtener datos de la venta
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Venta principal
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_activas'))
        
        # CORREGIDO: Convertir fecha a formato string para input type="date"
        if venta['fecha_venta']:
            from datetime import datetime
            if isinstance(venta['fecha_venta'], datetime):
                venta['fecha_venta_str'] = venta['fecha_venta'].strftime('%Y-%m-%d')
            elif hasattr(venta['fecha_venta'], 'strftime'):  # date object
                venta['fecha_venta_str'] = venta['fecha_venta'].strftime('%Y-%m-%d')
            else:
                venta['fecha_venta_str'] = str(venta['fecha_venta'])
        else:
            venta['fecha_venta_str'] = ''
        
        # Items de la venta
        cursor.execute('''
            SELECT iv.*, 
                   COALESCE(pb.nombre, pc.nombre) as nombre_producto
            FROM items_venta iv
            LEFT JOIN productos_base pb ON iv.sku = pb.sku
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            WHERE iv.venta_id = %s
        ''', (venta_id,))
        items = cursor.fetchall()
        
        # Productos disponibles para agregar
        cursor.execute('''
            SELECT sku, nombre, tipo FROM productos_base
            UNION
            SELECT sku, nombre, 'sommier' as tipo FROM productos_compuestos WHERE activo = 1
            ORDER BY nombre
        ''')
        productos = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return render_template('editar_venta.html', 
                             venta=venta, 
                             items=items,
                             productos=productos)
    
    # POST - Guardar cambios
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener items ANTES del cambio (para comparar)
        cursor.execute('SELECT sku, cantidad FROM items_venta WHERE venta_id = %s', (venta_id,))
        items_anteriores = {item['sku']: item['cantidad'] for item in cursor.fetchall()}
        
        # ========================================
        # 1. ACTUALIZAR DATOS GENERALES
        # ========================================
        numero_venta = request.form.get('numero_venta')
        fecha_venta = request.form.get('fecha_venta')
        canal = request.form.get('canal')
        mla_code = request.form.get('mla_code', '').strip()
        nombre_cliente = request.form.get('nombre_cliente', '').strip()
        telefono_cliente = request.form.get('telefono_cliente', '')
        
        # Entrega
        tipo_entrega = request.form.get('tipo_entrega')
        direccion_entrega = request.form.get('direccion_entrega', '')
        metodo_envio = request.form.get('metodo_envio', '')
        zona_envio = request.form.get('zona_envio', '')
        
        # Calcular ubicación de despacho
        if metodo_envio == 'Full':
            ubicacion_despacho = 'FULL'
        else:
            ubicacion_despacho = 'DEP'
        
        responsable_entrega = request.form.get('responsable_entrega', '')
        costo_flete = float(request.form.get('costo_flete', 0))
        
        # ========================================
        # 2. PRODUCTOS Y CÁLCULO DE TOTAL
        # ========================================
        # Borrar items existentes
        cursor.execute('DELETE FROM items_venta WHERE venta_id = %s', (venta_id,))
        
        # Insertar items nuevos y calcular total
        productos = request.form.to_dict(flat=False)
        items_nuevos = {}
        importe_total = 0  # CORREGIDO: Calcular desde productos
        
        for key in productos.keys():
            if key.startswith('productos[') and key.endswith('[sku]'):
                index = key.split('[')[1].split(']')[0]
                sku = productos.get(f'productos[{index}][sku]', [None])[0]
                cantidad = int(productos.get(f'productos[{index}][cantidad]', [0])[0])
                precio = float(productos.get(f'productos[{index}][precio]', [0])[0])
                
                if sku and cantidad > 0:
                    # Insertar item
                    cursor.execute('''
                        INSERT INTO items_venta (venta_id, sku, cantidad, precio_unitario)
                        VALUES (%s, %s, %s, %s)
                    ''', (venta_id, sku, cantidad, precio))
                    
                    items_nuevos[sku] = cantidad
                    
                    # CORREGIDO: Sumar al total
                    importe_total += (cantidad * precio)
        
        # ========================================
        # 3. PAGO (INDEPENDIENTE DEL TOTAL)
        # ========================================
        metodo_pago = request.form.get('metodo_pago')
        # CORREGIDO: Los pagos son independientes del total calculado
        pago_mercadopago = float(request.form.get('pago_mercadopago', 0))
        pago_efectivo = float(request.form.get('pago_efectivo', 0))
        importe_abonado = pago_mercadopago + pago_efectivo
        
        # Estado pago
        if importe_abonado >= importe_total:
            estado_pago = 'pagado'
        elif importe_abonado > 0:
            estado_pago = 'pago_parcial'
        else:
            estado_pago = 'pago_pendiente'
        
        # Notas
        notas = request.form.get('notas', '')
        
        # ========================================
        # 4. ACTUALIZAR VENTA
        # ========================================
        cursor.execute('''
            UPDATE ventas SET
                numero_venta = %s,
                fecha_venta = %s,
                canal = %s,
                mla_code = %s,
                nombre_cliente = %s,
                telefono_cliente = %s,
                tipo_entrega = %s,
                metodo_envio = %s,
                ubicacion_despacho = %s,
                zona_envio = %s,
                direccion_entrega = %s,
                responsable_entrega = %s,
                costo_flete = %s,
                metodo_pago = %s,
                importe_total = %s,
                importe_abonado = %s,
                pago_mercadopago = %s,
                pago_efectivo = %s,
                estado_pago = %s,
                notas = %s
            WHERE id = %s
        ''', (
            numero_venta, fecha_venta, canal, mla_code,
            nombre_cliente, telefono_cliente,
            tipo_entrega, metodo_envio, ubicacion_despacho,
            zona_envio, direccion_entrega, responsable_entrega,
            costo_flete, metodo_pago, importe_total, importe_abonado,
            pago_mercadopago, pago_efectivo,
            estado_pago, notas,
            venta_id
        ))
        
        # ========================================
        # 5. DETECTAR ALERTAS (Solo si cambiaron items)
        # ========================================
        items_cambiaron = items_anteriores != items_nuevos
        
        if items_cambiaron:
            try:
                items_vendidos_lista = [{'sku': sku, 'cantidad': cant} for sku, cant in items_nuevos.items()]
                if items_vendidos_lista:
                    productos_sin_stock = detectar_alertas_stock_bajo(cursor, items_vendidos_lista, venta_id)
                    
                    if productos_sin_stock:
                        # Mostrar alerta simplificada
                        productos_base = [p for p in productos_sin_stock if p.get('tipo_producto') == 'base']
                        if productos_base:
                            nombres = ', '.join([p['nombre'] for p in productos_base[:3]])
                            flash(f'⚠️ Productos sin stock: {nombres}', 'warning')
            except Exception as e:
                print(f"Error al detectar alertas: {str(e)}")
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash(f'✅ Venta {numero_venta} actualizada correctamente. Total: ${importe_total:,.0f}', 'success')
        return redirect(url_for('ventas_activas'))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        import traceback
        traceback.print_exc()
        flash(f'❌ Error al actualizar venta: {str(e)}', 'error')
        return redirect(url_for('ventas_activas'))



# ============================================================================
# FUNCIÓN AUXILIAR: DEVOLVER STOCK (OPUESTO A DESCONTAR)
# ============================================================================

def devolver_stock_item(cursor, item, ubicacion_despacho):
    """
    Devuelve stock de un item (lo opuesto a descontar_stock_item)
    Considera:
    - Ubicación de despacho (DEP o FULL)
    - Descomposición de combos
    - Bases grandes (160, 180, 200) devuelven 2 bases chicas
    """
    sku = item['sku']
    cantidad = item['cantidad']
    
    # Verificar si es un combo
    cursor.execute('SELECT id FROM productos_compuestos WHERE sku = %s', (sku,))
    combo = cursor.fetchone()
    
    if combo:
        # Es combo: descomponer y devolver componentes
        cursor.execute('''
            SELECT pb.sku, pb.tipo, c.cantidad_necesaria
            FROM componentes c
            JOIN productos_base pb ON c.producto_base_id = pb.id
            WHERE c.producto_compuesto_id = %s
        ''', (combo['id'],))
        componentes = cursor.fetchall()
        
        for comp in componentes:
            sku_comp = comp['sku']
            cant_comp = comp['cantidad_necesaria'] * cantidad
            tipo_comp = comp['tipo']
            
            # Devolver componente según ubicación
            devolver_stock_simple(cursor, sku_comp, cant_comp, tipo_comp, ubicacion_despacho)
    
    else:
        # Es producto simple
        cursor.execute('SELECT tipo FROM productos_base WHERE sku = %s', (sku,))
        prod = cursor.fetchone()
        tipo = prod['tipo'] if prod else 'colchon'
        
        devolver_stock_simple(cursor, sku, cantidad, tipo, ubicacion_despacho)


def devolver_stock_simple(cursor, sku, cantidad, tipo, ubicacion_despacho):
    """
    Devuelve stock de un producto simple según ubicación (SUMA en lugar de RESTAR)
    """
    # COMPAC: tiene _DEP y _FULL
    if '_DEP' in sku or '_FULL' in sku:
        if ubicacion_despacho == 'FULL':
            # Devolver a _FULL
            sku_real = sku.replace('_DEP', '_FULL')
        else:
            # Devolver a _DEP
            sku_real = sku.replace('_FULL', '_DEP')
        
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual + %s 
            WHERE sku = %s
        ''', (cantidad, sku_real))
    
    # ALMOHADAS: tienen stock_actual (DEP) y stock_full (FULL)
    elif tipo == 'almohada':
        if ubicacion_despacho == 'FULL':
            cursor.execute('''
                UPDATE productos_base 
                SET stock_full = stock_full + %s 
                WHERE sku = %s
            ''', (cantidad, sku))
        else:
            cursor.execute('''
                UPDATE productos_base 
                SET stock_actual = stock_actual + %s 
                WHERE sku = %s
            ''', (cantidad, sku))
    
    # BASES CHICAS (80200, 90200, 100200): devolver directamente
    elif tipo == 'base' and any(x in sku for x in ['80200', '90200', '100200']):
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual + %s 
            WHERE sku = %s
        ''', (cantidad, sku))
    
    # BASES GRANDES (160, 180, 200): devuelven 2 bases chicas
    elif tipo == 'base' and any(x in sku for x in ['160', '180', '200']):
        # Determinar SKU de bases chicas
        if '160' in sku:
            sku_chica = sku.replace('160', '80200')
            cant_bases = cantidad * 2
        elif '180' in sku:
            sku_chica = sku.replace('180', '90200')
            cant_bases = cantidad * 2
        elif '200' in sku:
            sku_chica = sku.replace('200', '100200')
            cant_bases = cantidad * 2
        else:
            return
        
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual + %s 
            WHERE sku = %s
        ''', (cant_bases, sku_chica))
    
    # OTROS: devolver a stock_actual
    else:
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual + %s 
            WHERE sku = %s
        ''', (cantidad, sku))



# ============================================================================
# FUNCIÓN AUXILIAR: DESCONTAR STOCK
# ============================================================================

def descontar_stock_item(cursor, item, ubicacion_despacho):
    """
    Descuenta stock de un item considerando:
    - Ubicación de despacho (DEP o FULL)
    - Descomposición de combos
    - Bases grandes (160, 180, 200) descuentan 2 bases chicas
    """
    sku = item['sku']
    cantidad = item['cantidad']
    
    # Verificar si es un combo
    cursor.execute('SELECT id FROM productos_compuestos WHERE sku = %s', (sku,))
    combo = cursor.fetchone()
    
    if combo:
        # Es combo: descomponer y descontar componentes
        cursor.execute('''
            SELECT pb.sku, pb.tipo, c.cantidad_necesaria
            FROM componentes c
            JOIN productos_base pb ON c.producto_base_id = pb.id
            WHERE c.producto_compuesto_id = %s
        ''', (combo['id'],))
        componentes = cursor.fetchall()
        
        for comp in componentes:
            sku_comp = comp['sku']
            cant_comp = comp['cantidad_necesaria'] * cantidad
            tipo_comp = comp['tipo']
            
            # Descontar componente según ubicación
            descontar_stock_simple(cursor, sku_comp, cant_comp, tipo_comp, ubicacion_despacho)
    
    else:
        # Es producto simple
        cursor.execute('SELECT tipo FROM productos_base WHERE sku = %s', (sku,))
        prod = cursor.fetchone()
        tipo = prod['tipo'] if prod else 'colchon'
        
        descontar_stock_simple(cursor, sku, cantidad, tipo, ubicacion_despacho)


def descontar_stock_simple(cursor, sku, cantidad, tipo, ubicacion_despacho):
    """Descuenta stock de un producto simple según ubicación"""
    
    # COMPAC: tiene _DEP y _FULL
    if '_DEP' in sku or '_FULL' in sku:
        if ubicacion_despacho == 'FULL':
            sku_real = sku.replace('_DEP', '_FULL')
        else:
            sku_real = sku.replace('_FULL', '_DEP')
        
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual - %s 
            WHERE sku = %s
        ''', (cantidad, sku_real))
    
    # ALMOHADAS: tienen stock_actual (DEP) y stock_full (FULL)
    elif tipo == 'almohada':
        if ubicacion_despacho == 'FULL':
            cursor.execute('''
                UPDATE productos_base 
                SET stock_full = stock_full - %s 
                WHERE sku = %s
            ''', (cantidad, sku))
        else:
            cursor.execute('''
                UPDATE productos_base 
                SET stock_actual = stock_actual - %s 
                WHERE sku = %s
            ''', (cantidad, sku))
    
    # BASES CHICAS (80200, 90200, 100200): descontar directamente
    elif tipo == 'base' and any(x in sku for x in ['80200', '90200', '100200']):
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual - %s 
            WHERE sku = %s
        ''', (cantidad, sku))
    
    # BASES GRANDES (160, 180, 200): descuentan 2 bases chicas
    elif tipo == 'base' and any(x in sku for x in ['160', '180', '200']):
        if '160' in sku:
            sku_chica = sku.replace('160', '80200')
            cant_bases = cantidad * 2
        elif '180' in sku:
            sku_chica = sku.replace('180', '90200')
            cant_bases = cantidad * 2
        elif '200' in sku:
            sku_chica = sku.replace('200', '100200')
            cant_bases = cantidad * 2
        
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual - %s 
            WHERE sku = %s
        ''', (cant_bases, sku_chica))
    
    # OTROS: descontar de stock_actual
    else:
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = stock_actual - %s 
            WHERE sku = %s
        ''', (cantidad, sku))




# ─── FUNCIÓN HELPER: Actualizar stock en ML ───
def actualizar_stock_ml(mla_id, cantidad, access_token):
    """
    Actualizar stock de una publicación en Mercado Libre
    
    Args:
        mla_id: ID de la publicación (ej: MLA603027006)
        cantidad: Nueva cantidad de stock (0 para pausar ventas)
        access_token: Token de ML
    
    Returns:
        (success: bool, message: str)
    """
    try:
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        data = {
            "available_quantity": cantidad
        }
        
        response = requests.put(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers,
            json=data
        )
        
        if response.status_code == 200:
            return True, f"Stock actualizado a {cantidad} en ML"
        else:
            error_data = response.json()
            error_msg = error_data.get('message', 'Error desconocido')
            return False, f"Error ML: {error_msg}"
    
    except Exception as e:
        return False, f"Error: {str(e)}"


# ─── FUNCIÓN HELPER: Pausar publicación en ML ───
def pausar_publicacion_ml(mla_id, access_token):
    """
    Pausar una publicación en Mercado Libre
    
    Args:
        mla_id: ID de la publicación
        access_token: Token de ML
    
    Returns:
        (success: bool, message: str)
    """
    try:
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        data = {
            "status": "paused"
        }
        
        response = requests.put(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers,
            json=data
        )
        
        if response.status_code == 200:
            return True, "Publicación pausada en ML"
        else:
            error_data = response.json()
            error_msg = error_data.get('message', 'Error desconocido')
            return False, f"Error ML: {error_msg}"
    
    except Exception as e:
        return False, f"Error: {str(e)}"


# ─── RUTA NUEVA: Sincronizar stock con ML desde alertas ───
@app.route('/alertas/<int:alerta_id>/sincronizar-ml', methods=['POST'])
@login_required
def sincronizar_ml_desde_alerta(alerta_id):
    """
    Poner stock en 0 en las publicaciones NORMALES (sin Z)
    Procesa solo la parte NORMAL de la alerta
    """
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay token de ML configurado', 'error')
        return redirect(url_for('alertas_ml'))
    
    try:
        # Obtener la alerta
        alerta = query_db('SELECT * FROM alertas_stock WHERE id = %s', (alerta_id,))
        if alerta:
            alerta = alerta[0]
        
        if not alerta:
            flash('❌ Alerta no encontrada', 'error')
            return redirect(url_for('alertas_ml'))
        
        sku = alerta['sku']
        
        # Obtener publicaciones NORMALES (sin Z)
        publicaciones = query_db(
            'SELECT * FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE',
            (sku,)
        )
        
        if not publicaciones:
            flash(f'⚠️ No hay publicaciones de ML mapeadas para el SKU {sku}', 'warning')
            return redirect(url_for('alertas_ml'))
        
        # Actualizar stock en cada publicación
        resultados = []
        errores = []
        
        for pub in publicaciones:
            mla_id = pub['mla_id']
            success, message = actualizar_stock_ml(mla_id, 0, access_token)
            
            if success:
                resultados.append(f"{mla_id}: {message}")
            else:
                errores.append(f"{mla_id}: {message}")
        
        # ✅ LÓGICA DE PROCESAMIENTO INDEPENDIENTE
        sku_con_z = f"{sku}Z"
        tiene_variante_z = len(query_db(
            'SELECT 1 FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE LIMIT 1',
            (sku_con_z,)
        )) > 0
        
        tipo_procesado_actual = alerta.get('tipo_procesado')
        
        if tiene_variante_z:
            # Tiene variante Z - solo marcar que procesamos la parte normal
            if tipo_procesado_actual == 'z':
                # Ya se había procesado Z → ahora marcar como ambos y cerrar alerta
                execute_db(
                    "UPDATE alertas_stock SET tipo_procesado = 'ambos', estado = 'procesada', fecha_procesada = NOW() WHERE id = %s",
                    (alerta_id,)
                )
            else:
                # Solo marcar que se procesó la parte normal
                execute_db(
                    "UPDATE alertas_stock SET tipo_procesado = 'normal' WHERE id = %s",
                    (alerta_id,)
                )
        else:
            # No tiene variante Z - cerrar la alerta directamente
            execute_db(
                "UPDATE alertas_stock SET tipo_procesado = 'ambos', estado = 'procesada', fecha_procesada = NOW() WHERE id = %s",
                (alerta_id,)
            )
        
        # Mostrar resultados
        if resultados:
            flash(f'✅ Stock actualizado en ML: {", ".join(resultados)}', 'success')
        
        if errores:
            flash(f'❌ Errores: {", ".join(errores)}', 'error')
    
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('alertas_ml'))


# ─── RUTA ACTUALIZADA: alertas_ml con info de publicaciones ───
@app.route('/alertas')
@login_required
def alertas_ml():
    """Ver alertas de stock pendientes con info de publicaciones ML (normales y con Z)"""
    alertas = []
    try:
        alertas_raw = query_db('''
            SELECT * FROM alertas_stock 
            WHERE estado = 'pendiente'
            ORDER BY stock_disponible ASC, fecha_creacion DESC
        ''')
        
        # Enriquecer cada alerta con info de publicaciones ML
        for alerta in alertas_raw:
            sku = alerta['sku']
            
            # Publicaciones normales (sin Z)
            publicaciones = query_db(
                'SELECT mla_id, titulo_ml FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE',
                (sku,)
            )
            
            # Publicaciones con variante Z
            sku_con_z = f"{sku}Z"
            publicaciones_z = query_db(
                'SELECT mla_id, titulo_ml FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE',
                (sku_con_z,)
            )
            
            alerta['publicaciones_ml'] = publicaciones
            alerta['tiene_ml'] = len(publicaciones) > 0
            alerta['publicaciones_ml_z'] = publicaciones_z
            alerta['tiene_ml_z'] = len(publicaciones_z) > 0
            alertas.append(alerta)
    
    except Exception as e:
        flash(f'Error al cargar alertas: {str(e)}', 'error')
    
    return render_template('alertas.html', alertas=alertas)



@app.route('/alertas/<int:alerta_id>/procesar', methods=['POST'])
@login_required
def marcar_alerta_procesada(alerta_id):
    """Marcar una alerta como procesada"""
    try:
        execute_db(
            "UPDATE alertas_stock SET estado = 'procesada', fecha_procesada = NOW() WHERE id = %s",
            (alerta_id,)
        )
        flash('✅ Alerta marcada como procesada', 'success')
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
    
    return redirect(url_for('alertas_ml'))


@app.route('/alertas/marcar-todas-procesadas', methods=['POST'])
@login_required
def marcar_todas_procesadas():
    """Marcar TODAS las alertas pendientes como procesadas"""
    try:
        result = execute_db(
            "UPDATE alertas_stock SET estado = 'procesada', fecha_procesada = NOW() WHERE estado = 'pendiente'"
        )
        flash('✅ Todas las alertas fueron marcadas como procesadas', 'success')
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
    
    return redirect(url_for('alertas_ml'))

@app.route('/stock')
@login_required
def ver_stock():
    """Ver stock disponible con filtros - PRODUCTOS BASE + COMBOS"""
    productos = []
    filtro_estado = request.args.get('estado', 'TODOS')
    filtro_tipo = request.args.get('tipo', 'TODOS')
    filtro_modelo = request.args.get('modelo', 'TODOS')
    filtro_medida = request.args.get('medida', 'TODAS')
    buscar = request.args.get('buscar', '').strip()
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # ============================================
        # 1. VENTAS ACTIVAS
        # ============================================
        cursor.execute('''
            SELECT 
                COALESCE(pb_comp.sku, iv.sku) as sku,
                SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
            FROM items_venta iv
            JOIN ventas v ON iv.venta_id = v.id
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
            LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
            WHERE v.estado_entrega = 'pendiente'
            GROUP BY sku
        ''')
        ventas_activas = cursor.fetchall()
        ventas_dict = {v['sku']: int(v['vendido']) for v in ventas_activas}
        
        # ============================================
        # 2. PRODUCTOS BASE
        # ============================================
        query_productos = """
            SELECT 
                sku, nombre, tipo, medida, modelo,
                stock_actual, COALESCE(stock_full, 0) as stock_full
            FROM productos_base
            WHERE 1=1
        """
        params = []
        
        if filtro_tipo != 'TODOS' and filtro_tipo in ('colchon', 'base', 'almohada'):
            query_productos += " AND tipo = %s"
            params.append(filtro_tipo)
        
        if filtro_medida != 'TODAS':
            query_productos += " AND medida = %s"
            params.append(filtro_medida)
        
        if filtro_modelo != 'TODOS':
            query_productos += " AND (nombre LIKE %s OR modelo LIKE %s)"
            params.append(f'%{filtro_modelo}%')
            params.append(f'%{filtro_modelo}%')
        
        if buscar:
            query_productos += " AND (sku LIKE %s OR nombre LIKE %s)"
            params.append(f'%{buscar}%')
            params.append(f'%{buscar}%')
        
        cursor.execute(query_productos, tuple(params) if params else None)
        productos_base = cursor.fetchall()
        
        for prod in productos_base:
            sku = prod['sku']
            stock_fisico = prod['stock_actual'] + prod['stock_full']
            vendido = ventas_dict.get(sku, 0)
            stock_disponible = stock_fisico - vendido
            
            if stock_disponible <= 0:
                estado_stock = 'SIN_STOCK'
            elif stock_disponible <= 2:
                estado_stock = 'POCO_STOCK'
            else:
                estado_stock = 'DISPONIBLE'
            
            if filtro_estado != 'TODOS' and estado_stock != filtro_estado:
                continue
            
            productos.append({
                'sku': sku,
                'nombre': prod['nombre'],
                'tipo': prod['tipo'],
                'medida': prod['medida'],
                'modelo': prod['modelo'],
                'stock_fisico': stock_fisico,
                'stock_disponible': stock_disponible,
                'estado_stock': estado_stock,
                'es_combo': False
            })
        
        # ============================================
        # 3. COMBOS
        # ============================================
        if filtro_tipo == 'TODOS' or filtro_tipo == 'sommier':
            # Consultar solo columnas que existen
            query_combos = """
                SELECT id, sku, nombre
                FROM productos_compuestos
                WHERE activo = 1
            """
            params_combos = []
            
            # Filtros básicos
            if filtro_modelo != 'TODOS':
                query_combos += " AND nombre LIKE %s"
                params_combos.append(f'%{filtro_modelo}%')
            
            if buscar:
                query_combos += " AND (sku LIKE %s OR nombre LIKE %s)"
                params_combos.append(f'%{buscar}%')
                params_combos.append(f'%{buscar}%')
            
            cursor.execute(query_combos, tuple(params_combos) if params_combos else None)
            combos = cursor.fetchall()
            
            for combo in combos:
                combo_id = combo['id']
                combo_sku = combo['sku']
                combo_nombre = combo['nombre']
                
                # Extraer medida del SKU (ej: SEX140 -> 140)
                medida_combo = None
                if combo_sku[-3:].isdigit():
                    medida_num = combo_sku[-3:]
                    if medida_num in ['080', '090', '100', '140', '150', '160', '180', '200']:
                        medida_combo = medida_num.lstrip('0') + 'x190'  # Ej: 140x190
                
                # Filtro por medida (si aplica)
                if filtro_medida != 'TODAS':
                    if not medida_combo or not medida_combo.startswith(filtro_medida.split('x')[0]):
                        continue
                
                # Obtener componentes
                cursor.execute('''
                    SELECT pb.sku, pb.stock_actual, COALESCE(pb.stock_full, 0) as stock_full,
                           c.cantidad_necesaria
                    FROM componentes c
                    JOIN productos_base pb ON c.producto_base_id = pb.id
                    WHERE c.producto_compuesto_id = %s
                ''', (combo_id,))
                
                componentes = cursor.fetchall()
                
                if not componentes:
                    continue
                
                # Calcular stock disponible
                stock_disponible_combo = float('inf')
                
                for comp in componentes:
                    comp_sku = comp['sku']
                    comp_stock_fisico = comp['stock_actual'] + comp['stock_full']
                    comp_vendido = ventas_dict.get(comp_sku, 0)
                    comp_stock_disponible = comp_stock_fisico - comp_vendido
                    combos_posibles = comp_stock_disponible // comp['cantidad_necesaria']
                    stock_disponible_combo = min(stock_disponible_combo, combos_posibles)
                
                stock_disponible_combo = int(stock_disponible_combo) if stock_disponible_combo != float('inf') else 0
                
                # Estado
                if stock_disponible_combo <= 0:
                    estado_stock = 'SIN_STOCK'
                elif stock_disponible_combo <= 2:
                    estado_stock = 'POCO_STOCK'
                else:
                    estado_stock = 'DISPONIBLE'
                
                # Filtro por estado
                if filtro_estado != 'TODOS' and estado_stock != filtro_estado:
                    continue
                
                productos.append({
                    'sku': combo_sku,
                    'nombre': combo_nombre,
                    'tipo': 'sommier',
                    'medida': medida_combo,
                    'modelo': None,
                    'stock_fisico': '-',
                    'stock_disponible': stock_disponible_combo,
                    'estado_stock': estado_stock,
                    'es_combo': True
                })
        
        # Ordenar
        productos.sort(key=lambda x: (x['tipo'], x['nombre'], x['medida'] or ''))
        
        cursor.close()
        conn.close()
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    
    return render_template('stock.html', 
                         productos=productos, 
                         filtro_estado=filtro_estado,
                         filtro_tipo=filtro_tipo,
                         filtro_modelo=filtro_modelo,
                         filtro_medida=filtro_medida,
                         buscar=buscar)


@app.route('/ventas/historicas')
@login_required
def ventas_historicas():
    """Lista de ventas históricas (entregadas y canceladas) con filtros"""
    try:
        # ========================================
        # OBTENER FILTROS
        # ========================================
        filtro_buscar = request.args.get('buscar', '').strip()
        filtro_estado = request.args.get('estado', '')  # '' = Todos, 'entregada', 'cancelada'
        filtro_periodo = request.args.get('periodo', 'todo')  # 'hoy', 'semana', 'mes', 'trimestre', 'todo'
        filtro_metodo_envio = request.args.get('metodo_envio', '')
        filtro_zona = request.args.get('zona', '')
        filtro_canal = request.args.get('canal', '')
        
        # ========================================
        # CONSTRUIR QUERY CON FILTROS
        # ========================================
        query = '''
            SELECT 
                id, numero_venta, fecha_venta, fecha_entrega, canal, mla_code,
                nombre_cliente, telefono_cliente,
                tipo_entrega, metodo_envio, ubicacion_despacho,
                zona_envio, direccion_entrega, costo_flete,
metodo_pago, importe_total, importe_abonado,
                pago_mercadopago, pago_efectivo,
                estado_entrega, estado_pago, notas,
    factura_generada, factura_fecha_generacion
            FROM ventas
            WHERE estado_entrega IN ('entregada', 'cancelada')
        '''
        params = []
        
        # Filtro: Estado (entregada, cancelada, o ambas)
        if filtro_estado:
            query += ' AND estado_entrega = %s'
            params.append(filtro_estado)
        
        # Filtro: Período (por fecha de entrega)
        if filtro_periodo == 'hoy':
            query += ' AND DATE(COALESCE(fecha_entrega, fecha_modificacion)) = CURDATE()'
        elif filtro_periodo == 'semana':
            query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 7 DAY)'
        elif filtro_periodo == 'mes':
            query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 30 DAY)'
        elif filtro_periodo == 'trimestre':
            query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 90 DAY)'
        
        # Filtro: Búsqueda de texto
        if filtro_buscar:
            query += '''
                AND (
                    mla_code LIKE %s 
                    OR nombre_cliente LIKE %s
                    OR id IN (
                        SELECT venta_id FROM items_venta WHERE sku LIKE %s
                    )
                )
            '''
            busqueda = f'%{filtro_buscar}%'
            params.extend([busqueda, busqueda, busqueda])
        
        # Filtro: Método de envío
        if filtro_metodo_envio:
            query += ' AND metodo_envio = %s'
            params.append(filtro_metodo_envio)
        
        # Filtro: Zona
        if filtro_zona:
            query += ' AND zona_envio = %s'
            params.append(filtro_zona)
        
        # Filtro: Canal
        if filtro_canal:
            query += ' AND canal = %s'
            params.append(filtro_canal)
        
        # Ordenar: Más recientes arriba (por fecha de entrega, o fecha_modificacion si no hay fecha_entrega)
        query += ' ORDER BY id DESC'
        
        # Ejecutar query
        ventas = query_db(query, tuple(params) if params else None)
        
        # ========================================
        # OBTENER ITEMS DE CADA VENTA
        # ========================================
        for venta in ventas:
            items = query_db('''
                SELECT 
                    iv.sku, 
                    iv.cantidad, 
                    iv.precio_unitario,
                    COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
                ORDER BY iv.id
            ''', (venta['id'],))
            venta['items'] = items
        
        # ========================================
        # CONTAR ENTREGADAS Y CANCELADAS
        # ========================================
        stats_query = '''
            SELECT 
                estado_entrega,
                COUNT(*) as total
            FROM ventas
            WHERE estado_entrega IN ('entregada', 'cancelada')
        '''
        if filtro_periodo != 'todo':
            if filtro_periodo == 'hoy':
                stats_query += ' AND DATE(COALESCE(fecha_entrega, fecha_modificacion)) = CURDATE()'
            elif filtro_periodo == 'semana':
                stats_query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 7 DAY)'
            elif filtro_periodo == 'mes':
                stats_query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 30 DAY)'
            elif filtro_periodo == 'trimestre':
                stats_query += ' AND COALESCE(fecha_entrega, fecha_modificacion) >= DATE_SUB(NOW(), INTERVAL 90 DAY)'
        
        stats_query += ' GROUP BY estado_entrega'
        stats = query_db(stats_query)
        
        entregadas = 0
        canceladas = 0
        for stat in stats:
            if stat['estado_entrega'] == 'entregada':
                entregadas = stat['total']
            elif stat['estado_entrega'] == 'cancelada':
                canceladas = stat['total']
        
        return render_template('ventas_historicas.html', 
                             ventas=ventas,
                             entregadas=entregadas,
                             canceladas=canceladas,
                             filtro_buscar=filtro_buscar,
                             filtro_estado=filtro_estado,
                             filtro_periodo=filtro_periodo,
                             filtro_metodo_envio=filtro_metodo_envio,
                             filtro_zona=filtro_zona,
                             filtro_canal=filtro_canal)
        
    except Exception as e:
        flash(f'Error al cargar ventas históricas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('index'))

@app.route('/ventas/historicas/<int:venta_id>/volver_activas', methods=['POST'])
@login_required
def historicas_volver_activas(venta_id):
    """Volver venta histórica (entregada o cancelada) a ventas activas"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener venta
        cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            flash('Venta no encontrada', 'error')
            return redirect(url_for('ventas_historicas'))
        
        estado_anterior = venta['estado_entrega']
        numero_venta = venta['numero_venta']
        
        # Verificar que sea histórica (entregada o cancelada)
        if estado_anterior not in ['entregada', 'cancelada']:
            flash(f'La venta {numero_venta} no es histórica', 'warning')
            return redirect(url_for('ventas_historicas'))
        
        # ========================================
        # LÓGICA SEGÚN ESTADO ANTERIOR
        # ========================================
        
        if estado_anterior == 'cancelada':
            # CANCELADA → ACTIVA
            # NO devolver stock (porque nunca se descontó)
            cursor.execute('''
                UPDATE ventas 
                SET estado_entrega = 'pendiente',
                    fecha_modificacion = NOW()
                WHERE id = %s
            ''', (venta_id,))
            
            mensaje = f'✅ Venta {numero_venta} devuelta a Ventas Activas (sin cambios en stock)'
        
        elif estado_anterior == 'entregada':
            # ENTREGADA → ACTIVA
            # SÍ devolver stock (porque se descontó cuando se entregó)
            cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
            items = cursor.fetchall()
            
            # Devolver stock de cada item
            for item in items:
                devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
            
            # Cambiar estado a pendiente
            cursor.execute('''
                UPDATE ventas 
                SET estado_entrega = 'pendiente',
                    fecha_modificacion = NOW()
                WHERE id = %s
            ''', (venta_id,))
            
            mensaje = f'✅ Venta {numero_venta} devuelta a Ventas Activas. Stock restaurado.'
        
        conn.commit()
        flash(mensaje, 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al volver a activas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    finally:
        cursor.close()
        conn.close()
    
    return redirect(url_for('ventas_historicas'))

@app.route('/ventas/historicas/volver-activas-multiple', methods=['POST'])
@login_required
def historicas_volver_activas_multiple():
    """
    Volver múltiples ventas históricas a ventas activas
    Usa la misma lógica que historicas_volver_activas() individual
    Mantiene filtros después de la acción
    """
    try:
        venta_ids = request.form.getlist('venta_ids')
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        ventas_procesadas = 0
        ventas_con_stock_devuelto = 0
        ventas_sin_stock = 0
        
        for venta_id in venta_ids:
            try:
                # Obtener venta
                cursor.execute('SELECT * FROM ventas WHERE id = %s', (venta_id,))
                venta = cursor.fetchone()
                
                if not venta:
                    continue
                
                estado_anterior = venta['estado_entrega']
                
                # Verificar que sea histórica
                if estado_anterior not in ['entregada', 'cancelada']:
                    continue
                
                # ========================================
                # LÓGICA SEGÚN ESTADO ANTERIOR
                # (igual que historicas_volver_activas)
                # ========================================
                
                if estado_anterior == 'cancelada':
                    # CANCELADA → ACTIVA
                    # NO devolver stock (porque nunca se descontó)
                    cursor.execute('''
                        UPDATE ventas 
                        SET estado_entrega = 'pendiente',
                            fecha_modificacion = NOW()
                        WHERE id = %s
                    ''', (venta_id,))
                    
                    ventas_procesadas += 1
                    ventas_sin_stock += 1
                
                elif estado_anterior == 'entregada':
                    # ENTREGADA → ACTIVA
                    # SÍ devolver stock (porque se descontó cuando se entregó)
                    cursor.execute('SELECT * FROM items_venta WHERE venta_id = %s', (venta_id,))
                    items = cursor.fetchall()
                    
                    # Devolver stock de cada item usando la función existente
                    for item in items:
                        devolver_stock_item(cursor, item, venta['ubicacion_despacho'])
                    
                    # Cambiar estado a pendiente
                    cursor.execute('''
                        UPDATE ventas 
                        SET estado_entrega = 'pendiente',
                            fecha_modificacion = NOW()
                        WHERE id = %s
                    ''', (venta_id,))
                    
                    ventas_procesadas += 1
                    ventas_con_stock_devuelto += 1
            
            except Exception as e:
                # Si falla una venta, continuar con las demás
                print(f"⚠️ Error al procesar venta {venta_id}: {str(e)}")
                continue
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # ========================================
        # MENSAJE DE ÉXITO
        # ========================================
        if ventas_procesadas == 0:
            flash('❌ No se pudieron procesar las ventas seleccionadas', 'error')
        else:
            mensaje = f'✅ {ventas_procesadas} venta(s) devueltas a Ventas Activas.'
            
            if ventas_con_stock_devuelto > 0:
                mensaje += f' Stock restaurado en {ventas_con_stock_devuelto} venta(s).'
            
            if ventas_sin_stock > 0:
                mensaje += f' {ventas_sin_stock} cancelada(s) sin cambios en stock.'
            
            flash(mensaje, 'success')
        
        # ========================================
        # ✅ MANTENER FILTROS
        # ========================================
        filtros = {}
        if request.form.get('buscar'):
            filtros['buscar'] = request.form.get('buscar')
        if request.form.get('estado'):
            filtros['estado'] = request.form.get('estado')
        if request.form.get('periodo'):
            filtros['periodo'] = request.form.get('periodo')
        if request.form.get('metodo_envio'):
            filtros['metodo_envio'] = request.form.get('metodo_envio')
        if request.form.get('zona'):
            filtros['zona'] = request.form.get('zona')
        if request.form.get('canal'):
            filtros['canal'] = request.form.get('canal')
        
        return redirect(url_for('ventas_historicas', **filtros))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))

# ============================================================================
# MAPEO DE PROVINCIA A CÓDIGO AFIP
# ============================================================================

def provincia_a_codigo(provincia_str):
    """Convierte nombre de provincia a código AFIP, sin importar tildes/mayúsculas"""
    import unicodedata

    def normalizar(s):
        if not s:
            return ''
        s = str(s).upper().strip()
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return s

    MAPA = {
        'CDAD. AUTONOMA DE BS AS': 1,
        'CIUDAD AUTONOMA DE BUENOS AIRES': 1,
        'CABA': 1,
        'CAPITAL FEDERAL': 1,
        'C.A.B.A.': 1,
        'CIUDAD DE BUENOS AIRES': 1,
        'BUENOS AIRES': 2,
        'PROVINCIA DE BUENOS AIRES': 2,
        'PBA': 2,
        'CATAMARCA': 3,
        'CORDOBA': 4,
        'CORRIENTES': 5,
        'ENTRE RIOS': 6,
        'JUJUY': 7,
        'MENDOZA': 8,
        'LA RIOJA': 9,
        'SALTA': 10,
        'SAN JUAN': 11,
        'SAN LUIS': 12,
        'SANTA FE': 13,
        'SANTIAGO DEL ESTERO': 14,
        'TUCUMAN': 15,
        'CHACO': 16,
        'CHUBUT': 17,
        'FORMOSA': 18,
        'MISIONES': 19,
        'NEUQUEN': 20,
        'LA PAMPA': 21,
        'RIO NEGRO': 22,
        'SANTA CRUZ': 23,
        'TIERRA DEL FUEGO': 24,
        'EXTERIOR': 25,
    }

    norm = normalizar(provincia_str)
    # Búsqueda exacta primero
    if norm in MAPA:
        return MAPA[norm]
    # Búsqueda parcial (ej: "Buenos Aires" puede venir como zona_envio)
    for key, codigo in MAPA.items():
        if key in norm or norm in key:
            return codigo
    # Default: Capital Federal
    return 1


@app.route('/ventas/historicas/<int:venta_id>/generar-factura-excel')
@login_required
def generar_factura_excel(venta_id):
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    try:
        venta = query_one('SELECT * FROM ventas WHERE id = %s', (venta_id,))

        if not venta:
            flash('❌ Venta no encontrada', 'error')
            return redirect(url_for('ventas_historicas'))

        items = list(query_db('''
            SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
            FROM items_venta iv
            LEFT JOIN productos_base pb ON iv.sku = pb.sku
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            WHERE iv.venta_id = %s
        ''', (venta_id,)))

        costo_flete = float(venta.get('costo_flete') or 0)

        incluir_flete_param = request.args.get('incluir_flete', 'false').lower()
        incluir_flete = (incluir_flete_param == 'true' and costo_flete > 0)

        cant_slots = len(items) + (1 if incluir_flete else 0)

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación"

        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia', 'rubro']
        for idx in range(1, cant_slots + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        headers.append('importe total')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row_data = []

        if venta.get('mla_code'):
            id_venta = venta['mla_code']
        else:
            id_venta = venta.get('factura_business_name') or venta['nombre_cliente']
        row_data.append(id_venta)

        row_data.append(venta.get('factura_taxpayer_type') or 'Consumidor Final')

        row_data.append(venta.get('factura_business_name') or venta['nombre_cliente'])

        if venta.get('factura_doc_number'):
            dni = str(venta['factura_doc_number'])
        elif venta.get('dni_cliente'):
            dni = str(venta['dni_cliente'])
        else:
            dni = '99999999'
        row_data.append(dni)

        if venta.get('factura_street'):
            direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
        else:
            direccion = venta.get('direccion_entrega') or ''
        row_data.append(direccion)

        provincia_str = (venta.get('factura_state') or venta.get('provincia_cliente') or
                         venta.get('zona_envio') or 'Capital Federal')
        row_data.append(provincia_a_codigo(provincia_str))

        rubro = 'F' if len(dni.replace('.', '').replace('-', '').strip()) == 11 else 'R'
        row_data.append(rubro)

        for item in items:
            row_data.append(item['sku'])
            row_data.append(int(item['cantidad']))
            row_data.append(float(item['precio_unitario']))

        if incluir_flete:
            row_data.append('FLETE')
            row_data.append(1)
            row_data.append(costo_flete)

        total_excel = float(venta['importe_total']) + (costo_flete if incluir_flete else 0)
        row_data.append(total_excel)

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=2, column=col_idx, value=value)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        nombre_archivo = f"factura_{venta['numero_venta'].replace('/', '-')}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        execute_db('''
            UPDATE ventas SET factura_generada = TRUE, factura_fecha_generacion = NOW()
            WHERE id = %s
        ''', (venta_id,))

        return response

    except Exception as e:
        flash(f'❌ Error al generar factura: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))


# ============================================================================
# FUNCIÓN 2: FACTURA MÚLTIPLE
# ============================================================================

@app.route('/ventas/historicas/facturar-multiple-excel')
@login_required
def facturar_multiple_excel():
    from flask import make_response
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO

    try:
        ids_str = request.args.get('ids', '')
        if not ids_str:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_historicas'))

        venta_ids = [int(id) for id in ids_str.split(',') if id.strip()]

        if not venta_ids:
            flash('❌ No se seleccionaron ventas válidas', 'error')
            return redirect(url_for('ventas_historicas'))

        ids_con_flete_str = request.args.get('ids_con_flete', '')
        ids_con_flete = set()
        if ids_con_flete_str:
            ids_con_flete = {int(i) for i in ids_con_flete_str.split(',') if i.strip()}

        placeholders = ', '.join(['%s'] * len(venta_ids))
        ventas = query_db(f'SELECT * FROM ventas WHERE id IN ({placeholders}) ORDER BY id DESC', tuple(venta_ids))

        if not ventas:
            flash('❌ No se encontraron ventas', 'error')
            return redirect(url_for('ventas_historicas'))

        ventas_preparadas = []
        max_slots = 0

        for venta in ventas:
            items = list(query_db('''
                SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
            ''', (venta['id'],)))

            costo_flete = float(venta.get('costo_flete') or 0)
            incluir_flete = (venta['id'] in ids_con_flete and costo_flete > 0)

            cant_slots = len(items) + (1 if incluir_flete else 0)
            if cant_slots > max_slots:
                max_slots = cant_slots

            ventas_preparadas.append({
                'venta': venta,
                'items': items,
                'costo_flete': costo_flete,
                'incluir_flete': incluir_flete,
                'cant_slots': cant_slots
            })

        wb = Workbook()
        ws = wb.active
        ws.title = "Facturación Múltiple"

        headers = ['id venta', 'categoria de iva', 'nombre', 'dni', 'direccion', 'provincia', 'rubro']
        for idx in range(1, max_slots + 1):
            headers.extend([f'sku{idx}', f'cant sku{idx}', f'importe sku{idx}'])
        headers.append('importe total')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row = 2

        for venta_prep in ventas_preparadas:
            venta = venta_prep['venta']
            items = venta_prep['items']
            costo_flete = venta_prep['costo_flete']
            incluir_flete = venta_prep['incluir_flete']
            cant_slots = venta_prep['cant_slots']

            row_data = []

            if venta.get('mla_code'):
                id_venta = venta['mla_code']
            else:
                id_venta = venta.get('factura_business_name') or venta['nombre_cliente']
            row_data.append(id_venta)

            row_data.append(venta.get('factura_taxpayer_type') or 'Consumidor Final')
            row_data.append(venta.get('factura_business_name') or venta['nombre_cliente'])

            if venta.get('factura_doc_number'):
                dni = str(venta['factura_doc_number'])
            elif venta.get('dni_cliente'):
                dni = str(venta['dni_cliente'])
            else:
                dni = '99999999'
            row_data.append(dni)

            if venta.get('factura_street'):
                direccion = f"{venta['factura_street']}, {venta.get('factura_city', '')}"
            else:
                direccion = venta.get('direccion_entrega') or ''
            row_data.append(direccion)

            provincia_str = (venta.get('factura_state') or venta.get('provincia_cliente') or
                             venta.get('zona_envio') or 'Capital Federal')
            row_data.append(provincia_a_codigo(provincia_str))

            rubro = 'F' if len(dni.replace('.', '').replace('-', '').strip()) == 11 else 'R'
            row_data.append(rubro)

            for item in items:
                row_data.append(item['sku'])
                row_data.append(int(item['cantidad']))
                row_data.append(float(item['precio_unitario']))

            if incluir_flete:
                row_data.append('FLETE')
                row_data.append(1)
                row_data.append(costo_flete)

            for _ in range(max_slots - cant_slots):
                row_data.extend(['', '', ''])

            total_excel = float(venta['importe_total']) + (costo_flete if incluir_flete else 0)
            row_data.append(total_excel)

            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_idx, value=value)

            current_row += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = make_response(excel_file.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"facturas_multiple_{fecha_str}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

        placeholders = ', '.join(['%s'] * len(venta_ids))
        execute_db(f'''
            UPDATE ventas SET factura_generada = TRUE, factura_fecha_generacion = NOW()
            WHERE id IN ({placeholders})
        ''', tuple(venta_ids))

        return response

    except Exception as e:
        flash(f'❌ Error al generar facturas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))




# ============================================================================
# VENTAS HISTÓRICAS - VOLVER A ACTIVAS
# Agregar en app.py después de la ruta de ventas_historicas()
# ============================================================================



@app.route('/ventas/historicas/facturar-multiple')
@login_required
def facturar_multiple():
    """
    Generar UN SOLO archivo .txt con TODAS las ventas seleccionadas
    """
    from flask import make_response
    from datetime import datetime
    
    try:
        # Obtener IDs
        ids_str = request.args.get('ids', '')
        if not ids_str:
            flash('❌ No se seleccionaron ventas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        venta_ids = [int(id) for id in ids_str.split(',') if id.strip()]
        
        if not venta_ids:
            flash('❌ No se seleccionaron ventas válidas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        # ============================================
        # OBTENER TODAS LAS VENTAS
        # ============================================
        placeholders = ', '.join(['%s'] * len(venta_ids))
        query = f'''
            SELECT * FROM ventas 
            WHERE id IN ({placeholders})
            ORDER BY fecha_venta DESC
        '''
        ventas = query_db(query, tuple(venta_ids))
        
        if not ventas:
            flash('❌ No se encontraron ventas', 'error')
            return redirect(url_for('ventas_historicas'))
        
        # ============================================
        # GENERAR CONTENIDO DEL ARCHIVO TXT
        # ============================================
        
        lineas = []
        lineas.append("="*80)
        lineas.append("FACTURACIÓN MÚLTIPLE - DATOS PARA FACTURAR")
        lineas.append("="*80)
        lineas.append(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        lineas.append(f"Total de ventas: {len(ventas)}")
        lineas.append("="*80)
        lineas.append("")
        
        total_general = 0
        
        # Procesar cada venta
        for idx, venta in enumerate(ventas, 1):
            lineas.append("")
            lineas.append("="*80)
            lineas.append(f"VENTA #{idx} - {venta['numero_venta']}")
            lineas.append("="*80)
            lineas.append("")
            
            # DATOS DE LA VENTA
            lineas.append(f"  Fecha: {venta['fecha_venta'].strftime('%d/%m/%Y')}")
            if venta.get('mla_code'):
                lineas.append(f"  ML Code: {venta['mla_code']}")
            lineas.append("")
            
            # DATOS DEL COMPRADOR
            lineas.append("-"*80)
            lineas.append("DATOS DEL COMPRADOR:")
            lineas.append("-"*80)
            
            if venta.get('factura_business_name'):
                # Tiene datos de facturación
                lineas.append(f"  Razón Social: {venta['factura_business_name']}")
                lineas.append(f"  {venta.get('factura_doc_type', 'Documento')}: {venta.get('factura_doc_number', 'N/A')}")
                lineas.append(f"  Condición IVA: {venta.get('factura_taxpayer_type', 'N/A')}")
                lineas.append("")
                lineas.append("  DOMICILIO FISCAL:")
                lineas.append(f"    Calle: {venta.get('factura_street', 'N/A')}")
                lineas.append(f"    Ciudad: {venta.get('factura_city', 'N/A')}")
                lineas.append(f"    Provincia: {venta.get('factura_state', 'N/A')}")
                lineas.append(f"    CP: {venta.get('factura_zip_code', 'N/A')}")
            else:
                # Consumidor Final
                lineas.append(f"  Nombre: {venta['nombre_cliente']}")
                if venta.get('telefono_cliente'):
                    lineas.append(f"  Teléfono: {venta['telefono_cliente']}")
                lineas.append(f"  Condición IVA: Consumidor Final")
                
                if venta.get('direccion_entrega'):
                    lineas.append("")
                    lineas.append("  DIRECCIÓN DE ENTREGA:")
                    lineas.append(f"    {venta['direccion_entrega']}")
                    if venta.get('zona_envio'):
                        lineas.append(f"    Zona: {venta['zona_envio']}")
            
            lineas.append("")
            
            # PRODUCTOS
            items = query_db('''
                SELECT iv.*, COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre_producto
                FROM items_venta iv
                LEFT JOIN productos_base pb ON iv.sku = pb.sku
                LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
                WHERE iv.venta_id = %s
            ''', (venta['id'],))
            
            lineas.append("-"*80)
            lineas.append("PRODUCTOS:")
            lineas.append("-"*80)
            lineas.append("")
            lineas.append(f"{'Cant':<6} {'SKU':<15} {'Descripción':<35} {'P.Unit':<12} {'Subtotal':<12}")
            lineas.append("-"*80)
            
            total_items = 0
            for item in items:
                cant = item['cantidad']
                sku = item['sku']
                nombre = item['nombre_producto']
                precio = item['precio_unitario']
                subtotal = cant * precio
                
                lineas.append(f"{cant:<6} {sku:<15} {nombre:<35} ${precio:<11,.2f} ${subtotal:<11,.2f}")
                total_items += subtotal
            
            lineas.append("-"*80)
            lineas.append("")
            
            # TOTALES
            lineas.append("TOTALES:")
            lineas.append(f"  Subtotal Productos: ${total_items:,.2f}")
            
            if venta.get('costo_flete') and venta['costo_flete'] > 0:
                lineas.append(f"  Costo de Envío: ${venta['costo_flete']:,.2f}")
                lineas.append(f"  TOTAL: ${venta['importe_total']:,.2f}")
            else:
                lineas.append(f"  TOTAL: ${venta['importe_total']:,.2f}")
            
            lineas.append("")
            lineas.append("-"*80)
            
            # MÉTODO DE PAGO
            lineas.append("MÉTODO DE PAGO:")
            lineas.append(f"  {venta.get('metodo_pago', 'N/A')}")
            if venta.get('pago_mercadopago') and venta['pago_mercadopago'] > 0:
                lineas.append(f"    Mercadopago: ${venta['pago_mercadopago']:,.2f}")
            if venta.get('pago_efectivo') and venta['pago_efectivo'] > 0:
                lineas.append(f"    Efectivo: ${venta['pago_efectivo']:,.2f}")
            
            lineas.append("")
            
            total_general += venta['importe_total']
        
        # RESUMEN FINAL
        lineas.append("")
        lineas.append("="*80)
        lineas.append("RESUMEN GENERAL")
        lineas.append("="*80)
        lineas.append(f"Total de ventas facturadas: {len(ventas)}")
        lineas.append(f"TOTAL GENERAL: ${total_general:,.2f}")
        lineas.append("="*80)
        
        # ============================================
        # CREAR RESPUESTA CON ARCHIVO
        # ============================================
        
        contenido = "\n".join(lineas)
        
        response = make_response(contenido)
        response.headers['Content-Type'] = 'text/plain; charset=utf-8'
        
        # Nombre del archivo
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"facturas_multiple_{fecha_str}.txt"
        response.headers['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        
        # ============================================
        # MARCAR TODAS COMO GENERADAS
        # ============================================
        
        placeholders = ', '.join(['%s'] * len(venta_ids))
        execute_db(f'''
            UPDATE ventas 
            SET factura_generada = TRUE,
                factura_fecha_generacion = NOW()
            WHERE id IN ({placeholders})
        ''', tuple(venta_ids))
        
        return response
        
    except Exception as e:
        flash(f'❌ Error al generar facturas: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_historicas'))




# ============================================================================
# FORMULARIOS VISUALES - SIN SQL
# ============================================================================

# ============================================================================
# API PARA CARRITO
# ============================================================================

@app.route('/api/productos')
@login_required
def api_productos():
    """API para el buscador de productos en templates"""
    from flask import jsonify
    
    # Productos base
    productos_base = query_db('SELECT sku, nombre, tipo, stock_actual FROM productos_base ORDER BY nombre')
    
    # Combos - tabla: productos_compuestos
    combos = query_db('SELECT sku, nombre FROM productos_compuestos ORDER BY nombre')
    
    todos = []
    
    for p in productos_base:
        todos.append({
            'sku': p['sku'],
            'nombre': p['nombre'],
            'tipo': p['tipo'],
            'stock_actual': p['stock_actual']
        })
    
    for c in combos:
        todos.append({
            'sku': c['sku'],
            'nombre': c['nombre'],
            'tipo': 'combo',
            'stock_actual': 0
        })
    
    return jsonify(todos)

@app.route('/cargar-stock', methods=['GET', 'POST'])
@login_required
def cargar_stock():
    """Formulario para cargar/agregar stock de productos"""
    from flask import jsonify
    
    if request.method == 'POST' and request.is_json:
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            data = request.get_json()
            items = data.get('items', [])
            
            for item in items:
                sku = item['sku']
                cantidad = item['cantidad']
                ubicacion = item['ubicacion']
                motivo = item.get('motivo', 'Carga de stock')
                
                # Obtener datos actuales del producto
                cursor.execute('SELECT nombre, stock_actual, COALESCE(stock_full, 0) as stock_full FROM productos_base WHERE sku = %s', (sku,))
                prod = cursor.fetchone()
                
                if not prod:
                    continue
                
                nombre_producto = prod['nombre']
                
                # Actualizar stock y registrar movimiento
                if ubicacion == 'stock_actual':
                    stock_anterior = prod['stock_actual']
                    stock_nuevo = stock_anterior + cantidad
                    
                    cursor.execute('UPDATE productos_base SET stock_actual = stock_actual + %s WHERE sku = %s', (cantidad, sku))
                    
                    # Registrar en historial
                    cursor.execute('''
                        INSERT INTO movimientos_stock 
                        (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ''', (sku, nombre_producto, 'carga', cantidad, stock_anterior, stock_nuevo, motivo))
                    
                else:  # stock_full
                    stock_anterior = prod['stock_full']
                    stock_nuevo = stock_anterior + cantidad
                    
                    cursor.execute('UPDATE productos_base SET stock_full = stock_full + %s WHERE sku = %s', (cantidad, sku))
                    
                    # Registrar en historial
                    cursor.execute('''
                        INSERT INTO movimientos_stock 
                        (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ''', (sku, nombre_producto, 'carga', cantidad, stock_anterior, stock_nuevo, motivo))
            
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({'success': True, 'message': f'Stock cargado: {len(items)} productos'})
        except Exception as e:
            conn.rollback()
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # Si es GET, mostrar formulario
    productos = []
    try:
        productos = query_db('''
            SELECT * FROM productos_base 
            ORDER BY 
                CASE tipo
                    WHEN 'colchon' THEN 1
                    WHEN 'base' THEN 2
                    WHEN 'almohada' THEN 3
                END,
                nombre
        ''')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return render_template('cargar_stock.html', productos=productos)

@app.route('/cargar-stock/guardar', methods=['POST'])
@login_required
def guardar_stock():
    """Agregar stock (SUMA al stock actual) y registrar movimientos"""
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # Procesar cada producto
        productos_cargados = 0
        for key, value in request.form.items():
            if key.startswith('agregar_'):
                sku = key.replace('agregar_', '')
                cantidad_agregar = int(value) if value else 0
                
                # Solo procesar si hay cantidad a agregar
                if cantidad_agregar > 0:
                    # Obtener stock ACTUAL de la base de datos
                    cursor.execute('SELECT stock_actual, nombre FROM productos_base WHERE sku = %s', (sku,))
                    resultado = cursor.fetchone()
                    
                    if resultado:
                        stock_anterior = resultado['stock_actual']
                        nombre_producto = resultado['nombre']
                        stock_nuevo = stock_anterior + cantidad_agregar
                        
                        # Actualizar stock (SUMAR)
                        cursor.execute('''
                            UPDATE productos_base 
                            SET stock_actual = stock_actual + %s 
                            WHERE sku = %s
                        ''', (cantidad_agregar, sku))
                        
                        # Registrar movimiento
                        cursor.execute('''
                            INSERT INTO movimientos_stock 
                            (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku, nombre_producto, 'carga', cantidad_agregar, stock_anterior, stock_nuevo, 'Carga de stock nuevo'))
                        
                        productos_cargados += 1
        
        conn.commit()
        
        if productos_cargados > 0:
            flash(f'✅ Stock cargado correctamente ({productos_cargados} productos)', 'success')
        else:
            flash('ℹ️ No se agregó stock. Ingresá cantidades mayores a 0.', 'info')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('cargar_stock'))


@app.route('/bajar-stock')
@login_required
def bajar_stock():
    """Formulario para dar de baja stock"""
    productos = []
    try:
        productos = query_db('''
            SELECT sku, nombre, tipo, stock_actual
            FROM productos_base 
            WHERE stock_actual > 0 
            AND tipo IN ('colchon', 'base', 'almohada')
            ORDER BY 
                CASE tipo
                    WHEN 'colchon' THEN 1
                    WHEN 'base' THEN 2
                    WHEN 'almohada' THEN 3
                END,
                nombre
        ''')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return render_template('bajar_stock.html', productos=productos)


@app.route('/bajar-stock/guardar', methods=['POST'])
@login_required
def bajar_stock_guardar():
    """Guardar bajas de stock y registrar movimientos"""
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        motivo = request.form.get('motivo', 'Baja de stock manual')
        
        # Obtener datos del formulario
        bajas = []
        for key, value in request.form.items():
            if key.startswith('baja_'):
                sku = key.replace('baja_', '')
                cantidad_baja = int(value) if value else 0
                if cantidad_baja > 0:
                    bajas.append((cantidad_baja, sku))
        
        # Descontar stock y registrar movimientos
        for cantidad_baja, sku in bajas:
            # Obtener stock anterior
            cursor.execute('SELECT stock_actual, nombre FROM productos_base WHERE sku = %s', (sku,))
            resultado = cursor.fetchone()
            stock_anterior = resultado['stock_actual'] if resultado else 0
            nombre_producto = resultado['nombre'] if resultado else ''
            stock_nuevo = stock_anterior - cantidad_baja
            
            # Actualizar stock
            cursor.execute('''
                UPDATE productos_base 
                SET stock_actual = stock_actual - %s 
                WHERE sku = %s
            ''', (cantidad_baja, sku))
            
            # Registrar movimiento
            cursor.execute('''
                INSERT INTO movimientos_stock 
                (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (sku, nombre_producto, 'baja', cantidad_baja, stock_anterior, stock_nuevo, motivo))
        
        conn.commit()
        
        motivo_msg = f' - {motivo}' if motivo else ''
        flash(f'✅ Stock dado de baja correctamente ({len(bajas)} productos){motivo_msg}', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error al dar de baja stock: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('bajar_stock'))


@app.route('/historial-stock')
@login_required
def historial_stock():
    """Ver historial de movimientos de stock"""
    from datetime import datetime, timedelta
    
    # Obtener filtros
    filtro_tipo = request.args.get('tipo', '')
    filtro_sku = request.args.get('sku', '').strip()
    filtro_fecha = request.args.get('fecha', 'todo')
    
    try:
        # Query base
        query = """
            SELECT 
                id,
                sku,
                nombre_producto,
                tipo_movimiento,
                cantidad,
                stock_anterior,
                stock_nuevo,
                motivo,
                usuario,
                fecha_movimiento
            FROM movimientos_stock
            WHERE 1=1
        """
        params = []
        
        # Aplicar filtros
        if filtro_tipo:
            query += " AND tipo_movimiento = %s"
            params.append(filtro_tipo)
        
        if filtro_sku:
            query += " AND (sku LIKE %s OR nombre_producto LIKE %s)"
            params.append(f'%{filtro_sku}%')
            params.append(f'%{filtro_sku}%')
        
        # Filtro de fecha
        if filtro_fecha == 'hoy':
            query += " AND DATE(fecha_movimiento) = CURDATE()"
        elif filtro_fecha == 'semana':
            query += " AND fecha_movimiento >= DATE_SUB(NOW(), INTERVAL 7 DAY)"
        elif filtro_fecha == 'mes':
            query += " AND fecha_movimiento >= DATE_SUB(NOW(), INTERVAL 30 DAY)"
        
        query += " ORDER BY fecha_movimiento DESC LIMIT 500"
        
        movimientos = query_db(query, tuple(params) if params else None)
        
        # Estadísticas
        stats_query = """
            SELECT 
                tipo_movimiento,
                COUNT(*) as total_movimientos,
                SUM(cantidad) as total_unidades
            FROM movimientos_stock
            GROUP BY tipo_movimiento
        """
        estadisticas = query_db(stats_query)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        movimientos = []
        estadisticas = []
    
    return render_template('historial_stock.html', 
                         movimientos=movimientos,
                         estadisticas=estadisticas,
                         filtro_tipo=filtro_tipo,
                         filtro_sku=filtro_sku,
                         filtro_fecha=filtro_fecha)


@app.route('/transferir-stock')
@login_required
def transferir_stock():
    """Formulario para transferir stock de Depósito a Full (Compac y Almohadas)"""
    productos_compac = []
    productos_almohadas = []
    
    try:
        # Productos Compac con ubicación _DEP que tengan stock
        productos_compac = query_db('''
            SELECT sku, nombre, stock_actual
            FROM productos_base 
            WHERE sku IN ('CCO80_DEP', 'CCO100_DEP', 'CCO140_DEP', 'CCO160_DEP',
                          'CCP80_DEP', 'CCP100_DEP', 'CCP140_DEP', 'CCP160_DEP')
            AND stock_actual > 0
            ORDER BY sku
        ''')
        
        # Almohadas que tengan stock en Depósito (stock_actual > 0)
        productos_almohadas = query_db('''
            SELECT sku, nombre, stock_actual, stock_full
            FROM productos_base 
            WHERE tipo = 'almohada'
            AND stock_actual > 0
            ORDER BY nombre
        ''')
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return render_template('transferir_stock.html', 
                         productos_compac=productos_compac,
                         productos_almohadas=productos_almohadas)


@app.route('/transferir-stock/guardar', methods=['POST'])
@login_required
def transferir_stock_guardar():
    """Procesar transferencia de stock de Depósito a Full (Compac y Almohadas)"""
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        transferencias = 0
        
        # Procesar COMPAC (usan _DEP y _FULL en SKU)
        for key, value in request.form.items():
            if key.startswith('transferir_compac_'):
                sku_dep = key.replace('transferir_compac_', '')
                cantidad_transferir = int(value) if value else 0
                
                if cantidad_transferir > 0:
                    sku_full = sku_dep.replace('_DEP', '_FULL')
                    
                    cursor.execute('SELECT stock_actual, nombre FROM productos_base WHERE sku = %s', (sku_dep,))
                    dep = cursor.fetchone()
                    cursor.execute('SELECT stock_actual, nombre FROM productos_base WHERE sku = %s', (sku_full,))
                    full = cursor.fetchone()
                    
                    if dep and full:
                        stock_dep_anterior = dep['stock_actual']
                        stock_full_anterior = full['stock_actual']
                        nombre_base = dep['nombre'].replace(' (Depósito)', '')
                        
                        if cantidad_transferir > stock_dep_anterior:
                            flash(f'❌ {sku_dep}: No hay suficiente stock (disponible: {stock_dep_anterior})', 'error')
                            continue
                        
                        # Descontar de depósito
                        cursor.execute('''
                            UPDATE productos_base 
                            SET stock_actual = stock_actual - %s 
                            WHERE sku = %s
                        ''', (cantidad_transferir, sku_dep))
                        
                        # Sumar a Full
                        cursor.execute('''
                            UPDATE productos_base 
                            SET stock_actual = stock_actual + %s 
                            WHERE sku = %s
                        ''', (cantidad_transferir, sku_full))
                        
                        motivo_baja = 'Transferencia a Full (' + sku_full + ')'
                        motivo_carga = 'Transferencia desde Deposito (' + sku_dep + ')'
                        
                        cursor.execute('''
                            INSERT INTO movimientos_stock 
                            (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku_dep, nombre_base, 'baja', cantidad_transferir, 
                              stock_dep_anterior, stock_dep_anterior - cantidad_transferir, 
                              motivo_baja))
                        
                        cursor.execute('''
                            INSERT INTO movimientos_stock 
                            (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku_full, nombre_base, 'carga', cantidad_transferir, 
                              stock_full_anterior, stock_full_anterior + cantidad_transferir, 
                              motivo_carga))
                        
                        transferencias += 1
        
        # Procesar ALMOHADAS (usan stock_actual y stock_full en misma fila)
        for key, value in request.form.items():
            if key.startswith('transferir_alm_'):
                sku_alm = key.replace('transferir_alm_', '')
                cantidad_transferir = int(value) if value else 0
                
                if cantidad_transferir > 0:
                    cursor.execute('SELECT stock_actual, stock_full, nombre FROM productos_base WHERE sku = %s', (sku_alm,))
                    alm = cursor.fetchone()
                    
                    if alm:
                        stock_dep_anterior = alm['stock_actual']
                        stock_full_anterior = alm.get('stock_full', 0)
                        nombre = alm['nombre']
                        
                        if cantidad_transferir > stock_dep_anterior:
                            flash(f'❌ {nombre}: No hay suficiente stock en Depósito (disponible: {stock_dep_anterior})', 'error')
                            continue
                        
                        # Transferir: resta de stock_actual, suma a stock_full
                        cursor.execute('''
                            UPDATE productos_base 
                            SET stock_actual = stock_actual - %s,
                                stock_full = stock_full + %s
                            WHERE sku = %s
                        ''', (cantidad_transferir, cantidad_transferir, sku_alm))
                        
                        motivo_baja = 'Transferencia a Full ML'
                        motivo_carga = 'Transferencia desde Deposito'
                        
                        cursor.execute('''
                            INSERT INTO movimientos_stock 
                            (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku_alm, nombre, 'baja', cantidad_transferir, 
                              stock_dep_anterior, stock_dep_anterior - cantidad_transferir, 
                              motivo_baja))
                        
                        cursor.execute('''
                            INSERT INTO movimientos_stock 
                            (sku, nombre_producto, tipo_movimiento, cantidad, stock_anterior, stock_nuevo, motivo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        ''', (sku_alm, nombre, 'carga', cantidad_transferir, 
                              stock_full_anterior, stock_full_anterior + cantidad_transferir, 
                              motivo_carga))
                        
                        transferencias += 1
        
        conn.commit()
        
        if transferencias > 0:
            flash(f'✅ Transferencia completada ({transferencias} productos)', 'success')
        else:
            flash('ℹ️ No se realizaron transferencias', 'info')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Error: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('transferir_stock'))


# ============================================================================
# HELPER PARA JSON (agregar después de las funciones de BD)
# ============================================================================

def decimal_to_float(obj):
    """Convertir Decimals a float para JSON serialization"""
    from decimal import Decimal
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, dict):
        return {k: decimal_to_float(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [decimal_to_float(item) for item in obj]
    return obj


# ============================================================================
# NUEVA VENTA CORREGIDO - Reemplazar desde línea 266 hasta antes de dashboard_visual
# ============================================================================

@app.route('/nueva-venta')
@login_required
def nueva_venta():
    """Formulario para registrar venta con stock disponible y ubicaciones"""
    from datetime import date
    import json
    
    try:
        # ========================================
        # 1. GENERAR SIGUIENTE NÚMERO DE VENTA
        # ========================================
        ultima_venta = query_db('SELECT numero_venta FROM ventas ORDER BY id DESC LIMIT 1')
        siguiente_numero = 'VENTA-001'
        
        if ultima_venta and len(ultima_venta) > 0:
            num_venta = ultima_venta[0].get('numero_venta')
            if num_venta and '-' in num_venta:
                try:
                    ultimo_num = int(num_venta.split('-')[1])
                    siguiente_numero = f'VENTA-{ultimo_num + 1:03d}'
                except (ValueError, IndexError):
                    siguiente_numero = 'VENTA-001'
        
        # ========================================
        # 2. OBTENER STOCK FÍSICO (con ubicaciones)
        # ========================================
        productos_base = query_db('''
            SELECT 
                sku, 
                nombre, 
                tipo, 
                stock_actual,
                COALESCE(stock_full, 0) as stock_full
            FROM productos_base 
            ORDER BY tipo, nombre
        ''')
        
        # ========================================
        # 3. OBTENER VENTAS ACTIVAS (DESCOMPONIENDO COMBOS)
        # ========================================
        ventas_activas = query_db('''
            SELECT 
                COALESCE(pb_comp.sku, iv.sku) as sku,
                SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
            FROM items_venta iv
            JOIN ventas v ON iv.venta_id = v.id
            LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
            LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
            LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
            WHERE v.estado_entrega = 'pendiente'
            GROUP BY sku
        ''')
        
        # Convertir a diccionario
        ventas_dict = {v['sku']: int(v['vendido']) for v in ventas_activas}
        
        # ========================================
        # 4. CALCULAR STOCK DISPONIBLE
        # ========================================
        productos_procesados = []
        
        for prod in productos_base:
            sku = prod['sku']
            vendido = ventas_dict.get(sku, 0)
            
            # Para productos con ubicaciones
            if '_DEP' in sku or '_FULL' in sku:
                # Compac: tiene _DEP y _FULL separados
                stock_fisico = int(prod['stock_actual'])
                stock_disponible_dep = stock_fisico - vendido
                
                productos_procesados.append({
                    'sku': sku,
                    'nombre': prod['nombre'],
                    'tipo': prod['tipo'],
                    'stock': stock_fisico,
                    'stock_disponible': stock_disponible_dep,
                    'tiene_ubicaciones': True,
                    'ubicacion': 'DEP' if '_DEP' in sku else 'FULL',
                    'precio': 0
                })
                
            elif prod['tipo'] == 'almohada':
                # Almohadas: tienen stock_actual (DEP) y stock_full (FULL)
                stock_dep = int(prod['stock_actual'])
                stock_full = int(prod['stock_full'])
                stock_total = stock_dep + stock_full
                stock_disponible = stock_total - vendido
                
                productos_procesados.append({
                    'sku': sku,
                    'nombre': prod['nombre'],
                    'tipo': prod['tipo'],
                    'stock': stock_total,
                    'stock_dep': stock_dep,
                    'stock_full': stock_full,
                    'stock_disponible': stock_disponible,
                    'tiene_ubicaciones': True,
                    'precio': 0
                })
                
            else:
                # Otros productos: solo stock_actual
                stock_fisico = int(prod['stock_actual'])
                stock_disponible = stock_fisico - vendido
                
                productos_procesados.append({
                    'sku': sku,
                    'nombre': prod['nombre'],
                    'tipo': prod['tipo'],
                    'stock': stock_fisico,
                    'stock_disponible': stock_disponible,
                    'tiene_ubicaciones': False,
                    'precio': 0
                })
        
        # ========================================
        # 5. OBTENER COMBOS/SOMMIERS
        # ========================================
        try:
            productos_combos = query_db('''
                SELECT 
                    sku, 
                    nombre, 
                    'combo' as tipo,
                    0 as stock,
                    0 as stock_disponible,
                    0 as precio
                FROM productos_compuestos
                WHERE activo = 1
                ORDER BY nombre
            ''')
            
            # Para cada combo, calcular su disponibilidad según componentes
            if productos_combos:
                for combo in productos_combos:
                    # Obtener componentes del combo usando IDs
                    componentes = query_db('''
                        SELECT pb.sku, c.cantidad_necesaria 
                        FROM componentes c
                        JOIN productos_base pb ON c.producto_base_id = pb.id
                        JOIN productos_compuestos pc ON c.producto_compuesto_id = pc.id
                        WHERE pc.sku = %s
                    ''', (combo['sku'],))
                    
                    # Calcular cuántos combos se pueden armar con el stock disponible
                    stock_disponible_combo = 999999  # Empezar con infinito
                    
                    for comp in componentes:
                        sku_comp = comp['sku']
                        cant_necesaria = int(comp['cantidad_necesaria'])
                        
                        # Buscar el stock disponible de este componente
                        prod_comp = next((p for p in productos_procesados if p['sku'] == sku_comp), None)
                        if prod_comp:
                            stock_disp_comp = prod_comp['stock_disponible']
                            # Cuántos combos se pueden hacer con este componente
                            combos_posibles = stock_disp_comp // cant_necesaria if cant_necesaria > 0 else 0
                            # El mínimo define cuántos combos se pueden armar
                            stock_disponible_combo = min(stock_disponible_combo, combos_posibles)
                        else:
                            # Si no existe el componente, no se puede armar el combo
                            stock_disponible_combo = 0
                            break
                    
                    # Si no hay componentes o todos dan infinito, poner 0
                    if stock_disponible_combo == 999999 or stock_disponible_combo < 0:
                        stock_disponible_combo = 0
                    
                    productos_procesados.append({
                        'sku': combo['sku'],
                        'nombre': combo['nombre'],
                        'tipo': combo['tipo'],
                        'stock': 0,  # Los combos no tienen stock físico
                        'stock_disponible': stock_disponible_combo,
                        'tiene_ubicaciones': False,
                        'precio': float(combo.get('precio', 0))
                    })
            
        except Exception as e:
            print(f"Nota: No se pudieron cargar combos - {str(e)}")
            import traceback
            traceback.print_exc()
        
        # ========================================
        # 6. CONVERTIR A JSON (SIN DECIMALS)
        # ========================================
        productos_json = json.dumps(productos_procesados)
        
        return render_template('nueva_venta.html',
                             siguiente_numero=siguiente_numero,
                             fecha_hoy=date.today().strftime('%Y-%m-%d'),
                             productos_json=productos_json)
                             
    except Exception as e:
        import traceback
        error_completo = traceback.format_exc()
        flash(f'Error al cargar Nueva Venta: {str(e)}', 'error')
        print(f"ERROR en nueva_venta:\n{error_completo}")
        return redirect(url_for('index'))



@app.route('/nueva-venta/guardar', methods=['POST'])
@login_required
def guardar_venta():
    """Guardar venta SIN descontar stock (solo registra la venta)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        from flask import session
        from datetime import datetime
        
        # ========================================
        # FECHA DE VENTA
        # ========================================
        fecha_venta_iso = session.get('ml_fecha_venta')
        if fecha_venta_iso:
            fecha_venta = datetime.fromisoformat(fecha_venta_iso)
            print(f"✅ Usando fecha de ML: {fecha_venta}")
        else:
            fecha_venta_form = request.form.get('fecha_venta')
            fecha_venta = datetime.strptime(fecha_venta_form, '%Y-%m-%d') if fecha_venta_form else datetime.now()
            print(f"✅ Usando fecha del formulario: {fecha_venta}")
        
        # ========================================
        # 1. DATOS GENERALES
        # ========================================
        numero_venta = request.form.get('numero_venta')
        canal = request.form.get('canal', 'Mercado Libre')
        mla_code = request.form.get('mla_code', '').strip()
        nombre_cliente = request.form.get('nombre_cliente', '').strip()
        
        if not nombre_cliente:
            nombre_cliente = mla_code if mla_code else 'Cliente sin especificar'
        
        telefono_cliente = request.form.get('telefono_cliente', '')
        
        # ========================================
        # 2. ENTREGA
        # ========================================
        tipo_entrega = request.form.get('tipo_entrega')
        direccion_entrega = request.form.get('direccion_entrega', '')
        metodo_envio = request.form.get('metodo_envio', '')
        zona_envio = request.form.get('zona_envio', '')
        
        if metodo_envio == 'Full':
            ubicacion_despacho = 'FULL'
        else:
            ubicacion_despacho = 'DEP'
        
        responsable_entrega = request.form.get('responsable_entrega', '')
        costo_flete = float(request.form.get('costo_flete', 0))
        
        # ========================================
        # 3. PRODUCTOS - calcular importe_total desde items reales
        # ========================================
        productos_form = request.form.to_dict(flat=False)
        importe_total = 0.0
        
        for key in productos_form.keys():
            if key.startswith('productos[') and key.endswith('[sku]'):
                index = key.split('[')[1].split(']')[0]
                sku = productos_form.get(f'productos[{index}][sku]', [None])[0]
                cantidad = int(productos_form.get(f'productos[{index}][cantidad]', [0])[0])
                precio = float(productos_form.get(f'productos[{index}][precio]', [0])[0])
                if sku and cantidad > 0:
                    importe_total += cantidad * precio
        
        print(f"✅ importe_total calculado desde items: ${importe_total}")
        
        # ========================================
        # 4. PAGO
        # ========================================
        metodo_pago = request.form.get('metodo_pago')
        pago_mercadopago = float(request.form.get('pago_mercadopago', 0))
        pago_efectivo = float(request.form.get('pago_efectivo', 0))
        
        # Para Flete Propio y Zippin de ML, el cliente paga productos + flete por MP
        if canal == 'Mercado Libre' and metodo_envio in ['Flete Propio', 'Zippin'] and costo_flete > 0:
            pago_mercadopago += costo_flete
            print(f"✅ Sumando flete a pago_mercadopago: +${costo_flete} → total ${pago_mercadopago}")
        
        importe_abonado = pago_mercadopago + pago_efectivo
        
        print(f"✅ pago_mercadopago: ${pago_mercadopago}")
        print(f"✅ pago_efectivo: ${pago_efectivo}")
        print(f"✅ importe_abonado: ${importe_abonado}")
        
        # ========================================
        # 5. DATOS DE FACTURACIÓN
        # ========================================
        ml_billing_data = session.get('ml_billing_data')
        billing_info = {
            'business_name': None,
            'doc_type': None,
            'doc_number': None,
            'taxpayer_type': None,
            'city': None,
            'street': None,
            'state': None,
            'zip_code': None
        }
        
        if ml_billing_data:
            billing_info = extraer_billing_info_ml(ml_billing_data)
        
        # ========================================
        # 6. OBSERVACIONES Y ESTADO
        # ========================================
        notas = request.form.get('notas', '')
        estado_entrega = 'pendiente'
        estado_pago = 'pago_pendiente' if importe_abonado < importe_total else 'pagado'
        
        # ========================================
        # 7. INSERTAR VENTA
        # ========================================
        cursor.execute('''
            INSERT INTO ventas (
                numero_venta, fecha_venta, canal, mla_code,
                nombre_cliente, telefono_cliente,
                tipo_entrega, metodo_envio, ubicacion_despacho,
                zona_envio, direccion_entrega, responsable_entrega,
                costo_flete, metodo_pago, importe_total, importe_abonado,
                pago_mercadopago, pago_efectivo,
                estado_entrega, estado_pago, notas,
                factura_business_name, factura_doc_type, factura_doc_number,
                factura_taxpayer_type, factura_city, factura_street,
                factura_state, factura_zip_code
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s
            )
        ''', (
            numero_venta, fecha_venta, canal, mla_code,
            nombre_cliente, telefono_cliente,
            tipo_entrega, metodo_envio, ubicacion_despacho,
            zona_envio, direccion_entrega, responsable_entrega,
            costo_flete, metodo_pago, importe_total, importe_abonado,
            pago_mercadopago, pago_efectivo,
            estado_entrega, estado_pago, notas,
            billing_info['business_name'],
            billing_info['doc_type'],
            billing_info['doc_number'],
            billing_info['taxpayer_type'],
            billing_info['city'],
            billing_info['street'],
            billing_info['state'],
            billing_info['zip_code']
        ))
        
        venta_id = cursor.lastrowid
        
        # ========================================
        # 8. GUARDAR PRODUCTOS
        # ========================================
        items_agregados = 0
        
        for key in productos_form.keys():
            if key.startswith('productos[') and key.endswith('[sku]'):
                index = key.split('[')[1].split(']')[0]
                sku = productos_form.get(f'productos[{index}][sku]', [None])[0]
                cantidad = int(productos_form.get(f'productos[{index}][cantidad]', [0])[0])
                precio = float(productos_form.get(f'productos[{index}][precio]', [0])[0])
                
                if sku and cantidad > 0:
                    cursor.execute('''
                        INSERT INTO items_venta (venta_id, sku, cantidad, precio_unitario)
                        VALUES (%s, %s, %s, %s)
                    ''', (venta_id, sku, cantidad, precio))
                    items_agregados += 1
        
        conn.commit()
        
        # ========================================
        # 9. DETECTAR ALERTAS DE STOCK
        # ========================================
        productos_sin_stock = []
        try:
            items_vendidos_lista = []
            for key in productos_form.keys():
                if key.startswith('productos[') and key.endswith('[sku]'):
                    index = key.split('[')[1].split(']')[0]
                    sku = productos_form.get(f'productos[{index}][sku]', [None])[0]
                    cantidad = int(productos_form.get(f'productos[{index}][cantidad]', [0])[0])
                    if sku and cantidad > 0:
                        items_vendidos_lista.append({'sku': sku, 'cantidad': cantidad})
            
            if items_vendidos_lista:
                productos_sin_stock = detectar_alertas_stock_bajo(cursor, items_vendidos_lista, venta_id)
        except Exception as e_alertas:
            print(f"⚠️ Error al detectar alertas: {str(e_alertas)}")
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # ========================================
        # 10. LIMPIAR SESIÓN DE ML
        # ========================================
        session.pop('ml_orden_id', None)
        session.pop('ml_items', None)
        session.pop('ml_comprador_nombre', None)
        session.pop('ml_comprador_nickname', None)
        session.pop('ml_shipping', None)
        session.pop('ml_fecha_venta', None)
        session.pop('ml_billing_data', None)
        
        # ========================================
        # 11. MENSAJE Y REDIRECCIÓN
        # ========================================
        if productos_sin_stock:
            productos_base = [p for p in productos_sin_stock if p.get('tipo_producto') == 'base']
            combos_afectados = [p for p in productos_sin_stock if p.get('tipo_producto') == 'combo']
            
            mensaje_html = f'''
                <div class="alert alert-success mb-3">
                    <strong>✅ Venta {numero_venta} registrada correctamente</strong>
                </div>
                <div class="alert alert-warning">
                    <h5><i class="bi bi-exclamation-triangle-fill"></i> ⚠️ Alerta de Stock ML</h5>
            '''
            
            if productos_base:
                mensaje_html += '<p class="mb-2"><strong>Productos base sin stock disponible:</strong></p><ul class="mb-3">'
                for prod in productos_base:
                    mensaje_html += f'''
                        <li><strong>{prod['nombre']}</strong> (SKU: {prod['sku']})<br>
                            <small>Stock físico: {prod['stock_fisico']} | Vendido: {prod['vendido']} | Disponible: <span class="text-danger">{prod['stock_disponible']}</span></small>
                        </li>
                    '''
                mensaje_html += '</ul>'
            
            if combos_afectados:
                mensaje_html += '<p class="mb-2"><strong>Combos/Sommiers que NO se pueden armar:</strong></p><ul class="mb-3">'
                for combo in combos_afectados:
                    mensaje_html += f'''
                        <li><strong>{combo['nombre']}</strong> (SKU: {combo['sku']})<br>
                            <small class="text-muted">Falta componente: {combo.get('componente_faltante', 'N/A')}</small>
                        </li>
                    '''
                mensaje_html += '</ul>'
            
            mensaje_html += '<p class="mb-0"><strong>Recordá:</strong> Pausá las publicaciones en ML o cargá más stock.</p></div>'
            flash(mensaje_html, 'alerta_stock')
        else:
            mensaje = f'✅ Venta {numero_venta} registrada'
            if ubicacion_despacho == 'FULL':
                mensaje += ' - Se despachará desde FULL ML'
            else:
                mensaje += ' - Se despachará desde Depósito'
            flash(mensaje, 'success')
        
        return redirect(url_for('ventas_activas'))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        import traceback
        error_completo = traceback.format_exc()
        print(f"ERROR al guardar venta:\n{error_completo}")
        flash(f'❌ Error al guardar venta: {str(e)}', 'error')
        return redirect(url_for('nueva_venta'))



# ============================================================================
# DASHBOARD VISUAL - CORREGIDO CON LÓGICA DE BASES GRANDES
# ============================================================================

@app.route('/dashboard-visual')
@login_required
def dashboard_visual():
    """Dashboard visual - Stock Físico y Ventas Activas con lógica de bases grandes"""
    
    try:
        # =================================
        # 1. OBTENER STOCK FÍSICO (con stock_full para almohadas)
        # =================================
        query_stock = """
            SELECT sku, stock_actual, stock_full, tipo
            FROM productos_base
            ORDER BY tipo, sku
        """
        productos_stock = query_db(query_stock)
        
        # Crear diccionario de stock
        stock_dict = {}
        stock_full_dict = {}
        for prod in productos_stock:
            stock_dict[prod['sku']] = prod['stock_actual'] or 0
            stock_full_dict[prod['sku']] = prod.get('stock_full', 0) or 0
        
        # =================================
        # 2. OBTENER VENTAS ACTIVAS (PENDIENTES)
        # =================================
        query_ventas_activas = """
    SELECT 
        COALESCE(pb_comp.sku, iv.sku) as sku,
        SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as cantidad_vendida
    FROM items_venta iv
    JOIN ventas v ON iv.venta_id = v.id
    LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
    LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
    LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
    WHERE v.estado_entrega = 'pendiente'
    GROUP BY sku
        """
        ventas_activas = query_db(query_ventas_activas)
        
        # Convertir ventas activas a diccionario
        ventas_dict = {}
        for venta in ventas_activas:
            ventas_dict[venta['sku']] = venta['cantidad_vendida']
        
        # =================================
        # 3. FUNCIÓN HELPER PARA OBTENER DATOS
        # =================================
        def get_datos(sku):
            """Obtener stock y vendido para un SKU, con lógica de bases grandes"""
            stock = stock_dict.get(sku, 0)
            vendido = ventas_dict.get(sku, 0)
            
            # LÓGICA DE BASES GRANDES (160, 180, 200)
            # Bases x200: 2 bases físicas = 1 en dashboard
            if sku in ['BASE_CHOC160', 'BASE_GRIS160', 'BASE_SUBL160']:
                sku_base = sku.replace('160', '80200')
                stock = stock_dict.get(sku_base, 0) // 2
                vendido = ventas_dict.get(sku_base, 0) // 2
            elif sku in ['BASE_CHOC180', 'BASE_GRIS180', 'BASE_SUBL180']:
                sku_base = sku.replace('180', '90200')
                stock = stock_dict.get(sku_base, 0) // 2
                vendido = ventas_dict.get(sku_base, 0) // 2
            elif sku in ['BASE_CHOC200', 'BASE_GRIS200', 'BASE_SUBL200']:
                sku_base = sku.replace('200', '100200')
                stock = stock_dict.get(sku_base, 0) // 2
                vendido = ventas_dict.get(sku_base, 0) // 2
            
            return {'stock': stock, 'vendido': vendido}
        
        # =================================
        # 4. LÍNEA ESPUMA
        # =================================
        datos_espuma = {}
        for medida in ['80', '90', '100', '140', '150', '160', '180', '200']:
            datos_espuma[medida] = {
                'CPR20': get_datos(f'CPR{medida}20'),
                'BASE_SAB': get_datos(f'BASE_SAB{medida}'),
                'CEX': get_datos(f'CEX{medida}'),
                'CEXP': get_datos(f'CEXP{medida}'),
                'BASE_CHOC': get_datos(f'BASE_CHOC{medida}'),
                'CPR23': get_datos(f'CPR{medida}23'),
                'CRE': get_datos(f'CRE{medida}'),
                'CREP': get_datos(f'CREP{medida}'),
                'BASE_GRIS': get_datos(f'BASE_GRIS{medida}')
            }
        
        # =================================
        # 5. LÍNEA RESORTE
        # =================================
        datos_resorte = {}
        for medida in ['80', '90', '100', '140', '150', '160', '180', '200']:
            datos_resorte[medida] = {
                'CSO': get_datos(f'CSO{medida}'),
                'BASE_SAB': get_datos(f'BASE_SAB{medida}'),
                'CDO': get_datos(f'CDO{medida}'),
                'CDOP': get_datos(f'CDOP{medida}'),
                'BASE_GRIS': get_datos(f'BASE_GRIS{medida}'),
                'CSUP': get_datos(f'CSUP{medida}'),
                'BASE_SUBL': get_datos(f'BASE_SUBL{medida}')
            }
        
        # =================================
        # 6. OTROS PRODUCTOS
        # =================================
        datos_otros = {}
        for medida in ['80', '90', '100', '140', '150', '160', '180', '200']:
            # Tropical: Solo 80, 90, 100
            tropical = {'stock': 0, 'vendido': 0}
            if medida in ['80', '90', '100']:
                tropical = get_datos(f'CTR{medida}')
            
            # Compac: 80, 100, 140, 160 (CON UBICACIONES _DEP y _FULL)
            compac_dep = {'stock': 0, 'vendido': 0}
            compac_full = {'stock': 0, 'vendido': 0}
            if medida in ['80', '100', '140', '160']:
                compac_dep = get_datos(f'CCO{medida}_DEP')
                compac_full = get_datos(f'CCO{medida}_FULL')
            
            # Compac Plus: 80, 100, 140, 160 (CON UBICACIONES _DEP y _FULL)
            compac_plus_dep = {'stock': 0, 'vendido': 0}
            compac_plus_full = {'stock': 0, 'vendido': 0}
            if medida in ['80', '100', '140', '160']:
                compac_plus_dep = get_datos(f'CCP{medida}_DEP')
                compac_plus_full = get_datos(f'CCP{medida}_FULL')
            
            datos_otros[medida] = {
                'TROPICAL': tropical,
                'COMPAC_DEP': compac_dep,
                'COMPAC_FULL': compac_full,
                'COMPAC_PLUS_DEP': compac_plus_dep,
                'COMPAC_PLUS_FULL': compac_plus_full
            }
        
        # =================================
        # 7. ALMOHADAS (CON UBICACIONES DEP Y FULL)
        # =================================
        almohadas = []
        almohadas_skus = ['CERVICAL', 'CLASICA', 'DORAL', 'DUAL', 'EXCLUSIVE', 'PLATINO', 'RENOVATION', 'SUBLIME']
        
        for sku in almohadas_skus:
            stock_dep = stock_dict.get(sku, 0)
            stock_full = stock_full_dict.get(sku, 0)
            vendido = ventas_dict.get(sku, 0)
            
            # Mostrar si hay stock en cualquier ubicación o si hay ventas
            if stock_dep >= 0 or stock_full >= 0 or vendido >= 0:
                nombre_map = {
                    'CERVICAL': 'Almohada Visco Cervical',
                    'CLASICA': 'Almohada Visco Clásica',
                    'DORAL': 'Almohada Doral',
                    'DUAL': 'Almohada Dual Refreshing',
                    'EXCLUSIVE': 'Almohada Exclusive',
                    'PLATINO': 'Almohada Platino',
                    'RENOVATION': 'Almohada Renovation',
                    'SUBLIME': 'Almohada Sublime'
                }
                almohadas.append({
                    'nombre': nombre_map.get(sku, sku),
                    'stock_dep': stock_dep,
                    'stock_full': stock_full,
                    'vendido': vendido  # Total de ventas (no separadas por ubicación)
                })
        
        # =================================
        # 8. FECHA ACTUAL
        # =================================
        fecha_actual = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        # =================================
        # 9. RENDERIZAR TEMPLATE
        # =================================
        return render_template('dashboard_visual.html',
                             datos_espuma=datos_espuma,
                             datos_resorte=datos_resorte,
                             datos_otros=datos_otros,
                             almohadas=almohadas,
                             fecha_actual=fecha_actual)
    
    except Exception as e:
        flash(f'Error al cargar dashboard visual: {str(e)}', 'error')
        return redirect(url_for('index'))


# ============================================================================
# CONFIGURACIÓN - RESETEO DE SISTEMA
# ============================================================================

@app.route('/configuracion/resetear')
@login_required
def resetear_sistema():
    """Página para resetear el sistema (requiere contraseña)"""
    return render_template('resetear_sistema.html')


@app.route('/configuracion/resetear/ejecutar', methods=['POST'])
@login_required
def ejecutar_reseteo():
    """Ejecutar reseteo completo del sistema"""
    from flask import jsonify
    
    # Verificar contraseña
    password = request.form.get('password', '')
    confirmar = request.form.get('confirmar', '')
    
    # Contraseña configurada (CAMBIAR ESTO por tu contraseña)
    PASSWORD_CORRECTA = '32267845'
    
    if password != PASSWORD_CORRECTA:
        flash('❌ Contraseña incorrecta', 'error')
        return redirect(url_for('resetear_sistema'))
    
    if confirmar != 'RESETEAR':
        flash('❌ Debes escribir "RESETEAR" para confirmar', 'error')
        return redirect(url_for('resetear_sistema'))
    
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        
        # ==================================
        # 1. BORRAR TODAS LAS VENTAS
        # ==================================
        cursor.execute('DELETE FROM items_venta')
        items_borrados = cursor.rowcount
        
        cursor.execute('DELETE FROM ventas')
        ventas_borradas = cursor.rowcount
        
        # ==================================
        # 2. BORRAR MOVIMIENTOS DE STOCK
        # ==================================
        cursor.execute('DELETE FROM movimientos_stock')
        movimientos_borrados = cursor.rowcount
        
        # ==================================
        # 3. RESETEAR STOCK A 0
        # ==================================
        cursor.execute('''
            UPDATE productos_base 
            SET stock_actual = 0, 
                stock_full = 0
        ''')
        productos_reseteados = cursor.rowcount
        
        # ==================================
        # 4. BORRAR ALERTAS DE STOCK
        # ==================================
        try:
            cursor.execute('DELETE FROM alertas_stock')
            alertas_borradas = cursor.rowcount
        except:
            alertas_borradas = 0
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # Mensaje de confirmación
        mensaje = f'''
        ✅ <strong>Sistema reseteado correctamente</strong><br><br>
        📊 <strong>Operaciones realizadas:</strong><br>
        • {ventas_borradas} ventas eliminadas<br>
        • {items_borrados} items de venta eliminados<br>
        • {movimientos_borrados} movimientos de stock eliminados<br>
        • {productos_reseteados} productos reseteados a stock 0<br>
        • {alertas_borradas} alertas eliminadas<br><br>
        🎯 El sistema está listo para comenzar de cero.
        '''
        
        flash(mensaje, 'success')
        return redirect(url_for('index'))
        
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        flash(f'❌ Error al resetear: {str(e)}', 'error')
        return redirect(url_for('resetear_sistema'))

# ============================================================================
# FUNCIONES AUXILIARES: MERCADO LIBRE
# Agregar ANTES del if __name__ == '__main__' al final de app.py
# ============================================================================

import time  # Agregar este import si no lo tenés

def refresh_ml_token():
    """Renovar el access_token usando el refresh_token guardado en DB"""
    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'ml_token'")
        if not row:
            return None
        data = json.loads(row['valor'])

        refresh_token = data.get('refresh_token')
        client_id = data.get('client_id')
        client_secret = data.get('client_secret')

        if not refresh_token or not client_id or not client_secret:
            print("⚠️  No hay refresh_token o credenciales guardadas")
            return None

        print("🔄 Renovando access_token de ML...")
        response = requests.post(
            "https://api.mercadolibre.com/oauth/token",
            data={
                "grant_type": "refresh_token",
                "client_id": client_id,
                "client_secret": client_secret,
                "refresh_token": refresh_token
            }
        )

        if response.status_code == 200:
            new_data = response.json()
            data['access_token'] = new_data.get('access_token')
            data['refresh_token'] = new_data.get('refresh_token', refresh_token)
            data['expires_at'] = time.time() + new_data.get('expires_in', 21600) - 300
            execute_db(
                "INSERT INTO configuracion (clave, valor) VALUES ('ml_token', %s) "
                "ON DUPLICATE KEY UPDATE valor = %s, actualizado_at = NOW()",
                (json.dumps(data), json.dumps(data))
            )
            print("✅ Token renovado automáticamente!")
            return data['access_token']
        else:
            print(f"❌ Error renovando token: {response.status_code} - {response.json()}")
            return None
    except Exception as e:
        print(f"Error en refresh_ml_token: {e}")
        return None

def cargar_ml_token():
    """Cargar token ML desde la base de datos. Si está vencido, lo renueva."""
    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'ml_token'")
        if not row:
            return None
        data = json.loads(row['valor'])
        access_token = data.get('access_token')
        expires_at = data.get('expires_at', 0)
        if expires_at and time.time() > expires_at:
            print("⚠️  Token ML vencido, renovando automáticamente...")
            access_token = refresh_ml_token()
        return access_token
    except Exception as e:
        print(f"Error cargando token ML: {e}")
        return None


def guardar_ml_token(token_data):
    """Guardar token ML en la base de datos (persiste en Railway)"""
    try:
        # Preservar refresh_token y credenciales si ya existen en DB
        existing_json = query_one("SELECT valor FROM configuracion WHERE clave = 'ml_token'")
        if existing_json:
            existing = json.loads(existing_json['valor'])
            if 'refresh_token' not in token_data and 'refresh_token' in existing:
                token_data['refresh_token'] = existing['refresh_token']
            if 'client_id' not in token_data and 'client_id' in existing:
                token_data['client_id'] = existing['client_id']
            if 'client_secret' not in token_data and 'client_secret' in existing:
                token_data['client_secret'] = existing['client_secret']

        if 'refresh_token' not in token_data:
            token_data.pop('expires_at', None)

        execute_db(
            "INSERT INTO configuracion (clave, valor) VALUES ('ml_token', %s) "
            "ON DUPLICATE KEY UPDATE valor = %s, actualizado_at = NOW()",
            (json.dumps(token_data), json.dumps(token_data))
        )
        return True
    except Exception as e:
        print(f"Error guardando token ML: {e}")
        return False


def obtener_ordenes_ml(access_token, limit=20):
    """
    Obtener órdenes de Mercado Libre
    Retorna: (success, data_o_error)
    """
    headers = {'Authorization': f'Bearer {access_token}'}
    
    try:
        # 1. Obtener USER_ID
        user_response = requests.get('https://api.mercadolibre.com/users/me', headers=headers)
        
        if user_response.status_code != 200:
            return False, f"Error obteniendo usuario: {user_response.status_code}"
        
        user_id = user_response.json()['id']
        
        # 2. Buscar órdenes como vendedor
        orders_url = f"https://api.mercadolibre.com/orders/search?seller={user_id}&sort=date_desc&limit={limit}"
        orders_response = requests.get(orders_url, headers=headers)
        
        if orders_response.status_code != 200:
            return False, f"Error obteniendo órdenes: {orders_response.status_code}"
        
        orders_data = orders_response.json()
        return True, orders_data['results']
        
    except Exception as e:
        return False, str(e)


def obtener_shipping_details(access_token, shipping_id):
    """
    Obtener detalles del envío desde ML
    Retorna: (success, data_o_error)
    """
    if not shipping_id:
        return False, "No hay shipping_id"
    
    headers = {'Authorization': f'Bearer {access_token}'}
    
    try:
        response = requests.get(f'https://api.mercadolibre.com/shipments/{shipping_id}', headers=headers)
        
        if response.status_code == 200:
            return True, response.json()
        else:
            return False, f"Error {response.status_code}"
    except Exception as e:
        return False, str(e)

# ─── FUNCIÓN 4: Mapeo automático al importar - quita la Z del SKU ───
def normalizar_sku_ml(sku_ml):
    """
    Quitar la Z del final del SKU de ML si existe
    Ejemplo: CEX140Z → CEX140
    """
    if sku_ml and sku_ml.endswith('Z'):
        return sku_ml[:-1]  # Quitar último carácter
    return sku_ml


# REEMPLAZAR la función procesar_orden_ml completa
def procesar_orden_ml(orden):
    """
    Procesar orden de ML SIN obtener detalles de shipping
    Usar al LISTAR órdenes (más rápido)
    CAPTURA FECHA REAL DE VENTA y COSTO DE ENVÍO
    """
    # Fecha REAL de la venta en ML
    fecha = datetime.fromisoformat(orden['date_created'].replace('Z', '+00:00'))
    
    # Items/Productos
    items = []
    for item in orden['order_items']:
        items.append({
            'sku': item['item'].get('seller_sku', ''),
            'titulo': item['item']['title'],
            'cantidad': item['quantity'],
            'precio': item['unit_price']
        })
    
    # Comprador
    buyer = orden.get('buyer', {})
    comprador_nombre = f"{buyer.get('first_name', '')} {buyer.get('last_name', '')}".strip()
    comprador_nickname = buyer.get('nickname', '')
    
    # Total
    total = orden['total_amount']
    
    # Estado
    estado = orden['status']
    
    # Shipping (solo ID, sin detalles) + COSTO DE ENVÍO
    shipping = orden.get('shipping', {})
    shipping_id = shipping.get('id', '')
    costo_envio = shipping.get('shipping_cost', 0)  # ✅ NUEVO: Capturar costo de envío
    
    shipping_data = {
        'tiene_envio': bool(shipping_id),
        'shipping_id': shipping_id,
        'costo_envio': costo_envio,  # ✅ NUEVO
        'metodo_envio': '',
        'direccion': '',
        'ciudad': '',
        'provincia': '',
        'codigo_postal': '',
        'zona': ''
    }
    
    return {
        'id': orden['id'],
        'fecha': fecha,  # ✅ Fecha real de ML
        'comprador_nombre': comprador_nombre,
        'comprador_nickname': comprador_nickname,
        'items': items,
        'total': total,
        'estado': estado,
        'shipping': shipping_data
    }

def obtener_shipping_completo(shipping_id, access_token):
    """
    Obtener detalles completos de shipping desde ML
    MAPEO CORREGIDO según tipos reales de ML
    ✅ NUEVO: Captura COSTO DE ENVÍO del shipment
    """
    shipping_data = {
        'tiene_envio': True,
        'shipping_id': shipping_id,
        'metodo_envio': '',
        'metodo_envio_ml': '',
        'logistic_type_ml': '',
        'costo_envio': 0,  # ✅ NUEVO: Inicializar
        'direccion': '',
        'ciudad': '',
        'provincia': '',
        'codigo_postal': '',
        'zona': ''
    }
    
    if not shipping_id or not access_token:
        return shipping_data
    
    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        response = requests.get(f'https://api.mercadolibre.com/shipments/{shipping_id}', headers=headers)
        
        if response.status_code != 200:
            print(f"⚠️ Error al obtener shipment {shipping_id}: {response.status_code}")
            return shipping_data
        
        shipment = response.json()
        
        # ✅ NUEVO: Capturar COSTO DE ENVÍO
        # Puede estar en varios lugares según el tipo de envío
        costo_envio = 0
        
        # Opción 1: base_cost
        if 'base_cost' in shipment:
            costo_envio = shipment.get('base_cost', 0)
            print(f"💰 Costo envío (base_cost): ${costo_envio}")
        
        # Opción 2: shipping_cost
        elif 'shipping_cost' in shipment:
            costo_envio = shipment.get('shipping_cost', 0)
            print(f"💰 Costo envío (shipping_cost): ${costo_envio}")
        
        # Opción 3: cost
        elif 'cost' in shipment:
            costo_envio = shipment.get('cost', 0)
            print(f"💰 Costo envío (cost): ${costo_envio}")
        
        # Opción 4: shipping_option -> cost
        elif 'shipping_option' in shipment:
            shipping_option = shipment.get('shipping_option', {})
            if 'cost' in shipping_option:
                costo_envio = shipping_option.get('cost', 0)
                print(f"💰 Costo envío (shipping_option.cost): ${costo_envio}")
        
        shipping_data['costo_envio'] = costo_envio
        
        # Método de envío
        shipping_option = shipment.get('shipping_option', {})
        shipping_mode = shipping_option.get('shipping_method_id', '')
        logistic_type = shipment.get('logistic_type', '')
        
        # 🔍 DEBUGGING
        print(f"\n🚚 SHIPPING ID: {shipping_id}")
        print(f"📦 shipping_method_id: {shipping_mode}")
        print(f"📦 logistic_type: {logistic_type}")
        print(f"💰 COSTO TOTAL: ${costo_envio}")
        
        # Guardar valores originales
        shipping_data['metodo_envio_ml'] = shipping_mode
        shipping_data['logistic_type_ml'] = logistic_type
        
        # 🔧 MAPEO CORREGIDO según logs reales
        if logistic_type == 'fulfillment':
            shipping_data['metodo_envio'] = 'Full'
            print(f"✅ MAPEADO A: Full")
        
        elif logistic_type == 'self_service':
            shipping_data['metodo_envio'] = 'Flex'
            print(f"✅ MAPEADO A: Flex")
        
        elif logistic_type == 'xd_drop_off':
            shipping_data['metodo_envio'] = 'Mercadoenvios'
            print(f"✅ MAPEADO A: Mercadoenvios")
        
        elif logistic_type == 'cross_docking':
            shipping_data['metodo_envio'] = 'Mercadoenvios'
            print(f"✅ MAPEADO A: Flex")
        
        elif logistic_type == 'default':
            # Default depende de la zona (Flete propio o Zippin)
            # Por ahora dejamos como Flete Propio y luego se puede ajustar manualmente
            shipping_data['metodo_envio'] = 'Flete Propio'
            print(f"⚠️ MAPEADO A: Flete Propio (default - ajustar según zona)")
        
        elif 'mercadoenvios' in str(shipping_mode).lower():
            shipping_data['metodo_envio'] = 'Mercadoenvios'
            print(f"✅ MAPEADO A: Mercadoenvios")
        
        else:
            shipping_data['metodo_envio'] = 'Mercadoenvios'
            print(f"⚠️ MAPEADO A: Mercadoenvios (default)")
        
        # Dirección
        receiver_address = shipment.get('receiver_address', {})
        
        if receiver_address:
            address_line = receiver_address.get('address_line', '')
            street_name = receiver_address.get('street_name', '')
            street_number = receiver_address.get('street_number', '')
            floor = receiver_address.get('floor', '')
            apartment = receiver_address.get('apartment', '')
            
            if address_line:
                shipping_data['direccion'] = address_line
            elif street_name and street_number:
                direccion = f"{street_name} {street_number}"
                if floor:
                    direccion += f" Piso {floor}"
                if apartment:
                    direccion += f" Depto {apartment}"
                shipping_data['direccion'] = direccion
            
            # Ciudad y provincia
            city = receiver_address.get('city', {})
            state = receiver_address.get('state', {})
            
            if isinstance(city, dict):
                shipping_data['ciudad'] = str(city.get('name', ''))
            else:
                shipping_data['ciudad'] = str(city) if city else ''
            
            if isinstance(state, dict):
                shipping_data['provincia'] = str(state.get('name', ''))
            else:
                shipping_data['provincia'] = str(state) if state else ''
            
            shipping_data['codigo_postal'] = str(receiver_address.get('zip_code', ''))
            
            # Inferir zona
            if shipping_data['ciudad']:
                ciudad_lower = shipping_data['ciudad'].lower()
                provincia_lower = shipping_data['provincia'].lower()
                
                if 'capital federal' in ciudad_lower or 'ciudad' in ciudad_lower or 'caba' in ciudad_lower or 'autonoma' in provincia_lower:
                    shipping_data['zona'] = 'Capital'
                elif any(x in ciudad_lower for x in ['plata', 'quilmes', 'avellaneda', 'berazategui', 'florencio varela', 'lanus']):
                    shipping_data['zona'] = 'Sur'
                elif any(x in ciudad_lower for x in ['san isidro', 'tigre', 'pilar', 'escobar', 'san fernando']):
                    shipping_data['zona'] = 'Norte-Noroeste'
                elif any(x in ciudad_lower for x in ['moron', 'merlo', 'ituzaingo', 'hurlingham', 'moreno']):
                    shipping_data['zona'] = 'Oeste'
    
    except Exception as e:
        print(f"⚠️ Error al procesar shipping {shipping_id}: {str(e)}")
        import traceback
        traceback.print_exc()
    
    return shipping_data


def verificar_sku_en_bd(sku):
    """
    Verificar si un SKU existe en la base de datos
    Retorna: (existe, tipo_producto, nombre)
    """
    # Buscar en productos_base
    prod = query_one('SELECT tipo, nombre FROM productos_base WHERE sku = %s', (sku,))
    if prod:
        return True, prod['tipo'], prod['nombre']
    
    # Buscar en productos_compuestos
    combo = query_one('SELECT nombre FROM productos_compuestos WHERE sku = %s', (sku,))
    if combo:
        return True, 'combo', combo['nombre']
    
    return False, None, None


# ─── FUNCIÓN 1: Actualizar handling_time (tiempo de demora) en ML ───
def actualizar_handling_time_ml(mla_id, dias, access_token):
    """
    Actualizar el tiempo de disponibilidad (handling_time) en ML
    
    Args:
        mla_id: ID de la publicación
        dias: Cantidad de días de demora
        access_token: Token de ML
    
    Returns:
        (success: bool, message: str)
    """
    try:
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        # IMPORTANTE: Primero traer la publicación actual
        response_get = requests.get(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers
        )
        
        if response_get.status_code != 200:
            return False, "Error obteniendo publicación"
        
        # Actualizar solo handling_time
        data = {
            "sale_terms": [
                {
                    "id": "MANUFACTURING_TIME",
                    "value_name": f"{dias} días"
                }
            ]
        }
        
        response = requests.put(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers,
            json=data
        )
        
        if response.status_code == 200:
            return True, f"Demora configurada a {dias} días"
        else:
            error_data = response.json()
            error_msg = error_data.get('message', 'Error desconocido')
            return False, f"Error ML: {error_msg}"
    
    except Exception as e:
        return False, f"Error: {str(e)}"

# ============================================================================
# FUNCIÓN HELPER PARA QUITAR DEMORA (basada en la que funciona)
# ============================================================================

def quitar_handling_time_ml(mla_id, access_token):
    """
    Quitar el tiempo de disponibilidad (handling_time) en ML
    CON LOGGING DETALLADO
    """
    try:
        headers = {
            'Authorization': f'Bearer {access_token}',
            'Content-Type': 'application/json'
        }
        
        # GET de la publicación
        response_get = requests.get(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers
        )
        
        if response_get.status_code != 200:
            return False, f"Error obteniendo publicación: {response_get.status_code}"
        
        item_data = response_get.json()
        
        # VER QUÉ SALE_TERMS TIENE
        sale_terms_antes = item_data.get('sale_terms', [])
        print(f"\n=== DEBUG {mla_id} ===")
        print(f"sale_terms ANTES: {sale_terms_antes}")
        
        # Filtrar para quitar MANUFACTURING_TIME
        sale_terms_despues = [
            term for term in sale_terms_antes
            if term.get('id') != 'MANUFACTURING_TIME'
        ]
        
        print(f"sale_terms DESPUÉS: {sale_terms_despues}")
        
        # Enviar actualización
        data = {
            "sale_terms": sale_terms_despues
        }
        
        print(f"Enviando: {data}")
        
        response = requests.put(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers,
            json=data
        )
        
        print(f"Response status: {response.status_code}")
        print(f"Response body: {response.text[:500]}")  # Primeros 500 chars
        
        if response.status_code == 200:
            # VERIFICAR si realmente se quitó
            response_check = requests.get(
                f'https://api.mercadolibre.com/items/{mla_id}',
                headers=headers
            )
            
            if response_check.status_code == 200:
                item_actualizado = response_check.json()
                sale_terms_final = item_actualizado.get('sale_terms', [])
                print(f"sale_terms FINAL (después de actualizar): {sale_terms_final}")
                
                # Verificar si todavía tiene MANUFACTURING_TIME
                tiene_demora = any(
                    term.get('id') == 'MANUFACTURING_TIME' 
                    for term in sale_terms_final
                )
                
                if tiene_demora:
                    return False, f"ML aceptó el cambio pero la demora sigue ahí"
                else:
                    return True, f"Demora quitada correctamente de {mla_id}"
            
            return True, f"Actualizado (sin verificar)"
        else:
            error_data = response.json() if response.text else {}
            error_msg = error_data.get('message', 'Error desconocido')
            return False, f"Error ML: {error_msg}"
    
    except Exception as e:
        print(f"Exception: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, f"Error: {str(e)}"



# ============================================================================
# RUTAS: MERCADO LIBRE
# Agregar DESPUÉS de la ruta /ventas/activas en app.py
# ============================================================================

# ─── FUNCIÓN 2: Sincronizar variantes con Z (demora) ───
@app.route('/alertas/<int:alerta_id>/configurar-demora-ml', methods=['POST'])
@login_required
def configurar_demora_ml_desde_alerta(alerta_id):
    """
    Configurar días de demora en las publicaciones CON Z
    Procesa solo la parte CON Z de la alerta
    """
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay token de ML configurado', 'error')
        return redirect(url_for('alertas_ml'))
    
    # Obtener días de demora del formulario
    dias_demora = request.form.get('dias_demora', type=int)
    
    if not dias_demora or dias_demora < 1 or dias_demora > 90:
        flash('❌ Los días de demora deben estar entre 1 y 90', 'error')
        return redirect(url_for('alertas_ml'))
    
    try:
        # Obtener la alerta
        alerta = query_db('SELECT * FROM alertas_stock WHERE id = %s', (alerta_id,))
        if alerta:
            alerta = alerta[0]
        
        if not alerta:
            flash('❌ Alerta no encontrada', 'error')
            return redirect(url_for('alertas_ml'))
        
        sku = alerta['sku']
        sku_con_z = f"{sku}Z"
        
        # Obtener publicaciones CON Z
        publicaciones = query_db(
            'SELECT * FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE',
            (sku_con_z,)
        )
        
        if not publicaciones:
            flash(f'⚠️ No hay publicaciones con Z mapeadas para {sku_con_z}', 'warning')
            return redirect(url_for('alertas_ml'))
        
        # Actualizar demora en cada publicación
        resultados = []
        errores = []
        
        for pub in publicaciones:
            mla_id = pub['mla_id']
            success, message = actualizar_handling_time_ml(mla_id, dias_demora, access_token)
            
            if success:
                resultados.append(f"{mla_id}: {message}")
            else:
                errores.append(f"{mla_id}: {message}")
        
        # ✅ LÓGICA DE PROCESAMIENTO INDEPENDIENTE
        tiene_variante_normal = len(query_db(
            'SELECT 1 FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE LIMIT 1',
            (sku,)
        )) > 0
        
        tipo_procesado_actual = alerta.get('tipo_procesado')
        
        if tiene_variante_normal:
            # Tiene variante normal - solo marcar que procesamos la parte Z
            if tipo_procesado_actual == 'normal':
                # Ya se había procesado normal → ahora marcar como ambos y cerrar alerta
                execute_db(
                    "UPDATE alertas_stock SET tipo_procesado = 'ambos', estado = 'procesada', fecha_procesada = NOW() WHERE id = %s",
                    (alerta_id,)
                )
            else:
                # Solo marcar que se procesó la parte Z
                execute_db(
                    "UPDATE alertas_stock SET tipo_procesado = 'z' WHERE id = %s",
                    (alerta_id,)
                )
        else:
            # No tiene variante normal - cerrar la alerta directamente
            execute_db(
                "UPDATE alertas_stock SET tipo_procesado = 'ambos', estado = 'procesada', fecha_procesada = NOW() WHERE id = %s",
                (alerta_id,)
            )
        
        # Mostrar resultados
        if resultados:
            flash(f'✅ Demora configurada en ML: {", ".join(resultados)}', 'success')
        
        if errores:
            flash(f'❌ Errores: {", ".join(errores)}', 'error')
    
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('alertas_ml'))

@app.route('/ventas/ml/configurar_token', methods=['GET', 'POST'])
@login_required
def ml_configurar_token():
    """
    Página para configurar/actualizar el token de ML
    Maneja 3 acciones:
    - GET: Mostrar página con URL de autorización
    - POST action=canjear_code: Canjear el code por access_token
    - POST action=token_manual: Guardar token manual (como antes)
    """
    
    # Credenciales de la app ML
    CLIENT_ID = os.getenv('ML_APP_ID')
    CLIENT_SECRET = os.getenv('ML_SECRET_KEY')
    REDIRECT_URI = os.getenv('ML_REDIRECT_URI')
    
    # Generar URL de autorización
    url_autorizacion = (
        f"https://auth.mercadolibre.com.ar/authorization"
        f"?response_type=code"
        f"&client_id={CLIENT_ID}"
        f"&redirect_uri={REDIRECT_URI}"
        f"&scope=offline_access"
    )
    
    if request.method == 'POST':
        accion = request.form.get('action', 'token_manual')
        
        # ─── OPCIÓN 1: Canjear code por token automáticamente ───
        if accion == 'canjear_code':
            code = request.form.get('code', '').strip()
            
            if not code:
                flash('❌ Debes ingresar el CODE de la URL de redirección', 'error')
                return redirect(url_for('ml_configurar_token'))
            
            try:
                response = requests.post(
                    "https://api.mercadolibre.com/oauth/token",
                    data={
                        "grant_type": "authorization_code",
                        "client_id": CLIENT_ID,
                        "client_secret": CLIENT_SECRET,
                        "code": code,
                        "redirect_uri": REDIRECT_URI
                    }
                )
                
                if response.status_code == 200:
                    data = response.json()
                    
                    token_data = {
                        "access_token": data.get("access_token"),
                        "refresh_token": data.get("refresh_token"),  # None si ML no lo da
                        "expires_at": time.time() + data.get("expires_in", 21600) - 300,
                        "client_id": CLIENT_ID,
                        "client_secret": CLIENT_SECRET
                    }
                    
                    if guardar_ml_token(token_data):
                        if token_data.get("refresh_token"):
                            flash('✅ Token configurado con auto-renovación activada', 'success')
                        else:
                            flash('✅ Token configurado correctamente (válido por 6 horas)', 'success')
                        return redirect(url_for('ventas_activas'))
                    else:
                        flash('❌ Error al guardar el token', 'error')
                
                else:
                    error_msg = response.json().get('message', 'Error desconocido')
                    flash(f'❌ Error al canjear el code: {error_msg}', 'error')
                    
            except Exception as e:
                flash(f'❌ Error: {str(e)}', 'error')
            
            return redirect(url_for('ml_configurar_token'))
        
        # ─── OPCIÓN 2: Pegar token manual (como antes) ───
        else:
            access_token = request.form.get('access_token', '').strip()
            
            if not access_token:
                flash('❌ Debes ingresar un ACCESS_TOKEN', 'error')
                return redirect(url_for('ml_configurar_token'))
            
            token_data = {
                'access_token': access_token,
                'fecha_configuracion': datetime.now().isoformat()
            }
            
            if guardar_ml_token(token_data):
                flash('✅ Token configurado correctamente', 'success')
                return redirect(url_for('ventas_activas'))
            else:
                flash('❌ Error al guardar token', 'error')
    
    # Verificar estado actual del token
    token_actual = cargar_ml_token()
    
    # Ver si tiene auto-refresh
    tiene_refresh = False
    horas_restantes = None
    try:
        token_path = 'config/ml_token.json'
        if os.path.exists(token_path):
            with open(token_path, 'r') as f:
                data = json.load(f)
                tiene_refresh = bool(data.get('refresh_token'))
                expires_at = data.get('expires_at')
                if expires_at:
                    segundos = expires_at - time.time()
                    if segundos > 0:
                        horas_restantes = round(segundos / 3600, 1)
    except:
        pass
    
    return render_template('ml_configurar_token.html',
                          token_actual=token_actual,
                          url_autorizacion=url_autorizacion,
                          tiene_refresh=tiene_refresh,
                          horas_restantes=horas_restantes)



@app.route('/ventas/ml/importar')
@login_required
def ml_importar_ordenes():
    """
    Traer órdenes de ML - FILTRO ARREGLADO
    """
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay ACCESS_TOKEN configurado. Configuralo primero.', 'error')
        return redirect(url_for('ml_configurar_token'))
    
    # Obtener órdenes de ML
    success, result = obtener_ordenes_ml(access_token, limit=50)
    
    if not success:
        flash(f'❌ Error al obtener órdenes de ML: {result}', 'error')
        return redirect(url_for('ventas_activas'))
    
    # 🔧 FILTRO ARREGLADO
    ordenes_importadas = set()
    try:
        # Buscar en TODAS las ventas (no solo las que empiezan con ML-)
        # porque el numero_venta puede estar en cualquier formato
        todas_ventas = query_db('''
            SELECT id, numero_venta, fecha_venta, canal, mla_code 
            FROM ventas 
            ORDER BY fecha_venta DESC 
            LIMIT 200
        ''')
        
        print(f"\n📊 DEBUG FILTRO DE DUPLICADOS:")
        print(f"Total ventas en BD (últimas 200): {len(todas_ventas)}")
        
        # Revisar cada venta y ver si el numero_venta contiene un ID de ML
        for venta in todas_ventas:
            numero = venta['numero_venta']
            
            # Extraer números del numero_venta
            # Puede ser: "ML-2000015174126338", "2000015174126338", etc.
            import re
            numeros = re.findall(r'\d+', numero)
            
            for num in numeros:
                # Los IDs de ML tienen 16 dígitos y empiezan con 2000
                if len(num) == 16 and num.startswith('2000'):
                    ordenes_importadas.add(num)
                    print(f"   ✅ Encontrado: {numero} → ID ML: {num}")
        
        print(f"\n✅ Total IDs de ML ya importados: {len(ordenes_importadas)}\n")
        
    except Exception as e:
        print(f"⚠️ Error al obtener órdenes importadas: {e}")
        import traceback
        traceback.print_exc()
    
    # Procesar órdenes
    ordenes_procesadas = []
    ordenes_filtradas = 0
    
    for orden in result:
        orden_id = str(orden['id'])
        
        print(f"🔍 Orden ML: {orden_id}")
        
        if orden_id in ordenes_importadas:
            print(f"   ⏭️ YA IMPORTADA - Saltando")
            ordenes_filtradas += 1
            continue
        else:
            print(f"   ✅ NUEVA - Agregando")
        
        # Solo mostrar órdenes pagadas
        if orden['status'] in ['paid']:
            orden_data = procesar_orden_ml(orden)
            
            # Verificar SKU
            for item in orden_data['items']:
                sku = item['sku']
                if sku:
                    existe, tipo, nombre = verificar_sku_en_bd(sku)
                    item['existe_en_bd'] = existe
                    item['tipo_producto'] = tipo
                    item['nombre_bd'] = nombre
                else:
                    item['existe_en_bd'] = False
                    item['tipo_producto'] = None
                    item['nombre_bd'] = None
            
            ordenes_procesadas.append(orden_data)
    
    print(f"\n📊 RESUMEN:")
    print(f"   Órdenes en ML: {len(result)}")
    print(f"   Órdenes filtradas: {ordenes_filtradas}")
    print(f"   Órdenes a mostrar: {len(ordenes_procesadas)}\n")
    
    return render_template('ml_importar_ordenes.html', ordenes=ordenes_procesadas)


@app.route('/ventas/ml/seleccionar/<orden_id>')
@login_required
def ml_seleccionar_orden(orden_id):
    """
    Seleccionar orden - Con normalización automática de SKU (quita Z)
    GUARDA FECHA REAL DE VENTA Y DATOS DE FACTURACIÓN EN SESIÓN
    ✅ CORREGIDO: No sobrescribe costo_envio (ya viene del shipment)
    """
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay ACCESS_TOKEN configurado', 'error')
        return redirect(url_for('ml_configurar_token'))
    
    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        response = requests.get(f'https://api.mercadolibre.com/orders/{orden_id}', headers=headers)
        
        if response.status_code != 200:
            flash('❌ Error al obtener orden de ML', 'error')
            return redirect(url_for('ventas_activas'))
        
        orden = response.json()
        orden_data = procesar_orden_ml(orden)
        
        # OBTENER SHIPPING COMPLETO
        if orden_data['shipping']['shipping_id']:
            shipping_completo = obtener_shipping_completo(
                orden_data['shipping']['shipping_id'],
                access_token
            )
            # ✅ CORREGIDO: Ya no sobrescribimos costo_envio
            # porque obtener_shipping_completo() ya lo captura del shipment
            orden_data['shipping'] = shipping_completo
        
        # ✅ NUEVO: OBTENER BILLING INFO (FACTURACIÓN)
        billing_data = None
        try:
            billing_response = requests.get(
                f'https://api.mercadolibre.com/orders/{orden_id}/billing_info',
                headers=headers
            )
            
            if billing_response.status_code == 200:
                billing_data = billing_response.json()
                print(f"✅ Billing info obtenida para orden {orden_id}")
            else:
                print(f"ℹ️ No hay billing info para orden {orden_id} (Consumidor Final)")
        
        except Exception as e:
            print(f"⚠️ Error al obtener billing info: {str(e)}")
        
        # Verificar mapeo de SKU CON NORMALIZACIÓN AUTOMÁTICA
        items_sin_mapear = []
        items_mapeados = []
        
        for item in orden_data['items']:
            sku_ml_original = item['sku']
            if sku_ml_original:
                existe, tipo, nombre = verificar_sku_en_bd(sku_ml_original)
                sku_a_usar = sku_ml_original
                
                if not existe and sku_ml_original.endswith('Z'):
                    sku_normalizado = sku_ml_original[:-1]
                    existe, tipo, nombre = verificar_sku_en_bd(sku_normalizado)
                    if existe:
                        sku_a_usar = sku_normalizado
                        print(f"✅ Mapeo automático: {sku_ml_original} → {sku_normalizado}")
                
                if existe:
                    items_mapeados.append({
                        'sku_ml': sku_ml_original,
                        'sku_bd': sku_a_usar,
                        'titulo': item['titulo'],
                        'cantidad': item['cantidad'],
                        'precio': item['precio'],
                        'nombre_bd': nombre
                    })
                else:
                    items_sin_mapear.append(item)
            else:
                items_sin_mapear.append(item)
        
        # Mapeo manual si hay productos sin mapear
        if items_sin_mapear:
            productos_bd = query_db('SELECT sku, nombre, tipo FROM productos_base ORDER BY nombre')
            combos_bd = query_db('SELECT sku, nombre FROM productos_compuestos ORDER BY nombre')
            
            return render_template('ml_mapear_productos.html',
                                 orden_id=orden_id,
                                 items_sin_mapear=items_sin_mapear,
                                 items_mapeados=items_mapeados,
                                 productos_bd=productos_bd,
                                 combos_bd=combos_bd,
                                 orden_data=orden_data)
        
        # ✅ GUARDAR EN SESIÓN CON FECHA REAL DE ML Y BILLING INFO
        from flask import session
        session['ml_orden_id'] = orden_id
        session['ml_items'] = items_mapeados
        session['ml_comprador_nombre'] = orden_data['comprador_nombre']
        session['ml_comprador_nickname'] = orden_data['comprador_nickname']
        session['ml_shipping'] = orden_data['shipping']
        session['ml_fecha_venta'] = orden_data['fecha'].isoformat()
        
        # ✅ NUEVO: Guardar billing info en sesión
        if billing_data:
            session['ml_billing_data'] = billing_data
        else:
            session['ml_billing_data'] = None
        
        flash('✅ Productos mapeados correctamente', 'success')
        return redirect(url_for('nueva_venta_desde_ml'))
        
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('ventas_activas'))


@app.route('/ventas/ml/mapear', methods=['POST'])
@login_required
def ml_guardar_mapeo():
    """
    Guardar mapeo - Obtiene shipping completo y billing info
    Con normalización automática de SKU (quita Z)
    GUARDA FECHA REAL DE VENTA Y DATOS DE FACTURACIÓN EN SESIÓN
    ✅ CORREGIDO: No sobrescribe costo_envio
    """
    orden_id = request.form.get('orden_id')
    items_mapeados = json.loads(request.form.get('items_mapeados', '[]'))
    items_form = request.form.getlist('item_sku_ml')
    
    for i, sku_ml in enumerate(items_form):
        sku_bd = request.form.get(f'mapeo_{i}')
        titulo = request.form.get(f'titulo_{i}')
        cantidad = int(request.form.get(f'cantidad_{i}'))
        precio = float(request.form.get(f'precio_{i}'))
        
        if sku_bd:
            existe, tipo, nombre = verificar_sku_en_bd(sku_bd)
            if existe:
                items_mapeados.append({
                    'sku_ml': sku_ml,
                    'sku_bd': sku_bd,
                    'titulo': titulo,
                    'cantidad': cantidad,
                    'precio': precio,
                    'nombre_bd': nombre
                })
    
    access_token = cargar_ml_token()
    
    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        response = requests.get(f'https://api.mercadolibre.com/orders/{orden_id}', headers=headers)
        
        if response.status_code == 200:
            orden = response.json()
            orden_data = procesar_orden_ml(orden)
            
            # OBTENER SHIPPING COMPLETO
            if orden_data['shipping']['shipping_id']:
                shipping_completo = obtener_shipping_completo(
                    orden_data['shipping']['shipping_id'],
                    access_token
                )
                # ✅ CORREGIDO: Ya no sobrescribimos costo_envio
                orden_data['shipping'] = shipping_completo
            
            # ✅ NUEVO: OBTENER BILLING INFO
            billing_data = None
            try:
                billing_response = requests.get(
                    f'https://api.mercadolibre.com/orders/{orden_id}/billing_info',
                    headers=headers
                )
                
                if billing_response.status_code == 200:
                    billing_data = billing_response.json()
            except:
                pass
            
            # ✅ GUARDAR EN SESIÓN CON FECHA REAL Y BILLING
            from flask import session
            session['ml_orden_id'] = orden_id
            session['ml_items'] = items_mapeados
            session['ml_comprador_nombre'] = orden_data['comprador_nombre']
            session['ml_comprador_nickname'] = orden_data['comprador_nickname']
            session['ml_shipping'] = orden_data['shipping']
            session['ml_fecha_venta'] = orden_data['fecha'].isoformat()
            session['ml_billing_data'] = billing_data
        else:
            from flask import session
            session['ml_orden_id'] = orden_id
            session['ml_items'] = items_mapeados
            session['ml_comprador_nombre'] = ''
            session['ml_comprador_nickname'] = ''
            session['ml_shipping'] = {}
            session['ml_fecha_venta'] = datetime.now().isoformat()
            session['ml_billing_data'] = None
    
    except Exception as e:
        from flask import session
        session['ml_orden_id'] = orden_id
        session['ml_items'] = items_mapeados
        session['ml_comprador_nombre'] = ''
        session['ml_comprador_nickname'] = ''
        session['ml_shipping'] = {}
        session['ml_fecha_venta'] = datetime.now().isoformat()
        session['ml_billing_data'] = None
        import traceback
        traceback.print_exc()
    
    flash('✅ Productos mapeados correctamente', 'success')
    return redirect(url_for('nueva_venta_desde_ml'))



@app.route('/ventas/nueva/ml')
@login_required
def nueva_venta_desde_ml():
    """
    Crear nueva venta con datos precargados desde ML
    ✅ CORREGIDO: Pasa ml_shipping al template
    """
    from flask import session
    
    print("\n" + "="*70)
    print("🔍 DEBUG: Cargando nueva venta desde ML")
    print("="*70)
    
    if 'ml_items' not in session:
        flash('❌ No hay datos de ML para importar', 'error')
        return redirect(url_for('ventas_activas'))
    
    ml_items = session.get('ml_items', [])
    ml_orden_id = session.get('ml_orden_id', '')
    ml_shipping = session.get('ml_shipping', {})
    
    print(f"📦 Orden ID: {ml_orden_id}")
    print(f"🛍️ Items: {len(ml_items)} productos")
    print(f"🚚 Shipping data en sesión: {ml_shipping}")
    
    # Obtener datos necesarios para el formulario
    productos = query_db('SELECT * FROM productos_base ORDER BY nombre')
    combos = query_db('SELECT * FROM productos_compuestos ORDER BY nombre')
    
    print("="*70 + "\n")
    
    return render_template('nueva_venta_ml.html',
                         productos=productos,
                         combos=combos,
                         ml_items=ml_items,
                         ml_orden_id=ml_orden_id,
                         ml_shipping=ml_shipping)  # ✅ NUEVO: Pasar ml_shipping



import requests
import json
from pprint import pprint

def debug_orden_ml_completa(orden_id, access_token):
    """
    Ver TODOS los datos que trae ML de una orden
    Incluyendo billing_info, buyer, shipping, payments, etc.
    """
    headers = {'Authorization': f'Bearer {access_token}'}
    
    print("\n" + "="*80)
    print(f"🔍 DEBUG COMPLETO - ORDEN ML: {orden_id}")
    print("="*80 + "\n")
    
    # ============================================
    # 1. ORDEN COMPLETA
    # ============================================
    print("📦 1. INFORMACIÓN COMPLETA DE LA ORDEN:")
    print("-" * 80)
    
    try:
        response = requests.get(
            f'https://api.mercadolibre.com/orders/{orden_id}',
            headers=headers
        )
        
        if response.status_code != 200:
            print(f"❌ Error {response.status_code}: {response.text}")
            return
        
        orden = response.json()
        
        # Guardar JSON completo en archivo
        with open(f'debug_orden_{orden_id}.json', 'w', encoding='utf-8') as f:
            json.dump(orden, f, indent=2, ensure_ascii=False)
        
        print(f"✅ JSON completo guardado en: debug_orden_{orden_id}.json")
        print()
        
        # Mostrar estructura principal
        print("📋 ESTRUCTURA PRINCIPAL:")
        for key in orden.keys():
            valor = orden[key]
            tipo = type(valor).__name__
            
            if isinstance(valor, (list, dict)):
                if isinstance(valor, list):
                    longitud = f"[{len(valor)} items]"
                else:
                    longitud = f"{{{len(valor)} keys}}"
                print(f"   • {key}: {tipo} {longitud}")
            else:
                print(f"   • {key}: {tipo} = {valor}")
        
        print()
        
    except Exception as e:
        print(f"❌ Error al obtener orden: {str(e)}")
        return
    
    # ============================================
    # 2. DATOS DEL COMPRADOR (BUYER)
    # ============================================
    print("\n👤 2. DATOS DEL COMPRADOR (buyer):")
    print("-" * 80)
    
    buyer = orden.get('buyer', {})
    
    if buyer:
        print("📋 Campos disponibles en 'buyer':")
        pprint(buyer, width=120)
    else:
        print("⚠️ No hay datos de buyer")
    
    print()
    
    # ============================================
    # 3. BILLING INFO (FACTURACIÓN)
    # ============================================
    print("\n🧾 3. INFORMACIÓN DE FACTURACIÓN (billing_info):")
    print("-" * 80)
    
    try:
        billing_response = requests.get(
            f'https://api.mercadolibre.com/orders/{orden_id}/billing_info',
            headers=headers
        )
        
        if billing_response.status_code == 200:
            billing = billing_response.json()
            
            print("✅ BILLING INFO DISPONIBLE:")
            print()
            
            # Guardar JSON de billing
            with open(f'debug_billing_{orden_id}.json', 'w', encoding='utf-8') as f:
                json.dump(billing, f, indent=2, ensure_ascii=False)
            
            print(f"✅ JSON guardado en: debug_billing_{orden_id}.json")
            print()
            
            print("📋 DATOS DE FACTURACIÓN:")
            pprint(billing, width=120)
            
        elif billing_response.status_code == 404:
            print("⚠️ NO HAY BILLING INFO (404 - Not Found)")
            print()
            print("Esto es normal si:")
            print("   • No sos Responsable Inscripto")
            print("   • La venta no requiere factura")
            print("   • El comprador es Consumidor Final sin CUIT")
        
        else:
            print(f"⚠️ Error {billing_response.status_code}: {billing_response.text}")
    
    except Exception as e:
        print(f"❌ Error al obtener billing_info: {str(e)}")
    
    print()
    
    # ============================================
    # 4. SHIPPING (ENVÍO)
    # ============================================
    print("\n🚚 4. DATOS DE ENVÍO (shipping):")
    print("-" * 80)
    
    shipping = orden.get('shipping', {})
    
    if shipping:
        print("📋 Campos disponibles en 'shipping':")
        pprint(shipping, width=120)
        
        # Si hay shipping_id, obtener detalles completos
        shipping_id = shipping.get('id')
        if shipping_id:
            print()
            print(f"📦 Obteniendo detalles completos del envío {shipping_id}...")
            
            try:
                ship_response = requests.get(
                    f'https://api.mercadolibre.com/shipments/{shipping_id}',
                    headers=headers
                )
                
                if ship_response.status_code == 200:
                    shipment = ship_response.json()
                    
                    with open(f'debug_shipment_{orden_id}.json', 'w', encoding='utf-8') as f:
                        json.dump(shipment, f, indent=2, ensure_ascii=False)
                    
                    print(f"✅ Shipment guardado en: debug_shipment_{orden_id}.json")
                    print()
                    
                    print("📍 DIRECCIÓN DE ENVÍO:")
                    receiver = shipment.get('receiver_address', {})
                    pprint(receiver, width=120)
                
            except Exception as e:
                print(f"⚠️ Error al obtener shipment: {str(e)}")
    else:
        print("⚠️ No hay datos de shipping")
    
    print()
    
    # ============================================
    # 5. PAYMENTS (PAGOS)
    # ============================================
    print("\n💳 5. INFORMACIÓN DE PAGOS (payments):")
    print("-" * 80)
    
    payments = orden.get('payments', [])
    
    if payments:
        for i, pago in enumerate(payments, 1):
            print(f"\n💰 Pago #{i}:")
            pprint(pago, width=120)
    else:
        print("⚠️ No hay datos de pagos")
    
    print()
    
    # ============================================
    # 6. ORDER ITEMS (PRODUCTOS)
    # ============================================
    print("\n📦 6. PRODUCTOS DE LA ORDEN (order_items):")
    print("-" * 80)
    
    items = orden.get('order_items', [])
    
    if items:
        for i, item in enumerate(items, 1):
            print(f"\n🛒 Item #{i}:")
            pprint(item, width=120)
    else:
        print("⚠️ No hay items")
    
    print()
    
    # ============================================
    # 7. RESUMEN
    # ============================================
    print("\n" + "="*80)
    print("📊 RESUMEN - DATOS ÚTILES PARA FACTURACIÓN:")
    print("="*80 + "\n")
    
    print("✅ DATOS DISPONIBLES:")
    print()
    
    # Buyer
    print("👤 COMPRADOR (buyer):")
    print(f"   • ID: {buyer.get('id', 'N/A')}")
    print(f"   • Nickname: {buyer.get('nickname', 'N/A')}")
    print(f"   • Email: {buyer.get('email', 'N/A')}")
    print(f"   • First name: {buyer.get('first_name', 'N/A')}")
    print(f"   • Last name: {buyer.get('last_name', 'N/A')}")
    print(f"   • Phone: {buyer.get('phone', {}).get('number', 'N/A')}")
    print()
    
    # Billing (si existe)
    try:
        if billing_response.status_code == 200:
            billing = billing_response.json()
            print("🧾 FACTURACIÓN (billing_info):")
            print(f"   • Doc type: {billing.get('doc_type', 'N/A')}")
            print(f"   • Doc number: {billing.get('doc_number', 'N/A')}")
            
            # Buscar más datos
            if 'additional_info' in billing:
                print(f"   • Additional info: {billing.get('additional_info')}")
            
            print()
        else:
            print("🧾 FACTURACIÓN: No disponible (consumidor final)")
            print()
    except:
        pass
    
    # Totales
    print("💰 TOTALES:")
    print(f"   • Total amount: ${orden.get('total_amount', 0)}")
    print(f"   • Paid amount: ${orden.get('paid_amount', 0)}")
    
    if payments:
        fee = payments[0].get('marketplace_fee', 0)
        print(f"   • Marketplace fee: ${fee}")
        print(f"   • Neto vendedor: ${orden.get('total_amount', 0) - fee}")
    
    print()
    
    # Dirección
    if shipping and 'receiver_address' in shipment:
        addr = shipment.get('receiver_address', {})
        print("📍 DIRECCIÓN:")
        print(f"   • Address line: {addr.get('address_line', 'N/A')}")
        print(f"   • City: {addr.get('city', {}).get('name', 'N/A')}")
        print(f"   • State: {addr.get('state', {}).get('name', 'N/A')}")
        print(f"   • Zip code: {addr.get('zip_code', 'N/A')}")
        print()
    
    print("="*80)
    print("✅ DEBUG COMPLETO - Revisá los archivos JSON generados")
    print("="*80 + "\n")

def extraer_billing_info_ml(billing_data):
    """
    Extraer datos de facturación del formato de ML
    
    ML devuelve los datos en formato:
    {
      "billing_info": {
        "additional_info": [
          {"type": "BUSINESS_NAME", "value": "FERNANDEZ GONZALO JAVIER"},
          {"type": "DOC_TYPE", "value": "CUIT"},
          ...
        ]
      }
    }
    
    Returns:
        dict con los datos extraídos
    """
    datos = {
        'business_name': None,
        'doc_type': None,
        'doc_number': None,
        'taxpayer_type': None,
        'city': None,
        'street': None,
        'state': None,
        'zip_code': None
    }
    
    try:
        # Obtener billing_info
        billing_info = billing_data.get('billing_info', {})
        
        # Doc type y number están en el nivel principal
        datos['doc_type'] = billing_info.get('doc_type')
        datos['doc_number'] = billing_info.get('doc_number')
        
        # Los demás datos están en additional_info
        additional_info = billing_info.get('additional_info', [])
        
        # Crear diccionario para búsqueda fácil
        info_dict = {}
        for item in additional_info:
            tipo = item.get('type')
            valor = item.get('value')
            if tipo and valor:
                info_dict[tipo] = valor
        
        # Extraer campos
        datos['business_name'] = info_dict.get('BUSINESS_NAME')
        datos['taxpayer_type'] = info_dict.get('TAXPAYER_TYPE_ID')
        datos['city'] = info_dict.get('CITY_NAME')
        datos['street'] = info_dict.get('STREET_NAME')
        datos['state'] = info_dict.get('STATE_NAME')
        datos['zip_code'] = info_dict.get('ZIP_CODE')
        
        # Si no hay doc_type en el nivel principal, intentar de additional_info
        if not datos['doc_type']:
            datos['doc_type'] = info_dict.get('DOC_TYPE')
        if not datos['doc_number']:
            datos['doc_number'] = info_dict.get('DOC_NUMBER')
        
        print(f"✅ Datos de facturación extraídos:")
        print(f"   • Razón social: {datos['business_name']}")
        print(f"   • {datos['doc_type']}: {datos['doc_number']}")
        print(f"   • Condición IVA: {datos['taxpayer_type']}")
        
    except Exception as e:
        print(f"⚠️ Error al extraer billing_info: {str(e)}")
    
    return datos



# ============================================================================
# RUTA TEMPORAL PARA USAR EN FLASK
# ============================================================================

@app.route('/debug/ml/<orden_id>')
@login_required
def debug_ml_orden(orden_id):
    """Ver qué datos trae ML de una orden"""
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('No hay token de ML', 'error')
        return redirect(url_for('ventas_activas'))
    
    headers = {'Authorization': f'Bearer {access_token}'}
    
    try:
        # Obtener orden
        response = requests.get(f'https://api.mercadolibre.com/orders/{orden_id}', headers=headers)
        
        if response.status_code != 200:
            return f"<h2>Error {response.status_code}</h2>"
        
        orden = response.json()
        buyer = orden.get('buyer', {})
        
        # Intentar obtener billing
        billing_html = ""
        try:
            billing_response = requests.get(
                f'https://api.mercadolibre.com/orders/{orden_id}/billing_info',
                headers=headers
            )
            
            if billing_response.status_code == 200:
                billing = billing_response.json()
                import json
                billing_json = json.dumps(billing, indent=2, ensure_ascii=False)
                billing_html = f"<h3>TIENE DATOS DE FACTURACION:</h3><pre>{billing_json}</pre>"
            else:
                billing_html = "<h3>NO TIENE DATOS DE FACTURACION (Consumidor Final)</h3>"
        except:
            billing_html = "<h3>Error al obtener billing</h3>"
        
        # Construir HTML de respuesta
        html_response = """
        <html>
        <head>
            <title>Debug ML Orden</title>
            <style>
                body {{ font-family: monospace; padding: 20px; background: #1e1e1e; color: #d4d4d4; }}
                pre {{ background: #2d2d2d; padding: 15px; border-radius: 5px; overflow-x: auto; }}
                h2 {{ color: #4fc3f7; }}
                h3 {{ color: #ce9178; }}
            </style>
        </head>
        <body>
            <h2>DEBUG ORDEN ML: {orden_id}</h2>
            
            <h3>COMPRADOR:</h3>
            <p>Nombre: {nombre}</p>
            <p>Nickname: {nickname}</p>
            <p>Email: {email}</p>
            
            {billing_info}
            
            <br><br>
            <a href="/ventas/activas" style="color: #4fc3f7;">Volver a Ventas Activas</a>
        </body>
        </html>
        """.format(
            orden_id=orden_id,
            nombre=f"{buyer.get('first_name', '')} {buyer.get('last_name', '')}",
            nickname=buyer.get('nickname', ''),
            email=buyer.get('email', 'No disponible'),
            billing_info=billing_html
        )
        
        return html_response
        
    except Exception as e:
        import traceback
        error_completo = traceback.format_exc()
        return f"<pre>Error: {error_completo}</pre>"

# ============================================================================
# FUNCIÓN OBTENER DATOS ML - CON STATUS REAL
# ============================================================================

CAMPAÑAS_CUOTAS = {
    'pcj-co-funded': 'Cuota Simple',
    '3x_campaign':  '3 cuotas s/interés',
    '6x_campaign':  '6 cuotas s/interés',
    '9x_campaign':  '9 cuotas s/interés',
    '12x_campaign': '12 cuotas s/interés',
}

def obtener_datos_ml(mla_id, access_token):
    """
    Consulta datos actuales de una publicación ML.
    Devuelve: titulo, stock, status, demora, precio, listing_type
    """
    CAMPAÑAS_CUOTAS = {
        'pcj-co-funded': 'Cuota Simple',
        '3x_campaign':   '3 cuotas s/interés',
        '6x_campaign':   '6 cuotas s/interés',
        '9x_campaign':   '9 cuotas s/interés',
        '12x_campaign':  '12 cuotas s/interés',
    }

    try:
        r = ml_request('get', f'https://api.mercadolibre.com/items/{mla_id}', access_token)

        if r.status_code != 200:
            return {
                'titulo': mla_id, 'stock': 0, 'status': 'unknown',
                'demora': None, 'precio': None, 'listing_type': None
            }

        data = r.json()

        demora = None
        campaign = None
        for term in data.get('sale_terms', []):
            if term.get('id') == 'MANUFACTURING_TIME':
                demora = term.get('value_name')
            if term.get('id') == 'INSTALLMENTS_CAMPAIGN':
                campaign = (term.get('value_name') or '').split('|')[0].strip()

        listing_type_id = data.get('listing_type_id', '')
        if listing_type_id == 'gold_special':
            financiacion = CAMPAÑAS_CUOTAS.get(campaign, 'Sin cuotas propias') if campaign else 'Sin cuotas propias'
        elif listing_type_id == 'gold_pro':
            financiacion = CAMPAÑAS_CUOTAS.get(campaign, '6 cuotas s/interés')
        else:
            financiacion = listing_type_id or '-'

        return {
            'titulo':       data.get('title', mla_id),
            'stock':        data.get('available_quantity', 0),
            'status':       data.get('status', 'unknown'),
            'demora':       demora,
            'precio':       data.get('price'),
            'listing_type': financiacion,
        }

    except Exception as e:
        print(f"Error obteniendo datos ML de {mla_id}: {e}")
        return {
            'titulo': mla_id, 'stock': 0, 'status': 'unknown',
            'demora': None, 'precio': None, 'listing_type': None
        }
# ============================================================================
# RUTAS CARGAR STOCK ML - CORREGIDAS CON COLUMNAS REALES
# ============================================================================

@app.route('/cargar-stock-ml', methods=['GET'])
@login_required
def cargar_stock_ml():
    """Mostrar página para cargar stock en ML"""
    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'porcentajes_ml'")
        porcentajes = json.loads(row['valor']) if row else PORCENTAJES_ML_DEFAULT
    except:
        porcentajes = PORCENTAJES_ML_DEFAULT

    return render_template('cargar_stock_ml.html',
                           sku_buscado=None,
                           publicaciones=[],
                           es_sku_con_z=False,
                           mensaje=None,
                           mensaje_tipo=None,
                           porcentajes=porcentajes)

# ============================================================================
# RUTA BUSCAR SKU - ACTUALIZADA CON STATUS REAL DE ML
# ============================================================================


def ml_request(method, url, access_token, json_data=None, params=None, max_retries=4):
    """
    Helper para requests a ML con rate limiting global + retry exponencial.
    Toda llamada a ML debe pasar por acá.
    """
    global _ml_last_request

    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}

    for attempt in range(max_retries):
        # Rate limiting global: esperar si la última request fue muy reciente
        with _ml_rate_lock:
            now = time.time()
            elapsed = now - _ml_last_request
            if elapsed < _ML_MIN_INTERVAL:
                time.sleep(_ML_MIN_INTERVAL - elapsed)
            _ml_last_request = time.time()

        try:
            if method == 'get':
                r = requests.get(url, headers=headers, params=params, timeout=15)
            else:
                r = requests.put(url, headers=headers, json=json_data, timeout=15)

            if r.status_code == 429:
                wait = min(5 * (2 ** attempt), 60)  # 5s, 10s, 20s, 40s
                print(f"⚠️ 429 ML - esperando {wait}s (intento {attempt+1}/{max_retries})")
                time.sleep(wait)
                continue

            return r

        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(3)
                continue
            raise

    return r

def obtener_datos_ml_batch(mla_ids, access_token):
    """
    Consulta datos de múltiples publicaciones ML en UNA sola llamada (batch).
    Devuelve dict: { mla_id: { titulo, stock, status, demora, precio, listing_type } }
    """
    if not mla_ids:
        return {}

    CAMPAÑAS_CUOTAS = {
        'pcj-co-funded': 'Cuota Simple',
        '3x_campaign':   '3 cuotas s/interés',
        '6x_campaign':   '6 cuotas s/interés',
        '9x_campaign':   '9 cuotas s/interés',
        '12x_campaign':  '12 cuotas s/interés',
    }

    resultado = {}
    try:
        r = ml_request('get', 'https://api.mercadolibre.com/items', access_token,
                       params={'ids': ','.join(mla_ids)})

        if r.status_code != 200:
            for mla_id in mla_ids:
                resultado[mla_id] = {
                    'titulo': mla_id, 'stock': 0, 'status': 'unknown',
                    'demora': None, 'precio': None, 'listing_type': None, 'status_raw': 'unknown'
                }
            return resultado

        items = r.json()
        for item in items:
            if item.get('code') != 200:
                mla_id = item.get('body', {}).get('id', '')
                resultado[mla_id] = {
                    'titulo': mla_id, 'stock': 0, 'status': 'unknown',
                    'demora': None, 'precio': None, 'listing_type': None, 'status_raw': 'unknown'
                }
                continue

            data = item['body']
            mla_id = data.get('id', '')

            demora = None
            campaign = None
            for term in data.get('sale_terms', []):
                if term.get('id') == 'MANUFACTURING_TIME':
                    demora = term.get('value_name')
                if term.get('id') == 'INSTALLMENTS_CAMPAIGN':
                    campaign = (term.get('value_name') or '').split('|')[0].strip()

            listing_type_id = data.get('listing_type_id', '')
            if listing_type_id == 'gold_special':
                financiacion = CAMPAÑAS_CUOTAS.get(campaign, 'Sin cuotas propias') if campaign else 'Sin cuotas propias'
            elif listing_type_id == 'gold_pro':
                financiacion = CAMPAÑAS_CUOTAS.get(campaign, '6 cuotas s/interés')
            else:
                financiacion = listing_type_id

            resultado[mla_id] = {
                'titulo':       data.get('title', mla_id),
                'stock':        data.get('available_quantity', 0),
                'status':       data.get('status', 'unknown'),
                'status_raw':   data.get('status', 'unknown'),
                'demora':       demora,
                'precio':       data.get('price'),
                'listing_type': financiacion
            }

    except Exception as e:
        print(f"Error en obtener_datos_ml_batch: {e}")
        for mla_id in mla_ids:
            resultado[mla_id] = {
                'titulo': mla_id, 'stock': 0, 'status': 'unknown',
                'demora': None, 'precio': None, 'listing_type': None, 'status_raw': 'unknown'
            }
    return resultado


@app.route('/buscar-sku-ml', methods=['POST'])
@login_required
def buscar_sku_ml():
    sku_buscado = request.form.get('sku_buscar', '').strip().upper()

    if not sku_buscado:
        flash('Debes ingresar un SKU', 'warning')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()

    # Cargar porcentajes
    try:
        row = query_one("SELECT valor FROM configuracion WHERE clave = 'porcentajes_ml'")
        porcentajes = json.loads(row['valor']) if row else PORCENTAJES_ML_DEFAULT
    except:
        porcentajes = PORCENTAJES_ML_DEFAULT

    if not access_token:
        flash('❌ No hay token de ML configurado', 'warning')
        return render_template('cargar_stock_ml.html',
                               sku_buscado=sku_buscado,
                               publicaciones=[],
                               es_sku_con_z=sku_buscado.endswith('Z'),
                               porcentajes=porcentajes)

    mla_directo = None
    if sku_buscado.startswith('MLA'):
        mla_directo = sku_buscado
    elif sku_buscado.isdigit():
        mla_directo = f'MLA{sku_buscado}'

    if mla_directo:
        mla_ids = [mla_directo]
        es_sku_con_z = False
    else:
        es_sku_con_z = sku_buscado.endswith('Z')
        r = ml_request('get',
            f'https://api.mercadolibre.com/users/{ML_SELLER_ID}/items/search',
            access_token, params={'seller_sku': sku_buscado})
        if r.status_code != 200:
            flash(f'❌ Error consultando ML: {r.status_code}', 'danger')
            return render_template('cargar_stock_ml.html',
                                   sku_buscado=sku_buscado,
                                   publicaciones=[],
                                   es_sku_con_z=es_sku_con_z,
                                   porcentajes=porcentajes)
        mla_ids = r.json().get('results', [])

    if not mla_ids:
        flash(f'No se encontraron publicaciones para "{sku_buscado}"', 'warning')
        return render_template('cargar_stock_ml.html',
                               sku_buscado=sku_buscado,
                               publicaciones=[],
                               es_sku_con_z=es_sku_con_z,
                               porcentajes=porcentajes)

    estado_map = {
        'active': 'Activa', 'paused': 'Pausada', 'closed': 'Cerrada',
        'under_review': 'En revisión', 'inactive': 'Inactiva'
    }

    # Orden de tipos de publicación
    ORDEN_TIPOS = {
        'Sin cuotas propias': 0,
        'Cuota Simple':       1,
        '3 cuotas s/interés': 2,
        '6 cuotas s/interés': 3,
        '9 cuotas s/interés': 4,
        '12 cuotas s/interés': 5,
    }

    datos_batch = obtener_datos_ml_batch(mla_ids, access_token)
    publicaciones = []
    for mla_id in mla_ids:
        datos_ml = datos_batch.get(mla_id, {
            'titulo': mla_id, 'stock': 0, 'status': 'unknown',
            'demora': None, 'precio': None, 'listing_type': None, 'status_raw': 'unknown'
        })
        status_ml = datos_ml.get('status', 'unknown')
        publicaciones.append({
            'mla':          mla_id,
            'titulo':       datos_ml['titulo'],
            'stock_actual': datos_ml['stock'],
            'demora':       datos_ml.get('demora'),
            'precio':       datos_ml.get('precio'),
            'listing_type': datos_ml.get('listing_type'),
            'estado':       estado_map.get(status_ml, status_ml.capitalize()),
            'status_raw':   status_ml
        })

    # Ordenar por tipo de publicación
    publicaciones.sort(key=lambda p: ORDEN_TIPOS.get(p.get('listing_type', ''), 99))

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku_buscado,
                           publicaciones=publicaciones,
                           es_sku_con_z=es_sku_con_z,
                           mensaje=None,
                           mensaje_tipo=None,
                           porcentajes=porcentajes)
# ============================================================================
# 4. RUTA: Cambiar precio — INDIVIDUAL
# ============================================================================

@app.route('/cambiar-precio-mla', methods=['POST'])
@login_required
def cambiar_precio_mla():
    """Cambiar el precio de una publicación específica"""
    mla    = request.form.get('mla')
    sku    = request.form.get('sku')
    precio = request.form.get('precio', '').strip()

    if not mla or not sku or not precio:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    try:
        precio_float = float(precio)
        if precio_float <= 0:
            raise ValueError()
    except ValueError:
        flash('❌ Precio inválido', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    import requests as req
    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
    payload = {'price': precio_float}

    r = ml_request('put', f'https://api.mercadolibre.com/items/{mla}', access_token, json_data=payload)

    if r.status_code == 200:
        flash(f'✅ Precio actualizado a ${precio_float:,.0f} en {mla}', 'success')
    else:
        try:
            err = r.json()
        except:
            err = r.text
        flash(f'❌ Error ML {r.status_code} en {mla}: {err}', 'danger')

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token,
                               pubs_actuales=request.form.get('pubs_json'),
                               actualizar_mla=mla, campo='precio', valor=precio_float),
                           es_sku_con_z=sku.endswith('Z'))


# ============================================================================
# 5. RUTA: Cambiar precio — MASIVO
# ============================================================================

@app.route('/cambiar-precio-masivo', methods=['POST'])
@login_required
def cambiar_precio_masivo():
    """Cambiar el precio de todas las publicaciones de un SKU"""
    sku    = request.form.get('sku')
    precio = request.form.get('precio', '').strip()

    if not sku or not precio:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    try:
        precio_float = float(precio)
        if precio_float <= 0:
            raise ValueError()
    except ValueError:
        flash('❌ Precio inválido', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    mlas = query_db(
        "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,)
    )

    import requests as req
    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
    payload = {'price': precio_float}

    exitos, errores = 0, []
    for row in mlas:
        r = ml_request('put', f'https://api.mercadolibre.com/items/{row["mla_id"]}', access_token, json_data=payload)
        if r.status_code == 200:
            exitos += 1
        else:
            errores.append(f'{row["mla_id"]}: {r.status_code}')
        time.sleep(2)

    if exitos:
        flash(f'✅ Precio actualizado a ${precio_float:,.0f} en {exitos} publicación{"es" if exitos > 1 else ""}', 'success')
    for msg in errores[:3]:
        flash(f'❌ {msg}', 'danger')

    # Actualizar precio en todas las pubs localmente
    pubs_json = request.form.get('pubs_json')
    import json as _j
    try:
        pubs = _j.loads(pubs_json) if pubs_json else None
        if pubs:
            for p in pubs: p['precio'] = precio_float
    except: pubs = None
    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token, pubs_actuales=pubs),
                           es_sku_con_z=sku.endswith('Z'))

@app.route('/cambiar-precios-individuales', methods=['POST'])
@login_required
def cambiar_precios_individuales():
    """Actualizar precios individuales de múltiples MLAs de una vez"""
    import requests as req
    import json

    sku = request.form.get('sku')
    precios_json = request.form.get('precios_json', '[]')

    try:
        precios = json.loads(precios_json)
    except:
        flash('❌ Error al procesar los precios', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    if not precios:
        flash('❌ No se recibieron precios', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}

    exitos, errores = 0, []
    for item in precios:
        mla = item.get('mla')
        precio = item.get('precio')
        try:
            precio_float = float(precio)
            if precio_float <= 0:
                raise ValueError()
        except:
            errores.append(f'{mla}: precio inválido')
            continue

        r = ml_request('put', f'https://api.mercadolibre.com/items/{mla}', access_token, json_data={'price': precio_float})
        if r.status_code == 200:
            exitos += 1
        else:
            errores.append(f'{mla}: error {r.status_code}')
        time.sleep(2)

    if exitos:
        flash(f'✅ {exitos} precio{"s" if exitos > 1 else ""} actualizado{"s" if exitos > 1 else ""} correctamente', 'success')
    for msg in errores[:3]:
        flash(f'❌ {msg}', 'danger')

    # Actualizar precios individuales localmente
    pubs_json = request.form.get('pubs_json')
    import json as _j
    try:
        pubs = _j.loads(pubs_json) if pubs_json else None
        if pubs:
            precios_dict = {item['mla']: float(item['precio']) for item in precios}
            for p in pubs:
                if p['mla'] in precios_dict:
                    p['precio'] = precios_dict[p['mla']]
    except: pubs = None
    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token, pubs_actuales=pubs),
                           es_sku_con_z=sku.endswith('Z') if sku else False)



# ============================================================================
# RUTAS NUEVAS: Bajar stock a 0 y Cargar demora
# Agregar en app.py junto a las rutas de cargar_stock_ml
# ============================================================================

def _recargar_publicaciones(sku, access_token, pubs_actuales=None, actualizar_mla=None, campo=None, valor=None):
    """
    Helper: devuelve lista de publicaciones.
    Si se pasan pubs_actuales, las usa directamente y solo actualiza el campo indicado.
    Si no, consulta ML (solo al buscar por primera vez).
    """
    import json as _json

    estado_map = {
        'active': 'Activa', 'paused': 'Pausada', 'closed': 'Cerrada',
        'under_review': 'En revisión', 'inactive': 'Inactiva'
    }

    # Si tenemos datos actuales del form, usarlos sin llamar a ML
    if pubs_actuales:
        try:
            if isinstance(pubs_actuales, str):
                pubs_lista = _json.loads(pubs_actuales)
            else:
                pubs_lista = pubs_actuales
            # Aplicar el cambio puntual si se indicó
            if actualizar_mla and campo:
                for pub in pubs_lista:
                    if pub.get('mla') == actualizar_mla:
                        pub[campo] = valor
            return pubs_lista
        except:
            pass  # Si falla el parse, caer al batch

    # Sin datos actuales: consultar ML (solo pasa al buscar por primera vez)
    publicaciones = query_db(
        "SELECT mla_id, titulo_ml, activo FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
        (sku,)
    )
    pubs_lista = []
    if access_token and publicaciones:
        mla_ids = [row['mla_id'] for row in publicaciones]
        datos_batch = obtener_datos_ml_batch(mla_ids, access_token)
        for row in publicaciones:
            datos_ml = datos_batch.get(row['mla_id'], {
                'titulo': row['mla_id'], 'stock': 0, 'status': 'unknown',
                'demora': None, 'precio': None, 'listing_type': None, 'status_raw': 'unknown'
            })
            status_ml = datos_ml.get('status', 'unknown')
            pubs_lista.append({
                'mla':          row['mla_id'],
                'titulo':       datos_ml['titulo'],
                'stock_actual': datos_ml['stock'],
                'demora':       datos_ml.get('demora'),
                'precio':       datos_ml.get('precio'),
                'listing_type': datos_ml.get('listing_type'),
                'estado':       estado_map.get(status_ml, status_ml.capitalize()),
                'status_raw':   status_ml
            })
    else:
        for row in publicaciones:
            pubs_lista.append({
                'mla':          row['mla_id'],
                'titulo':       row['titulo_ml'] or 'Sin título',
                'stock_actual': '-', 'demora': None,
                'precio':       None, 'listing_type': None,
                'estado':       'Activa' if row['activo'] else 'Pausada',
                'status_raw':   'active' if row['activo'] else 'paused'
            })
    return pubs_lista




# ─── Bajar stock a 0 — INDIVIDUAL ────────────────────────────────────────────

@app.route('/bajar-stock-mla-cero', methods=['POST'])
@login_required
def bajar_stock_mla_cero():
    """Poner stock en 0 en una publicación específica"""
    mla = request.form.get('mla')
    sku = request.form.get('sku')

    if not mla or not sku:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    success, message = actualizar_stock_ml(mla, 0, access_token)

    if success:
        flash(f'✅ Stock bajado a 0 en {mla}', 'success')
    else:
        flash(f'❌ {message}', 'danger')

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token,
                               pubs_actuales=request.form.get('pubs_json'),
                               actualizar_mla=mla, campo='stock_actual', valor=0),
                           es_sku_con_z=sku.endswith('Z'))


# ─── Bajar stock a 0 — MASIVO ────────────────────────────────────────────────

@app.route('/bajar-stock-cero-masivo', methods=['POST'])
@login_required
def bajar_stock_cero_masivo():
    """Poner stock en 0 en todas las publicaciones de un SKU"""
    sku = request.form.get('sku')
    if not sku:
        flash('Falta el SKU', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    mlas = query_db(
        "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,)
    )

    exitos, errores = 0, []
    for row in mlas:
        ok, msg = actualizar_stock_ml(row['mla_id'], 0, access_token)
        if ok:
            exitos += 1
        else:
            errores.append(msg)
        time.sleep(2)

    if exitos:
        flash(f'✅ Stock bajado a 0 en {exitos} publicación{"es" if exitos > 1 else ""}', 'success')
    for msg in errores[:3]:
        flash(f'❌ {msg}', 'danger')

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token, pubs_actuales=request.form.get('pubs_json')),
                           es_sku_con_z=sku.endswith('Z'))


# ─── Cargar demora — INDIVIDUAL ──────────────────────────────────────────────

@app.route('/cargar-demora-mla', methods=['POST'])
@login_required
def cargar_demora_mla():
    """Poner X días de MANUFACTURING_TIME en una publicación"""
    mla  = request.form.get('mla')
    sku  = request.form.get('sku')
    dias = request.form.get('dias', '').strip()

    if not mla or not sku or not dias:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    import requests as req
    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
    payload = {"sale_terms": [{"id": "MANUFACTURING_TIME", "value_name": f"{dias} días"}]}

    r = ml_request('put', f'https://api.mercadolibre.com/items/{mla}', access_token, json_data=payload)

    if r.status_code == 200:
        flash(f'✅ Demora de {dias} días cargada en {mla}', 'success')
    else:
        try:
            err = r.json()
        except:
            err = r.text
        flash(f'❌ Error ML {r.status_code} en {mla}: {err}', 'danger')

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token, pubs_actuales=request.form.get('pubs_json')),
                           es_sku_con_z=sku.endswith('Z'))


# ─── Cargar demora — MASIVO ───────────────────────────────────────────────────

@app.route('/cargar-demora-masivo', methods=['POST'])
@login_required
def cargar_demora_masivo():
    """Poner X días de MANUFACTURING_TIME en todas las publicaciones de un SKU"""
    sku  = request.form.get('sku')
    dias = request.form.get('dias', '').strip()

    if not sku or not dias:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    mlas = query_db(
        "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,)
    )

    import requests as req
    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}
    payload = {"sale_terms": [{"id": "MANUFACTURING_TIME", "value_name": f"{dias} días"}]}

    exitos, errores = 0, []
    for row in mlas:
        r = ml_request('put', f'https://api.mercadolibre.com/items/{row["mla_id"]}', access_token, json_data=payload)
        if r.status_code == 200:
            exitos += 1
        else:
            errores.append(f'{row["mla_id"]}: {r.status_code}')
        time.sleep(2)

    if exitos:
        flash(f'✅ {dias} días de demora cargados en {exitos} publicación{"es" if exitos > 1 else ""}', 'success')
    for msg in errores[:3]:
        flash(f'❌ {msg}', 'danger')

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=_recargar_publicaciones(sku, access_token, pubs_actuales=request.form.get('pubs_json')),
                           es_sku_con_z=sku.endswith('Z'))



# ============================================================================
# RUTAS QUITAR DEMORA - CON STATUS REAL DE ML
# ============================================================================

# ============================================================================
# FUNCIÓN HELPER - Quitar MANUFACTURING_TIME completamente (envía null a ML)
# Reemplaza a quitar_handling_time_ml y actualizar_handling_time_ml
# ============================================================================

def quitar_manufacturing_time_ml(mla_id, access_token):
    """
    Elimina completamente el MANUFACTURING_TIME de una publicación ML.
    Envía value_id: null y value_name: null — ambos requeridos por la API.
    Retorna (True, mensaje) o (False, error)
    """
    import requests as req

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    payload = {
        "sale_terms": [
            {
                "id": "MANUFACTURING_TIME",
                "value_id": None,
                "value_name": None
            }
        ]
    }

    try:
        r = req.put(
            f'https://api.mercadolibre.com/items/{mla_id}',
            headers=headers,
            json=payload
        )

        if r.status_code == 200:
            data = r.json()
            # Verificar que quedó en None
            mt_nuevo = None
            for term in data.get('sale_terms', []):
                if term.get('id') == 'MANUFACTURING_TIME':
                    mt_nuevo = term.get('value_name')
                    break

            if mt_nuevo is None:
                return True, f'Demora eliminada en {mla_id}'
            else:
                return True, f'Demora actualizada a {mt_nuevo} en {mla_id}'
        else:
            try:
                err = r.json()
            except:
                err = r.text
            return False, f'Error ML {r.status_code} en {mla_id}: {err}'

    except Exception as e:
        return False, f'Excepción en {mla_id}: {str(e)}'


# ============================================================================
# RUTA: Quitar demora de UNA publicación
# ============================================================================

@app.route('/quitar-demora-mla', methods=['POST'])
@login_required
def quitar_demora_mla():
    """Eliminar MANUFACTURING_TIME de una publicación específica"""

    mla = request.form.get('mla')
    sku = request.form.get('sku')

    if not mla or not sku:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    success, message = quitar_manufacturing_time_ml(mla, access_token)

    if success:
        flash(f'✅ {message}', 'success')
    else:
        flash(f'❌ {message}', 'danger')

    # Recargar publicaciones con datos actualizados de ML
    publicaciones = query_db(
        "SELECT mla_id, titulo_ml, activo FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
        (sku,)
    )

    pubs_lista = []
    access_token_refresh = cargar_ml_token()

    for row in publicaciones:
        if access_token_refresh:
            datos_ml = obtener_datos_ml(row['mla_id'], access_token_refresh)
            status_ml = datos_ml.get('status', 'unknown')
            estado_map = {
                'active': 'Activa', 'paused': 'Pausada', 'closed': 'Cerrada',
                'under_review': 'En revisión', 'inactive': 'Inactiva'
            }
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': datos_ml['titulo'],
                'stock_actual': datos_ml['stock'],
                'demora': datos_ml.get('demora'),
                'estado': estado_map.get(status_ml, status_ml.capitalize()),
                'status_raw': status_ml
            })
        else:
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': row['titulo_ml'] or 'Sin título',
                'stock_actual': '-',
                'demora': None,
                'estado': 'Activa' if row['activo'] else 'Pausada',
                'status_raw': 'active' if row['activo'] else 'paused'
            })

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=pubs_lista,
                           es_sku_con_z=sku.endswith('Z'))


# ============================================================================
# RUTA: Quitar demora de TODAS las publicaciones de un SKU
# ============================================================================

@app.route('/quitar-demora-masivo', methods=['POST'])
@login_required
def quitar_demora_masivo():
    """Eliminar MANUFACTURING_TIME de todas las publicaciones de un SKU"""
    sku = request.form.get('sku')
    if not sku:
        flash('Falta el SKU', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    access_token = cargar_ml_token()
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))

    mlas = query_db(
        "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,)
    )

    exitos, errores, mensajes_error = 0, 0, []

    for row in mlas:
        success, message = quitar_manufacturing_time_ml(row['mla_id'], access_token)
        if success:
            exitos += 1
        else:
            errores += 1
            mensajes_error.append(message)
        time.sleep(2)

    if exitos > 0:
        flash(f'✅ Demora eliminada en {exitos} publicación{"es" if exitos > 1 else ""}', 'success')
    if errores > 0:
        flash(f'⚠️ {errores} publicación{"es" if errores > 1 else ""} con errores', 'warning')
        for msg in mensajes_error[:3]:
            flash(msg, 'warning')

    # Recargar con UNA sola llamada batch
    publicaciones = query_db(
        "SELECT mla_id, titulo_ml, activo FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE", (sku,)
    )
    estado_map = {
        'active': 'Activa', 'paused': 'Pausada', 'closed': 'Cerrada',
        'under_review': 'En revisión', 'inactive': 'Inactiva'
    }
    pubs_lista = []
    if publicaciones:
        mla_ids = [row['mla_id'] for row in publicaciones]
        datos_batch = obtener_datos_ml_batch(mla_ids, access_token)
        for row in publicaciones:
            datos_ml = datos_batch.get(row['mla_id'], {
                'titulo': row['titulo_ml'] or row['mla_id'],
                'stock': '-', 'status': 'unknown', 'demora': None,
                'precio': None, 'listing_type': None
            })
            status_ml = datos_ml.get('status', 'unknown')
            pubs_lista.append({
                'mla':          row['mla_id'],
                'titulo':       datos_ml.get('titulo', row['titulo_ml'] or row['mla_id']),
                'stock_actual': datos_ml.get('stock', '-'),
                'demora':       datos_ml.get('demora'),
                'precio':       datos_ml.get('precio'),
                'listing_type': datos_ml.get('listing_type'),
                'estado':       estado_map.get(status_ml, status_ml.capitalize()),
                'status_raw':   status_ml
            })

    return render_template('cargar_stock_ml.html',
                           sku_buscado=sku,
                           publicaciones=pubs_lista,
                           es_sku_con_z=sku.endswith('Z'))


# ============================================================================
# RUTAS CARGAR STOCK - CON STATUS REAL DE ML
# ============================================================================

@app.route('/cargar-stock-mla', methods=['POST'])
@login_required
def cargar_stock_mla():
    """Cargar stock en una publicación específica"""
    
    mla = request.form.get('mla')
    sku = request.form.get('sku')
    stock_nuevo = request.form.get('stock_nuevo')
    
    if not mla or not sku or not stock_nuevo:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))
    
    # Obtener token dinámico
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))
    
    try:
        stock_nuevo = int(stock_nuevo)
        
        # Usar la función helper actualizar_stock_ml
        success, message = actualizar_stock_ml(mla, stock_nuevo, access_token)
        
        if success:
            flash(f'✅ {message}', 'success')
        else:
            flash(f'❌ {message}', 'danger')
    
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'danger')
    
    # Volver a mostrar resultados CON DATOS ACTUALIZADOS DE ML
    publicaciones = query_db(
        "SELECT mla_id, titulo_ml, activo FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
        (sku,)
    )
    
    pubs_lista = []
    access_token_refresh = cargar_ml_token()
    
    for row in publicaciones:
        if access_token_refresh:
            datos_ml = obtener_datos_ml(row['mla_id'], access_token_refresh)
            
            # Mapear status de ML a español
            status_ml = datos_ml.get('status', 'unknown')
            
            if status_ml == 'active':
                estado_texto = 'Activa'
            elif status_ml == 'paused':
                estado_texto = 'Pausada'
            elif status_ml == 'closed':
                estado_texto = 'Cerrada'
            elif status_ml == 'under_review':
                estado_texto = 'En revisión'
            elif status_ml == 'inactive':
                estado_texto = 'Inactiva'
            else:
                estado_texto = status_ml.capitalize()
            
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': datos_ml['titulo'],
                'stock_actual': datos_ml['stock'],
                'demora': datos_ml.get('demora'),
                'estado': estado_texto,
                'status_raw': status_ml
            })
        else:
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': row['titulo_ml'] or 'Sin título',
                'stock_actual': '-',
                'demora': None,
                'estado': 'Activa' if row['activo'] else 'Pausada',
                'status_raw': 'active' if row['activo'] else 'paused'
            })
    
    return render_template('cargar_stock_ml.html',
                         sku_buscado=sku,
                         publicaciones=pubs_lista,
                         es_sku_con_z=sku.endswith('Z'))


@app.route('/cargar-stock-masivo', methods=['POST'])
@login_required
def cargar_stock_masivo():
    """Cargar el mismo stock en todas las publicaciones de un SKU"""
    
    sku = request.form.get('sku')
    stock_nuevo = request.form.get('stock_nuevo')
    
    if not sku or not stock_nuevo:
        flash('Faltan datos', 'danger')
        return redirect(url_for('cargar_stock_ml'))
    
    # Obtener token dinámico
    access_token = cargar_ml_token()
    
    if not access_token:
        flash('❌ No hay token de ML configurado', 'danger')
        return redirect(url_for('cargar_stock_ml'))
    
    try:
        stock_nuevo = int(stock_nuevo)
        
        # Obtener todas las publicaciones del SKU
        mlas = query_db(
            "SELECT mla_id FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
            (sku,)
        )
        
        exitos = 0
        errores = 0
        mensajes_error = []
        
        for row in mlas:
            mla = row['mla_id']
            
            # Usar la función helper actualizar_stock_ml
            success, message = actualizar_stock_ml(mla, stock_nuevo, access_token)
            
            if success:
                exitos += 1
            else:
                errores += 1
                mensajes_error.append(f"{mla}: {message}")
            time.sleep(2)
        
        if exitos > 0:
            flash(f'✅ Stock cargado en {exitos} publicaciones: {stock_nuevo} unidades', 'success')
        if errores > 0:
            flash(f'⚠️ {errores} publicaciones con errores', 'warning')
            if mensajes_error:
                for msg in mensajes_error[:3]:  # Mostrar solo los primeros 3 errores
                    flash(msg, 'warning')
    
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'danger')
    
    # Volver a mostrar resultados CON DATOS ACTUALIZADOS DE ML
    publicaciones = query_db(
        "SELECT mla_id, titulo_ml, activo FROM sku_mla_mapeo WHERE sku = %s AND activo = TRUE",
        (sku,)
    )
    
    pubs_lista = []
    access_token_refresh = cargar_ml_token()
    
    for row in publicaciones:
        if access_token_refresh:
            datos_ml = obtener_datos_ml(row['mla_id'], access_token_refresh)
            
            # Mapear status de ML a español
            status_ml = datos_ml.get('status', 'unknown')
            
            if status_ml == 'active':
                estado_texto = 'Activa'
            elif status_ml == 'paused':
                estado_texto = 'Pausada'
            elif status_ml == 'closed':
                estado_texto = 'Cerrada'
            elif status_ml == 'under_review':
                estado_texto = 'En revisión'
            elif status_ml == 'inactive':
                estado_texto = 'Inactiva'
            else:
                estado_texto = status_ml.capitalize()
            
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': datos_ml['titulo'],
                'stock_actual': datos_ml['stock'],
                'demora': datos_ml.get('demora'),
                'estado': estado_texto,
                'status_raw': status_ml
            })
        else:
            pubs_lista.append({
                'mla': row['mla_id'],
                'titulo': row['titulo_ml'] or 'Sin título',
                'stock_actual': '-',
                'demora': None,
                'estado': 'Activa' if row['activo'] else 'Pausada',
                'status_raw': 'active' if row['activo'] else 'paused'
            })
    
    return render_template('cargar_stock_ml.html',
                         sku_buscado=sku,
                         publicaciones=pubs_lista,
                         es_sku_con_z=sku.endswith('Z'))


# ============================================================================
# AUDITORÍA ML - SECCIONES INDEPENDIENTES
# Reemplazar el bloque completo de auditoría en app.py
# ============================================================================

# ============================================================================
# HELPER: CALCULAR STOCK DISPONIBLE (compartido por los 3 tipos de auditoría)
# ============================================================================

def calcular_stock_por_sku():
    """
    Calcula stock disponible para todos los SKUs (base + combos).
    Devuelve dict: { sku: { nombre, stock_fisico, stock_disponible } }
    """
    # 1. Obtener stock físico de productos base
    productos_base_query = query_db('''
        SELECT 
            sku, 
            nombre, 
            tipo, 
            stock_actual,
            COALESCE(stock_full, 0) as stock_full
        FROM productos_base 
        ORDER BY tipo, nombre
    ''')
    
    # 2. Ventas activas descomponiendo combos
    ventas_activas = query_db('''
        SELECT 
            COALESCE(pb_comp.sku, iv.sku) as sku,
            SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
        FROM items_venta iv
        JOIN ventas v ON iv.venta_id = v.id
        LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
        LEFT JOIN componentes c ON c.producto_compuesto_id = pc.id
        LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
        WHERE v.estado_entrega = 'pendiente'
        GROUP BY sku
    ''')
    
    ventas_dict = {v['sku']: int(v['vendido']) for v in ventas_activas}
    
    # 3. Calcular stock disponible por SKU base
    stock_por_sku = {}
    
    for prod in productos_base_query:
        sku = prod['sku']
        vendido = ventas_dict.get(sku, 0)
        
        if '_DEP' in sku or '_FULL' in sku:
            stock_fisico = int(prod['stock_actual'])
            stock_disponible = stock_fisico - vendido
        elif prod['tipo'] == 'almohada':
            stock_dep = int(prod['stock_actual'])
            stock_full = int(prod['stock_full'])
            stock_fisico = stock_dep + stock_full
            stock_disponible = stock_fisico - vendido
        else:
            stock_fisico = int(prod['stock_actual'])
            stock_disponible = stock_fisico - vendido
        
        stock_por_sku[sku] = {
            'nombre': prod['nombre'],
            'stock_fisico': stock_fisico,
            'stock_disponible': stock_disponible
        }
    
    # 4. Calcular stock disponible de combos
    try:
        productos_combos = query_db('''
            SELECT sku, nombre
            FROM productos_compuestos
            WHERE activo = 1
            ORDER BY nombre
        ''')
        
        if productos_combos:
            for combo in productos_combos:
                sku_combo = combo['sku']
                componentes = query_db('''
                    SELECT pb.sku, c.cantidad_necesaria 
                    FROM componentes c
                    JOIN productos_base pb ON c.producto_base_id = pb.id
                    JOIN productos_compuestos pc ON c.producto_compuesto_id = pc.id
                    WHERE pc.sku = %s
                ''', (sku_combo,))
                
                stock_disponible_combo = 999999
                for comp in componentes:
                    sku_comp = comp['sku']
                    cant_necesaria = int(comp['cantidad_necesaria'])
                    prod_comp = stock_por_sku.get(sku_comp)
                    if prod_comp:
                        combos_posibles = prod_comp['stock_disponible'] // cant_necesaria if cant_necesaria > 0 else 0
                        stock_disponible_combo = min(stock_disponible_combo, combos_posibles)
                    else:
                        stock_disponible_combo = 0
                        break
                
                if stock_disponible_combo == 999999 or stock_disponible_combo < 0:
                    stock_disponible_combo = 0
                
                stock_por_sku[sku_combo] = {
                    'nombre': combo['nombre'],
                    'stock_fisico': 0,
                    'stock_disponible': stock_disponible_combo
                }
    except Exception as e:
        print(f"Error calculando combos: {str(e)}")
    
    return stock_por_sku


# ============================================================================
# PÁGINA PRINCIPAL DE AUDITORÍA (sin barrido, solo estructura)
# ============================================================================

@app.route('/auditoria-ml', methods=['GET'])
@login_required
def auditoria_ml():
    """Renderiza la página de auditoría. Los datos se cargan vía AJAX por sección."""
    return render_template('auditoria_ml.html')


# ============================================================================
# ENDPOINT AJAX: EJECUTAR AUDITORÍA POR TIPO
# GET /auditoria-ml/run/<tipo>
# tipo: 'pausadas_sin_stock' | 'pausadas_con_stock' | 'demoras'
# ============================================================================

@app.route('/auditoria-ml/run/<tipo>', methods=['GET'])
@login_required
def auditoria_ml_run(tipo):
    """
    Ejecuta un tipo específico de auditoría y devuelve JSON con los resultados.
    """
    if tipo not in ['pausadas_sin_stock', 'pausadas_con_stock', 'demoras']:
        return jsonify({'error': 'Tipo de auditoría inválido'}), 400
    
    access_token = cargar_ml_token()
    if not access_token:
        return jsonify({'error': 'No hay token de ML configurado'}), 401
    
    try:
        # Calcular stock local
        stock_por_sku = calcular_stock_por_sku()
        
        # Obtener publicaciones relevantes de la BD
        if tipo == 'demoras':
            # Solo SKUs que terminan en Z
            publicaciones_db = query_db("""
                SELECT mla_id, sku, titulo_ml 
                FROM sku_mla_mapeo 
                WHERE activo = TRUE AND sku LIKE '%%Z'
                ORDER BY sku
            """)
        else:
            publicaciones_db = query_db("""
                SELECT mla_id, sku, titulo_ml 
                FROM sku_mla_mapeo 
                WHERE activo = TRUE
                ORDER BY sku
            """)
        
        resultados = []
        
        for pub in publicaciones_db:
            mla_id = pub['mla_id']
            sku = pub['sku']
            
            # Obtener stock local del SKU (con fallback sin Z)
            stock_info = stock_por_sku.get(sku)
            if not stock_info and sku.endswith('Z'):
                stock_info = stock_por_sku.get(sku[:-1])
            
            if not stock_info:
                continue
            
            stock_disponible = stock_info['stock_disponible']
            
            # Solo consultar ML si hay stock local relevante
            if tipo in ['pausadas_sin_stock', 'pausadas_con_stock'] and stock_disponible <= 0:
                continue
            if tipo == 'demoras' and stock_disponible <= 0:
                continue
            
            # Consultar datos de ML
            datos_ml = obtener_datos_ml(mla_id, access_token)
            status_ml = datos_ml.get('status', 'unknown')
            stock_ml = datos_ml.get('stock', 0)
            demora_ml = datos_ml.get('demora')
            
            item_base = {
                'mla': mla_id,
                'sku': sku,
                'titulo': datos_ml.get('titulo', pub.get('titulo_ml', '')),
                'stock_disponible': stock_disponible,
                'stock_ml': stock_ml,
                'status': status_ml
            }
            
            if tipo == 'pausadas_sin_stock':
                if status_ml == 'paused' and stock_ml == 0 and stock_disponible > 0:
                    resultados.append(item_base)
            
            elif tipo == 'pausadas_con_stock':
                if status_ml == 'paused' and stock_ml > 0 and stock_disponible > 0:
                    resultados.append(item_base)
            
            elif tipo == 'demoras':
                if demora_ml and demora_ml != 'Sin especificar':
                    try:
                        import re
                        numeros = re.findall(r'\d+', str(demora_ml))
                        if numeros and int(numeros[0]) > 0:
                            item_base['demora'] = demora_ml
                            resultados.append(item_base)
                    except Exception as e:
                        print(f"Error parseando demora '{demora_ml}': {e}")
        
        print(f"✅ Auditoría '{tipo}': {len(resultados)} resultados")
        return jsonify({'tipo': tipo, 'resultados': resultados, 'total': len(resultados)})
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# ACCIONES - Devuelven JSON para que el frontend refresque solo la sección
# ============================================================================

@app.route('/auditoria-ml/activar', methods=['POST'])
@login_required
def auditoria_activar_publicaciones():
    """Activar (despausar) publicaciones seleccionadas. Devuelve JSON."""
    
    mlas_seleccionadas = request.json.get('mlas', []) if request.is_json else request.form.getlist('mlas[]')
    
    if not mlas_seleccionadas:
        return jsonify({'error': 'No se seleccionaron publicaciones'}), 400
    
    access_token = cargar_ml_token()
    if not access_token:
        return jsonify({'error': 'No hay token de ML configurado'}), 401
    
    exitos = 0
    errores = []
    
    for mla in mlas_seleccionadas:
        try:
            url = f'https://api.mercadolibre.com/items/{mla}'
            headers = {'Authorization': f'Bearer {access_token}'}
            data = {'status': 'active'}
            
            response = ml_request('put', url, access_token, json_data=data)
            
            if response.status_code == 200:
                exitos += 1
            else:
                errores.append(f'{mla}: {response.status_code}')
            time.sleep(2)
        except Exception as e:
            errores.append(f'{mla}: {str(e)}')
    
    return jsonify({'exitos': exitos, 'errores': errores, 'total': len(mlas_seleccionadas)})


@app.route('/auditoria-ml/cargar-stock', methods=['POST'])
@login_required
def auditoria_cargar_stock():
    """Cargar stock en publicaciones seleccionadas. Devuelve JSON."""
    
    # Acepta JSON: { mla_stock: ["MLA123:5", "MLA456:3"] }
    # O form: mla_stock[] = "MLA123:5"
    if request.is_json:
        mlas_data = request.json.get('mla_stock', [])
    else:
        mlas_data = request.form.getlist('mla_stock')
    
    if not mlas_data:
        return jsonify({'error': 'No se seleccionaron publicaciones'}), 400
    
    access_token = cargar_ml_token()
    if not access_token:
        return jsonify({'error': 'No hay token de ML configurado'}), 401
    
    exitos = 0
    errores = []
    
    for item in mlas_data:
        try:
            mla, stock_str = item.split(':')
            stock = int(stock_str)
            success, message = actualizar_stock_ml(mla, stock, access_token)
            if success:
                exitos += 1
            else:
                errores.append(f'{mla}: {message}')
            time.sleep(2)
        except Exception as e:
            errores.append(f'{item}: {str(e)}')
    
    return jsonify({'exitos': exitos, 'errores': errores, 'total': len(mlas_data)})


@app.route('/auditoria-ml/reducir-demora', methods=['POST'])
@login_required
def auditoria_reducir_demora():
    """Quitar demora completamente en publicaciones seleccionadas. Devuelve JSON."""
    
    mlas_seleccionadas = request.json.get('mlas', []) if request.is_json else request.form.getlist('mlas_demora[]')
    
    if not mlas_seleccionadas:
        return jsonify({'error': 'No se seleccionaron publicaciones'}), 400
    
    access_token = cargar_ml_token()
    if not access_token:
        return jsonify({'error': 'No hay token de ML configurado'}), 401
    
    exitos = 0
    errores = []
    
    for mla in mlas_seleccionadas:
        try:
            success, message = quitar_manufacturing_time_ml(mla, access_token)
            if success:
                exitos += 1
            else:
                errores.append(f'{mla}: {message}')
            time.sleep(2)
        except Exception as e:
            errores.append(f'{mla}: {str(e)}')
    
    return jsonify({'exitos': exitos, 'errores': errores, 'total': len(mlas_seleccionadas)})
# ============================================================================
# ESTADÍSTICAS DE VENTAS
# ============================================================================

@app.route('/estadisticas')
@login_required
def estadisticas():
    from datetime import datetime, timedelta

    # ========================================
    # FILTROS
    # ========================================
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    filtro_canal = request.args.get('canal', '')
    filtro_metodo = request.args.get('metodo_envio', '')
    filtro_zona = request.args.get('zona', '')

    # Default: último mes
    if not fecha_desde:
        fecha_desde = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not fecha_hasta:
        fecha_hasta = datetime.now().strftime('%Y-%m-%d')

    # ========================================
    # BASE DE FILTROS (reutilizable)
    # ========================================
    base_where = """
        WHERE v.estado_entrega != 'cancelada'
        AND DATE(v.fecha_venta) BETWEEN %s AND %s
    """
    base_params = [fecha_desde, fecha_hasta]

    if filtro_canal == 'ML':
        base_where += " AND v.canal = 'Mercado Libre'"
    elif filtro_canal == 'no_ml':
        base_where += " AND v.canal != 'Mercado Libre'"

    if filtro_metodo:
        base_where += " AND v.metodo_envio = %s"
        base_params.append(filtro_metodo)

    if filtro_zona:
        base_where += " AND v.zona_envio = %s"
        base_params.append(filtro_zona)

    # ========================================
    # 1. MÉTRICAS RESUMEN
    # ========================================
    resumen = query_one(f"""
        SELECT
            COUNT(*) as total_ventas,
            COALESCE(SUM(v.importe_total), 0) as total_facturado,
            COALESCE(AVG(v.importe_total), 0) as ticket_promedio,
            COALESCE(SUM(iv.cantidad), 0) as total_unidades
        FROM ventas v
        LEFT JOIN items_venta iv ON iv.venta_id = v.id
        {base_where}
    """, tuple(base_params))

    # ========================================
    # 2. VENTAS POR DÍA (para gráfico de línea)
    # ========================================
    ventas_por_dia = query_db(f"""
        SELECT
            DATE(v.fecha_venta) as dia,
            COUNT(*) as cantidad,
            SUM(v.importe_total) as total
        FROM ventas v
        {base_where}
        GROUP BY DATE(v.fecha_venta)
        ORDER BY dia
    """, tuple(base_params))

    # ========================================
    # 3. DESGLOSE POR CANAL (para torta)
    # ========================================
    por_canal = query_db(f"""
        SELECT
            CASE WHEN v.canal = 'Mercado Libre' THEN 'MercadoLibre' ELSE 'Venta Directa' END as canal_label,
            COUNT(*) as cantidad,
            SUM(v.importe_total) as total
        FROM ventas v
        {base_where}
        GROUP BY canal_label
    """, tuple(base_params))

    # ========================================
    # 4. DESGLOSE POR MÉTODO DE ENVÍO (para torta)
    # ========================================
    por_metodo = query_db(f"""
        SELECT
            COALESCE(v.metodo_envio, 'Sin especificar') as metodo,
            COUNT(*) as cantidad,
            SUM(v.importe_total) as total
        FROM ventas v
        {base_where}
        GROUP BY v.metodo_envio
        ORDER BY cantidad DESC
    """, tuple(base_params))

    # ========================================
    # 5. TOP PRODUCTOS MÁS VENDIDOS
    # ========================================
    top_productos = query_db(f"""
        SELECT
            iv.sku,
            COALESCE(pb.nombre, pc.nombre, iv.sku) as nombre,
            SUM(iv.cantidad) as total_unidades,
            SUM(iv.cantidad * iv.precio_unitario) as total_facturado
        FROM ventas v
        JOIN items_venta iv ON iv.venta_id = v.id
        LEFT JOIN productos_base pb ON iv.sku = pb.sku
        LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
        {base_where}
        GROUP BY iv.sku, nombre
        ORDER BY total_unidades DESC
        LIMIT 15
    """, tuple(base_params))

    # ========================================
    # 6. FILTROS DISPONIBLES (para los selects)
    # ========================================
    metodos_disponibles = query_db("""
        SELECT DISTINCT metodo_envio FROM ventas
        WHERE metodo_envio IS NOT NULL AND metodo_envio != ''
        ORDER BY metodo_envio
    """)

    zonas_disponibles = query_db("""
        SELECT DISTINCT zona_envio FROM ventas
        WHERE zona_envio IS NOT NULL AND zona_envio != ''
        ORDER BY zona_envio
    """)

    return render_template('estadisticas.html',
        resumen=resumen,
        ventas_por_dia=ventas_por_dia,
        por_canal=por_canal,
        por_metodo=por_metodo,
        top_productos=top_productos,
        metodos_disponibles=metodos_disponibles,
        zonas_disponibles=zonas_disponibles,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        filtro_canal=filtro_canal,
        filtro_metodo=filtro_metodo,
        filtro_zona=filtro_zona,
    )


# ============================================================================
# EXPORTAR REPOSICIÓN A EXCEL
# ============================================================================

@app.route('/estadisticas/exportar-reposicion')
@login_required
def exportar_reposicion():
    from flask import make_response
    from datetime import datetime, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO

    fecha_desde = request.args.get('fecha_desde', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    fecha_hasta = request.args.get('fecha_hasta', datetime.now().strftime('%Y-%m-%d'))
    filtro_canal = request.args.get('canal', '')
    filtro_metodo = request.args.get('metodo_envio', '')
    filtro_zona = request.args.get('zona', '')

    base_where = """
        WHERE v.estado_entrega != 'cancelada'
        AND DATE(v.fecha_venta) BETWEEN %s AND %s
    """
    base_params = [fecha_desde, fecha_hasta]

    if filtro_canal == 'ML':
        base_where += " AND v.canal = 'Mercado Libre'"
    elif filtro_canal == 'no_ml':
        base_where += " AND v.canal != 'Mercado Libre'"
    if filtro_metodo:
        base_where += " AND v.metodo_envio = %s"
        base_params.append(filtro_metodo)
    if filtro_zona:
        base_where += " AND v.zona_envio = %s"
        base_params.append(filtro_zona)

    # Stock disponible = stock_actual - pendiente de entrega (descomponiendo combos igual que nueva_venta)
    vendido_rows = query_db("""
        SELECT
            COALESCE(pb_comp.sku, iv.sku) as sku,
            SUM(iv.cantidad * COALESCE(c.cantidad_necesaria, 1)) as vendido
        FROM items_venta iv
        JOIN ventas v ON iv.venta_id = v.id
        LEFT JOIN productos_compuestos pc ON iv.sku = pc.sku
        LEFT JOIN componentes c ON pc.id = c.producto_compuesto_id
        LEFT JOIN productos_base pb_comp ON c.producto_base_id = pb_comp.id
        WHERE v.estado_entrega NOT IN ('entregada', 'cancelada')
        GROUP BY COALESCE(pb_comp.sku, iv.sku)
    """)
    vendido_map = {r['sku']: int(r['vendido']) for r in vendido_rows}

    stock_fisico_rows = query_db("SELECT sku, stock_actual FROM productos_base")
    stock_disponible_map = {
        r['sku']: int(r['stock_actual']) - vendido_map.get(r['sku'], 0)
        for r in stock_fisico_rows
    }

    # Productos base vendidos directamente (no son combos)
    base_directos = query_db(f"""
        SELECT
            pb.sku,
            pb.nombre,
            SUM(iv.cantidad) as cantidad_vendida
        FROM ventas v
        JOIN items_venta iv ON iv.venta_id = v.id
        JOIN productos_base pb ON iv.sku = pb.sku
        {base_where}
        GROUP BY pb.sku, pb.nombre
    """, tuple(base_params))

    # Combos → explotar en componentes base
    combos = query_db(f"""
        SELECT
            pb.sku,
            pb.nombre,
            SUM(iv.cantidad * c.cantidad_necesaria) as cantidad_vendida
        FROM ventas v
        JOIN items_venta iv ON iv.venta_id = v.id
        JOIN productos_compuestos pc ON iv.sku = pc.sku
        JOIN componentes c ON c.producto_compuesto_id = pc.id
        JOIN productos_base pb ON pb.id = c.producto_base_id
        {base_where}
        GROUP BY pb.sku, pb.nombre
    """, tuple(base_params))

    # Consolidar sumando directos + combos
    totales = {}
    for row in list(base_directos) + list(combos):
        sku = row['sku']
        if sku in totales:
            totales[sku]['cantidad_vendida'] += int(row['cantidad_vendida'])
        else:
            totales[sku] = {
                'sku': sku,
                'nombre': row['nombre'],
                'cantidad_vendida': int(row['cantidad_vendida']),
                'stock_disponible': stock_disponible_map.get(sku, 0)
            }

    productos = sorted(totales.values(), key=lambda x: x['cantidad_vendida'], reverse=True)

    # CREAR EXCEL
    wb = Workbook()
    ws = wb.active
    ws.title = "Reposición"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Título
    ws.merge_cells('A1:D1')
    titulo = ws['A1']
    titulo.value = f"Reposición de Stock — {fecha_desde} al {fecha_hasta}"
    titulo.font = Font(bold=True, color='FFFFFF', name='Arial', size=12)
    titulo.fill = header_fill
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Headers
    headers = ['SKU', 'Descripción', 'Unidades Vendidas', 'Stock Disponible']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Datos
    for i, p in enumerate(productos):
        row = i + 3
        fill = alt_fill if i % 2 == 0 else PatternFill(fill_type=None)
        valores = [p['sku'], p['nombre'], int(p['cantidad_vendida']), int(p['stock_disponible'])]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name='Arial', size=10)
            cell.fill = fill
            cell.border = border
            if col in (3, 4):
                cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.freeze_panes = 'A3'

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = make_response(excel_file.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    fecha_str = datetime.now().strftime('%Y%m%d')
    response.headers['Content-Disposition'] = f'attachment; filename=reposicion_{fecha_str}.xlsx'
    return response

# ============================================================================
# TEST: GESTIÓN DE MANUFACTURING_TIME EN PUBLICACIONES ML
# Ruta segura y aislada - no toca nada del sistema existente
# ============================================================================

@app.route('/test/manufacturing-time', methods=['GET'])
@login_required
def test_manufacturing_time():
    """
    Página de prueba para ver y modificar el MANUFACTURING_TIME
    de publicaciones ML con sufijo Z
    """
    return render_template('test_manufacturing_time.html')


@app.route('/test/manufacturing-time/ver', methods=['POST'])
@login_required
def ver_manufacturing_time():
    """
    Consulta el estado actual de una publicación:
    - Muestra el MANUFACTURING_TIME actual
    - Muestra available_quantity
    - Solo lectura, no modifica nada
    """
    import requests as req

    item_id = request.form.get('item_id', '').strip().upper()
    if not item_id:
        return {'error': 'Falta item_id'}, 400

    access_token = cargar_ml_token()
    if not access_token:
        return {'error': 'No hay token ML activo'}, 400

    headers = {'Authorization': f'Bearer {access_token}'}

    try:
        r = req.get(f'https://api.mercadolibre.com/items/{item_id}', headers=headers)
        if r.status_code != 200:
            return {'error': f'ML devolvió {r.status_code}: {r.text}'}, 400

        data = r.json()

        # Extraer MANUFACTURING_TIME de sale_terms
        manufacturing_time = None
        for term in data.get('sale_terms', []):
            if term.get('id') == 'MANUFACTURING_TIME':
                manufacturing_time = term.get('value_name')
                break

        return {
            'ok': True,
            'item_id': item_id,
            'title': data.get('title'),
            'status': data.get('status'),
            'available_quantity': data.get('available_quantity'),
            'manufacturing_time': manufacturing_time,  # None = sin demora
            'permalink': data.get('permalink'),
        }

    except Exception as e:
        return {'error': str(e)}, 500


@app.route('/test/manufacturing-time/quitar', methods=['POST'])
@login_required
def quitar_manufacturing_time():
    """
    Elimina el MANUFACTURING_TIME de una publicación enviando null.
    Según la API de ML:
    PUT /items/{item_id}
    { "sale_terms": [{ "id": "MANUFACTURING_TIME", "value_id": null, "value_name": null }] }
    """
    import requests as req

    item_id = request.form.get('item_id', '').strip().upper()
    if not item_id:
        return {'error': 'Falta item_id'}, 400

    access_token = cargar_ml_token()
    if not access_token:
        return {'error': 'No hay token ML activo'}, 400

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    payload = {
        "sale_terms": [
            {
                "id": "MANUFACTURING_TIME",
                "value_id": None,
                "value_name": None
            }
        ]
    }

    try:
        r = req.put(
            f'https://api.mercadolibre.com/items/{item_id}',
            headers=headers,
            json=payload
        )

        if r.status_code == 200:
            data = r.json()
            # Verificar que quedó en None
            manufacturing_time_nuevo = None
            for term in data.get('sale_terms', []):
                if term.get('id') == 'MANUFACTURING_TIME':
                    manufacturing_time_nuevo = term.get('value_name')
                    break

            return {
                'ok': True,
                'item_id': item_id,
                'title': data.get('title'),
                'manufacturing_time_anterior': request.form.get('mt_anterior'),
                'manufacturing_time_nuevo': manufacturing_time_nuevo,
                'mensaje': '✅ Demora eliminada correctamente' if not manufacturing_time_nuevo else '⚠️ No se pudo eliminar'
            }
        else:
            return {
                'ok': False,
                'error': f'ML devolvió {r.status_code}',
                'detalle': r.json()
            }, 400

    except Exception as e:
        return {'error': str(e)}, 500


@app.route('/test/manufacturing-time/poner', methods=['POST'])
@login_required
def poner_manufacturing_time():
    """
    Pone o restaura el MANUFACTURING_TIME a un valor específico.
    Útil para revertir si algo sale mal.
    """
    import requests as req

    item_id = request.form.get('item_id', '').strip().upper()
    dias = request.form.get('dias', '').strip()

    if not item_id or not dias:
        return {'error': 'Faltan datos'}, 400

    access_token = cargar_ml_token()
    if not access_token:
        return {'error': 'No hay token ML activo'}, 400

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    payload = {
        "sale_terms": [
            {
                "id": "MANUFACTURING_TIME",
                "value_name": f"{dias} días"
            }
        ]
    }

    try:
        r = req.put(
            f'https://api.mercadolibre.com/items/{item_id}',
            headers=headers,
            json=payload
        )

        if r.status_code == 200:
            data = r.json()
            manufacturing_time_nuevo = None
            for term in data.get('sale_terms', []):
                if term.get('id') == 'MANUFACTURING_TIME':
                    manufacturing_time_nuevo = term.get('value_name')
                    break

            return {
                'ok': True,
                'item_id': item_id,
                'manufacturing_time_nuevo': manufacturing_time_nuevo,
                'mensaje': f'✅ Demora restaurada a {dias} días'
            }
        else:
            return {
                'ok': False,
                'error': f'ML devolvió {r.status_code}',
                'detalle': r.json()
            }, 400

    except Exception as e:
        return {'error': str(e)}, 500


@app.route('/debug-mla')
@login_required
def debug_mla():
    import requests
    access_token = cargar_ml_token()
    mla_id = 'MLA1538119895'
    headers = {'Authorization': f'Bearer {access_token}'}
    r = requests.get(f'https://api.mercadolibre.com/items/{mla_id}', headers=headers)
    data = r.json()
    
    listing_type_id = data.get('listing_type_id')
    campaign = None
    for term in data.get('sale_terms', []):
        if term.get('id') == 'INSTALLMENTS_CAMPAIGN':
            campaign = term
            break
    
    return {
        'listing_type_id': listing_type_id,
        'installments_campaign': campaign
    }


@app.route('/ml/callback')
@login_required
def ml_callback():
    """Recibe el code de ML y lo canjea automáticamente por el token"""
    code = request.args.get('code', '').strip()
    if not code:
        flash('❌ No se recibió el code de MercadoLibre', 'error')
        return redirect(url_for('ml_configurar_token'))
    
    CLIENT_ID = os.getenv('ML_APP_ID')
    CLIENT_SECRET = os.getenv('ML_SECRET_KEY')
    REDIRECT_URI = os.getenv('ML_REDIRECT_URI')
    
    try:
        response = requests.post(
            "https://api.mercadolibre.com/oauth/token",
            data={
                "grant_type": "authorization_code",
                "client_id": CLIENT_ID,
                "client_secret": CLIENT_SECRET,
                "code": code,
                "redirect_uri": REDIRECT_URI
            }
        )
        if response.status_code == 200:
            data = response.json()
            token_data = {
                "access_token": data.get("access_token"),
                "refresh_token": data.get("refresh_token"),
                "expires_at": time.time() + data.get("expires_in", 21600) - 300,
                "client_id": CLIENT_ID,
                "client_secret": CLIENT_SECRET
            }
            if guardar_ml_token(token_data):
                flash('✅ Token configurado con auto-renovación activada', 'success')
                return redirect(url_for('ventas_activas'))
            else:
                flash('❌ Error al guardar el token', 'error')
        else:
            error_msg = response.json().get('message', 'Error desconocido')
            flash(f'❌ Error al canjear el code: {error_msg}', 'error')
    except Exception as e:
        flash(f'❌ Error: {str(e)}', 'error')
    
    return redirect(url_for('ml_configurar_token'))


# ============================================================================
# INICIAR APLICACIÓN
# ============================================================================

if __name__ == '__main__':
    # Configuración para red local
    app.run(
        host='0.0.0.0',      # Escuchar en todas las interfaces de red
        port=5000,            # Puerto del servidor
        debug=False,          # Desactivar debug en producción
        threaded=True         # Permitir múltiples usuarios simultáneos
    )
